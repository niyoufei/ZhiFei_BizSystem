from __future__ import annotations

import copy
import csv
import hashlib
import html as html_lib
import io
import json
import logging
import os
import re
import shutil
import subprocess
import tempfile
import threading
import unicodedata
from collections import Counter
from datetime import datetime, timezone

# 加载 .env，使 SPARK_APIPASSWORD、OPENAI_API_KEY 等生效
try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass

from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import quote_plus
from uuid import uuid4

try:
    import pymupdf
except Exception:
    pymupdf = None
try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None
try:
    from docx import Document
except Exception:
    Document = None
try:
    from PIL import Image
except Exception:
    Image = None
try:
    import pytesseract
except Exception:
    pytesseract = None
from fastapi import (
    APIRouter,
    Depends,
    FastAPI,
    File,
    Form,
    Header,
    HTTPException,
    Query,
    Request,
    UploadFile,
)
from fastapi.exception_handlers import http_exception_handler
from fastapi.responses import RedirectResponse, Response
from starlette.exceptions import HTTPException as StarletteHTTPException

from app.auth import get_auth_status, verify_api_key
from app.cache import (
    cache_score_result,
    clear_score_cache,
    get_cache_stats,
    get_cached_score,
)
from app.config import get_config_status, load_config, reload_config
from app.engine.adaptive import (
    apply_adaptive_patch,
    apply_rubric_patch,
    build_adaptive_patch,
    build_adaptive_suggestions,
)
from app.engine.anchors import (
    build_project_requirements_from_anchors,
    extract_project_anchors_from_text,
)
from app.engine.calibrator import (
    build_feature_row,
    calc_metrics,
    cross_validate_calibrator,
    predict_with_model,
    train_best_calibrator_auto,
    train_isotonic1d_calibrator,
    train_linear1d_calibrator,
    train_offset_calibrator,
    train_ridge_calibrator,
)
from app.engine.compare import build_compare_narrative
from app.engine.dimensions import DIMENSIONS
from app.engine.evaluation import evaluate_project_variants
from app.engine.evolution import build_evolution_report
from app.engine.feature_distillation import (
    select_top_logic_skeletons,
    update_feature_confidence,
)
from app.engine.history import (
    analyze_trend,
    get_history,
)
from app.engine.history import (
    record_score as record_history_score,
)
from app.engine.insights import build_project_insights
from app.engine.learning import build_learning_profile
from app.engine.llm_evolution import (
    enhance_evolution_report_with_llm,
    get_llm_backend_status,
    preview_evolution_report_with_ollama,
)
from app.engine.reflection import (
    build_calibration_samples,
    build_delta_cases,
    evaluate_patch_shadow,
    mine_patch_package,
)
from app.engine.scorer import score_text
from app.engine.surrogate_learning import calibrate_weights, compute_time_decay_weight
from app.engine.v2_scorer import compute_v2_rule_total, score_text_v2
from app.i18n import DEFAULT_LOCALE, SUPPORTED_LOCALES, t
from app.metrics import get_metrics, record_score, update_project_stats
from app.rate_limit import get_rate_limit_status, setup_rate_limiting
from app.schemas import (
    RESPONSES_401,
    RESPONSES_404,
    RESPONSES_409,
    RESPONSES_422,
    RESPONSES_NO_PROFILE,
    RESPONSES_NO_SUBMISSIONS,
    AdaptiveApplyResult,
    AdaptivePatch,
    AdaptiveSuggestions,
    AdaptiveValidation,
    AnalysisBundleResponse,
    CacheClearResponse,
    CacheStatsResponse,
    CalibrationSampleRecord,
    CalibratorDeployRequest,
    CalibratorModelRecord,
    CalibratorPredictResponse,
    CalibratorTrainRequest,
    CompareNarrative,
    CompareReport,
    CompilationInstructions,
    ConfigReloadResponse,
    ConfigStatusResponse,
    ConstraintPack,
    DataHygieneResponse,
    DeltaCaseRecord,
    EvaluationSummaryResponse,
    EvidenceTraceMarkdownResponse,
    EvidenceTraceResponse,
    EvolutionHealthResponse,
    EvolutionReport,
    ExpertProfileRecord,
    ExpertProfileUpdate,
    GroundTruthBatchResponse,
    GroundTruthCreate,
    GroundTruthFromSubmissionCreate,
    GroundTruthRecord,
    HealthResponse,
    InsightsReport,
    LatestReportResponse,
    LearningProfile,
    LLMBackendStatus,
    MaterialDepthReportMarkdownResponse,
    MaterialDepthReportResponse,
    MaterialKnowledgeProfileMarkdownResponse,
    MaterialKnowledgeProfileResponse,
    MaterialRecord,
    OllamaEvolutionPreviewResponse,
    PatchDeploymentRecord,
    PatchDeployRequest,
    PatchMineRequest,
    PatchPackageRecord,
    PatchShadowEvalResponse,
    ProjectAnchorRecord,
    ProjectContextIn,
    ProjectContextOut,
    ProjectCreate,
    ProjectEvaluationResponse,
    ProjectExpertProfileResponse,
    ProjectMeceAuditResponse,
    ProjectPreScoreListResponse,
    ProjectRecord,
    ProjectRequirementRecord,
    ProjectScoreHistory,
    QingTianResultCreate,
    QingTianResultRecord,
    ReadyResponse,
    ReflectionAutoRunResponse,
    RescoreRequest,
    RescoreResponse,
    ScoreReport,
    ScoreRequest,
    ScoringBasisResponse,
    ScoringFactorsMarkdownResponse,
    ScoringFactorsResponse,
    ScoringReadinessResponse,
    SelfCheckResponse,
    SubmissionRecord,
    TrendAnalysis,
    WritingGuidance,
)
from app.storage import (
    MATERIALS_DIR,
    ensure_data_dirs,
    load_calibration_models,
    load_calibration_samples,
    load_delta_cases,
    load_evidence_units,
    load_evolution_reports,
    load_expert_profiles,
    load_ground_truth,
    load_learning_profiles,
    load_materials,
    load_patch_deployments,
    load_patch_packages,
    load_project_anchors,
    load_project_context,
    load_project_requirements,
    load_projects,
    load_qingtian_results,
    load_score_history,
    load_score_reports,
    load_submissions,
    save_calibration_models,
    save_calibration_samples,
    save_delta_cases,
    save_evidence_units,
    save_evolution_reports,
    save_expert_profiles,
    save_ground_truth,
    save_learning_profiles,
    save_materials,
    save_patch_deployments,
    save_patch_packages,
    save_project_anchors,
    save_project_context,
    save_project_requirements,
    save_projects,
    save_qingtian_results,
    save_score_history,
    save_score_reports,
    save_submissions,
)

logger = logging.getLogger(__name__)

# OpenAPI 标签定义
OPENAPI_TAGS = [
    {
        "name": "健康检查",
        "description": "容器化部署健康检查端点（Kubernetes liveness/readiness probes）",
    },
    {
        "name": "监控指标",
        "description": "Prometheus 格式的运行时指标导出",
    },
    {
        "name": "系统状态",
        "description": "系统认证与限流状态查询",
    },
    {
        "name": "评分",
        "description": "施工组织设计文档评分核心功能",
    },
    {
        "name": "项目管理",
        "description": "项目创建、列表、材料上传等项目全生命周期管理",
    },
    {
        "name": "施组提交",
        "description": "施工组织设计文档上传与评分记录",
    },
    {
        "name": "对比分析",
        "description": "多次提交的横向对比与统计分析",
    },
    {
        "name": "自适应优化",
        "description": "基于历史数据的评分规则自适应调整",
    },
    {
        "name": "洞察与学习",
        "description": "项目洞察报告与学习画像生成",
    },
    {
        "name": "历史与趋势",
        "description": "评分历史记录查询与趋势分析",
    },
]

DIMENSION_IDS = sorted(DIMENSIONS.keys())
DEFAULT_REGION = "合肥"
DEFAULT_QINGTIAN_MODEL_VERSION = "qingtian-2026.02"
DEFAULT_SCORING_ENGINE_LOCKED = "v2.0.0"
DEFAULT_CALIBRATOR_LOCKED = None
DEFAULT_RULE_SCORE_WEIGHT = 0.7
DEFAULT_LLM_SCORE_WEIGHT = 0.3
DEFAULT_LLM_DELTA_CAP = 35.0
DEFAULT_SCORE_SCALE_MAX = 100
DEFAULT_ENFORCE_GB_REDLINE = False
DEFAULT_NORM_RULE_VERSION = "v1_m=0.5+a/10_norm=sum"
DEFAULT_ENFORCE_MATERIAL_GATE = True
DEFAULT_REQUIRED_MATERIAL_TYPES = ["tender_qa", "boq", "drawing"]
DEFAULT_MIN_PARSED_CHARS_BY_TYPE = {
    "tender_qa": 12000,
    "boq": 2000,
    "drawing": 1500,
    "site_photo": 200,
}
DEFAULT_ENFORCE_MATERIAL_DEPTH_GATE = False
DEFAULT_MIN_PARSED_CHUNKS_BY_TYPE = {
    "tender_qa": 12,
    "boq": 3,
    "drawing": 3,
    "site_photo": 1,
}
DEFAULT_MIN_TOTAL_PARSED_CHUNKS = 20
DEFAULT_MIN_NUMERIC_TERMS_BY_TYPE = {
    "tender_qa": 6,
    "boq": 8,
    "drawing": 4,
    "site_photo": 0,
}
DEFAULT_MIN_TOTAL_PARSED_CHARS = 18000
DEFAULT_MAX_MATERIAL_PARSE_FAIL_RATIO = 0.6
DEFAULT_BLOCK_ON_ANY_MATERIAL_PARSE_FAILURE = True
DEFAULT_ENFORCE_MATERIAL_UTILIZATION_GATE = True
DEFAULT_MATERIAL_UTILIZATION_GATE_MODE = "block"
DEFAULT_MIN_MATERIAL_RETRIEVAL_HIT_RATE = 0.25
DEFAULT_MIN_MATERIAL_CONSISTENCY_HIT_RATE = 0.25
DEFAULT_MIN_MATERIAL_RETRIEVAL_TOTAL = 0
DEFAULT_MAX_UNCOVERED_REQUIRED_TYPES = 0
DEFAULT_MIN_REQUIRED_TYPE_PRESENCE_RATE = 0.0
DEFAULT_MIN_REQUIRED_TYPE_COVERAGE_RATE = 0.67
DEFAULT_MATERIAL_RETRIEVAL_TOP_K = 18
DEFAULT_MATERIAL_RETRIEVAL_PER_TYPE_QUOTA = 2
DEFAULT_MATERIAL_RETRIEVAL_PER_FILE_QUOTA = 3
DEFAULT_MIN_MATERIAL_RETRIEVAL_FILE_COVERAGE_RATE = 0.0
DEFAULT_ENFORCE_UPLOADED_TYPE_COVERAGE = True
DEFAULT_MIN_UPLOADED_TYPE_COVERAGE_RATE = 1.0
DEFAULT_PDF_TEXT_MIN_CHARS_FOR_OCR = 200
DEFAULT_PDF_OCR_MAX_PAGES = 30
DEFAULT_MATERIAL_INDEX_CACHE_SIZE = 12
DEFAULT_CALIBRATION_MIN_SAMPLES = 20
DEFAULT_CALIBRATION_FRESHNESS_DAYS = 90
DEFAULT_CALIBRATION_MIN_RECENT_RATIO = 0.6
PROFILE_LOCKED_STATUSES = {"submitted_to_qingtian"}
DEFAULT_CHAPTER_REQUIREMENTS = {
    "required_sections": [
        "重难点及危大工程（对应维度07）",
        "进度保障措施（对应维度09）",
        "安全生产与文明施工（对应维度02/03）",
        "质量保障体系（对应维度08）",
    ],
    "required_charts_images": [
        "进度计划或横道图",
        "危大工程或重难点分析示意图（如适用）",
        "组织架构或责任分工表",
    ],
    "mandatory_elements": [
        "控制参数或阈值",
        "执行频次（如日报/周检）",
        "责任岗位或责任人",
        "验收或检查动作（报验/旁站/签认等）",
    ],
    "forbidden_patterns": [
        "仅使用「保证」「严格落实」「确保」等空泛承诺而无量化或动作",
        "措施类描述缺少参数/频次/责任/验收中至少两类",
    ],
}

_MATERIAL_INDEX_CACHE: Dict[str, Dict[str, object]] = {}
_MATERIAL_INDEX_CACHE_ORDER: List[str] = []
_MATERIAL_INDEX_CACHE_LOCK = threading.RLock()


def _material_row_cache_token(row: Dict[str, object]) -> Dict[str, object]:
    path_text = str(row.get("path") or "").strip()
    stat_size = -1
    stat_mtime_ns = -1
    if path_text:
        p = Path(path_text)
        if p.exists():
            try:
                st = p.stat()
                stat_size = int(st.st_size)
                stat_mtime_ns = int(getattr(st, "st_mtime_ns", int(st.st_mtime * 1_000_000_000)))
            except Exception:
                stat_size = -1
                stat_mtime_ns = -1
    return {
        "id": str(row.get("id") or ""),
        "material_type": _normalize_material_type(
            row.get("material_type"), filename=row.get("filename")
        ),
        "filename": str(row.get("filename") or ""),
        "path": path_text,
        "created_at": str(row.get("created_at") or ""),
        "updated_at": str(row.get("updated_at") or ""),
        "size": stat_size,
        "mtime_ns": stat_mtime_ns,
    }


def _compute_material_index_signature(project_id: str, rows: List[Dict[str, object]]) -> str:
    payload = {
        "project_id": project_id,
        "rows": [_material_row_cache_token(row) for row in rows],
    }
    digest = hashlib.sha1(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8"))
    return digest.hexdigest()


def _touch_material_index_cache(project_id: str) -> None:
    if project_id in _MATERIAL_INDEX_CACHE_ORDER:
        _MATERIAL_INDEX_CACHE_ORDER.remove(project_id)
    _MATERIAL_INDEX_CACHE_ORDER.append(project_id)
    while len(_MATERIAL_INDEX_CACHE_ORDER) > int(DEFAULT_MATERIAL_INDEX_CACHE_SIZE):
        stale = _MATERIAL_INDEX_CACHE_ORDER.pop(0)
        _MATERIAL_INDEX_CACHE.pop(stale, None)


def _invalidate_material_index_cache(project_id: Optional[str] = None) -> None:
    with _MATERIAL_INDEX_CACHE_LOCK:
        if project_id is None:
            _MATERIAL_INDEX_CACHE.clear()
            _MATERIAL_INDEX_CACHE_ORDER.clear()
            return
        _MATERIAL_INDEX_CACHE.pop(project_id, None)
        if project_id in _MATERIAL_INDEX_CACHE_ORDER:
            _MATERIAL_INDEX_CACHE_ORDER.remove(project_id)


def _build_project_material_index(project_id: str) -> Dict[str, object]:
    rows = [m for m in load_materials() if str(m.get("project_id")) == str(project_id)]
    rows.sort(key=lambda x: str(x.get("created_at") or ""), reverse=True)
    signature = _compute_material_index_signature(project_id, rows)
    with _MATERIAL_INDEX_CACHE_LOCK:
        cached = _MATERIAL_INDEX_CACHE.get(project_id)
        if cached and str(cached.get("signature") or "") == signature:
            _touch_material_index_cache(project_id)
            return cached

    counts_by_type: Dict[str, int] = {}
    parsed_ok_by_type: Dict[str, int] = {}
    parsed_fail_by_type: Dict[str, int] = {}
    chars_by_type: Dict[str, int] = {}
    chunks_by_type: Dict[str, int] = {}
    numeric_terms_by_type: Dict[str, int] = {}
    lexical_terms_by_type: Dict[str, int] = {}
    parsed_fail_details: List[Dict[str, str]] = []
    parsed_ok_files = 0
    parsed_failed_files = 0
    total_parsed_chunks = 0
    total_numeric_terms = 0
    total_lexical_terms = 0
    files: List[Dict[str, object]] = []
    available_types: List[str] = []
    available_filenames: List[str] = []

    for row in rows:
        filename = str(row.get("filename") or "").strip()
        mat_type = _normalize_material_type(row.get("material_type"), filename=filename)
        if mat_type and mat_type not in available_types:
            available_types.append(mat_type)
        if filename and filename not in available_filenames:
            available_filenames.append(filename)
        counts_by_type[mat_type] = counts_by_type.get(mat_type, 0) + 1

        entry: Dict[str, object] = {
            "id": str(row.get("id") or ""),
            "project_id": project_id,
            "material_type": mat_type,
            "filename": filename,
            "path": str(row.get("path") or "").strip(),
            "created_at": str(row.get("created_at") or ""),
            "row": dict(row),
            "parsed_ok": False,
            "chars": 0,
            "chunks": [],
            "numeric_terms_norm": [],
            "lexical_terms": [],
            "text": "",
            "parse_error": "",
        }

        path = str(row.get("path") or "").strip()
        if not path:
            parsed_failed_files += 1
            parsed_fail_by_type[mat_type] = parsed_fail_by_type.get(mat_type, 0) + 1
            entry["parse_error"] = "missing_path"
            parsed_fail_details.append(
                {"filename": filename, "material_type": mat_type, "reason": "missing_path"}
            )
            files.append(entry)
            continue

        p = Path(path)
        if not p.exists():
            parsed_failed_files += 1
            parsed_fail_by_type[mat_type] = parsed_fail_by_type.get(mat_type, 0) + 1
            entry["parse_error"] = "path_not_exists"
            parsed_fail_details.append(
                {"filename": filename, "material_type": mat_type, "reason": "path_not_exists"}
            )
            files.append(entry)
            continue

        try:
            content = p.read_bytes()
            text = _read_uploaded_file_content(content, p.name)
            text_value = str(text or "")
            chars = len(text_value.strip())
            chunks = _split_material_text_chunks(text_value, max_chars=900)
            numeric_terms_norm = sorted(
                {
                    token
                    for token in (
                        _normalize_numeric_token(item)
                        for item in _extract_numeric_terms(text_value, max_terms=260)
                    )
                    if token
                }
            )
            lexical_terms = _extract_terms(text_value, max_terms=220)
            entry.update(
                {
                    "parsed_ok": True,
                    "text": text_value,
                    "chars": chars,
                    "chunks": chunks,
                    "numeric_terms_norm": numeric_terms_norm,
                    "lexical_terms": lexical_terms,
                }
            )
            if mat_type == "boq":
                entry["boq_structured_summary"] = _build_boq_structured_summary(
                    content,
                    p.name,
                    parsed_text=text_value,
                )
            parsed_ok_files += 1
            parsed_ok_by_type[mat_type] = parsed_ok_by_type.get(mat_type, 0) + 1
            chars_by_type[mat_type] = chars_by_type.get(mat_type, 0) + chars
            chunks_by_type[mat_type] = chunks_by_type.get(mat_type, 0) + len(chunks)
            numeric_terms_by_type[mat_type] = numeric_terms_by_type.get(mat_type, 0) + len(
                numeric_terms_norm
            )
            lexical_terms_by_type[mat_type] = lexical_terms_by_type.get(mat_type, 0) + len(
                lexical_terms
            )
            total_parsed_chunks += len(chunks)
            total_numeric_terms += len(numeric_terms_norm)
            total_lexical_terms += len(lexical_terms)
        except Exception as exc:
            parsed_failed_files += 1
            parsed_fail_by_type[mat_type] = parsed_fail_by_type.get(mat_type, 0) + 1
            reason = f"{type(exc).__name__}: {exc}"
            entry["parse_error"] = reason
            parsed_fail_details.append(
                {"filename": filename, "material_type": mat_type, "reason": reason}
            )
        files.append(entry)

    total_files = len(rows)
    total_chars = int(sum(chars_by_type.values()))
    fail_ratio = (float(parsed_failed_files) / float(total_files)) if total_files > 0 else 0.0
    quality_snapshot = {
        "project_id": project_id,
        "total_files": total_files,
        "counts_by_type": counts_by_type,
        "parsed_ok_files": parsed_ok_files,
        "parsed_failed_files": parsed_failed_files,
        "parsed_ok_by_type": parsed_ok_by_type,
        "parsed_fail_by_type": parsed_fail_by_type,
        "chars_by_type": chars_by_type,
        "chunks_by_type": chunks_by_type,
        "numeric_terms_by_type": numeric_terms_by_type,
        "lexical_terms_by_type": lexical_terms_by_type,
        "total_parsed_chars": total_chars,
        "total_parsed_chunks": int(total_parsed_chunks),
        "total_numeric_terms": int(total_numeric_terms),
        "total_lexical_terms": int(total_lexical_terms),
        "parse_fail_ratio": round(fail_ratio, 4),
        "parsed_fail_details": parsed_fail_details[:20],
        "available_types": available_types,
        "available_filenames": available_filenames[:120],
    }
    index: Dict[str, object] = {
        "project_id": project_id,
        "signature": signature,
        "rows": [dict(r) for r in rows],
        "files": files,
        "available_types": available_types,
        "available_filenames": available_filenames,
        "quality_snapshot": quality_snapshot,
        "built_at": _now_iso(),
    }
    with _MATERIAL_INDEX_CACHE_LOCK:
        _MATERIAL_INDEX_CACHE[project_id] = index
        _touch_material_index_cache(project_id)
        return index


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _default_weights_raw() -> Dict[str, int]:
    return {dim_id: 5 for dim_id in DIMENSION_IDS}


def _normalize_weights(weights_raw: Dict[str, int]) -> Dict[str, float]:
    multipliers = {
        dim_id: 0.5 + (int(weights_raw.get(dim_id, 5)) / 10.0) for dim_id in DIMENSION_IDS
    }
    total = sum(multipliers.values()) or 1.0
    return {dim_id: multipliers[dim_id] / total for dim_id in DIMENSION_IDS}


def _weights_raw_from_norm(weights_norm: Dict[str, float]) -> Dict[str, int]:
    """
    将归一化权重反推为 0..10 的关注度整数（用于专家配置落库与前端滑杆展示）。
    """
    total_dims = max(1, len(DIMENSION_IDS))
    out: Dict[str, int] = {}
    for dim_id in DIMENSION_IDS:
        w = max(0.0, float(weights_norm.get(dim_id, 1.0 / total_dims)))
        multiplier = w * total_dims
        raw = int(round((multiplier - 0.5) * 10.0))
        out[dim_id] = max(0, min(10, raw))
    return out


def _coerce_weights_raw(weights_raw: Dict[str, int]) -> Dict[str, int]:
    merged = dict(_default_weights_raw())
    for dim_id in DIMENSION_IDS:
        raw_value = weights_raw.get(dim_id, merged[dim_id])
        try:
            value = int(raw_value)
        except Exception:
            raise HTTPException(status_code=422, detail=f"维度 {dim_id} 的关注度必须是整数(0..10)")
        if value < 0 or value > 10:
            raise HTTPException(status_code=422, detail=f"维度 {dim_id} 的关注度超出范围(0..10)")
        merged[dim_id] = value
    return merged


def _new_expert_profile(name: str, weights_raw: Dict[str, int]) -> Dict[str, object]:
    now = _now_iso()
    normalized = _normalize_weights(weights_raw)
    return {
        "id": str(uuid4()),
        "name": name,
        "weights_raw": weights_raw,
        "weights_norm": normalized,
        "norm_rule_version": DEFAULT_NORM_RULE_VERSION,
        "created_at": now,
        "updated_at": now,
    }


def _ensure_project_v2_fields(
    project: Dict[str, object], *, include_engine_defaults: bool = True
) -> bool:
    changed = False
    if not project.get("created_at"):
        project["created_at"] = _now_iso()
        changed = True
    if not project.get("region"):
        project["region"] = DEFAULT_REGION
        changed = True
    if not project.get("qingtian_model_version"):
        project["qingtian_model_version"] = DEFAULT_QINGTIAN_MODEL_VERSION
        changed = True
    if include_engine_defaults and not project.get("scoring_engine_version_locked"):
        project["scoring_engine_version_locked"] = DEFAULT_SCORING_ENGINE_LOCKED
        changed = True
    # 校准器锁定版本默认留空，避免误套历史/跨项目模型。
    if include_engine_defaults and "calibrator_version_locked" not in project:
        project["calibrator_version_locked"] = DEFAULT_CALIBRATOR_LOCKED
        changed = True
    if not project.get("status"):
        project["status"] = "scoring_preparation"
        changed = True
    if not project.get("updated_at"):
        project["updated_at"] = project.get("created_at") or _now_iso()
        changed = True
    meta = project.get("meta")
    if not isinstance(meta, dict):
        project["meta"] = {}
        meta = project["meta"]
        changed = True
    strict_mode = bool(meta.get("strict_material_mode", True))

    def _set_min_rate(key: str, value: float) -> None:
        nonlocal changed
        current = _to_float_or_none(meta.get(key))
        if current is None or float(current) < float(value):
            meta[key] = float(value)
            changed = True

    def _set_min_int(key: str, value: int) -> None:
        nonlocal changed
        current = _to_float_or_none(meta.get(key))
        if current is None or int(round(float(current))) < int(value):
            meta[key] = int(value)
            changed = True

    if "score_scale_max" not in meta:
        meta["score_scale_max"] = DEFAULT_SCORE_SCALE_MAX
        changed = True
    if "enforce_gb_redline" not in meta:
        meta["enforce_gb_redline"] = bool(DEFAULT_ENFORCE_GB_REDLINE)
        changed = True
    if "enforce_material_gate" not in meta:
        meta["enforce_material_gate"] = DEFAULT_ENFORCE_MATERIAL_GATE
        changed = True
    required_material_types = meta.get("required_material_types")
    if not isinstance(required_material_types, list):
        meta["required_material_types"] = list(DEFAULT_REQUIRED_MATERIAL_TYPES)
        changed = True
    else:
        normalized_required: List[str] = []
        for item in required_material_types:
            key = _normalize_material_type(item)
            if key and key not in normalized_required:
                normalized_required.append(key)
        for required_key in DEFAULT_REQUIRED_MATERIAL_TYPES:
            if required_key not in normalized_required:
                normalized_required.append(required_key)
                changed = True
        if normalized_required != required_material_types:
            meta["required_material_types"] = normalized_required
            changed = True
    if "min_parsed_chars_by_type" not in meta:
        meta["min_parsed_chars_by_type"] = dict(DEFAULT_MIN_PARSED_CHARS_BY_TYPE)
        changed = True
    if "enforce_material_depth_gate" not in meta:
        meta["enforce_material_depth_gate"] = bool(DEFAULT_ENFORCE_MATERIAL_DEPTH_GATE)
        changed = True
    elif strict_mode and not bool(meta.get("enforce_material_depth_gate")):
        meta["enforce_material_depth_gate"] = True
        changed = True
    if "min_parsed_chunks_by_type" not in meta:
        meta["min_parsed_chunks_by_type"] = dict(DEFAULT_MIN_PARSED_CHUNKS_BY_TYPE)
        changed = True
    if "min_numeric_terms_by_type" not in meta:
        meta["min_numeric_terms_by_type"] = dict(DEFAULT_MIN_NUMERIC_TERMS_BY_TYPE)
        changed = True
    if "min_total_parsed_chunks" not in meta:
        meta["min_total_parsed_chunks"] = int(DEFAULT_MIN_TOTAL_PARSED_CHUNKS)
        changed = True
    if "min_total_parsed_chars" not in meta:
        meta["min_total_parsed_chars"] = int(DEFAULT_MIN_TOTAL_PARSED_CHARS)
        changed = True
    if "max_material_parse_fail_ratio" not in meta:
        meta["max_material_parse_fail_ratio"] = float(DEFAULT_MAX_MATERIAL_PARSE_FAIL_RATIO)
        changed = True
    if "block_on_any_material_parse_failure" not in meta:
        meta["block_on_any_material_parse_failure"] = bool(
            DEFAULT_BLOCK_ON_ANY_MATERIAL_PARSE_FAILURE
        )
        changed = True
    if "enforce_material_utilization_gate" not in meta:
        meta["enforce_material_utilization_gate"] = bool(DEFAULT_ENFORCE_MATERIAL_UTILIZATION_GATE)
        changed = True
    if "material_utilization_gate_mode" not in meta:
        meta["material_utilization_gate_mode"] = DEFAULT_MATERIAL_UTILIZATION_GATE_MODE
        changed = True
    if "min_material_retrieval_hit_rate" not in meta:
        meta["min_material_retrieval_hit_rate"] = float(DEFAULT_MIN_MATERIAL_RETRIEVAL_HIT_RATE)
        changed = True
    if "min_material_consistency_hit_rate" not in meta:
        meta["min_material_consistency_hit_rate"] = float(DEFAULT_MIN_MATERIAL_CONSISTENCY_HIT_RATE)
        changed = True
    if "min_material_retrieval_total" not in meta:
        meta["min_material_retrieval_total"] = int(DEFAULT_MIN_MATERIAL_RETRIEVAL_TOTAL)
        changed = True
    if "max_uncovered_required_types" not in meta:
        meta["max_uncovered_required_types"] = int(DEFAULT_MAX_UNCOVERED_REQUIRED_TYPES)
        changed = True
    if "min_required_type_presence_rate" not in meta:
        meta["min_required_type_presence_rate"] = float(DEFAULT_MIN_REQUIRED_TYPE_PRESENCE_RATE)
        changed = True
    if "min_required_type_coverage_rate" not in meta:
        meta["min_required_type_coverage_rate"] = float(DEFAULT_MIN_REQUIRED_TYPE_COVERAGE_RATE)
        changed = True
    if "material_retrieval_top_k" not in meta:
        meta["material_retrieval_top_k"] = int(DEFAULT_MATERIAL_RETRIEVAL_TOP_K)
        changed = True
    if "material_retrieval_per_type_quota" not in meta:
        meta["material_retrieval_per_type_quota"] = int(DEFAULT_MATERIAL_RETRIEVAL_PER_TYPE_QUOTA)
        changed = True
    if "material_retrieval_per_file_quota" not in meta:
        meta["material_retrieval_per_file_quota"] = int(DEFAULT_MATERIAL_RETRIEVAL_PER_FILE_QUOTA)
        changed = True
    if "min_material_retrieval_file_coverage_rate" not in meta:
        meta["min_material_retrieval_file_coverage_rate"] = float(
            DEFAULT_MIN_MATERIAL_RETRIEVAL_FILE_COVERAGE_RATE
        )
        changed = True
    if "enforce_uploaded_type_coverage" not in meta:
        meta["enforce_uploaded_type_coverage"] = bool(DEFAULT_ENFORCE_UPLOADED_TYPE_COVERAGE)
        changed = True
    if "min_uploaded_type_coverage_rate" not in meta:
        meta["min_uploaded_type_coverage_rate"] = float(DEFAULT_MIN_UPLOADED_TYPE_COVERAGE_RATE)
        changed = True
    if "evolution_weight_min_samples" not in meta:
        meta["evolution_weight_min_samples"] = 3
        changed = True
    if "evolution_weight_max_age_days" not in meta:
        meta["evolution_weight_max_age_days"] = DEFAULT_CALIBRATION_FRESHNESS_DAYS
        changed = True
    if strict_mode:
        _set_min_int("min_material_retrieval_total", DEFAULT_MIN_MATERIAL_RETRIEVAL_TOTAL)
        _set_min_rate(
            "min_material_retrieval_file_coverage_rate",
            DEFAULT_MIN_MATERIAL_RETRIEVAL_FILE_COVERAGE_RATE,
        )
        _set_min_rate("min_required_type_coverage_rate", DEFAULT_MIN_REQUIRED_TYPE_COVERAGE_RATE)
    return changed


def _assert_project_profile_operation_unlocked(
    project: Dict[str, object], force_unlock: bool
) -> None:
    status = str(project.get("status") or "").strip()
    if status in PROFILE_LOCKED_STATUSES and not force_unlock:
        raise HTTPException(
            status_code=409,
            detail="项目已进入青天评标阶段，默认锁定专家配置与重算。请确认后使用 force_unlock=true 重试。",
        )


def _ensure_project_expert_profile(
    project: Dict[str, object],
    all_profiles: List[Dict[str, object]],
) -> tuple[Dict[str, object], bool]:
    profile_id = str(project.get("expert_profile_id") or "")
    if profile_id:
        for profile in all_profiles:
            if profile.get("id") == profile_id:
                return profile, False

    profile_name = f"{project.get('name', '项目')} 默认配置"
    created = _new_expert_profile(profile_name, _default_weights_raw())
    all_profiles.append(created)
    project["expert_profile_id"] = created["id"]
    project["updated_at"] = _now_iso()
    return created, True


def _recover_missing_project_from_artifacts(
    project_id: str, projects: List[Dict[str, object]]
) -> Optional[Dict[str, object]]:
    pid = str(project_id or "").strip()
    if not pid:
        return None
    for p in projects:
        if str(p.get("id") or "") == pid:
            return p

    submissions = [s for s in load_submissions() if str(s.get("project_id") or "") == pid]
    materials = [m for m in load_materials() if str(m.get("project_id") or "") == pid]
    ground_truth = [g for g in load_ground_truth() if str(g.get("project_id") or "") == pid]
    evo_reports = load_evolution_reports()
    has_evolution = pid in evo_reports

    if not submissions and not materials and not ground_truth and not has_evolution:
        return None

    name_seed = ""
    for row in materials + submissions:
        filename = str(row.get("filename") or "").strip()
        if filename:
            name_seed = filename
            break
    if name_seed:
        stem = name_seed.rsplit(".", 1)[0].strip()
        recovered_name = (stem or name_seed) + "（恢复）"
    else:
        recovered_name = f"恢复项目_{pid[:8]}"

    time_points: List[str] = []
    for row in submissions:
        created_at = str(row.get("created_at") or "").strip()
        updated_at = str(row.get("updated_at") or "").strip()
        if created_at:
            time_points.append(created_at)
        if updated_at:
            time_points.append(updated_at)
    for row in materials + ground_truth:
        created_at = str(row.get("created_at") or "").strip()
        if created_at:
            time_points.append(created_at)
    evo_updated_at = str((evo_reports.get(pid) or {}).get("updated_at") or "").strip()
    if evo_updated_at:
        time_points.append(evo_updated_at)
    created_at = min(time_points) if time_points else _now_iso()
    updated_at = max(time_points) if time_points else _now_iso()

    score_scale_max = DEFAULT_SCORE_SCALE_MAX
    for s in submissions:
        report = s.get("report")
        if not isinstance(report, dict):
            continue
        meta = report.get("meta")
        if not isinstance(meta, dict):
            continue
        raw = meta.get("score_scale_max")
        if str(raw) == "5":
            score_scale_max = 5
            break
        if str(raw) == "100":
            score_scale_max = 100

    recovered = {
        "id": pid,
        "name": recovered_name,
        "meta": {"score_scale_max": score_scale_max},
        "region": DEFAULT_REGION,
        "expert_profile_id": None,
        "qingtian_model_version": DEFAULT_QINGTIAN_MODEL_VERSION,
        "scoring_engine_version_locked": DEFAULT_SCORING_ENGINE_LOCKED,
        "calibrator_version_locked": DEFAULT_CALIBRATOR_LOCKED,
        "status": "scoring_preparation",
        "created_at": created_at,
        "updated_at": updated_at,
    }
    _ensure_project_v2_fields(recovered)
    projects.append(recovered)
    save_projects(projects)
    return recovered


def _recover_latest_orphan_project(
    projects: List[Dict[str, object]],
) -> Optional[Dict[str, object]]:
    existing_ids = {str(p.get("id") or "") for p in projects}
    latest_pid = ""
    latest_at = ""

    for row in load_submissions():
        pid = str(row.get("project_id") or "").strip()
        if not pid or pid in existing_ids:
            continue
        ts = str(row.get("updated_at") or row.get("created_at") or "").strip()
        if ts and ts > latest_at:
            latest_at = ts
            latest_pid = pid
    for row in load_materials():
        pid = str(row.get("project_id") or "").strip()
        if not pid or pid in existing_ids:
            continue
        ts = str(row.get("created_at") or "").strip()
        if ts and ts > latest_at:
            latest_at = ts
            latest_pid = pid
    if not latest_pid:
        return None
    return _recover_missing_project_from_artifacts(latest_pid, projects)


def _find_project(project_id: str, projects: List[Dict[str, object]]) -> Dict[str, object]:
    for p in projects:
        if p.get("id") == project_id:
            return p
    if not os.environ.get("PYTEST_CURRENT_TEST"):
        recovered = _recover_missing_project_from_artifacts(project_id, projects)
        if recovered is not None:
            return recovered
    raise HTTPException(status_code=404, detail="项目不存在")


def _find_submission(submission_id: str, submissions: List[Dict[str, object]]) -> Dict[str, object]:
    for s in submissions:
        if str(s.get("id")) == submission_id:
            return s
    raise HTTPException(status_code=404, detail="施组提交记录不存在")


def _latest_project_submission(
    project_id: str,
    submissions: List[Dict[str, object]],
    *,
    prefer_scored: bool = True,
) -> Optional[Dict[str, object]]:
    rows = [s for s in submissions if str(s.get("project_id") or "") == project_id]
    if not rows:
        return None
    rows = sorted(rows, key=lambda x: str(x.get("created_at") or ""), reverse=True)
    if prefer_scored:
        for row in rows:
            if _submission_is_scored(row):
                return row
    return rows[0]


def _weights_norm_to_dimension_multipliers(weights_norm: Dict[str, float]) -> Dict[str, float]:
    baseline = 1.0 / len(DIMENSION_IDS)
    if baseline <= 0:
        return {}
    return {
        dim_id: float(weights_norm.get(dim_id, baseline)) / baseline for dim_id in DIMENSION_IDS
    }


def _resolve_project_scoring_context(
    project_id: str,
) -> tuple[Dict[str, float], Optional[Dict[str, object]], Dict[str, object]]:
    projects = load_projects()
    project = _find_project(project_id, projects)
    _ensure_project_v2_fields(project, include_engine_defaults=False)

    profiles = load_expert_profiles()
    profile = None
    if project.get("expert_profile_id"):
        for item in profiles:
            if item.get("id") == project.get("expert_profile_id"):
                profile = item
                break

    # 进化权重优先，但必须满足最小样本量与时效性，避免概念漂移导致误导评分。
    reports = load_evolution_reports()
    evo = reports.get(project_id) or {}
    se = evo.get("scoring_evolution") or {}
    mult = se.get("dimension_multipliers") or {}
    if mult:
        meta = project.get("meta") if isinstance(project.get("meta"), dict) else {}
        min_samples = max(1, int(_to_float_or_none(meta.get("evolution_weight_min_samples")) or 3))
        max_age_days = max(
            1.0,
            float(
                _to_float_or_none(meta.get("evolution_weight_max_age_days"))
                or DEFAULT_CALIBRATION_FRESHNESS_DAYS
            ),
        )
        sample_count = int(_to_float_or_none(evo.get("sample_count")) or 0)
        updated_at_dt = _parse_iso_datetime(evo.get("updated_at") or evo.get("created_at"))
        age_days = None
        if updated_at_dt is not None:
            age_days = max(
                0.0,
                (
                    datetime.now(timezone.utc) - updated_at_dt.astimezone(timezone.utc)
                ).total_seconds()
                / 86400.0,
            )
        evolved_usable = (
            sample_count >= min_samples
            and age_days is not None
            and float(age_days) <= float(max_age_days)
        )
        if evolved_usable:
            return dict(mult), None, project
        logger.info(
            "skip stale_or_thin_evolution_weights project_id=%s sample_count=%s min_samples=%s age_days=%s max_age_days=%s",
            project_id,
            sample_count,
            min_samples,
            round(float(age_days), 2) if age_days is not None else None,
            max_age_days,
        )

    if profile and isinstance(profile.get("weights_norm"), dict):
        multipliers = _weights_norm_to_dimension_multipliers(profile.get("weights_norm", {}))
        return multipliers, profile, project

    for p in load_learning_profiles():
        if p.get("project_id") == project_id:
            return dict(p.get("dimension_multipliers") or {}), None, project
    return {}, None, project


def _parse_iso_datetime(value: object) -> Optional[datetime]:
    text = str(value or "").strip()
    if not text:
        return None
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(text)
    except Exception:
        return None


def _latest_record_time(
    rows: List[Dict[str, object]], field: str = "created_at"
) -> Optional[datetime]:
    latest: Optional[datetime] = None
    for row in rows:
        if not isinstance(row, dict):
            continue
        dt = _parse_iso_datetime(row.get(field))
        if dt is None:
            continue
        if latest is None or dt > latest:
            latest = dt
    return latest


def _load_project_context_text(project_id: str) -> str:
    ctx = load_project_context().get(project_id) or {}
    return str(ctx.get("text") or "").strip()


def _build_constraints_source_text(project_id: str) -> str:
    materials_text = _merge_materials_text(project_id).strip()
    context_text = _load_project_context_text(project_id)
    if not context_text:
        return materials_text
    marker = "--- 项目上下文（投喂包/自定义指令） ---"
    if materials_text:
        return f"{materials_text}\n\n{marker}\n{context_text}".strip()
    return f"{marker}\n{context_text}".strip()


def _constraints_need_rebuild(
    project_id: str,
    anchors: List[Dict[str, object]],
    requirements: List[Dict[str, object]],
) -> bool:
    source_rows = [m for m in load_materials() if str(m.get("project_id")) == project_id]
    source_latest = _latest_record_time(source_rows, field="created_at")
    ctx_data = load_project_context().get(project_id) or {}
    ctx_updated = _parse_iso_datetime(ctx_data.get("updated_at"))
    if ctx_updated is not None and (source_latest is None or ctx_updated > source_latest):
        source_latest = ctx_updated
    if source_latest is None:
        return not anchors or not requirements
    constraints_latest = _latest_record_time(anchors, field="created_at")
    req_latest = _latest_record_time(requirements, field="created_at")
    if req_latest is not None and (constraints_latest is None or req_latest > constraints_latest):
        constraints_latest = req_latest
    if constraints_latest is None:
        return True
    return source_latest > constraints_latest


def _to_text_items(raw: object, *, max_items: int = 24) -> List[str]:
    items: List[str] = []
    if isinstance(raw, list):
        for v in raw:
            s = str(v or "").strip()
            if not s:
                continue
            items.append(s)
    elif isinstance(raw, str):
        for line in raw.replace("；", "\n").replace(";", "\n").splitlines():
            s = line.strip()
            if s:
                items.append(s)
    dedup: List[str] = []
    seen: set[str] = set()
    for item in items:
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        dedup.append(item)
        if len(dedup) >= max(1, int(max_items)):
            break
    return dedup


def _resolve_material_retrieval_config(project: Dict[str, object]) -> Dict[str, int]:
    meta = project.get("meta") if isinstance(project.get("meta"), dict) else {}

    def _resolve_int(raw: object, default: int, *, lower: int, upper: int) -> int:
        numeric = _to_float_or_none(raw)
        if numeric is None:
            numeric = float(default)
        return max(lower, min(upper, int(round(float(numeric)))))

    top_k = _resolve_int(
        meta.get("material_retrieval_top_k"),
        DEFAULT_MATERIAL_RETRIEVAL_TOP_K,
        lower=6,
        upper=48,
    )
    per_type_quota = _resolve_int(
        meta.get("material_retrieval_per_type_quota"),
        DEFAULT_MATERIAL_RETRIEVAL_PER_TYPE_QUOTA,
        lower=1,
        upper=6,
    )
    per_file_quota = _resolve_int(
        meta.get("material_retrieval_per_file_quota"),
        DEFAULT_MATERIAL_RETRIEVAL_PER_FILE_QUOTA,
        lower=1,
        upper=8,
    )
    return {
        "top_k": top_k,
        "per_type_quota": per_type_quota,
        "per_file_quota": per_file_quota,
    }


def _compute_dynamic_retrieval_budget(
    project_id: str,
    base_cfg: Dict[str, int],
    material_rows: List[Dict[str, object]],
    *,
    available_material_types: Optional[List[str]] = None,
) -> Dict[str, object]:
    """
    按资料体量动态放大检索预算，确保资料越多时检索“读得更深”。
    - top_k：随文件数、文件总大小、资料类型数增长（封顶48）。
    - per_type_quota：当类型丰富时，至少保证每类2条证据。
    - per_file_quota：当文件较多时，允许单文件更深采样，降低偶然漏命中。
    """
    base_top_k = max(6, min(48, int(base_cfg.get("top_k") or DEFAULT_MATERIAL_RETRIEVAL_TOP_K)))
    base_per_type = max(
        1,
        min(
            6,
            int(
                base_cfg.get(
                    "per_type_quota",
                    DEFAULT_MATERIAL_RETRIEVAL_PER_TYPE_QUOTA,
                )
                or DEFAULT_MATERIAL_RETRIEVAL_PER_TYPE_QUOTA
            ),
        ),
    )
    base_per_file = max(
        1,
        min(
            8,
            int(
                base_cfg.get(
                    "per_file_quota",
                    DEFAULT_MATERIAL_RETRIEVAL_PER_FILE_QUOTA,
                )
                or DEFAULT_MATERIAL_RETRIEVAL_PER_FILE_QUOTA
            ),
        ),
    )

    types: List[str] = []
    if isinstance(available_material_types, list):
        for item in available_material_types:
            key = _normalize_material_type(item)
            if key and key not in types:
                types.append(key)
    if not types:
        for row in material_rows:
            key = _normalize_material_type(row.get("material_type"), filename=row.get("filename"))
            if key and key not in types:
                types.append(key)

    file_count = 0
    total_size_bytes = 0
    for row in material_rows:
        if str(row.get("project_id") or "") != str(project_id):
            continue
        file_count += 1
        path = str(row.get("path") or "").strip()
        if not path:
            continue
        try:
            total_size_bytes += max(0, int(Path(path).stat().st_size))
        except Exception:
            continue
    total_size_mb = round(float(total_size_bytes) / (1024.0 * 1024.0), 3)

    type_count = len(types)
    # 类型覆盖底线：至少满足 “每类配额” 的总量。
    type_floor = max(6, int(type_count) * int(base_per_type))
    # 体量增强项：文件数和总体积越大，越需要更深检索。
    file_boost = max(0, min(12, file_count - 3))
    size_boost = max(0, min(12, int(total_size_mb // 20)))
    diversity_boost = 2 if type_count >= 4 else (1 if type_count >= 3 else 0)
    volume_boost = file_boost + size_boost + diversity_boost
    top_k = max(base_top_k, type_floor, type_floor + min(12, volume_boost))
    top_k = max(6, min(48, int(top_k)))

    per_type_quota = base_per_type
    if type_count >= 4:
        per_type_quota = max(per_type_quota, 2)
    per_type_quota = max(1, min(6, int(per_type_quota)))

    per_file_quota = base_per_file
    if file_count >= 12:
        per_file_quota = max(per_file_quota, 4)
    elif file_count >= 6:
        per_file_quota = max(per_file_quota, 3)
    per_file_quota = max(1, min(8, int(per_file_quota)))

    reasons: List[str] = []
    if top_k > base_top_k:
        reasons.append(f"top_k:{base_top_k}->{top_k}")
    if per_type_quota > base_per_type:
        reasons.append(f"per_type_quota:{base_per_type}->{per_type_quota}")
    if per_file_quota > base_per_file:
        reasons.append(f"per_file_quota:{base_per_file}->{per_file_quota}")
    if not reasons:
        reasons.append("base_budget_kept")

    return {
        "top_k": top_k,
        "per_type_quota": per_type_quota,
        "per_file_quota": per_file_quota,
        "base_top_k": base_top_k,
        "base_per_type_quota": base_per_type,
        "base_per_file_quota": base_per_file,
        "material_file_count": file_count,
        "material_type_count": type_count,
        "material_total_size_mb": total_size_mb,
        "budget_reasons": reasons,
    }


def _build_material_query_features(
    *,
    project_id: str,
    submission_text: str,
    required_sections: List[str],
    required_charts: List[str],
    mandatory_elements: List[str],
    custom_text_items: List[str],
    context_text: str,
) -> Dict[str, object]:
    seed_lines: List[str] = []
    seed_lines.extend(required_sections)
    seed_lines.extend(required_charts)
    seed_lines.extend(mandatory_elements)
    seed_lines.extend(custom_text_items)
    if context_text:
        seed_lines.append(context_text[:4000])

    project_anchors = [a for a in load_project_anchors() if str(a.get("project_id")) == project_id]
    for anchor in project_anchors[:80]:
        key = str(anchor.get("anchor_key") or "").strip()
        if key:
            seed_lines.append(key)
        value = anchor.get("anchor_value")
        if isinstance(value, list):
            seed_lines.extend(str(v) for v in value[:6] if str(v).strip())
        elif isinstance(value, str) and value.strip():
            seed_lines.append(value.strip())
        value_num = anchor.get("value_num")
        if value_num is not None:
            seed_lines.append(str(value_num))

    project_requirements = [
        r for r in load_project_requirements() if str(r.get("project_id")) == project_id
    ]
    for req in project_requirements[:120]:
        label = str(req.get("req_label") or "").strip()
        if label:
            seed_lines.append(label)
        patterns = req.get("patterns") if isinstance(req.get("patterns"), dict) else {}
        for key in ("keywords", "hints", "must_hit_terms"):
            values = patterns.get(key)
            if isinstance(values, list):
                seed_lines.extend(str(v) for v in values[:8] if str(v).strip())
            elif isinstance(values, str) and values.strip():
                seed_lines.append(values.strip())

    seed_text = "\n".join(seed_lines)
    text_terms = _extract_terms(seed_text, max_terms=100)
    numeric_terms = _extract_numeric_terms(seed_text, max_terms=40)
    submission_numeric_terms = _extract_numeric_terms(submission_text, max_terms=24)
    for token in submission_numeric_terms:
        if token not in numeric_terms:
            numeric_terms.append(token)

    return {
        "query_terms": text_terms,
        "query_numeric_terms": numeric_terms[:60],
        "seed_lines_count": len(seed_lines),
        "project_anchor_count": len(project_anchors),
        "project_requirement_count": len(project_requirements),
    }


def _build_runtime_custom_requirements(
    project_id: str,
    *,
    project: Dict[str, object],
    submission_text: str = "",
) -> tuple[List[Dict[str, object]], Dict[str, object]]:
    evo = load_evolution_reports().get(project_id) or {}
    compilation = (
        evo.get("compilation_instructions")
        if isinstance(evo.get("compilation_instructions"), dict)
        else {}
    )
    required_sections = _to_text_items(compilation.get("required_sections"), max_items=20)
    required_charts = _to_text_items(compilation.get("required_charts_images"), max_items=20)
    mandatory_elements = _to_text_items(compilation.get("mandatory_elements"), max_items=24)

    custom_text_items: List[str] = []
    meta_obj = project.get("meta") if isinstance(project.get("meta"), dict) else {}
    custom_text_items.extend(
        _to_text_items((meta_obj or {}).get("custom_scoring_instructions"), max_items=20)
    )
    # 允许在项目上下文里用“指令:”行注入运行时评分提示，避免另开配置入口。
    context_text = _load_project_context_text(project_id)
    for line in context_text.splitlines():
        s = line.strip()
        if not s or len(s) > 120:
            continue
        if any(mark in s for mark in ("指令", "要求", "必须", "禁止", "关注")):
            custom_text_items.append(s)
    custom_text_items = _to_text_items(custom_text_items, max_items=20)

    runtime_requirements: List[Dict[str, object]] = []
    created_at = _now_iso()
    req_index = 0

    def _append_req(
        *,
        dim_id: str,
        label: str,
        hints: List[str],
        mandatory: bool,
        weight: float,
        kind: str,
    ) -> None:
        nonlocal req_index
        if not hints:
            return
        req_index += 1
        runtime_requirements.append(
            {
                "id": f"runtime-{project_id[:8]}-{kind}-{req_index}",
                "project_id": project_id,
                "dimension_id": dim_id,
                "req_label": label,
                "req_type": "semantic",
                "patterns": {"hints": hints},
                "mandatory": mandatory,
                "weight": float(weight),
                "source_anchor_id": None,
                "source_pack_id": "runtime_custom",
                "source_pack_version": "v2-runtime",
                "priority": 90.0,
                "override_key": f"runtime::{kind}::{req_index}",
                "lint": {},
                "created_at": created_at,
            }
        )

    for item in required_sections:
        _append_req(
            dim_id="01",
            label=f"编制系统指令-必备章节：{item}",
            hints=[item],
            mandatory=True,
            weight=1.1,
            kind="section",
        )
    for item in required_charts:
        _append_req(
            dim_id="12",
            label=f"编制系统指令-图表要求：{item}",
            hints=[item],
            mandatory=False,
            weight=0.8,
            kind="chart",
        )
    for item in mandatory_elements:
        _append_req(
            dim_id="09",
            label=f"编制系统指令-必备要素：{item}",
            hints=[item],
            mandatory=True,
            weight=1.0,
            kind="element",
        )
    for item in custom_text_items:
        _append_req(
            dim_id="01",
            label=f"自定义评分指令：{item}",
            hints=[item],
            mandatory=False,
            weight=0.6,
            kind="custom",
        )

    retrieval_cfg = _resolve_material_retrieval_config(project)
    query_features = _build_material_query_features(
        project_id=project_id,
        submission_text=submission_text,
        required_sections=required_sections,
        required_charts=required_charts,
        mandatory_elements=mandatory_elements,
        custom_text_items=custom_text_items,
        context_text=context_text,
    )
    material_index = _build_project_material_index(project_id)
    material_rows = list(material_index.get("rows") or [])
    available_material_types = [
        str(x) for x in (material_index.get("available_types") or []) if str(x).strip()
    ]
    available_material_filenames = [
        str(x) for x in (material_index.get("available_filenames") or []) if str(x).strip()
    ]
    retrieval_budget = _compute_dynamic_retrieval_budget(
        project_id,
        retrieval_cfg,
        material_rows,
        available_material_types=available_material_types,
    )
    retrieval_top_k = int(
        retrieval_budget.get("top_k")
        or retrieval_cfg.get("top_k")
        or DEFAULT_MATERIAL_RETRIEVAL_TOP_K
    )
    retrieval_per_type_quota = int(
        retrieval_budget.get("per_type_quota")
        or retrieval_cfg.get("per_type_quota")
        or DEFAULT_MATERIAL_RETRIEVAL_PER_TYPE_QUOTA
    )
    retrieval_per_file_quota = int(
        retrieval_budget.get("per_file_quota")
        or retrieval_cfg.get("per_file_quota")
        or DEFAULT_MATERIAL_RETRIEVAL_PER_FILE_QUOTA
    )
    retrieval_chunks = _select_material_retrieval_chunks(
        project_id,
        submission_text,
        top_k=retrieval_top_k,
        per_type_quota=retrieval_per_type_quota,
        per_file_quota=retrieval_per_file_quota,
        query_terms_extra=query_features.get("query_terms"),
        query_numeric_terms=query_features.get("query_numeric_terms"),
        material_index=material_index,
    )
    retrieval_selected_filenames: List[str] = []
    for chunk in retrieval_chunks:
        filename_text = str(chunk.get("filename") or "").strip()
        if filename_text and filename_text not in retrieval_selected_filenames:
            retrieval_selected_filenames.append(filename_text)
    retrieval_types = sorted(
        {
            str(c.get("material_type") or "")
            for c in retrieval_chunks
            if str(c.get("material_type") or "").strip()
        }
    )
    retrieval_selected_via_counts: Dict[str, int] = {}
    for chunk in retrieval_chunks:
        selected_via = str(chunk.get("selected_via") or "").strip() or "unknown"
        retrieval_selected_via_counts[selected_via] = (
            int(retrieval_selected_via_counts.get(selected_via, 0)) + 1
        )
    retrieval_requirements = _build_material_retrieval_requirements(project_id, retrieval_chunks)
    consistency_requirements = _build_material_consistency_requirements(
        project_id,
        retrieval_chunks,
        available_material_types=available_material_types,
    )
    runtime_requirements.extend(retrieval_requirements)
    runtime_requirements.extend(consistency_requirements)

    meta = {
        "required_sections": len(required_sections),
        "required_charts_images": len(required_charts),
        "mandatory_elements": len(mandatory_elements),
        "custom_instruction_lines": len(custom_text_items),
        "runtime_custom_requirements": len(runtime_requirements),
        "project_context_chars": len(context_text),
        "material_retrieval_chunks": len(retrieval_chunks),
        "material_retrieval_requirements": len(retrieval_requirements),
        "material_consistency_requirements": len(consistency_requirements),
        "material_retrieval_top_k": retrieval_top_k,
        "material_retrieval_per_type_quota": retrieval_per_type_quota,
        "material_retrieval_per_file_quota": retrieval_per_file_quota,
        "material_retrieval_base_top_k": int(
            retrieval_budget.get("base_top_k")
            or retrieval_cfg.get("top_k")
            or DEFAULT_MATERIAL_RETRIEVAL_TOP_K
        ),
        "material_retrieval_base_per_type_quota": int(
            retrieval_budget.get("base_per_type_quota")
            or retrieval_cfg.get("per_type_quota")
            or DEFAULT_MATERIAL_RETRIEVAL_PER_TYPE_QUOTA
        ),
        "material_retrieval_base_per_file_quota": int(
            retrieval_budget.get("base_per_file_quota")
            or retrieval_cfg.get("per_file_quota")
            or DEFAULT_MATERIAL_RETRIEVAL_PER_FILE_QUOTA
        ),
        "material_retrieval_budget_reasons": list(retrieval_budget.get("budget_reasons") or []),
        "material_query_seed_lines": int(query_features.get("seed_lines_count") or 0),
        "material_query_terms_count": len(query_features.get("query_terms") or []),
        "material_query_numeric_terms_count": len(query_features.get("query_numeric_terms") or []),
        "material_query_terms_preview": list((query_features.get("query_terms") or [])[:12]),
        "material_query_numeric_terms_preview": list(
            (query_features.get("query_numeric_terms") or [])[:12]
        ),
        "material_query_anchor_count": int(query_features.get("project_anchor_count") or 0),
        "material_query_requirement_count": int(
            query_features.get("project_requirement_count") or 0
        ),
        "material_available_files": len(material_rows),
        "material_available_filenames": available_material_filenames[:120],
        "material_available_types": available_material_types,
        "material_total_size_mb": float(retrieval_budget.get("material_total_size_mb") or 0.0),
        "material_type_count": int(retrieval_budget.get("material_type_count") or 0),
        "material_retrieval_types": retrieval_types,
        "material_retrieval_selected_via_counts": retrieval_selected_via_counts,
        "material_retrieval_selected_filenames": retrieval_selected_filenames[:120],
        "material_retrieval_missing_types": [
            t for t in available_material_types if t not in retrieval_types
        ],
        "material_retrieval_preview": [
            {
                "material_type": c.get("material_type"),
                "filename": c.get("filename"),
                "chunk_id": c.get("chunk_id"),
                "dimension_id": c.get("dimension_id"),
                "score": c.get("score"),
                "matched_terms": c.get("matched_terms"),
                "matched_numeric_terms": c.get("matched_numeric_terms"),
                "selected_via": c.get("selected_via"),
            }
            for c in retrieval_chunks[:8]
        ],
        "material_consistency_preview": [
            {
                "material_type": req.get("material_type"),
                "dimension_id": req.get("dimension_id"),
                "mandatory": req.get("mandatory"),
                "terms": ((req.get("patterns") or {}).get("must_hit_terms") or [])[:6],
                "numbers": ((req.get("patterns") or {}).get("must_hit_numbers") or [])[:4],
            }
            for req in consistency_requirements[:6]
        ],
    }
    return runtime_requirements, meta


def _infer_weights_source(project_id: str, profile_snapshot: Optional[Dict[str, object]]) -> str:
    if profile_snapshot:
        return "expert_profile"
    evo = load_evolution_reports().get(project_id) or {}
    evo_mult = (evo.get("scoring_evolution") or {}).get("dimension_multipliers") or {}
    if isinstance(evo_mult, dict) and evo_mult:
        return "evolution"
    if any(str(p.get("project_id")) == project_id for p in load_learning_profiles()):
        return "learning_profile"
    return "default_uniform"


def _build_scoring_input_injection_meta(
    *,
    project_id: str,
    text: str,
    anchors_count: int,
    base_requirements_count: int,
    runtime_custom_requirements_count: int,
    weights_norm: Dict[str, float],
    profile_snapshot: Optional[Dict[str, object]],
    constraints_rebuilt: bool,
    runtime_req_meta: Dict[str, object],
    material_quality_snapshot: Optional[Dict[str, object]] = None,
) -> Dict[str, object]:
    material_rows = [m for m in load_materials() if str(m.get("project_id")) == project_id]
    material_type_counts: Dict[str, int] = {}
    for row in material_rows:
        key = _normalize_material_type(row.get("material_type"), filename=row.get("filename"))
        material_type_counts[key] = material_type_counts.get(key, 0) + 1
    snapshot = material_quality_snapshot if isinstance(material_quality_snapshot, dict) else None
    if snapshot:
        snapshot_type_counts = snapshot.get("counts_by_type")
        if isinstance(snapshot_type_counts, dict) and snapshot_type_counts:
            material_type_counts = {
                str(k): int(_to_float_or_none(v) or 0) for k, v in snapshot_type_counts.items()
            }
        materials_count = int(_to_float_or_none(snapshot.get("total_files")) or len(material_rows))
        total_parsed_chars = int(_to_float_or_none(snapshot.get("total_parsed_chars")) or 0)
        parse_fail_ratio = float(_to_float_or_none(snapshot.get("parse_fail_ratio")) or 0.0)
    else:
        materials_count = len(material_rows)
        total_parsed_chars = 0
        parse_fail_ratio = 0.0
    context_text = _load_project_context_text(project_id)
    gate = snapshot.get("gate") if isinstance(snapshot, dict) else {}
    gate_passed = bool((gate or {}).get("passed", True))
    return {
        "mece_inputs": {
            "project_materials_extracted": materials_count > 0,
            "shigong_parsed": bool(str(text or "").strip()),
            "bid_requirements_loaded": base_requirements_count > 0,
            "attention_16d_weights_injected": bool(weights_norm),
            "custom_instructions_injected": runtime_custom_requirements_count > 0
            or bool(context_text.strip()),
            "materials_quality_gate_passed": gate_passed,
        },
        "materials_count": materials_count,
        "material_type_counts": material_type_counts,
        "materials_total_parsed_chars": total_parsed_chars,
        "materials_parse_fail_ratio": round(parse_fail_ratio, 4),
        "project_context_chars": len(context_text),
        "anchors_count": int(anchors_count),
        "requirements_count": int(base_requirements_count),
        "runtime_custom_requirements_count": int(runtime_custom_requirements_count),
        "weights_source": _infer_weights_source(project_id, profile_snapshot),
        "weights_sum": round(sum(float(weights_norm.get(d, 0.0)) for d in DIMENSION_IDS), 6),
        "constraints_rebuilt": bool(constraints_rebuilt),
        "runtime_instruction_breakdown": runtime_req_meta,
        "material_gate": gate if isinstance(gate, dict) else {},
    }


def _build_material_utilization_summary(
    report: Dict[str, object],
    runtime_req_meta: Dict[str, object],
) -> Dict[str, object]:
    req_hits = (
        report.get("requirement_hits") if isinstance(report.get("requirement_hits"), list) else []
    )
    req_hits = req_hits if isinstance(req_hits, list) else []
    by_type: Dict[str, Dict[str, object]] = {}
    retrieval_total = 0
    retrieval_hit = 0
    consistency_total = 0
    consistency_hit = 0
    fallback_total = 0
    fallback_hit = 0
    retrieval_file_total_set: set[str] = set()
    retrieval_file_hit_set: set[str] = set()
    retrieval_total_via_counts: Dict[str, int] = {}
    retrieval_hit_via_counts: Dict[str, int] = {}

    def _bucket(material_type: str) -> Dict[str, object]:
        if material_type not in by_type:
            by_type[material_type] = {
                "retrieval_total": 0,
                "retrieval_hit": 0,
                "consistency_total": 0,
                "consistency_hit": 0,
                "fallback_total": 0,
                "fallback_hit": 0,
                "sample_labels": [],
            }
        return by_type[material_type]

    retrieval_files_raw = runtime_req_meta.get("material_retrieval_selected_filenames")
    if isinstance(retrieval_files_raw, list):
        for item in retrieval_files_raw:
            filename_text = str(item or "").strip()
            if filename_text:
                retrieval_file_total_set.add(filename_text)
    selected_via_raw = runtime_req_meta.get("material_retrieval_selected_via_counts")
    retrieval_selected_via_counts: Dict[str, int] = {}
    if isinstance(selected_via_raw, dict):
        for key, value in selected_via_raw.items():
            mode = str(key or "").strip() or "unknown"
            count = int(_to_float_or_none(value) or 0)
            if count > 0:
                retrieval_selected_via_counts[mode] = (
                    int(retrieval_selected_via_counts.get(mode, 0)) + count
                )

    for row in req_hits:
        if not isinstance(row, dict):
            continue
        source_pack_id = str(row.get("source_pack_id") or "")
        if source_pack_id not in {"runtime_material_rag", "runtime_material_consistency"}:
            continue
        material_type = _normalize_material_type(row.get("material_type"))
        hit = bool(row.get("hit"))
        source_mode = str(row.get("source_mode") or "").strip()
        bucket = _bucket(material_type)
        source_filename = str(row.get("source_filename") or "").strip()
        if not source_filename:
            chunk_id = str(row.get("chunk_id") or "").strip()
            if "#c" in chunk_id:
                source_filename = chunk_id.split("#c", 1)[0].strip()

        if source_pack_id == "runtime_material_rag":
            retrieval_total += 1
            bucket["retrieval_total"] = int(bucket.get("retrieval_total", 0)) + 1
            via_key = source_mode or "unknown"
            retrieval_total_via_counts[via_key] = (
                int(retrieval_total_via_counts.get(via_key, 0)) + 1
            )
            if source_filename:
                retrieval_file_total_set.add(source_filename)
            if hit:
                retrieval_hit += 1
                bucket["retrieval_hit"] = int(bucket.get("retrieval_hit", 0)) + 1
                retrieval_hit_via_counts[via_key] = (
                    int(retrieval_hit_via_counts.get(via_key, 0)) + 1
                )
                if source_filename:
                    retrieval_file_hit_set.add(source_filename)
        else:
            consistency_total += 1
            bucket["consistency_total"] = int(bucket.get("consistency_total", 0)) + 1
            if hit:
                consistency_hit += 1
                bucket["consistency_hit"] = int(bucket.get("consistency_hit", 0)) + 1

        if source_mode == "fallback_keywords":
            bucket["fallback_total"] = int(bucket.get("fallback_total", 0)) + 1
            fallback_total += 1
            if hit:
                bucket["fallback_hit"] = int(bucket.get("fallback_hit", 0)) + 1
                fallback_hit += 1

        label = str(row.get("label") or "").strip()
        if label:
            sample_labels = bucket.get("sample_labels")
            if (
                isinstance(sample_labels, list)
                and label not in sample_labels
                and len(sample_labels) < 3
            ):
                sample_labels.append(label)

    available_types_raw = runtime_req_meta.get("material_available_types")
    available_types = (
        [_normalize_material_type(x) for x in available_types_raw]
        if isinstance(available_types_raw, list)
        else []
    )
    uncovered_types = []
    for material_type in available_types:
        bucket = by_type.get(material_type) or {}
        any_hit = int(bucket.get("retrieval_hit", 0)) + int(bucket.get("consistency_hit", 0))
        if any_hit <= 0:
            uncovered_types.append(material_type)

    def _rate(hit_cnt: int, total_cnt: int) -> Optional[float]:
        if total_cnt <= 0:
            return None
        return round(float(hit_cnt) / float(total_cnt), 4)

    unhit_files = sorted(retrieval_file_total_set - retrieval_file_hit_set)

    return {
        "retrieval_total": retrieval_total,
        "retrieval_hit": retrieval_hit,
        "retrieval_hit_rate": _rate(retrieval_hit, retrieval_total),
        "retrieval_file_total": len(retrieval_file_total_set),
        "retrieval_file_hit": len(retrieval_file_hit_set),
        "retrieval_file_coverage_rate": _rate(
            len(retrieval_file_hit_set),
            len(retrieval_file_total_set),
        ),
        "retrieval_selected_filenames": sorted(retrieval_file_total_set)[:120],
        "retrieval_hit_filenames": sorted(retrieval_file_hit_set)[:120],
        "retrieval_unhit_filenames": unhit_files[:120],
        "retrieval_unhit_file_count": len(unhit_files),
        "retrieval_top_k": int(
            _to_float_or_none(runtime_req_meta.get("material_retrieval_top_k")) or 0
        ),
        "retrieval_per_type_quota": int(
            _to_float_or_none(runtime_req_meta.get("material_retrieval_per_type_quota")) or 0
        ),
        "retrieval_per_file_quota": int(
            _to_float_or_none(runtime_req_meta.get("material_retrieval_per_file_quota")) or 0
        ),
        "retrieval_base_top_k": int(
            _to_float_or_none(runtime_req_meta.get("material_retrieval_base_top_k")) or 0
        ),
        "retrieval_base_per_type_quota": int(
            _to_float_or_none(runtime_req_meta.get("material_retrieval_base_per_type_quota")) or 0
        ),
        "retrieval_base_per_file_quota": int(
            _to_float_or_none(runtime_req_meta.get("material_retrieval_base_per_file_quota")) or 0
        ),
        "retrieval_budget_reasons": list(
            runtime_req_meta.get("material_retrieval_budget_reasons")
            if isinstance(runtime_req_meta.get("material_retrieval_budget_reasons"), list)
            else []
        ),
        "material_total_size_mb": float(
            _to_float_or_none(runtime_req_meta.get("material_total_size_mb")) or 0.0
        ),
        "material_type_count": int(
            _to_float_or_none(runtime_req_meta.get("material_type_count")) or 0
        ),
        "material_file_count": int(
            _to_float_or_none(runtime_req_meta.get("material_available_files")) or 0
        ),
        "retrieval_selected_via_counts": retrieval_selected_via_counts,
        "retrieval_total_via_counts": retrieval_total_via_counts,
        "retrieval_hit_via_counts": retrieval_hit_via_counts,
        "consistency_total": consistency_total,
        "consistency_hit": consistency_hit,
        "consistency_hit_rate": _rate(consistency_hit, consistency_total),
        "fallback_total": fallback_total,
        "fallback_hit": fallback_hit,
        "fallback_hit_rate": _rate(fallback_hit, fallback_total),
        "query_terms_count": int(
            _to_float_or_none(runtime_req_meta.get("material_query_terms_count")) or 0
        ),
        "query_numeric_terms_count": int(
            _to_float_or_none(runtime_req_meta.get("material_query_numeric_terms_count")) or 0
        ),
        "by_type": by_type,
        "available_types": available_types,
        "uncovered_types": uncovered_types,
    }


def _build_evidence_trace_summary(report: Dict[str, object]) -> Dict[str, object]:
    req_hits_raw = report.get("requirement_hits")
    req_hits = req_hits_raw if isinstance(req_hits_raw, list) else []
    total_requirements = 0
    total_hits = 0
    mandatory_total = 0
    mandatory_hit = 0
    runtime_hits = 0
    source_pack_counter: Counter[str] = Counter()
    source_file_hits: List[str] = []
    preview_rows: List[Dict[str, object]] = []

    for item in req_hits:
        if not isinstance(item, dict):
            continue
        total_requirements += 1
        mandatory = bool(item.get("mandatory"))
        if mandatory:
            mandatory_total += 1
        if not bool(item.get("hit")):
            continue
        total_hits += 1
        if mandatory:
            mandatory_hit += 1

        source_pack = str(item.get("source_pack_id") or "").strip() or "unknown"
        source_pack_counter[source_pack] += 1
        if source_pack in {"runtime_material_rag", "runtime_material_consistency"}:
            runtime_hits += 1

        source_filename = str(item.get("source_filename") or "").strip()
        if not source_filename:
            chunk_id = str(item.get("chunk_id") or "").strip()
            if "#c" in chunk_id:
                source_filename = chunk_id.split("#c", 1)[0].strip()
        if source_filename and source_filename not in source_file_hits:
            source_file_hits.append(source_filename)

        if len(preview_rows) < 16:
            preview_rows.append(
                {
                    "dimension_id": str(item.get("dimension_id") or ""),
                    "label": str(item.get("label") or ""),
                    "reason": str(item.get("reason") or ""),
                    "mandatory": mandatory,
                    "source_pack_id": source_pack,
                    "material_type": str(item.get("material_type") or ""),
                    "source_filename": source_filename,
                    "chunk_id": str(item.get("chunk_id") or ""),
                }
            )

    mandatory_hit_rate = (
        round(float(mandatory_hit) / float(mandatory_total), 4) if mandatory_total > 0 else None
    )
    overall_hit_rate = (
        round(float(total_hits) / float(total_requirements), 4) if total_requirements > 0 else None
    )
    return {
        "total_requirements": total_requirements,
        "total_hits": total_hits,
        "overall_hit_rate": overall_hit_rate,
        "mandatory_total": mandatory_total,
        "mandatory_hit": mandatory_hit,
        "mandatory_hit_rate": mandatory_hit_rate,
        "runtime_material_hits": runtime_hits,
        "source_files_hit": source_file_hits[:120],
        "source_files_hit_count": len(source_file_hits),
        "source_pack_hit_counts": dict(source_pack_counter),
        "preview": preview_rows,
    }


def _parse_reason_ratio(reason: str, marker: str) -> tuple[int, int]:
    text = str(reason or "")
    m = re.search(rf"{re.escape(marker)}(\d+)/(\d+)", text)
    if not m:
        return 0, 0
    try:
        left = int(m.group(1))
        right = int(m.group(2))
        return max(0, left), max(0, right)
    except Exception:
        return 0, 0


def _build_material_conflict_summary_from_report(report: Dict[str, object]) -> Dict[str, object]:
    req_hits = (
        report.get("requirement_hits") if isinstance(report.get("requirement_hits"), list) else []
    )
    consistency = (
        report.get("material_consistency")
        if isinstance(report.get("material_consistency"), dict)
        else {}
    )
    by_type_raw = (
        consistency.get("by_material_type")
        if isinstance(consistency.get("by_material_type"), dict)
        else {}
    )

    by_type: Dict[str, Dict[str, object]] = {}
    for key, row in by_type_raw.items():
        mat_type = _normalize_material_type(key)
        row_dict = row if isinstance(row, dict) else {}
        by_type[mat_type] = {
            "material_type": mat_type,
            "material_type_label": _material_type_label(mat_type),
            "total": int(_to_float_or_none(row_dict.get("total")) or 0),
            "hit": int(_to_float_or_none(row_dict.get("hit")) or 0),
            "mandatory_total": int(_to_float_or_none(row_dict.get("mandatory_total")) or 0),
            "mandatory_hit": int(_to_float_or_none(row_dict.get("mandatory_hit")) or 0),
            "hit_rate": _to_float_or_none(row_dict.get("hit_rate")),
        }

    conflicts: List[Dict[str, object]] = []
    for item in req_hits:
        if not isinstance(item, dict):
            continue
        if bool(item.get("hit")):
            continue
        source_pack = str(item.get("source_pack_id") or "")
        if source_pack != "runtime_material_consistency":
            continue
        mat_type = _normalize_material_type(item.get("material_type"))
        bucket = by_type.setdefault(
            mat_type,
            {
                "material_type": mat_type,
                "material_type_label": _material_type_label(mat_type),
                "total": 0,
                "hit": 0,
                "mandatory_total": 0,
                "mandatory_hit": 0,
                "hit_rate": None,
            },
        )
        bucket["total"] = int(bucket.get("total", 0)) + 1
        mandatory = bool(item.get("mandatory"))
        if mandatory:
            bucket["mandatory_total"] = int(bucket.get("mandatory_total", 0)) + 1

        reason = str(item.get("reason") or "")
        term_hit, term_need = _parse_reason_ratio(reason, "t")
        num_hit, num_need = _parse_reason_ratio(reason, "n")
        numeric_conflict = num_need > 0 and num_hit < num_need
        term_conflict = term_need > 0 and term_hit < term_need
        severity = "high" if (mandatory or numeric_conflict) else "medium"
        conflict_kind = (
            "numeric_mismatch"
            if numeric_conflict
            else ("term_coverage_missing" if term_conflict else "material_consistency_missing")
        )
        conflicts.append(
            {
                "severity": severity,
                "conflict_kind": conflict_kind,
                "dimension_id": str(item.get("dimension_id") or ""),
                "material_type": mat_type,
                "material_type_label": _material_type_label(mat_type),
                "label": str(item.get("label") or ""),
                "reason": reason,
                "mandatory": mandatory,
                "source_filename": str(item.get("source_filename") or ""),
                "chunk_id": str(item.get("chunk_id") or ""),
                "source_mode": str(item.get("source_mode") or ""),
                "term_hit": term_hit,
                "term_need": term_need,
                "num_hit": num_hit,
                "num_need": num_need,
            }
        )

    for row in by_type.values():
        total = int(row.get("total", 0))
        hit = int(row.get("hit", 0))
        if total > 0 and row.get("hit_rate") is None:
            row["hit_rate"] = round(float(hit) / float(total), 4)

    high_cnt = sum(1 for x in conflicts if str(x.get("severity")) == "high")
    medium_cnt = sum(1 for x in conflicts if str(x.get("severity")) == "medium")
    recommendations: List[str] = []
    if conflicts:
        recommendations.append("施组与上传资料存在一致性缺口，建议按冲突项逐条补齐。")
    for row in sorted(
        by_type.values(),
        key=lambda x: float(_to_float_or_none(x.get("hit_rate")) or 0.0),
    ):
        hit_rate = _to_float_or_none(row.get("hit_rate"))
        total = int(row.get("total", 0))
        if hit_rate is None or total <= 0:
            continue
        if hit_rate < 0.35:
            recommendations.append(
                f"{row.get('material_type_label')}一致性命中率偏低（{hit_rate * 100:.1f}%），建议补充明确的量化约束与章节引用。"
            )

    return {
        "has_conflicts": bool(conflicts),
        "conflict_count": len(conflicts),
        "high_severity_count": high_cnt,
        "medium_severity_count": medium_cnt,
        "by_material_type": sorted(
            list(by_type.values()),
            key=lambda x: str(x.get("material_type") or ""),
        ),
        "conflicts": conflicts[:60],
        "recommendations": recommendations[:10],
    }


def _build_submission_material_conflicts(
    *,
    project_id: str,
    submission: Dict[str, object],
) -> Dict[str, object]:
    report = submission.get("report") if isinstance(submission.get("report"), dict) else {}
    base = _build_material_conflict_summary_from_report(report)
    submission_text = str(submission.get("text") or "")
    submission_numbers = {
        token
        for token in (
            _normalize_numeric_token(item)
            for item in _extract_numeric_terms(submission_text, max_terms=320)
        )
        if token
    }

    material_retrieval = (
        (report.get("meta") or {}).get("material_retrieval")
        if isinstance((report.get("meta") or {}), dict)
        else {}
    )
    consistency_preview = (
        material_retrieval.get("consistency_preview")
        if isinstance(material_retrieval, dict)
        and isinstance(material_retrieval.get("consistency_preview"), list)
        else []
    )
    key_material_numbers: List[str] = []
    for row in consistency_preview:
        if not isinstance(row, dict):
            continue
        for raw in row.get("numbers") or []:
            token = _normalize_numeric_token(raw)
            if token and token not in key_material_numbers:
                key_material_numbers.append(token)

    merged_material_text = _merge_materials_text(project_id)
    material_numbers = {
        token
        for token in (
            _normalize_numeric_token(item)
            for item in _extract_numeric_terms(merged_material_text, max_terms=420)
        )
        if token
    }
    if not key_material_numbers:
        key_material_numbers = sorted(list(material_numbers))[:40]
    missing_key_numbers = [n for n in key_material_numbers if n not in submission_numbers][:30]

    numeric_base = len(key_material_numbers)
    numeric_hit = len([n for n in key_material_numbers if n in submission_numbers])
    numeric_coverage_rate = (
        round(float(numeric_hit) / float(numeric_base), 4) if numeric_base > 0 else None
    )
    numeric_conflict = numeric_base > 0 and numeric_hit < max(1, int(numeric_base * 0.25))
    if numeric_conflict:
        conflicts = base.get("conflicts") if isinstance(base.get("conflicts"), list) else []
        conflicts.append(
            {
                "severity": "high",
                "conflict_kind": "key_numeric_constraint_missing",
                "dimension_id": "",
                "material_type": "cross_material",
                "material_type_label": "跨资料",
                "label": "关键数字约束未在施组中体现",
                "reason": f"关键数字命中 {numeric_hit}/{numeric_base}",
                "mandatory": True,
                "source_filename": "",
                "chunk_id": "",
                "source_mode": "numeric_cross_check",
                "term_hit": 0,
                "term_need": 0,
                "num_hit": numeric_hit,
                "num_need": numeric_base,
                "missing_numbers": missing_key_numbers[:12],
            }
        )
        base["conflicts"] = conflicts[:60]
        base["conflict_count"] = len(base["conflicts"])
        base["high_severity_count"] = int(base.get("high_severity_count", 0)) + 1

    recommendations = (
        base.get("recommendations") if isinstance(base.get("recommendations"), list) else []
    )
    if numeric_coverage_rate is not None and numeric_coverage_rate < 0.35:
        recommendations.append(
            f"关键数字约束覆盖率偏低（{numeric_coverage_rate * 100:.1f}%），建议在施组中补充工期/工程量/阈值等量化指标。"
        )
    base["recommendations"] = recommendations[:12]
    base["numeric_cross_check"] = {
        "key_numbers_total": numeric_base,
        "key_numbers_hit": numeric_hit,
        "key_numbers_coverage_rate": numeric_coverage_rate,
        "missing_key_numbers": missing_key_numbers,
        "submission_number_count": len(submission_numbers),
        "material_number_count": len(material_numbers),
    }
    return base


def _build_submission_evidence_trace_report(
    *,
    project_id: str,
    submission: Dict[str, object],
) -> Dict[str, object]:
    submission_id = str(submission.get("id") or "")
    filename = str(submission.get("filename") or "")
    report = submission.get("report") if isinstance(submission.get("report"), dict) else {}
    report_meta = report.get("meta") if isinstance(report.get("meta"), dict) else {}
    summary = (
        report_meta.get("evidence_trace")
        if isinstance(report_meta.get("evidence_trace"), dict)
        else _build_evidence_trace_summary(report)
    )
    req_hits = (
        report.get("requirement_hits") if isinstance(report.get("requirement_hits"), list) else []
    )

    by_dim_map: Dict[str, Dict[str, object]] = {}
    requirement_rows: List[Dict[str, object]] = []
    for item in req_hits:
        if not isinstance(item, dict):
            continue
        dim_id = str(item.get("dimension_id") or "")
        if dim_id:
            bucket = by_dim_map.setdefault(
                dim_id,
                {
                    "dimension_id": dim_id,
                    "dimension_name": str((DIMENSIONS.get(dim_id) or {}).get("name") or dim_id),
                    "total": 0,
                    "hit": 0,
                    "mandatory_total": 0,
                    "mandatory_hit": 0,
                },
            )
            bucket["total"] = int(bucket.get("total", 0)) + 1
            if bool(item.get("hit")):
                bucket["hit"] = int(bucket.get("hit", 0)) + 1
            if bool(item.get("mandatory")):
                bucket["mandatory_total"] = int(bucket.get("mandatory_total", 0)) + 1
                if bool(item.get("hit")):
                    bucket["mandatory_hit"] = int(bucket.get("mandatory_hit", 0)) + 1
        requirement_rows.append(
            {
                "dimension_id": dim_id,
                "dimension_name": str((DIMENSIONS.get(dim_id) or {}).get("name") or dim_id),
                "label": str(item.get("label") or ""),
                "hit": bool(item.get("hit")),
                "mandatory": bool(item.get("mandatory")),
                "reason": str(item.get("reason") or ""),
                "source_pack_id": str(item.get("source_pack_id") or ""),
                "material_type": str(item.get("material_type") or ""),
                "source_filename": str(item.get("source_filename") or ""),
                "chunk_id": str(item.get("chunk_id") or ""),
                "source_mode": str(item.get("source_mode") or ""),
            }
        )

    by_dimension_rows: List[Dict[str, object]] = []
    for dim_id, row in by_dim_map.items():
        total = int(row.get("total", 0))
        hit = int(row.get("hit", 0))
        mandatory_total = int(row.get("mandatory_total", 0))
        mandatory_hit = int(row.get("mandatory_hit", 0))
        by_dimension_rows.append(
            {
                **row,
                "hit_rate": round(float(hit) / float(total), 4) if total > 0 else None,
                "mandatory_hit_rate": round(float(mandatory_hit) / float(mandatory_total), 4)
                if mandatory_total > 0
                else None,
            }
        )
    by_dimension_rows.sort(key=lambda x: str(x.get("dimension_id") or ""))

    evidence_units_rows: List[Dict[str, object]] = []
    for unit in load_evidence_units():
        if str(unit.get("submission_id") or "") != submission_id:
            continue
        evidence_units_rows.append(
            {
                "id": str(unit.get("id") or ""),
                "dimension_id": str(unit.get("dimension_id") or ""),
                "dimension_name": str(
                    (DIMENSIONS.get(str(unit.get("dimension_id") or "")) or {}).get("name")
                    or str(unit.get("dimension_id") or "")
                ),
                "source_locator": str(
                    unit.get("source_locator")
                    or unit.get("locator")
                    or unit.get("anchor_locator")
                    or ""
                ),
                "source_filename": str(unit.get("source_filename") or ""),
                "confidence": _to_float_or_none(unit.get("confidence")),
                "text_snippet": str(
                    unit.get("text_snippet") or unit.get("text") or unit.get("unit_text") or ""
                )[:220],
            }
        )
    evidence_units_rows.sort(
        key=lambda x: float(_to_float_or_none(x.get("confidence")) or 0.0),
        reverse=True,
    )

    material_conflicts = _build_submission_material_conflicts(
        project_id=project_id,
        submission=submission,
    )
    recommendations: List[str] = []
    total_hits = int(_to_float_or_none(summary.get("total_hits")) or 0)
    total_reqs = int(_to_float_or_none(summary.get("total_requirements")) or 0)
    if total_reqs > 0 and total_hits <= 0:
        recommendations.append("当前评分未命中有效证据锚点，建议补充与资料一致的可检索表述。")
    if bool(material_conflicts.get("has_conflicts")):
        recommendations.extend(
            [str(x) for x in (material_conflicts.get("recommendations") or []) if str(x).strip()]
        )

    return {
        "project_id": project_id,
        "submission_id": submission_id,
        "filename": filename,
        "generated_at": _now_iso(),
        "summary": summary,
        "by_dimension": by_dimension_rows,
        "requirement_hits": requirement_rows[:180],
        "evidence_units": evidence_units_rows[:120],
        "material_conflicts": material_conflicts,
        "recommendations": recommendations[:16],
    }


def _build_submission_scoring_basis_report(
    *,
    project_id: str,
    submission: Dict[str, object],
) -> Dict[str, object]:
    """构建评分依据审计：展示评分时注入的输入与资料命中链路。"""
    submission_id = str(submission.get("id") or "")
    filename = str(submission.get("filename") or "")
    report = submission.get("report") if isinstance(submission.get("report"), dict) else {}
    meta = report.get("meta") if isinstance(report.get("meta"), dict) else {}
    input_injection = (
        meta.get("input_injection") if isinstance(meta.get("input_injection"), dict) else {}
    )
    material_quality = (
        meta.get("material_quality") if isinstance(meta.get("material_quality"), dict) else {}
    )
    if not material_quality:
        material_quality = _build_material_quality_snapshot(project_id)
    material_retrieval = (
        meta.get("material_retrieval") if isinstance(meta.get("material_retrieval"), dict) else {}
    )
    material_utilization = (
        meta.get("material_utilization")
        if isinstance(meta.get("material_utilization"), dict)
        else {}
    )
    material_utilization_gate = (
        meta.get("material_utilization_gate")
        if isinstance(meta.get("material_utilization_gate"), dict)
        else {}
    )
    evidence_trace = (
        meta.get("evidence_trace") if isinstance(meta.get("evidence_trace"), dict) else {}
    )
    if not evidence_trace:
        evidence_trace = _build_evidence_trace_summary(report)

    recommendations: List[str] = []
    mece_inputs = (
        input_injection.get("mece_inputs")
        if isinstance(input_injection.get("mece_inputs"), dict)
        else {}
    )
    if mece_inputs and not bool(mece_inputs.get("materials_quality_gate_passed", True)):
        recommendations.append("资料门禁未通过：建议先完成“3) 项目资料”整改后再评分。")
    if material_utilization_gate:
        for reason in material_utilization_gate.get("reasons") or []:
            reason_text = str(reason).strip()
            if reason_text:
                recommendations.append(reason_text)
    if (_to_float_or_none(evidence_trace.get("total_requirements")) or 0) > 0 and (
        _to_float_or_none(evidence_trace.get("total_hits")) or 0
    ) <= 0:
        recommendations.append("评分未命中任何资料证据：请补充与清单/图纸/答疑一致的量化约束。")

    deduped_recommendations: List[str] = []
    for item in recommendations:
        text = str(item or "").strip()
        if text and text not in deduped_recommendations:
            deduped_recommendations.append(text)

    return {
        "project_id": project_id,
        "submission_id": submission_id,
        "filename": filename,
        "generated_at": _now_iso(),
        "scoring_status": str(report.get("scoring_status") or "unknown"),
        "mece_inputs": mece_inputs,
        "material_quality": material_quality,
        "material_retrieval": material_retrieval,
        "material_utilization": material_utilization,
        "material_utilization_gate": material_utilization_gate,
        "evidence_trace": evidence_trace,
        "recommendations": deduped_recommendations[:16],
    }


def _render_evidence_trace_markdown(payload: Dict[str, object]) -> str:
    summary = payload.get("summary") if isinstance(payload.get("summary"), dict) else {}
    by_dimension = (
        payload.get("by_dimension") if isinstance(payload.get("by_dimension"), list) else []
    )
    requirement_hits = (
        payload.get("requirement_hits") if isinstance(payload.get("requirement_hits"), list) else []
    )
    material_conflicts = (
        payload.get("material_conflicts")
        if isinstance(payload.get("material_conflicts"), dict)
        else {}
    )
    recommendations = (
        payload.get("recommendations") if isinstance(payload.get("recommendations"), list) else []
    )
    lines = [
        "# 评分证据追溯报告",
        "",
        f"- 项目ID：`{payload.get('project_id') or '-'}`",
        f"- 施组ID：`{payload.get('submission_id') or '-'}`",
        f"- 文件名：`{payload.get('filename') or '-'}`",
        f"- 生成时间：`{payload.get('generated_at') or '-'}`",
        "",
        "## 证据总览",
        "",
        f"- 要求总数：`{summary.get('total_requirements', 0)}`",
        f"- 命中总数：`{summary.get('total_hits', 0)}`",
        f"- 命中率：`{summary.get('overall_hit_rate')}`",
        f"- 强制项命中率：`{summary.get('mandatory_hit_rate')}`",
        f"- 命中文件数：`{summary.get('source_files_hit_count', 0)}`",
        "",
        "## 按维度命中",
        "",
        "| 维度 | total | hit | mandatory hit | hit_rate |",
        "|---|---:|---:|---:|---:|",
    ]
    for row in by_dimension:
        if not isinstance(row, dict):
            continue
        lines.append(
            f"| {row.get('dimension_id', '-')} {row.get('dimension_name', '')} "
            f"| {row.get('total', 0)} "
            f"| {row.get('hit', 0)} "
            f"| {row.get('mandatory_hit', 0)}/{row.get('mandatory_total', 0)} "
            f"| {row.get('hit_rate', '-') if row.get('hit_rate') is not None else '-'} |"
        )
    lines.extend(["", "## 一致性冲突", ""])
    lines.append(f"- 是否存在冲突：`{bool(material_conflicts.get('has_conflicts'))}`")
    lines.append(f"- 冲突数量：`{material_conflicts.get('conflict_count', 0)}`")
    numeric_cross_check = (
        material_conflicts.get("numeric_cross_check")
        if isinstance(material_conflicts.get("numeric_cross_check"), dict)
        else {}
    )
    if numeric_cross_check:
        lines.append(
            "- 关键数字命中：`"
            + f"{numeric_cross_check.get('key_numbers_hit', 0)}/{numeric_cross_check.get('key_numbers_total', 0)}"
            + "`"
        )

    lines.extend(["", "## 未命中要求（Top30）", ""])
    lines.append("| 维度 | 标签 | 原因 | 来源文件 |")
    lines.append("|---|---|---|---|")
    miss_rows = [x for x in requirement_hits if isinstance(x, dict) and not bool(x.get("hit"))][:30]
    for row in miss_rows:
        lines.append(
            f"| {row.get('dimension_id', '-')} "
            f"| {str(row.get('label') or '').replace('|', ' ')} "
            f"| {str(row.get('reason') or '').replace('|', ' / ')} "
            f"| {row.get('source_filename') or '-'} |"
        )

    lines.extend(["", "## 建议动作", ""])
    if recommendations:
        for item in recommendations:
            text = str(item or "").strip()
            if text:
                lines.append(f"- {text}")
    else:
        lines.append("- 当前证据链条完整，无新增建议。")
    return "\n".join(lines).strip()


def _aggregate_material_utilization_summaries(
    summaries: List[Dict[str, object]],
) -> Dict[str, object]:
    """聚合多份施组评分的资料利用统计，便于前端直接展示。"""
    by_type: Dict[str, Dict[str, int]] = {}
    available_types: List[str] = []
    uncovered_types: List[str] = []
    retrieval_total = 0
    retrieval_hit = 0
    consistency_total = 0
    consistency_hit = 0
    fallback_total = 0
    fallback_hit = 0
    retrieval_file_total = 0
    retrieval_file_hit = 0
    retrieval_selected_filenames: set[str] = set()
    retrieval_hit_filenames: set[str] = set()
    retrieval_selected_via_counts: Dict[str, int] = {}
    retrieval_total_via_counts: Dict[str, int] = {}
    retrieval_hit_via_counts: Dict[str, int] = {}
    retrieval_top_k = 0
    retrieval_per_type_quota = 0
    retrieval_per_file_quota = 0
    retrieval_base_top_k = 0
    retrieval_base_per_type_quota = 0
    retrieval_base_per_file_quota = 0
    material_total_size_mb = 0.0
    material_type_count = 0
    material_file_count = 0
    retrieval_budget_reasons: List[str] = []
    query_terms_count = 0
    query_numeric_terms_count = 0

    def _rate(hit_cnt: int, total_cnt: int) -> Optional[float]:
        if total_cnt <= 0:
            return None
        return round(float(hit_cnt) / float(total_cnt), 4)

    def _ensure_bucket(material_type: str) -> Dict[str, int]:
        if material_type not in by_type:
            by_type[material_type] = {
                "retrieval_total": 0,
                "retrieval_hit": 0,
                "consistency_total": 0,
                "consistency_hit": 0,
                "fallback_total": 0,
                "fallback_hit": 0,
            }
        return by_type[material_type]

    for raw in summaries:
        if not isinstance(raw, dict):
            continue
        retrieval_total += int(_to_float_or_none(raw.get("retrieval_total")) or 0)
        retrieval_hit += int(_to_float_or_none(raw.get("retrieval_hit")) or 0)
        retrieval_file_total += int(_to_float_or_none(raw.get("retrieval_file_total")) or 0)
        retrieval_file_hit += int(_to_float_or_none(raw.get("retrieval_file_hit")) or 0)
        selected_files_raw = raw.get("retrieval_selected_filenames")
        if isinstance(selected_files_raw, list):
            for item in selected_files_raw:
                filename = str(item or "").strip()
                if filename:
                    retrieval_selected_filenames.add(filename)
        hit_files_raw = raw.get("retrieval_hit_filenames")
        if isinstance(hit_files_raw, list):
            for item in hit_files_raw:
                filename = str(item or "").strip()
                if filename:
                    retrieval_hit_filenames.add(filename)
        retrieval_top_k = max(
            retrieval_top_k,
            int(_to_float_or_none(raw.get("retrieval_top_k")) or 0),
        )
        retrieval_per_type_quota = max(
            retrieval_per_type_quota,
            int(_to_float_or_none(raw.get("retrieval_per_type_quota")) or 0),
        )
        retrieval_per_file_quota = max(
            retrieval_per_file_quota,
            int(_to_float_or_none(raw.get("retrieval_per_file_quota")) or 0),
        )
        retrieval_base_top_k = max(
            retrieval_base_top_k,
            int(_to_float_or_none(raw.get("retrieval_base_top_k")) or 0),
        )
        retrieval_base_per_type_quota = max(
            retrieval_base_per_type_quota,
            int(_to_float_or_none(raw.get("retrieval_base_per_type_quota")) or 0),
        )
        retrieval_base_per_file_quota = max(
            retrieval_base_per_file_quota,
            int(_to_float_or_none(raw.get("retrieval_base_per_file_quota")) or 0),
        )
        material_total_size_mb = max(
            material_total_size_mb,
            float(_to_float_or_none(raw.get("material_total_size_mb")) or 0.0),
        )
        material_type_count = max(
            material_type_count,
            int(_to_float_or_none(raw.get("material_type_count")) or 0),
        )
        material_file_count = max(
            material_file_count,
            int(_to_float_or_none(raw.get("material_file_count")) or 0),
        )
        reasons_raw = raw.get("retrieval_budget_reasons")
        if isinstance(reasons_raw, list):
            for item in reasons_raw:
                text = str(item or "").strip()
                if text and text not in retrieval_budget_reasons:
                    retrieval_budget_reasons.append(text)
        selected_via_raw = raw.get("retrieval_selected_via_counts")
        if isinstance(selected_via_raw, dict):
            for key, value in selected_via_raw.items():
                mode = str(key or "").strip() or "unknown"
                retrieval_selected_via_counts[mode] = int(
                    retrieval_selected_via_counts.get(mode, 0)
                ) + int(_to_float_or_none(value) or 0)
        total_via_raw = raw.get("retrieval_total_via_counts")
        if isinstance(total_via_raw, dict):
            for key, value in total_via_raw.items():
                mode = str(key or "").strip() or "unknown"
                retrieval_total_via_counts[mode] = int(
                    retrieval_total_via_counts.get(mode, 0)
                ) + int(_to_float_or_none(value) or 0)
        hit_via_raw = raw.get("retrieval_hit_via_counts")
        if isinstance(hit_via_raw, dict):
            for key, value in hit_via_raw.items():
                mode = str(key or "").strip() or "unknown"
                retrieval_hit_via_counts[mode] = int(retrieval_hit_via_counts.get(mode, 0)) + int(
                    _to_float_or_none(value) or 0
                )
        query_terms_count += int(_to_float_or_none(raw.get("query_terms_count")) or 0)
        query_numeric_terms_count += int(
            _to_float_or_none(raw.get("query_numeric_terms_count")) or 0
        )
        consistency_total += int(_to_float_or_none(raw.get("consistency_total")) or 0)
        consistency_hit += int(_to_float_or_none(raw.get("consistency_hit")) or 0)

        raw_types = raw.get("available_types")
        if isinstance(raw_types, list):
            for item in raw_types:
                key = _normalize_material_type(item)
                if key not in available_types:
                    available_types.append(key)
                    _ensure_bucket(key)

        raw_uncovered = raw.get("uncovered_types")
        if isinstance(raw_uncovered, list):
            for item in raw_uncovered:
                key = _normalize_material_type(item)
                if key not in uncovered_types:
                    uncovered_types.append(key)

        raw_by_type = raw.get("by_type")
        if not isinstance(raw_by_type, dict):
            continue
        for mat_type_raw, row in raw_by_type.items():
            key = _normalize_material_type(mat_type_raw)
            bucket = _ensure_bucket(key)
            row_dict = row if isinstance(row, dict) else {}
            rt = int(_to_float_or_none(row_dict.get("retrieval_total")) or 0)
            rh = int(_to_float_or_none(row_dict.get("retrieval_hit")) or 0)
            ct = int(_to_float_or_none(row_dict.get("consistency_total")) or 0)
            ch = int(_to_float_or_none(row_dict.get("consistency_hit")) or 0)
            ft = int(_to_float_or_none(row_dict.get("fallback_total")) or 0)
            fh = int(_to_float_or_none(row_dict.get("fallback_hit")) or 0)
            bucket["retrieval_total"] += rt
            bucket["retrieval_hit"] += rh
            bucket["consistency_total"] += ct
            bucket["consistency_hit"] += ch
            bucket["fallback_total"] += ft
            bucket["fallback_hit"] += fh
            fallback_total += ft
            fallback_hit += fh

    normalized_by_type: Dict[str, Dict[str, object]] = {}
    for mat_type, row in by_type.items():
        normalized_by_type[mat_type] = {
            "retrieval_total": row["retrieval_total"],
            "retrieval_hit": row["retrieval_hit"],
            "retrieval_hit_rate": _rate(row["retrieval_hit"], row["retrieval_total"]),
            "consistency_total": row["consistency_total"],
            "consistency_hit": row["consistency_hit"],
            "consistency_hit_rate": _rate(row["consistency_hit"], row["consistency_total"]),
            "fallback_total": row["fallback_total"],
            "fallback_hit": row["fallback_hit"],
            "fallback_hit_rate": _rate(row["fallback_hit"], row["fallback_total"]),
        }

    retrieval_unhit_filenames = sorted(retrieval_selected_filenames - retrieval_hit_filenames)

    return {
        "retrieval_total": retrieval_total,
        "retrieval_hit": retrieval_hit,
        "retrieval_hit_rate": _rate(retrieval_hit, retrieval_total),
        "retrieval_file_total": retrieval_file_total,
        "retrieval_file_hit": retrieval_file_hit,
        "retrieval_file_coverage_rate": _rate(retrieval_file_hit, retrieval_file_total),
        "retrieval_selected_filenames": sorted(retrieval_selected_filenames)[:120],
        "retrieval_hit_filenames": sorted(retrieval_hit_filenames)[:120],
        "retrieval_unhit_filenames": retrieval_unhit_filenames[:120],
        "retrieval_unhit_file_count": len(retrieval_unhit_filenames),
        "retrieval_top_k": retrieval_top_k,
        "retrieval_per_type_quota": retrieval_per_type_quota,
        "retrieval_per_file_quota": retrieval_per_file_quota,
        "retrieval_base_top_k": retrieval_base_top_k,
        "retrieval_base_per_type_quota": retrieval_base_per_type_quota,
        "retrieval_base_per_file_quota": retrieval_base_per_file_quota,
        "retrieval_budget_reasons": retrieval_budget_reasons[:12],
        "material_total_size_mb": round(material_total_size_mb, 3),
        "material_type_count": material_type_count,
        "material_file_count": material_file_count,
        "retrieval_selected_via_counts": retrieval_selected_via_counts,
        "retrieval_total_via_counts": retrieval_total_via_counts,
        "retrieval_hit_via_counts": retrieval_hit_via_counts,
        "consistency_total": consistency_total,
        "consistency_hit": consistency_hit,
        "consistency_hit_rate": _rate(consistency_hit, consistency_total),
        "fallback_total": fallback_total,
        "fallback_hit": fallback_hit,
        "fallback_hit_rate": _rate(fallback_hit, fallback_total),
        "query_terms_count": query_terms_count,
        "query_numeric_terms_count": query_numeric_terms_count,
        "by_type": normalized_by_type,
        "available_types": available_types,
        "uncovered_types": uncovered_types,
    }


def _build_material_utilization_alerts(
    summary: Optional[Dict[str, object]],
    material_gate: Optional[Dict[str, object]] = None,
) -> List[str]:
    """将资料利用统计转换为可读预警，提醒评分依据是否充分。"""
    data = summary if isinstance(summary, dict) else {}
    alerts: List[str] = []

    retrieval_total = int(_to_float_or_none(data.get("retrieval_total")) or 0)
    retrieval_hit_rate = _to_float_or_none(data.get("retrieval_hit_rate"))
    consistency_total = int(_to_float_or_none(data.get("consistency_total")) or 0)
    consistency_hit_rate = _to_float_or_none(data.get("consistency_hit_rate"))
    fallback_total = int(_to_float_or_none(data.get("fallback_total")) or 0)
    fallback_hit = int(_to_float_or_none(data.get("fallback_hit")) or 0)
    retrieval_file_total = int(_to_float_or_none(data.get("retrieval_file_total")) or 0)
    retrieval_file_coverage_rate = _to_float_or_none(data.get("retrieval_file_coverage_rate"))
    retrieval_unhit_count = int(_to_float_or_none(data.get("retrieval_unhit_file_count")) or 0)
    selected_via_counts = (
        data.get("retrieval_selected_via_counts")
        if isinstance(data.get("retrieval_selected_via_counts"), dict)
        else {}
    )

    if retrieval_total <= 0:
        alerts.append("评分未命中任何资料检索锚点，资料引用不足。")
    elif retrieval_hit_rate is not None and retrieval_total >= 4 and retrieval_hit_rate < 0.35:
        alerts.append(
            f"资料检索命中率偏低（{retrieval_hit_rate * 100:.1f}%），建议补充可检索的量化表述。"
        )

    if consistency_total > 0 and consistency_hit_rate is not None and consistency_hit_rate < 0.35:
        alerts.append(
            f"跨资料一致性命中率偏低（{consistency_hit_rate * 100:.1f}%），请核对工期/质量/危大工程等关键约束。"
        )

    if fallback_total > 0 and fallback_hit <= 0:
        alerts.append("仅触发了关键词兜底匹配且未命中，说明施组与项目资料耦合不足。")
    if selected_via_counts:
        selected_total = sum(
            max(0, int(_to_float_or_none(v) or 0)) for v in selected_via_counts.values()
        )
        if selected_total > 0:
            backfill_count = int(
                _to_float_or_none(selected_via_counts.get("type_backfill")) or 0
            ) + int(_to_float_or_none(selected_via_counts.get("global_backfill")) or 0)
            backfill_rate = float(backfill_count) / float(selected_total)
            if backfill_rate >= 0.4:
                alerts.append(
                    "资料检索存在较高比例的回填命中，建议在施组中增加对清单/图纸/答疑的显式章节引用。"
                )

    if (
        retrieval_file_total >= 3
        and retrieval_file_coverage_rate is not None
        and retrieval_file_coverage_rate < 0.4
    ):
        alerts.append(
            "资料文件覆盖率偏低（"
            + f"{retrieval_file_coverage_rate * 100:.1f}%），可能仅引用了少量文件。"
        )
    if retrieval_unhit_count > 0 and retrieval_file_total >= 3:
        alerts.append(
            f"有 {retrieval_unhit_count} 份已检索资料未形成命中证据，建议补充对应章节引用。"
        )

    uncovered_raw = data.get("uncovered_types")
    uncovered_types = uncovered_raw if isinstance(uncovered_raw, list) else []
    if uncovered_types:
        labels = [
            _material_type_label(_normalize_material_type(item)) for item in uncovered_types[:4]
        ]
        alerts.append("以下资料类型未形成有效评分证据：" + "、".join(labels))

    gate = material_gate if isinstance(material_gate, dict) else {}
    if gate and not bool(gate.get("passed", True)):
        issues = gate.get("issues") if isinstance(gate.get("issues"), list) else []
        issue_text = "；".join(str(x) for x in issues[:2]) if issues else "资料门禁未通过"
        alerts.append("资料门禁警告：" + issue_text)

    deduped: List[str] = []
    for item in alerts:
        text = str(item or "").strip()
        if text and text not in deduped:
            deduped.append(text)
    return deduped[:6]


def _resolve_material_utilization_policy(project: Dict[str, object]) -> Dict[str, object]:
    meta = project.get("meta") if isinstance(project.get("meta"), dict) else {}

    def _to_rate(value: object, default: float) -> float:
        numeric = _to_float_or_none(value)
        if numeric is None:
            numeric = default
        return min(1.0, max(0.0, float(numeric)))

    def _to_nonneg_int(value: object, default: int) -> int:
        numeric = _to_float_or_none(value)
        if numeric is None:
            return int(default)
        return max(0, int(round(float(numeric))))

    raw_mode = str(
        meta.get("material_utilization_gate_mode", DEFAULT_MATERIAL_UTILIZATION_GATE_MODE) or ""
    ).strip()
    mode = "warn" if raw_mode.lower() == "warn" else "block"

    return {
        "enabled": bool(
            meta.get("enforce_material_utilization_gate", DEFAULT_ENFORCE_MATERIAL_UTILIZATION_GATE)
        ),
        "mode": mode,
        "enforce_uploaded_type_coverage": bool(
            meta.get("enforce_uploaded_type_coverage", DEFAULT_ENFORCE_UPLOADED_TYPE_COVERAGE)
        ),
        "min_retrieval_total": _to_nonneg_int(
            meta.get("min_material_retrieval_total"),
            DEFAULT_MIN_MATERIAL_RETRIEVAL_TOTAL,
        ),
        "min_retrieval_file_coverage_rate": _to_rate(
            meta.get("min_material_retrieval_file_coverage_rate"),
            DEFAULT_MIN_MATERIAL_RETRIEVAL_FILE_COVERAGE_RATE,
        ),
        "min_retrieval_hit_rate": _to_rate(
            meta.get("min_material_retrieval_hit_rate"),
            DEFAULT_MIN_MATERIAL_RETRIEVAL_HIT_RATE,
        ),
        "min_consistency_hit_rate": _to_rate(
            meta.get("min_material_consistency_hit_rate"),
            DEFAULT_MIN_MATERIAL_CONSISTENCY_HIT_RATE,
        ),
        "max_uncovered_required_types": _to_nonneg_int(
            meta.get("max_uncovered_required_types"),
            DEFAULT_MAX_UNCOVERED_REQUIRED_TYPES,
        ),
        "min_required_type_presence_rate": _to_rate(
            meta.get("min_required_type_presence_rate"),
            DEFAULT_MIN_REQUIRED_TYPE_PRESENCE_RATE,
        ),
        "min_required_type_coverage_rate": _to_rate(
            meta.get("min_required_type_coverage_rate"),
            DEFAULT_MIN_REQUIRED_TYPE_COVERAGE_RATE,
        ),
        "min_uploaded_type_coverage_rate": _to_rate(
            meta.get("min_uploaded_type_coverage_rate"),
            DEFAULT_MIN_UPLOADED_TYPE_COVERAGE_RATE,
        ),
    }


def _evaluate_material_utilization_gate(
    summary: Optional[Dict[str, object]],
    *,
    policy: Optional[Dict[str, object]] = None,
    required_types: Optional[List[str]] = None,
) -> Dict[str, object]:
    data = summary if isinstance(summary, dict) else {}
    gate_policy = policy if isinstance(policy, dict) else {}
    enabled = bool(gate_policy.get("enabled", DEFAULT_ENFORCE_MATERIAL_UTILIZATION_GATE))
    mode = str(gate_policy.get("mode") or DEFAULT_MATERIAL_UTILIZATION_GATE_MODE).strip().lower()
    mode = "warn" if mode == "warn" else "block"

    retrieval_total = int(_to_float_or_none(data.get("retrieval_total")) or 0)
    retrieval_hit_rate = _to_float_or_none(data.get("retrieval_hit_rate"))
    retrieval_file_total = int(_to_float_or_none(data.get("retrieval_file_total")) or 0)
    retrieval_file_coverage_rate = _to_float_or_none(data.get("retrieval_file_coverage_rate"))
    consistency_total = int(_to_float_or_none(data.get("consistency_total")) or 0)
    consistency_hit_rate = _to_float_or_none(data.get("consistency_hit_rate"))

    available_raw = data.get("available_types")
    available_types = (
        [_normalize_material_type(x) for x in available_raw]
        if isinstance(available_raw, list)
        else []
    )
    uncovered_raw = data.get("uncovered_types")
    uncovered_types = (
        [_normalize_material_type(x) for x in uncovered_raw]
        if isinstance(uncovered_raw, list)
        else []
    )

    required_source = required_types if isinstance(required_types, list) else []
    normalized_required = []
    for item in required_source:
        key = _normalize_material_type(item)
        if key and key not in normalized_required:
            normalized_required.append(key)
    if not normalized_required:
        normalized_required = list(DEFAULT_REQUIRED_MATERIAL_TYPES)

    missing_required_upload = [t for t in normalized_required if t not in available_types]
    required_present = [t for t in normalized_required if t in available_types]
    uncovered_required = [t for t in required_present if t in uncovered_types]
    covered_required = [t for t in required_present if t not in uncovered_required]
    required_presence_rate = (
        round(float(len(required_present)) / float(len(normalized_required)), 4)
        if normalized_required
        else None
    )
    required_coverage_rate = (
        round(float(len(covered_required)) / float(len(required_present)), 4)
        if required_present
        else None
    )
    uncovered_uploaded_types = [t for t in available_types if t in uncovered_types]
    covered_uploaded_types = [t for t in available_types if t not in uncovered_uploaded_types]
    uploaded_type_coverage_rate = (
        round(float(len(covered_uploaded_types)) / float(len(available_types)), 4)
        if available_types
        else None
    )

    min_retrieval_total = max(
        0,
        int(
            round(
                float(
                    _to_float_or_none(gate_policy.get("min_retrieval_total"))
                    or DEFAULT_MIN_MATERIAL_RETRIEVAL_TOTAL
                )
            )
        ),
    )

    min_retrieval = min(
        1.0,
        max(
            0.0,
            float(
                _to_float_or_none(gate_policy.get("min_retrieval_hit_rate"))
                or DEFAULT_MIN_MATERIAL_RETRIEVAL_HIT_RATE
            ),
        ),
    )
    min_consistency = min(
        1.0,
        max(
            0.0,
            float(
                _to_float_or_none(gate_policy.get("min_consistency_hit_rate"))
                or DEFAULT_MIN_MATERIAL_CONSISTENCY_HIT_RATE
            ),
        ),
    )
    max_uncovered = max(
        0,
        int(
            round(
                float(
                    _to_float_or_none(gate_policy.get("max_uncovered_required_types"))
                    or DEFAULT_MAX_UNCOVERED_REQUIRED_TYPES
                )
            )
        ),
    )
    min_required_coverage = min(
        1.0,
        max(
            0.0,
            float(
                _to_float_or_none(gate_policy.get("min_required_type_coverage_rate"))
                or DEFAULT_MIN_REQUIRED_TYPE_COVERAGE_RATE
            ),
        ),
    )
    min_required_presence = min(
        1.0,
        max(
            0.0,
            float(
                _to_float_or_none(gate_policy.get("min_required_type_presence_rate"))
                or DEFAULT_MIN_REQUIRED_TYPE_PRESENCE_RATE
            ),
        ),
    )
    min_retrieval_file_coverage = min(
        1.0,
        max(
            0.0,
            float(
                _to_float_or_none(gate_policy.get("min_retrieval_file_coverage_rate"))
                or DEFAULT_MIN_MATERIAL_RETRIEVAL_FILE_COVERAGE_RATE
            ),
        ),
    )
    enforce_uploaded_type_coverage = bool(
        gate_policy.get("enforce_uploaded_type_coverage", DEFAULT_ENFORCE_UPLOADED_TYPE_COVERAGE)
    )
    min_uploaded_type_coverage = min(
        1.0,
        max(
            0.0,
            float(
                _to_float_or_none(gate_policy.get("min_uploaded_type_coverage_rate"))
                or DEFAULT_MIN_UPLOADED_TYPE_COVERAGE_RATE
            ),
        ),
    )

    reasons: List[str] = []
    if retrieval_total < min_retrieval_total:
        reasons.append(f"资料检索证据数量 {retrieval_total} 低于阈值 {min_retrieval_total}")
    if (
        retrieval_file_total > 0
        and retrieval_file_coverage_rate is not None
        and retrieval_file_coverage_rate < min_retrieval_file_coverage
    ):
        reasons.append(
            "资料检索文件覆盖率 "
            + f"{retrieval_file_coverage_rate:.1%} 低于阈值 {min_retrieval_file_coverage:.1%}"
        )
    if (
        retrieval_total > 0
        and retrieval_hit_rate is not None
        and retrieval_hit_rate < min_retrieval
    ):
        reasons.append(f"资料检索命中率 {retrieval_hit_rate:.1%} 低于阈值 {min_retrieval:.1%}")
    if (
        consistency_total > 0
        and consistency_hit_rate is not None
        and consistency_hit_rate < min_consistency
    ):
        reasons.append(
            f"跨资料一致性命中率 {consistency_hit_rate:.1%} 低于阈值 {min_consistency:.1%}"
        )
    if len(uncovered_required) > max_uncovered:
        labels = "、".join(_material_type_label(x) for x in uncovered_required)
        reasons.append(f"关键资料未形成证据：{labels}（允许未覆盖 {max_uncovered} 类）")
    if required_presence_rate is not None and required_presence_rate < min_required_presence:
        labels = "、".join(_material_type_label(x) for x in missing_required_upload)
        reasons.append(
            "关键资料上传覆盖率 "
            + f"{required_presence_rate:.1%} 低于阈值 {min_required_presence:.1%}"
            + (f"，缺少：{labels}" if labels else "")
        )
    if required_coverage_rate is not None and required_coverage_rate < min_required_coverage:
        reasons.append(
            f"关键资料覆盖率 {required_coverage_rate:.1%} 低于阈值 {min_required_coverage:.1%}"
        )
    if (
        enforce_uploaded_type_coverage
        and uploaded_type_coverage_rate is not None
        and uploaded_type_coverage_rate < min_uploaded_type_coverage
    ):
        labels = "、".join(_material_type_label(x) for x in uncovered_uploaded_types)
        reasons.append(
            "已上传资料类型覆盖率 "
            + f"{uploaded_type_coverage_rate:.1%} 低于阈值 {min_uploaded_type_coverage:.1%}"
            + (f"，未形成证据类型：{labels}" if labels else "")
        )

    failed = bool(reasons)
    blocked = bool(enabled and failed and mode == "block")
    warned = bool(enabled and failed and mode != "block")
    if not enabled:
        level = "disabled"
    elif blocked:
        level = "blocked"
    elif warned:
        level = "warn"
    else:
        level = "pass"

    return {
        "enabled": enabled,
        "mode": mode,
        "passed": not failed,
        "blocked": blocked,
        "warned": warned,
        "level": level,
        "reasons": reasons,
        "thresholds": {
            "min_retrieval_total": min_retrieval_total,
            "min_retrieval_file_coverage_rate": min_retrieval_file_coverage,
            "min_retrieval_hit_rate": min_retrieval,
            "min_consistency_hit_rate": min_consistency,
            "max_uncovered_required_types": max_uncovered,
            "min_required_type_presence_rate": min_required_presence,
            "min_required_type_coverage_rate": min_required_coverage,
            "min_uploaded_type_coverage_rate": min_uploaded_type_coverage,
        },
        "required_types": normalized_required,
        "required_types_missing_upload": missing_required_upload,
        "required_type_presence_rate": required_presence_rate,
        "required_types_present": required_present,
        "covered_required_types": covered_required,
        "uncovered_required_types": uncovered_required,
        "required_type_coverage_rate": required_coverage_rate,
        "enforce_uploaded_type_coverage": enforce_uploaded_type_coverage,
        "uploaded_types": available_types,
        "covered_uploaded_types": covered_uploaded_types,
        "uncovered_uploaded_types": uncovered_uploaded_types,
        "uploaded_type_coverage_rate": uploaded_type_coverage_rate,
        "metrics": {
            "retrieval_total": retrieval_total,
            "retrieval_hit_rate": retrieval_hit_rate,
            "retrieval_file_total": retrieval_file_total,
            "retrieval_file_coverage_rate": retrieval_file_coverage_rate,
            "consistency_total": consistency_total,
            "consistency_hit_rate": consistency_hit_rate,
        },
    }


def _aggregate_material_utilization_gates(gates: List[Dict[str, object]]) -> Dict[str, object]:
    if not gates:
        return {
            "enabled": False,
            "mode": DEFAULT_MATERIAL_UTILIZATION_GATE_MODE,
            "blocked_submissions": 0,
            "warn_submissions": 0,
            "pass_submissions": 0,
            "failed_submissions": 0,
            "failed_filenames": [],
        }
    enabled = any(bool(g.get("enabled")) for g in gates if isinstance(g, dict))
    mode = "block"
    if all(str(g.get("mode", "")).lower() == "warn" for g in gates if isinstance(g, dict)):
        mode = "warn"
    blocked_submissions = 0
    warn_submissions = 0
    pass_submissions = 0
    failed_submissions = 0
    failed_filenames: List[str] = []
    for g in gates:
        if not isinstance(g, dict):
            continue
        level = str(g.get("level") or "").strip().lower()
        passed = bool(g.get("passed", False))
        if passed:
            pass_submissions += 1
        else:
            failed_submissions += 1
        if level == "blocked":
            blocked_submissions += 1
        elif level == "warn":
            warn_submissions += 1
    return {
        "enabled": enabled,
        "mode": mode,
        "blocked_submissions": blocked_submissions,
        "warn_submissions": warn_submissions,
        "pass_submissions": pass_submissions,
        "failed_submissions": failed_submissions,
        "failed_filenames": failed_filenames,
    }


def _get_rubric_dim_cfg(rubric_dimensions: Dict[str, object], dim_id: str) -> Dict[str, object]:
    for k, v in (rubric_dimensions or {}).items():
        if _normalize_dimension_id(str(k)) == dim_id and isinstance(v, dict):
            return v
    return {}


def _build_scoring_factors_overview(project_id: Optional[str]) -> Dict[str, object]:
    config = load_config()
    rubric = config.rubric if isinstance(config.rubric, dict) else {}
    rubric_dims = rubric.get("dimensions") if isinstance(rubric.get("dimensions"), dict) else {}
    rubric_penalties = rubric.get("penalties") if isinstance(rubric.get("penalties"), dict) else {}

    dimensions: List[Dict[str, object]] = []
    for dim_id in DIMENSION_IDS:
        dim_meta = DIMENSIONS.get(dim_id, {})
        cfg = _get_rubric_dim_cfg(rubric_dims, dim_id)
        sub_items_raw = cfg.get("sub_items") if isinstance(cfg.get("sub_items"), list) else []
        sub_items = []
        for item in sub_items_raw:
            if not isinstance(item, dict):
                continue
            sub_items.append(
                {
                    "id": item.get("id"),
                    "name": item.get("name"),
                    "weight": float(item.get("weight", 0)),
                    "keywords_count": len(item.get("keywords") or []),
                    "regex_count": len(item.get("regex") or []),
                }
            )

        if dim_id in {"07", "09"}:
            model_desc = "5子项累计评分（每项通常2分，合计10分）"
        else:
            model_desc = "Coverage/Closure/Landing/Specificity 四项各2.5分，合计10分"

        dimensions.append(
            {
                "dimension_id": dim_id,
                "name": dim_meta.get("name", f"维度{dim_id}"),
                "module": dim_meta.get("module", ""),
                "max_score": float(cfg.get("max_score", 10)),
                "suggestion_threshold": float(cfg.get("suggestion_threshold", 6)),
                "suggested_gain": float(cfg.get("suggested_gain", 0)),
                "scoring_model": model_desc,
                "sub_items": sub_items,
            }
        )

    penalty_rules: List[Dict[str, object]] = []
    empty_cfg = (
        rubric_penalties.get("empty_promises")
        if isinstance(rubric_penalties.get("empty_promises"), dict)
        else {}
    )
    action_cfg = (
        rubric_penalties.get("action_missing")
        if isinstance(rubric_penalties.get("action_missing"), dict)
        else {}
    )
    penalty_rules.extend(
        [
            {
                "code": "P-EMPTY-001",
                "engine": "v1",
                "description": "空泛承诺扣分",
                "deduct_per_hit": float(empty_cfg.get("deduct", 0.5)),
                "max_deduct": float(empty_cfg.get("max_deduct", 3.0)),
            },
            {
                "code": "P-ACTION-001",
                "engine": "v1",
                "description": "措施缺动作要素扣分",
                "deduct_per_hit": float(action_cfg.get("deduct_per", 0.5)),
                "max_deduct": float(action_cfg.get("max_deduct", 5.0)),
            },
            {
                "code": "P-EMPTY-002",
                "engine": "v2",
                "description": "空泛承诺未绑定证据（缺参数/频次/责任/验收）",
                "deduct_per_hit": 0.5,
                "max_deduct": 3.0,
            },
            {
                "code": "P-ACTION-002",
                "engine": "v2",
                "description": "措施缺硬要素（至少需要责任岗位+验收动作）",
                "deduct_per_hit": 0.8,
                "max_deduct": 6.0,
            },
            {
                "code": "P-CONSIST-001",
                "engine": "v2",
                "description": "与锚点不一致（如工期冲突）",
                "deduct_per_hit": 2.0,
                "max_deduct": 6.0,
            },
        ]
    )

    chapter_requirements = dict(DEFAULT_CHAPTER_REQUIREMENTS)
    chapter_source = "default"
    if project_id:
        evo = load_evolution_reports().get(project_id) or {}
        ci = (
            evo.get("compilation_instructions")
            if isinstance(evo.get("compilation_instructions"), dict)
            else {}
        )
        if ci:
            chapter_requirements = {
                "required_sections": [str(x) for x in (ci.get("required_sections") or [])],
                "required_charts_images": [
                    str(x) for x in (ci.get("required_charts_images") or [])
                ],
                "mandatory_elements": [str(x) for x in (ci.get("mandatory_elements") or [])],
                "forbidden_patterns": [str(x) for x in (ci.get("forbidden_patterns") or [])],
            }
            chapter_source = "project_evolution"

    if project_id:
        anchors = [a for a in load_project_anchors() if str(a.get("project_id")) == project_id]
        consistency_anchors = sorted(
            {str(a.get("anchor_key")) for a in anchors if a.get("anchor_key")}
        )
    else:
        consistency_anchors = []
    if not consistency_anchors:
        consistency_anchors = [
            "contract_duration_days",
            "quality_standard",
            "dangerous_works_list",
            "key_milestones",
        ]

    req_sections = chapter_requirements.get("required_sections", [])
    req_charts = chapter_requirements.get("required_charts_images", [])
    req_elements = chapter_requirements.get("mandatory_elements", [])
    organization_markers = req_sections + req_charts
    capability_flags = {
        "organization_structure_required": any(
            ("组织架构" in s or "责任分工" in s) for s in organization_markers
        ),
        "chapter_content_completeness_required": bool(req_sections),
        "key_difficult_points_required": any(("重难点" in s or "危大" in s) for s in req_sections),
        "solutions_required": bool(req_elements),
        "graphic_content_required": bool(req_charts),
        "consistency_checks_enabled": True,
        "lint_checks_enabled": True,
    }

    return {
        "engine_version": "v2",
        "project_id": project_id,
        "dimension_count": len(DIMENSION_IDS),
        "dimensions": dimensions,
        "penalty_rules": penalty_rules,
        "lint_issue_codes": [
            "MissingRequirement",
            "AnchorMissing",
            "AnchorMismatch",
            "EmptyPromiseWithoutEvidence",
            "ActionMissingHardElements",
            "ClosureGap",
            "ConsistencyConflict",
        ],
        "consistency_anchors": consistency_anchors,
        "chapter_requirements": chapter_requirements,
        "capability_flags": capability_flags,
        "source": {
            "rubric_version": str(rubric.get("version", "unknown")),
            "chapter_requirements": chapter_source,
        },
        "updated_at": _now_iso(),
    }


def _render_scoring_factors_markdown(payload: Dict[str, object]) -> str:
    dims = payload.get("dimensions") or []
    penalties = payload.get("penalty_rules") or []
    chapter = payload.get("chapter_requirements") or {}
    flags = payload.get("capability_flags") or {}
    lines = [
        "# 评分体系总览",
        "",
        f"- 引擎版本：`{payload.get('engine_version', 'v2')}`",
        f"- 项目ID：`{payload.get('project_id') or '-'} `",
        f"- 维度数：`{payload.get('dimension_count', 0)}`",
        f"- 规则版本：`{(payload.get('source') or {}).get('rubric_version', '-')}`",
        "",
        "## 维度评分因子",
        "",
        "| 维度 | 名称 | 模块 | 满分 | 评分模型 |",
        "|---|---|---|---:|---|",
    ]
    for d in dims:
        lines.append(
            f"| {d.get('dimension_id')} | {d.get('name','')} | {d.get('module','')} | "
            f"{d.get('max_score', 10)} | {d.get('scoring_model','')} |"
        )
    lines.extend(
        [
            "",
            "## 扣分规则",
            "",
            "| 代码 | 引擎 | 单次扣分 | 最大扣分 | 说明 |",
            "|---|---|---:|---:|---|",
        ]
    )
    for p in penalties:
        lines.append(
            f"| {p.get('code')} | {p.get('engine')} | {p.get('deduct_per_hit')} | "
            f"{p.get('max_deduct')} | {p.get('description','')} |"
        )

    lines.extend(
        [
            "",
            "## 章节与编制要求",
            "",
            "### 必备章节",
        ]
    )
    for x in chapter.get("required_sections") or []:
        lines.append(f"- {x}")
    lines.append("")
    lines.append("### 必备图表/图片")
    for x in chapter.get("required_charts_images") or []:
        lines.append(f"- {x}")
    lines.append("")
    lines.append("### 必备要素")
    for x in chapter.get("mandatory_elements") or []:
        lines.append(f"- {x}")
    lines.append("")
    lines.append("### 禁止表述")
    for x in chapter.get("forbidden_patterns") or []:
        lines.append(f"- {x}")

    lines.extend(
        [
            "",
            "## 能力覆盖标识",
            "",
            f"- 组织机构要求：`{bool(flags.get('organization_structure_required'))}`",
            f"- 章节完整性要求：`{bool(flags.get('chapter_content_completeness_required'))}`",
            f"- 重难点要求：`{bool(flags.get('key_difficult_points_required'))}`",
            f"- 解决方案要素要求：`{bool(flags.get('solutions_required'))}`",
            f"- 图文要求：`{bool(flags.get('graphic_content_required'))}`",
            f"- 一致性校验启用：`{bool(flags.get('consistency_checks_enabled'))}`",
            f"- Lint校验启用：`{bool(flags.get('lint_checks_enabled'))}`",
        ]
    )
    return "\n".join(lines).strip()


def _render_project_analysis_bundle_markdown(
    *,
    project: Dict[str, object],
    factors_payload: Dict[str, object],
    evaluation_payload: Dict[str, object],
) -> str:
    factors_md = _render_scoring_factors_markdown(factors_payload)
    variants = evaluation_payload.get("variants") or {}
    acceptance = evaluation_payload.get("acceptance") or {}
    project_name = str(project.get("name") or project.get("id") or "")

    lines = [
        f"# 项目分析包：{project_name}",
        "",
        f"- 项目ID：`{project.get('id')}`",
        f"- 生成时间：`{_now_iso()}`",
        f"- 青天样本数：`{evaluation_payload.get('sample_count_qt', 0)}`",
        "",
        "## 验收指标（V1 / V2 / V2+Calib）",
        "",
        "| 版本 | 样本数 | MAE | RMSE | Spearman | 画像相似度 | 扣分命中率 |",
        "|---|---:|---:|---:|---:|---:|---:|",
    ]
    for key, name in (("v1", "V1"), ("v2", "V2"), ("v2_calib", "V2+Calib")):
        v = variants.get(key) or {}
        lines.append(
            f"| {name} | {v.get('sample_count', 0)} | {v.get('mae', 0)} | {v.get('rmse', 0)} | "
            f"{v.get('spearman', 0)} | {v.get('profile_similarity', 0)} | {v.get('penalty_hit_rate', 0)} |"
        )
    lines.extend(
        [
            "",
            "### 验收结论",
            "",
            f"- MAE/RMSE 优于 V1：`{bool(acceptance.get('mae_rmse_improved_vs_v1'))}`",
            f"- 排序相关性不劣于 V1：`{bool(acceptance.get('rank_corr_not_worse_vs_v1'))}`",
            f"- 画像相似度优于 V1：`{bool(acceptance.get('profile_similarity_improved_vs_v1'))}`",
            f"- 扣分命中率优于 V1：`{bool(acceptance.get('penalty_hit_rate_improved_vs_v1'))}`",
            "",
            "## 评分体系（当前生效）",
            "",
            factors_md,
        ]
    )
    return "\n".join(lines).strip()


def _build_score_report_snapshot(
    submission_id: str,
    project: Dict[str, object],
    report: Dict[str, object],
    profile_snapshot: Optional[Dict[str, object]],
    scoring_engine_version: str,
) -> Dict[str, object]:
    rule_dim_scores = report.get("rule_dim_scores")
    if not isinstance(rule_dim_scores, dict):
        rule_dim_scores = report.get("dimension_scores", {})
    return {
        "id": str(uuid4()),
        "submission_id": submission_id,
        "project_id": project.get("id"),
        "scoring_engine_version": scoring_engine_version,
        "expert_profile_snapshot": profile_snapshot or {},
        "rule_dim_scores": rule_dim_scores,
        "rule_total_score": float(report.get("rule_total_score", report.get("total_score", 0.0))),
        "dim_total_80": float(report.get("dim_total_80", 0.0)),
        "dim_total_90": float(report.get("dim_total_90", 0.0)),
        "consistency_bonus": float(report.get("consistency_bonus", 0.0)),
        "pred_dim_scores": report.get("pred_dim_scores"),
        "pred_total_score": report.get("pred_total_score"),
        "llm_total_score": report.get("llm_total_score"),
        "pred_confidence": report.get("pred_confidence"),
        "score_blend": report.get("score_blend"),
        "penalties": report.get("penalties", []),
        "lint_findings": report.get("lint_findings", []),
        "suggestions": report.get("suggestions", []),
        "material_consistency": report.get("material_consistency", {}),
        "created_at": _now_iso(),
    }


def _normalize_dimension_id(dim_id: str) -> str:
    value = str(dim_id or "").strip().upper()
    if value.startswith("D") and len(value) == 3 and value[1:].isdigit():
        value = value[1:]
    if value.isdigit() and len(value) == 1:
        value = f"0{value}"
    if value in DIMENSION_IDS:
        return value
    return "01"


def _weights_from_multipliers(multipliers: Dict[str, float]) -> Dict[str, float]:
    if not multipliers:
        return {dim_id: 1.0 / len(DIMENSION_IDS) for dim_id in DIMENSION_IDS}
    values: Dict[str, float] = {}
    for dim_id in DIMENSION_IDS:
        v = multipliers.get(dim_id)
        if v is None:
            v = multipliers.get(f"D{dim_id}")
        values[dim_id] = max(0.0, float(v) if v is not None else 1.0)
    total = sum(values.values()) or float(len(DIMENSION_IDS))
    return {dim_id: values[dim_id] / total for dim_id in DIMENSION_IDS}


def _determine_engine_version(project: Dict[str, object], requested: Optional[str] = None) -> str:
    value = str(requested or project.get("scoring_engine_version_locked") or "").strip().lower()
    if value.startswith("v2"):
        return "v2"
    return "v1"


def _char_locator_for_snippet(text: str, snippet: str) -> str:
    src = str(text or "")
    needle = " ".join(str(snippet or "").split()).strip()
    if not src or not needle:
        return ""
    idx = src.find(needle)
    if idx < 0 and len(needle) > 24:
        idx = src.find(needle[:24])
    if idx < 0:
        return ""
    end = min(len(src), idx + len(needle))
    return f"char:{idx}-{end}"


def _build_v2_evidence_by_dim(
    evidence_units: List[Dict[str, object]], text: str
) -> Dict[str, List[Dict[str, str]]]:
    grouped: Dict[str, List[Dict[str, object]]] = {dim_id: [] for dim_id in DIMENSION_IDS}
    for unit in evidence_units or []:
        if not isinstance(unit, dict):
            continue
        dim_id = _normalize_dimension_id(str(unit.get("dimension_primary") or "01"))
        grouped.setdefault(dim_id, []).append(unit)

    result: Dict[str, List[Dict[str, str]]] = {dim_id: [] for dim_id in DIMENSION_IDS}
    for dim_id in DIMENSION_IDS:
        units = grouped.get(dim_id, [])
        units_sorted = sorted(
            units,
            key=lambda u: (
                float(u.get("specificity_score", 0.0)),
                int(bool(u.get("tag_solution"))),
                int(bool(u.get("landing_accept"))),
                int(bool(u.get("landing_role"))),
            ),
            reverse=True,
        )
        seen: set[str] = set()
        rows: List[Dict[str, str]] = []
        for unit in units_sorted:
            snippet = " ".join(str(unit.get("text") or "").split()).strip()
            if not snippet:
                continue
            short = snippet[:220] + ("..." if len(snippet) > 220 else "")
            if short in seen:
                continue
            seen.add(short)
            locator = _char_locator_for_snippet(text, snippet) or str(unit.get("locator") or "")
            rows.append({"text_snippet": short, "locator": locator})
            if len(rows) >= 4:
                break
        result[dim_id] = rows
    return result


def _legacy_dimension_scores_from_rule(
    rule_dim_scores: Dict[str, object],
    *,
    evidence_by_dim: Optional[Dict[str, List[Dict[str, str]]]] = None,
) -> Dict[str, object]:
    result: Dict[str, object] = {}
    evidence_by_dim = evidence_by_dim or {}
    for dim_id in DIMENSION_IDS:
        item = rule_dim_scores.get(dim_id) or {}
        score = float(item.get("dim_score", 0.0))
        subs = item.get("subscores") or {}
        sub_scores = [
            {"name": str(k), "score": float(v), "hits": [], "evidence": []} for k, v in subs.items()
        ]
        meta = DIMENSIONS.get(dim_id) or {}
        result[dim_id] = {
            "id": dim_id,
            "name": meta.get("name", dim_id),
            "module": meta.get("module", ""),
            "score": round(score, 2),
            "max_score": 10.0,
            "hits": [],
            "evidence": evidence_by_dim.get(dim_id) or [],
            "sub_scores": sub_scores or None,
        }
    return result


def _rule_dim_scores_from_legacy(dimension_scores: Dict[str, object]) -> Dict[str, object]:
    result: Dict[str, object] = {}
    for key, raw in (dimension_scores or {}).items():
        dim_id = _normalize_dimension_id(str(key))
        item = raw if isinstance(raw, dict) else {}
        subs_raw = item.get("sub_scores") or []
        subscores: Dict[str, float] = {}
        if isinstance(subs_raw, list):
            for sub in subs_raw:
                if not isinstance(sub, dict):
                    continue
                name = str(sub.get("name") or "")
                if not name:
                    continue
                subscores[name] = round(float(sub.get("score", 0.0)), 2)
        result[dim_id] = {
            "dim_score": round(float(item.get("score", 0.0)), 2),
            "subscores": subscores,
            "coverage_rate": None,
            "evidence_count": len(item.get("evidence") or []),
        }
    for dim_id in DIMENSION_IDS:
        result.setdefault(
            dim_id,
            {
                "dim_score": 0.0,
                "subscores": {},
                "coverage_rate": None,
                "evidence_count": 0,
            },
        )
    return result


def _build_v2_report_payload(
    v2_result: Dict[str, object],
    *,
    text: str,
    project: Dict[str, object],
    profile_snapshot: Optional[Dict[str, object]],
    scoring_engine_version: str,
) -> Dict[str, object]:
    rule_dim_scores = v2_result.get("rule_dim_scores") or {}
    evidence_by_dim = _build_v2_evidence_by_dim(
        list(v2_result.get("evidence_units") or []),
        text,
    )
    report = {
        "total_score": float(v2_result.get("rule_total_score", 0.0)),
        "rule_total_score": float(v2_result.get("rule_total_score", 0.0)),
        "rule_dim_scores": rule_dim_scores,
        "dimension_scores": _legacy_dimension_scores_from_rule(
            rule_dim_scores,
            evidence_by_dim=evidence_by_dim,
        ),
        "dim_total_80": float(v2_result.get("dim_total_80", 0.0)),
        "dim_total_90": float(v2_result.get("dim_total_90", 0.0)),
        "consistency_bonus": float(v2_result.get("consistency_bonus", 0.0)),
        "consistency_checks": v2_result.get("consistency_checks", []),
        "pred_dim_scores": None,
        "pred_total_score": None,
        "llm_total_score": None,
        "pred_confidence": None,
        "score_blend": None,
        "penalties": v2_result.get("penalties", []),
        "lint_findings": v2_result.get("lint_findings", []),
        "suggestions": v2_result.get("suggestions", []),
        "probe_dimensions": v2_result.get("probe_dimensions", []),
        "pre_flight": v2_result.get("pre_flight", {}),
        "requirement_hits": v2_result.get("requirement_hits", []),
        "mandatory_req_hit_rate": v2_result.get("mandatory_req_hit_rate"),
        "material_consistency": v2_result.get("material_consistency", {}),
        "requirement_pack_versions": v2_result.get("requirement_pack_versions", []),
        "evidence_units_count": int(v2_result.get("evidence_units_count", 0) or 0),
        "meta": {
            "engine_version": "v2",
            "region": project.get("region", DEFAULT_REGION),
            "scoring_engine_version": scoring_engine_version,
            "requirement_pack_versions": v2_result.get("requirement_pack_versions", []),
        },
    }
    if profile_snapshot:
        report["meta"]["expert_profile_snapshot"] = profile_snapshot
        report["meta"]["expert_profile_id"] = profile_snapshot.get("id")
    return report


def _select_calibrator_model(project: Dict[str, object]) -> Optional[Dict[str, object]]:
    models = sorted(
        load_calibration_models(), key=lambda x: str(x.get("created_at", "")), reverse=True
    )
    if not models:
        return None
    project_id = str(project.get("id") or "")
    locked_version = str(project.get("calibrator_version_locked") or "")

    def _scope_project_id(model: Dict[str, object]) -> str:
        return str(((model.get("train_filter") or {}).get("project_id") or "")).strip()

    def _compatible(model: Dict[str, object]) -> bool:
        # 仅允许同项目训练出的校准器生效，避免跨项目污染总分。
        return _scope_project_id(model) == project_id

    if locked_version:
        for model in models:
            if str(model.get("calibrator_version") or "") == locked_version:
                return model if _compatible(model) else None
        return None

    for model in models:
        if bool(model.get("deployed")) and _compatible(model):
            return model
    return None


def _select_deployed_patch(project_id: str) -> Optional[Dict[str, object]]:
    packages = [
        p
        for p in load_patch_packages()
        if str(p.get("project_id")) == project_id and str(p.get("status")) == "deployed"
    ]
    if not packages:
        return None
    packages.sort(key=lambda x: str(x.get("updated_at", "")), reverse=True)
    return packages[0]


def _auto_govern_deployed_patch(
    *,
    project_id: str,
    delta_cases: List[Dict[str, object]],
) -> Dict[str, object]:
    """
    对当前已部署补丁做自动治理：
    - shadow 评估通过：保留部署
    - shadow 评估失败且样本充分：自动回滚，并尝试回退到 rollback_pointer
    """
    result: Dict[str, object] = {
        "checked": False,
        "project_id": project_id,
        "patch_id": None,
        "gate_passed": None,
        "sample_count": 0,
        "action": "skip",
        "reason": "no_deployed_patch",
        "rolled_back": False,
        "rollback_to_patch_id": None,
        "metrics_before_after": {},
        "deployment_record_ids": [],
    }
    if not delta_cases:
        result["reason"] = "no_delta_cases"
        return result

    packages = load_patch_packages()
    deployed = [
        p
        for p in packages
        if str(p.get("project_id")) == project_id and str(p.get("status")) == "deployed"
    ]
    if not deployed:
        return result

    deployed = sorted(deployed, key=lambda x: str(x.get("updated_at", "")), reverse=True)
    patch = deployed[0]
    patch_id = str(patch.get("id") or "")
    result["checked"] = True
    result["patch_id"] = patch_id

    shadow = evaluate_patch_shadow(patch=patch, delta_cases=delta_cases)
    metrics = shadow.get("metrics_before_after") or {}
    sample_count = int(_to_float_or_none(metrics.get("sample_count")) or len(delta_cases) or 0)
    gate_passed = bool(shadow.get("gate_passed"))
    result["sample_count"] = sample_count
    result["gate_passed"] = gate_passed
    result["metrics_before_after"] = metrics

    # 样本不足时不做自动回滚，避免少量噪声导致频繁抖动。
    min_rollback_samples = 3
    if gate_passed:
        result["action"] = "keep"
        result["reason"] = "shadow_passed"
        return result
    if sample_count < min_rollback_samples:
        result["action"] = "skip"
        result["reason"] = "insufficient_samples_for_rollback"
        return result

    now_iso = _now_iso()
    rollback_pointer = str(patch.get("rollback_pointer") or "").strip()
    rollback_target = None
    if rollback_pointer:
        rollback_target = next(
            (
                p
                for p in packages
                if str(p.get("id") or "") == rollback_pointer
                and str(p.get("project_id") or "") == project_id
            ),
            None,
        )

    for row in packages:
        if str(row.get("project_id") or "") != project_id:
            continue
        if str(row.get("status") or "") == "deployed":
            row["status"] = "shadow_pass"
            row["updated_at"] = now_iso

    patch["status"] = "rolled_back"
    patch["updated_at"] = now_iso

    rollback_to_patch_id: Optional[str] = None
    if rollback_target is not None:
        rollback_target["status"] = "deployed"
        rollback_target["updated_at"] = now_iso
        rollback_to_patch_id = str(rollback_target.get("id") or "")

    save_patch_packages(packages)

    deployment_record_ids: List[str] = []
    deploys = load_patch_deployments()
    rollback_record = {
        "id": str(uuid4()),
        "patch_id": patch_id,
        "project_id": project_id,
        "action": "auto_rollback",
        "deployed": False,
        "metrics_before_after": metrics,
        "rollback_to_version": rollback_to_patch_id or rollback_pointer or None,
        "created_at": now_iso,
    }
    deploys.append(rollback_record)
    deployment_record_ids.append(str(rollback_record["id"]))
    if rollback_to_patch_id:
        promote_record = {
            "id": str(uuid4()),
            "patch_id": rollback_to_patch_id,
            "project_id": project_id,
            "action": "auto_promote_rollback_pointer",
            "deployed": True,
            "metrics_before_after": metrics,
            "rollback_to_version": None,
            "created_at": now_iso,
        }
        deploys.append(promote_record)
        deployment_record_ids.append(str(promote_record["id"]))
    save_patch_deployments(deploys)

    result["action"] = "rollback"
    result["reason"] = "shadow_failed"
    result["rolled_back"] = True
    result["rollback_to_patch_id"] = rollback_to_patch_id
    result["deployment_record_ids"] = deployment_record_ids
    return result


def _apply_deployed_patch_to_report(project_id: str, report: Dict[str, object]) -> None:
    patch = _select_deployed_patch(project_id)
    if not patch:
        return
    payload = patch.get("patch_payload") or {}
    penalties = report.get("penalties")
    if not isinstance(penalties, list) or not penalties:
        report.setdefault("meta", {})
        report["meta"]["patch_id"] = patch.get("id")
        return

    multipliers = payload.get("penalty_multiplier") or {}
    old_penalty_total = sum(
        float(p.get("points", p.get("deduct", 0.0))) for p in penalties if isinstance(p, dict)
    )
    new_penalty_total = 0.0
    for penalty in penalties:
        if not isinstance(penalty, dict):
            continue
        code = str(penalty.get("code") or "")
        mul = float(multipliers.get(code, 1.0)) if code else 1.0
        if "points" in penalty and penalty.get("points") is not None:
            penalty["points"] = round(float(penalty.get("points", 0.0)) * mul, 2)
            new_penalty_total += float(penalty["points"])
        elif "deduct" in penalty and penalty.get("deduct") is not None:
            penalty["deduct"] = round(float(penalty.get("deduct", 0.0)) * mul, 2)
            new_penalty_total += float(penalty["deduct"])
        else:
            new_penalty_total += 0.0

    has_dim_components = ("dim_total_90" in report) or ("dim_total_80" in report)
    if has_dim_components:
        dim_total_80 = float(report.get("dim_total_80", 0.0))
        dim_total_90 = report.get("dim_total_90")
        if dim_total_90 is not None:
            # 统一回推为 dim_total_80 再走同一聚合公式，避免历史/新字段混用时口径漂移
            dim_total_80 = max(0.0, min(80.0, float(dim_total_90) * (80.0 / 90.0)))
        consistency_bonus = float(report.get("consistency_bonus", 0.0))
        new_rule_total, normalized_dim_total_90 = compute_v2_rule_total(
            dim_total_80=dim_total_80,
            consistency_bonus=consistency_bonus,
            penalty_points=new_penalty_total,
        )
        report["rule_total_score"] = new_rule_total
        report["total_score"] = new_rule_total
        report["dim_total_80"] = round(dim_total_80, 2)
        report["dim_total_90"] = normalized_dim_total_90
    else:
        old_total = float(report.get("rule_total_score", report.get("total_score", 0.0)))
        delta = new_penalty_total - old_penalty_total
        new_total = max(0.0, min(100.0, round(old_total - delta, 2)))
        report["rule_total_score"] = new_total
        report["total_score"] = new_total

    report.setdefault("meta", {})
    report["meta"]["patch_id"] = patch.get("id")
    report["meta"]["patch_status"] = patch.get("status")


def _clip_score(value: float, low: float = 0.0, high: float = 100.0) -> float:
    return max(low, min(high, float(value)))


def _normalize_score_scale_max(value: object, default: int = DEFAULT_SCORE_SCALE_MAX) -> int:
    try:
        numeric = int(float(value))
    except (TypeError, ValueError):
        numeric = int(default)
    return 5 if numeric == 5 else 100


def _resolve_project_score_scale_max(project: Dict[str, object]) -> int:
    meta = project.get("meta") if isinstance(project.get("meta"), dict) else {}
    return _normalize_score_scale_max(
        (meta or {}).get("score_scale_max"),
        default=DEFAULT_SCORE_SCALE_MAX,
    )


def _score_scale_label(score_scale_max: int) -> str:
    return "5分制" if int(score_scale_max) == 5 else "100分制"


def _convert_score_from_100(score: object, score_scale_max: int) -> Optional[float]:
    value = _to_float_or_none(score)
    if value is None:
        return None
    clipped = _clip_score(value, 0.0, 100.0)
    factor = float(_normalize_score_scale_max(score_scale_max)) / 100.0
    return round(clipped * factor, 2)


def _convert_score_to_100(score: object, score_scale_max: int) -> Optional[float]:
    value = _to_float_or_none(score)
    if value is None:
        return None
    scale = float(_normalize_score_scale_max(score_scale_max))
    if scale <= 0:
        return None
    clipped = _clip_score(value, 0.0, scale)
    normalized = clipped * (100.0 / scale)
    return round(_clip_score(normalized, 0.0, 100.0), 2)


def _resolve_score_blend_weights(project: Dict[str, object]) -> tuple[float, float, float]:
    meta = project.get("meta") if isinstance(project.get("meta"), dict) else {}
    blend_raw = meta.get("score_blend") if isinstance(meta, dict) else {}
    blend_cfg = blend_raw if isinstance(blend_raw, dict) else {}

    def _f(v: object, default: float) -> float:
        try:
            return float(v)
        except Exception:
            return default

    rule_w = _f(
        blend_cfg.get("rule_weight", blend_cfg.get("rule", DEFAULT_RULE_SCORE_WEIGHT)),
        DEFAULT_RULE_SCORE_WEIGHT,
    )
    llm_w = _f(
        blend_cfg.get("llm_weight", blend_cfg.get("llm", DEFAULT_LLM_SCORE_WEIGHT)),
        DEFAULT_LLM_SCORE_WEIGHT,
    )
    delta_cap = _f(blend_cfg.get("llm_delta_cap", DEFAULT_LLM_DELTA_CAP), DEFAULT_LLM_DELTA_CAP)

    rule_w = max(0.0, rule_w)
    llm_w = max(0.0, llm_w)
    total = rule_w + llm_w
    if total <= 1e-9:
        rule_w, llm_w = DEFAULT_RULE_SCORE_WEIGHT, DEFAULT_LLM_SCORE_WEIGHT
        total = rule_w + llm_w
    rule_w /= total
    llm_w /= total
    delta_cap = max(0.0, delta_cap)
    return rule_w, llm_w, delta_cap


def _resolve_dynamic_blend_adjustment(
    report: Dict[str, object],
) -> tuple[float, float, Dict[str, object]]:
    """
    根据证据覆盖质量动态调整融合权重：
    - 资料利用门禁 blocked/warn 时，显著降低 LLM 权重
    - 资料文件覆盖率/强制项命中率低时，进一步降权
    """
    meta = report.get("meta") if isinstance(report.get("meta"), dict) else {}
    util = (
        meta.get("material_utilization")
        if isinstance(meta.get("material_utilization"), dict)
        else {}
    )
    gate = (
        meta.get("material_utilization_gate")
        if isinstance(meta.get("material_utilization_gate"), dict)
        else {}
    )
    trace = meta.get("evidence_trace") if isinstance(meta.get("evidence_trace"), dict) else {}
    has_material_signal = bool(util) or bool(gate) or bool(trace)
    if not has_material_signal:
        return (
            1.0,
            1.0,
            {
                "coverage_scale": 1.0,
                "delta_cap_scale": 1.0,
                "reasons": [],
                "retrieval_file_coverage_rate": None,
                "retrieval_hit_rate": None,
                "mandatory_hit_rate": None,
                "source_files_hit_count": None,
                "signal_state": "no_material_signal",
            },
        )

    scale = 1.0
    reasons: List[str] = []

    if bool(gate.get("blocked")):
        scale *= 0.1
        reasons.append("material_gate_blocked")
    elif bool(gate.get("warned")):
        scale *= 0.45
        reasons.append("material_gate_warned")

    retrieval_file_cov = _to_float_or_none(util.get("retrieval_file_coverage_rate"))
    if retrieval_file_cov is not None and retrieval_file_cov < 0.7:
        ratio = max(0.2, float(retrieval_file_cov) / 0.7)
        scale *= ratio
        reasons.append(f"low_retrieval_file_coverage:{retrieval_file_cov:.3f}")

    retrieval_hit_rate = _to_float_or_none(util.get("retrieval_hit_rate"))
    if retrieval_hit_rate is not None and retrieval_hit_rate < 0.35:
        ratio = max(0.3, float(retrieval_hit_rate) / 0.35)
        scale *= ratio
        reasons.append(f"low_retrieval_hit_rate:{retrieval_hit_rate:.3f}")

    mandatory_hit_rate = _to_float_or_none(trace.get("mandatory_hit_rate"))
    if mandatory_hit_rate is not None and mandatory_hit_rate < 0.45:
        ratio = max(0.35, float(mandatory_hit_rate) / 0.45)
        scale *= ratio
        reasons.append(f"low_mandatory_hit_rate:{mandatory_hit_rate:.3f}")

    source_files_hit_count_raw = (
        _to_float_or_none(trace.get("source_files_hit_count"))
        if isinstance(trace, dict) and "source_files_hit_count" in trace
        else None
    )
    source_files_hit_count = (
        int(source_files_hit_count_raw) if source_files_hit_count_raw is not None else None
    )
    if source_files_hit_count is not None and source_files_hit_count <= 0:
        scale *= 0.25
        reasons.append("no_material_source_files_hit")

    scale = max(0.02, min(1.0, float(scale)))
    delta_cap_scale = max(0.2, min(1.0, 0.2 + scale))
    return (
        scale,
        delta_cap_scale,
        {
            "coverage_scale": round(scale, 4),
            "delta_cap_scale": round(delta_cap_scale, 4),
            "reasons": reasons,
            "retrieval_file_coverage_rate": retrieval_file_cov,
            "retrieval_hit_rate": retrieval_hit_rate,
            "mandatory_hit_rate": mandatory_hit_rate,
            "source_files_hit_count": source_files_hit_count,
            "signal_state": "material_signal_detected",
        },
    )


def _fuse_rule_and_llm_scores(
    *,
    rule_total: float,
    llm_total_raw: float,
    project: Dict[str, object],
    report: Optional[Dict[str, object]] = None,
) -> tuple[float, float, Dict[str, object]]:
    rule = _clip_score(rule_total)
    llm_raw = _clip_score(llm_total_raw)
    rule_w, llm_w, delta_cap = _resolve_score_blend_weights(project)
    coverage_meta: Dict[str, object] = {}
    if isinstance(report, dict):
        coverage_scale, delta_cap_scale, coverage_meta = _resolve_dynamic_blend_adjustment(report)
        llm_w *= coverage_scale
        total = rule_w + llm_w
        if total > 1e-9:
            rule_w /= total
            llm_w /= total
        else:
            rule_w, llm_w = 1.0, 0.0
        delta_cap *= delta_cap_scale
    llm_bounded = _clip_score(max(rule - delta_cap, min(rule + delta_cap, llm_raw)))
    fused = _clip_score(rule * rule_w + llm_bounded * llm_w)
    blend_info = {
        "rule_weight": round(rule_w, 4),
        "llm_weight": round(llm_w, 4),
        "llm_delta_cap": round(delta_cap, 2),
        "dynamic_coverage": coverage_meta,
    }
    return round(fused, 2), round(llm_bounded, 2), blend_info


def _apply_prediction_to_report(
    report: Dict[str, object],
    *,
    submission_like: Dict[str, object],
    project: Dict[str, object],
) -> Optional[str]:
    model = _select_calibrator_model(project)
    if not model:
        report["pred_total_score"] = None
        report["llm_total_score"] = None
        report["pred_confidence"] = None
        report["pred_dim_scores"] = None
        report["score_blend"] = None
        # `total_score` is the primary score used by UI/sorting.
        report["total_score"] = float(
            report.get("rule_total_score", report.get("total_score", 0.0))
        )
        submission_like["total_score"] = float(report.get("total_score", 0.0))
        return None
    artifact = model.get("model_artifact") or model.get("artifact") or {}
    if not isinstance(artifact, dict):
        report["pred_total_score"] = None
        report["llm_total_score"] = None
        report["pred_confidence"] = None
        report["pred_dim_scores"] = None
        report["score_blend"] = None
        report["total_score"] = float(
            report.get("rule_total_score", report.get("total_score", 0.0))
        )
        submission_like["total_score"] = float(report.get("total_score", 0.0))
        return None
    row = build_feature_row(report, submission=submission_like)
    try:
        pred, conf = predict_with_model(artifact, row.get("x_features") or {})
    except Exception as e:
        # 预测模型不应影响主流程可用性；失败时保留 rule 分并显式记录错误。
        report["pred_total_score"] = None
        report["llm_total_score"] = None
        report["pred_confidence"] = None
        report["pred_dim_scores"] = None
        report["score_blend"] = None
        report["total_score"] = float(
            report.get("rule_total_score", report.get("total_score", 0.0))
        )
        submission_like["total_score"] = float(report.get("total_score", 0.0))
        report.setdefault("meta", {})
        report["meta"]["calibrator_version"] = model.get("calibrator_version")
        report["meta"]["calibrator_error"] = f"{type(e).__name__}: {e}"
        return str(model.get("calibrator_version") or "")

    rule_total = float(report.get("rule_total_score", report.get("total_score", 0.0)))
    fused_total, llm_total, blend_info = _fuse_rule_and_llm_scores(
        rule_total=rule_total,
        llm_total_raw=float(pred),
        project=project,
        report=report,
    )
    sigma = float(_to_float_or_none(conf.get("sigma")) or 0.0)
    ci95_delta = 1.96 * sigma if sigma > 0 else 0.0
    ci95_lower = _clip_score(fused_total - ci95_delta)
    ci95_upper = _clip_score(fused_total + ci95_delta)
    report["pred_total_score"] = fused_total
    report["llm_total_score"] = llm_total
    report["pred_confidence"] = {
        **conf,
        "raw_llm_score": float(pred),
        "bounded_llm_score": llm_total,
        "fused_ci95_lower": round(ci95_lower, 2),
        "fused_ci95_upper": round(ci95_upper, 2),
        "fused_sigma": round(sigma, 2),
    }
    report["score_blend"] = blend_info
    report["pred_dim_scores"] = None
    report["total_score"] = float(fused_total)
    submission_like["total_score"] = float(fused_total)
    report.setdefault("meta", {})
    report["meta"]["calibrator_version"] = model.get("calibrator_version")
    return str(model.get("calibrator_version") or "")


def _to_float_or_none(value: Any) -> Optional[float]:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _resolve_submission_score_fields(
    submission: Dict[str, object],
    *,
    allow_pred_score: bool = True,
    score_scale_max: int = DEFAULT_SCORE_SCALE_MAX,
) -> Dict[str, object]:
    report = submission.get("report")
    pred_total = None
    rule_total = None
    if isinstance(report, dict):
        pred_total = _to_float_or_none(report.get("pred_total_score"))
        if not allow_pred_score:
            pred_total = None
        rule_total = _to_float_or_none(report.get("rule_total_score"))
        if rule_total is None:
            rule_total = _to_float_or_none(report.get("total_score"))
    fallback_total = _to_float_or_none(submission.get("total_score"))
    if rule_total is None:
        rule_total = fallback_total
    primary_total = pred_total if pred_total is not None else rule_total
    if primary_total is None:
        primary_total = fallback_total if fallback_total is not None else 0.0
    total_display = _convert_score_from_100(primary_total, score_scale_max)
    pred_display = _convert_score_from_100(pred_total, score_scale_max)
    rule_display = _convert_score_from_100(rule_total, score_scale_max)
    if total_display is None:
        total_display = 0.0
    return {
        "total_score": round(float(total_display), 2),
        "pred_total_score": round(float(pred_display), 2) if pred_display is not None else None,
        "rule_total_score": round(float(rule_display), 2) if rule_display is not None else None,
        "score_source": "pred" if pred_total is not None else "rule",
    }


def _report_is_blocked(report: Optional[Dict[str, object]]) -> bool:
    if not isinstance(report, dict):
        return False
    status = str(report.get("scoring_status") or "").strip().lower()
    if status == "blocked":
        return True
    meta = report.get("meta") if isinstance(report.get("meta"), dict) else {}
    return bool(meta.get("score_blocked_by_material_utilization"))


def _submission_is_scored(submission: Dict[str, object]) -> bool:
    report_obj = submission.get("report")
    if isinstance(report_obj, dict):
        if _report_is_blocked(report_obj):
            return False
        status = str(report_obj.get("scoring_status") or "").strip().lower()
        if status == "pending":
            return False
        if status == "scored":
            return True
        if _to_float_or_none(report_obj.get("rule_total_score")) is not None:
            return True
        if _to_float_or_none(report_obj.get("pred_total_score")) is not None:
            return True
        if _to_float_or_none(report_obj.get("total_score")) is not None:
            return True
    return _to_float_or_none(submission.get("total_score")) is not None


def _mark_report_scored(report: Dict[str, object], *, trigger: str) -> None:
    report["scoring_status"] = "scored"
    report["scoring_trigger"] = trigger
    report["scored_at"] = _now_iso()


def _build_pending_submission_report(
    *,
    project: Dict[str, object],
    scoring_engine_version: str,
) -> Dict[str, object]:
    return {
        "scoring_status": "pending",
        "scoring_trigger": "upload_only",
        "queued_at": _now_iso(),
        "total_score": None,
        "rule_total_score": None,
        "pred_total_score": None,
        "llm_total_score": None,
        "pred_confidence": None,
        "score_blend": None,
        "dimension_scores": {},
        "rule_dim_scores": {},
        "pred_dim_scores": None,
        "penalties": [],
        "lint_findings": [],
        "suggestions": [],
        "requirement_hits": [],
        "mandatory_req_hit_rate": None,
        "evidence_units_count": 0,
        "meta": {
            "engine_version": _determine_engine_version(project, scoring_engine_version),
            "region": project.get("region", DEFAULT_REGION),
            "scoring_engine_version": scoring_engine_version,
            "queued_for_scoring": True,
        },
    }


def _score_submission_for_project(
    *,
    submission_id: str,
    text: str,
    project_id: str,
    project: Dict[str, object],
    config: object,
    multipliers: Dict[str, float],
    profile_snapshot: Optional[Dict[str, object]],
    scoring_engine_version: str,
    anchors: Optional[List[Dict[str, object]]] = None,
    requirements: Optional[List[Dict[str, object]]] = None,
    material_quality_snapshot: Optional[Dict[str, object]] = None,
) -> tuple[Dict[str, object], List[Dict[str, object]]]:
    engine_version = _determine_engine_version(project, scoring_engine_version)
    if engine_version == "v2":
        anchors_from_payload = anchors is not None
        requirements_from_payload = requirements is not None
        anchors = (
            anchors
            if anchors is not None
            else [a for a in load_project_anchors() if str(a.get("project_id")) == project_id]
        )
        requirements = (
            requirements
            if requirements is not None
            else [r for r in load_project_requirements() if str(r.get("project_id")) == project_id]
        )
        constraints_rebuilt = False
        if not anchors or not requirements:
            anchors, requirements = _rebuild_project_anchors_and_requirements(project_id)
            constraints_rebuilt = True
        elif (
            not anchors_from_payload or not requirements_from_payload
        ) and _constraints_need_rebuild(project_id, anchors, requirements):
            anchors, requirements = _rebuild_project_anchors_and_requirements(project_id)
            constraints_rebuilt = True

        weights_norm = (
            dict(profile_snapshot.get("weights_norm") or {})
            if profile_snapshot
            else _weights_from_multipliers(multipliers)
        )
        runtime_custom_requirements, runtime_req_meta = _build_runtime_custom_requirements(
            project_id,
            project=project,
            submission_text=text,
        )
        effective_requirements = list(requirements) + list(runtime_custom_requirements)
        meta = project.get("meta") if isinstance(project.get("meta"), dict) else {}
        strict_pre_flight = bool(meta.get("enforce_gb_redline", DEFAULT_ENFORCE_GB_REDLINE))
        try:
            v2_result = score_text_v2(
                submission_id=submission_id,
                text=text,
                lexicon=config.lexicon,
                weights_norm=weights_norm,
                anchors=anchors,
                requirements=effective_requirements,
                strict_pre_flight=strict_pre_flight,
            )
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc))
        report = _build_v2_report_payload(
            v2_result,
            text=text,
            project=project,
            profile_snapshot=profile_snapshot,
            scoring_engine_version=scoring_engine_version,
        )
        snapshot_for_meta = (
            dict(material_quality_snapshot)
            if isinstance(material_quality_snapshot, dict)
            else _build_material_quality_snapshot(project_id)
        )
        report_meta = report.get("meta") if isinstance(report.get("meta"), dict) else {}
        report_meta["input_injection"] = _build_scoring_input_injection_meta(
            project_id=project_id,
            text=text,
            anchors_count=len(anchors),
            base_requirements_count=len(requirements),
            runtime_custom_requirements_count=len(runtime_custom_requirements),
            weights_norm=weights_norm,
            profile_snapshot=profile_snapshot,
            constraints_rebuilt=constraints_rebuilt,
            runtime_req_meta=runtime_req_meta,
            material_quality_snapshot=snapshot_for_meta,
        )
        report_meta["material_quality"] = snapshot_for_meta
        report_meta["material_retrieval"] = {
            "chunks": int(
                _to_float_or_none(runtime_req_meta.get("material_retrieval_chunks")) or 0
            ),
            "requirements": int(
                _to_float_or_none(runtime_req_meta.get("material_retrieval_requirements")) or 0
            ),
            "preview": runtime_req_meta.get("material_retrieval_preview") or [],
            "consistency_requirements": int(
                _to_float_or_none(runtime_req_meta.get("material_consistency_requirements")) or 0
            ),
            "consistency_preview": runtime_req_meta.get("material_consistency_preview") or [],
            "available_types": runtime_req_meta.get("material_available_types") or [],
            "retrieval_types": runtime_req_meta.get("material_retrieval_types") or [],
            "missing_types": runtime_req_meta.get("material_retrieval_missing_types") or [],
        }
        report_meta["material_utilization"] = _build_material_utilization_summary(
            report,
            runtime_req_meta,
        )
        gate_obj = snapshot_for_meta.get("gate")
        material_gate_cfg = _resolve_material_gate_config(project)
        material_required_types = (
            material_gate_cfg.get("required_types")
            if isinstance(material_gate_cfg.get("required_types"), list)
            else []
        )
        material_utilization_policy = _resolve_material_utilization_policy(project)
        utilization_gate = _evaluate_material_utilization_gate(
            report_meta.get("material_utilization")
            if isinstance(report_meta.get("material_utilization"), dict)
            else {},
            policy=material_utilization_policy,
            required_types=material_required_types,
        )
        if isinstance(gate_obj, dict):
            report_meta["material_gate"] = gate_obj
        report_meta["material_utilization_gate"] = utilization_gate
        report_meta["material_utilization_alerts"] = _build_material_utilization_alerts(
            report_meta.get("material_utilization")
            if isinstance(report_meta.get("material_utilization"), dict)
            else {},
            gate_obj if isinstance(gate_obj, dict) else {},
        )
        report_meta["evidence_trace"] = _build_evidence_trace_summary(report)
        if bool(utilization_gate.get("blocked")):
            report_meta["score_confidence_level"] = "low"
            report_meta["score_blocked_by_material_utilization"] = True
            alerts = (
                report_meta.get("material_utilization_alerts")
                if isinstance(report_meta.get("material_utilization_alerts"), list)
                else []
            )
            for reason in utilization_gate.get("reasons") or []:
                reason_text = str(reason).strip()
                if reason_text and reason_text not in alerts:
                    alerts.append("资料利用门禁：" + reason_text)
            report_meta["material_utilization_alerts"] = alerts[:8]
            report["scoring_status"] = "blocked"
            report["scoring_trigger"] = "material_utilization_gate"
            report["scored_at"] = _now_iso()
        elif bool(utilization_gate.get("warned")):
            report_meta["score_confidence_level"] = "medium"
        else:
            report_meta["score_confidence_level"] = "high"
        report["meta"] = report_meta
        _apply_deployed_patch_to_report(project_id, report)
        if _report_is_blocked(report):
            return report, list(v2_result.get("evidence_units") or [])
        submission_like = {"id": submission_id, "project_id": project_id, "text": text}
        _apply_prediction_to_report(report, submission_like=submission_like, project=project)
        _mark_report_scored(report, trigger="score_engine")
        return report, list(v2_result.get("evidence_units") or [])

    legacy = score_text(
        text,
        config.rubric,
        config.lexicon,
        dimension_multipliers=multipliers,
    ).model_dump()
    legacy.setdefault("rule_total_score", float(legacy.get("total_score", 0.0)))
    legacy.setdefault(
        "rule_dim_scores", _rule_dim_scores_from_legacy(legacy.get("dimension_scores", {}))
    )
    legacy.setdefault("pred_dim_scores", None)
    legacy.setdefault("pred_total_score", None)
    legacy.setdefault("pred_confidence", None)
    legacy.setdefault("lint_findings", [])
    legacy.setdefault("requirement_hits", [])
    legacy.setdefault("mandatory_req_hit_rate", None)
    legacy.setdefault("evidence_units_count", 0)
    legacy.setdefault("meta", {})
    legacy["meta"]["engine_version"] = "v1"
    legacy["meta"]["region"] = project.get("region", DEFAULT_REGION)
    legacy["meta"]["scoring_engine_version"] = scoring_engine_version
    if profile_snapshot:
        legacy["meta"]["expert_profile_snapshot"] = profile_snapshot
        legacy["meta"]["expert_profile_id"] = profile_snapshot.get("id")
    legacy["meta"]["evidence_trace"] = _build_evidence_trace_summary(legacy)
    _apply_deployed_patch_to_report(project_id, legacy)
    _mark_report_scored(legacy, trigger="score_engine")
    return legacy, []


def _replace_submission_evidence_units(
    all_units: List[Dict[str, object]],
    *,
    submission_id: str,
    new_units: List[Dict[str, object]],
) -> List[Dict[str, object]]:
    kept = [u for u in all_units if str(u.get("submission_id")) != submission_id]
    kept.extend(new_units)
    return kept


def _latest_records_by_submission(records: List[Dict[str, object]]) -> Dict[str, Dict[str, object]]:
    latest: Dict[str, Dict[str, object]] = {}
    for item in records:
        sid = str(item.get("submission_id") or "")
        if not sid:
            continue
        prev = latest.get(sid)
        if prev is None or str(item.get("created_at", "")) >= str(prev.get("created_at", "")):
            latest[sid] = item
    return latest


def _extract_auto_candidates(model_artifact: Dict[str, Any]) -> List[Dict[str, Any]]:
    best_selection = model_artifact.get("best_selection") or {}
    raw_candidates = best_selection.get("candidates") or []
    if not isinstance(raw_candidates, list):
        return []
    normalized: List[Dict[str, Any]] = []
    for item in raw_candidates:
        if not isinstance(item, dict):
            continue
        metrics = item.get("metrics") or {}
        cv_item = item.get("cv") or {}
        normalized.append(
            {
                "model_type": str(item.get("model_type") or ""),
                "ok": bool(item.get("ok")),
                "gate_passed": bool(item.get("gate_passed")),
                "cv_mae": metrics.get("cv_mae"),
                "cv_rmse": metrics.get("cv_rmse"),
                "cv_spearman": metrics.get("cv_spearman"),
                "cv_mode": cv_item.get("mode"),
                "cv_pred_count": cv_item.get("pred_count"),
            }
        )
    return normalized


def _build_calibrator_summary(
    *,
    model_type: Optional[str],
    calibrator_version: Optional[str],
    gate_passed: Optional[bool],
    cv_metrics: Optional[Dict[str, Any]] = None,
    baseline_metrics: Optional[Dict[str, Any]] = None,
    improve_threshold: Optional[float] = None,
    spearman_tolerance: Optional[float] = None,
    auto_candidates: Optional[List[Dict[str, Any]]] = None,
    sample_count: Optional[int] = None,
    skipped_reason: Optional[str] = None,
) -> Dict[str, Any]:
    gate_payload: Dict[str, Any] = {}
    if gate_passed is not None:
        gate_payload["passed"] = bool(gate_passed)
    if improve_threshold is not None:
        gate_payload["improve_threshold"] = round(float(improve_threshold), 4)
    if spearman_tolerance is not None:
        gate_payload["spearman_tolerance"] = float(spearman_tolerance)

    summary: Dict[str, Any] = {
        "calibrator_version": calibrator_version,
        "model_type": model_type,
        "gate_passed": gate_passed,
        "cv_metrics": cv_metrics or {},
        "baseline_metrics": baseline_metrics or {},
        "gate": gate_payload,
        "auto_candidates": auto_candidates or [],
    }
    if sample_count is not None:
        summary["sample_count"] = int(sample_count)
    if skipped_reason:
        summary["skipped_reason"] = skipped_reason
    return summary


def _refresh_project_reflection_objects(project_id: str) -> None:
    submissions = [s for s in load_submissions() if str(s.get("project_id")) == project_id]
    submissions_by_id = {str(s.get("id")): s for s in submissions}
    latest_reports = _latest_records_by_submission(
        [r for r in load_score_reports() if str(r.get("project_id")) == project_id]
    )
    projects = load_projects()
    project = next((p for p in projects if str(p.get("id")) == project_id), {})
    project_score_scale = _resolve_project_score_scale_max(project) if project else 100

    qingtian_results = load_qingtian_results()
    qingtian_changed = False
    scoped_qt: List[Dict[str, object]] = []
    for q in qingtian_results:
        sid = str(q.get("submission_id") or "")
        if sid not in submissions_by_id:
            continue
        raw_payload = q.get("raw_payload") if isinstance(q.get("raw_payload"), dict) else {}
        normalized_record = _ground_truth_record_for_learning(
            {
                "final_score": raw_payload.get("final_score"),
                "final_score_raw": raw_payload.get("final_score_raw"),
                "final_score_100": raw_payload.get("final_score_100"),
                "score_scale_max": raw_payload.get("score_scale_max"),
                "judge_scores": raw_payload.get("judge_scores") or [],
            },
            default_score_scale_max=project_score_scale,
        )
        normalized_qt_score = float(normalized_record.get("final_score", 0.0))
        old_qt_score = _to_float_or_none(q.get("qt_total_score"))
        if old_qt_score is None or abs(old_qt_score - normalized_qt_score) > 1e-6:
            q["qt_total_score"] = normalized_qt_score
            qingtian_changed = True
        merged_payload = dict(raw_payload or {})
        merged_payload["final_score_raw"] = normalized_record.get("final_score_raw")
        merged_payload["final_score_100"] = normalized_qt_score
        merged_payload["score_scale_max"] = normalized_record.get("score_scale_max")
        if merged_payload != raw_payload:
            q["raw_payload"] = merged_payload
            qingtian_changed = True
        scoped_qt.append(q)
    if qingtian_changed:
        save_qingtian_results(qingtian_results)

    latest_qt = _latest_records_by_submission(scoped_qt)

    delta_cases = build_delta_cases(
        project_id=project_id,
        latest_reports_by_submission=latest_reports,
        latest_qingtian_by_submission=latest_qt,
    )
    all_delta = [d for d in load_delta_cases() if str(d.get("project_id")) != project_id]
    all_delta.extend(delta_cases)
    save_delta_cases(all_delta)

    samples = build_calibration_samples(
        project_id=project_id,
        latest_reports_by_submission=latest_reports,
        latest_qingtian_by_submission=latest_qt,
        submissions_by_id=submissions_by_id,
    )
    all_samples = [s for s in load_calibration_samples() if str(s.get("project_id")) != project_id]
    all_samples.extend(samples)
    save_calibration_samples(all_samples)


def _build_feedback_records_for_project(project_id: str) -> List[Dict[str, object]]:
    projects = load_projects()
    project = next((p for p in projects if str(p.get("id")) == project_id), None)
    if project is None:
        return []
    project_score_scale = _resolve_project_score_scale_max(project)
    submissions = [s for s in load_submissions() if str(s.get("project_id")) == project_id]
    submissions_by_id: Dict[str, Dict[str, object]] = {str(s.get("id")): s for s in submissions}

    feedback_records: List[Dict[str, object]] = []
    ground_truth_rows = [r for r in load_ground_truth() if str(r.get("project_id")) == project_id]
    for row in ground_truth_rows:
        if not isinstance(row, dict):
            continue
        judge_scores = row.get("judge_scores")
        if not isinstance(judge_scores, list) or len(judge_scores) not in (5, 7):
            continue
        source_submission_id = str(row.get("source_submission_id") or "").strip()
        sub = submissions_by_id.get(source_submission_id) if source_submission_id else None
        if sub is None:
            gt_text = str(row.get("shigong_text") or "").strip()
            if gt_text:
                sub = next(
                    (
                        s
                        for s in submissions
                        if str(s.get("text") or "").strip() == gt_text and _submission_is_scored(s)
                    ),
                    None,
                )
        if sub is None:
            continue

        report = sub.get("report") if isinstance(sub.get("report"), dict) else {}
        pred_raw = _to_float_or_none(report.get("pred_total_score"))
        if pred_raw is None:
            pred_raw = _to_float_or_none(report.get("rule_total_score"))
        if pred_raw is None:
            pred_raw = _to_float_or_none(report.get("total_score"))
        if pred_raw is None:
            pred_raw = _to_float_or_none(sub.get("total_score"))
        if pred_raw is None:
            continue

        predicted_total_100 = _convert_score_to_100(float(pred_raw), project_score_scale)
        if predicted_total_100 is None:
            continue

        row_scale = _normalize_score_scale_max(
            row.get("score_scale_max"),
            default=project_score_scale,
        )
        tags_by_judge = row.get("qualitative_tags_by_judge")
        judge_feedbacks: List[Dict[str, object]] = []
        for idx, score_raw in enumerate(judge_scores):
            score_value = _to_float_or_none(score_raw)
            if score_value is None:
                score_value = 0.0
            score_100 = _convert_score_to_100(score_value, row_scale)
            if score_100 is None:
                score_100 = 0.0
            tags: List[str] = []
            if isinstance(tags_by_judge, list) and idx < len(tags_by_judge):
                candidate = tags_by_judge[idx]
                if isinstance(candidate, list):
                    tags = [str(x).strip() for x in candidate if str(x).strip()]
            judge_feedbacks.append(
                {
                    "judge_index": idx + 1,
                    "score": round(float(score_100), 4),
                    "qualitative_tags": tags,
                }
            )

        normalized_row = _ground_truth_record_for_learning(
            row,
            default_score_scale_max=project_score_scale,
        )
        feedback_records.append(
            {
                "id": str(row.get("id") or ""),
                "project_id": project_id,
                "submission_id": str(sub.get("id") or ""),
                "predicted_total_score": round(float(predicted_total_100), 4),
                "final_total_score": round(float(normalized_row.get("final_score", 0.0)), 4),
                "judge_feedbacks": judge_feedbacks,
                "created_at": str(row.get("created_at") or _now_iso()),
            }
        )
    return feedback_records


def _auto_update_from_delta_cases(project_id: str) -> Dict[str, object]:
    projects = load_projects()
    project = next((p for p in projects if str(p.get("id")) == project_id), None)
    if project is None:
        return {"updated": False, "reason": "project_not_found"}

    delta_cases = [d for d in load_delta_cases() if str(d.get("project_id")) == project_id]
    if len(delta_cases) < 2:
        return {
            "updated": False,
            "reason": "insufficient_delta_cases",
            "sample_count": len(delta_cases),
        }

    dim_stats: Dict[str, Dict[str, float]] = {
        dim_id: {"sum": 0.0, "count": 0.0} for dim_id in DIMENSION_IDS
    }
    for case in delta_cases:
        dim_errors = case.get("dim_errors") or {}
        if not isinstance(dim_errors, dict):
            continue
        for dim_id, raw_err in dim_errors.items():
            did = _normalize_dimension_id(str(dim_id))
            try:
                err = float(raw_err)
            except Exception:
                continue
            dim_stats[did]["sum"] += err
            dim_stats[did]["count"] += 1.0

    candidates: List[Dict[str, object]] = []
    for dim_id in DIMENSION_IDS:
        count = int(dim_stats[dim_id]["count"])
        if count < 2:
            continue
        mean_error = dim_stats[dim_id]["sum"] / float(count)
        if abs(mean_error) < 0.8:
            continue
        candidates.append(
            {
                "dim_id": dim_id,
                "count": count,
                "mean_error": round(mean_error, 3),
                "abs_mean_error": abs(mean_error),
                "step": -1 if mean_error > 0 else 1,
            }
        )

    if not candidates:
        return {
            "updated": False,
            "reason": "no_adjustable_dimensions",
            "sample_count": len(delta_cases),
        }

    candidates.sort(key=lambda x: float(x.get("abs_mean_error") or 0.0), reverse=True)
    top_dims = candidates[:4]

    profiles = load_expert_profiles()
    profile, created = _ensure_project_expert_profile(project, profiles)
    if created:
        save_expert_profiles(profiles)
        save_projects(projects)

    weights_raw = _coerce_weights_raw(dict(profile.get("weights_raw") or _default_weights_raw()))
    new_weights_raw = dict(weights_raw)
    changed_dims: List[Dict[str, object]] = []
    for item in top_dims:
        dim_id = str(item.get("dim_id"))
        step = int(item.get("step") or 0)
        before = int(new_weights_raw.get(dim_id, 5))
        after = max(0, min(10, before + step))
        if after == before:
            continue
        new_weights_raw[dim_id] = after
        changed_dims.append(
            {
                "dim_id": dim_id,
                "before": before,
                "after": after,
                "mean_error": item.get("mean_error"),
                "count": item.get("count"),
            }
        )

    if not changed_dims:
        return {"updated": False, "reason": "weights_unchanged", "sample_count": len(delta_cases)}

    auto_name = (
        f"{project.get('name', '项目')}_auto_feedback_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    )
    new_profile = _new_expert_profile(auto_name, new_weights_raw)
    profiles.append(new_profile)
    save_expert_profiles(profiles)

    project["expert_profile_id"] = new_profile["id"]
    project["updated_at"] = _now_iso()
    save_projects(projects)

    return {
        "updated": True,
        "sample_count": len(delta_cases),
        "changed_dims": changed_dims,
        "new_profile_id": new_profile["id"],
        "strategy": "delta_case_fallback",
        "new_weights_norm": dict(new_profile.get("weights_norm") or {}),
        "new_dimension_multipliers": _weights_norm_to_dimension_multipliers(
            dict(new_profile.get("weights_norm") or {})
        ),
    }


def _auto_update_project_weights_from_delta_cases(project_id: str) -> Dict[str, object]:
    """
    优先使用「总分+标签」定向反演（模块一/二）；
    若反馈样本不足再回退到 DELTA_CASE 规则。
    """
    projects = load_projects()
    project = next((p for p in projects if str(p.get("id")) == project_id), None)
    if project is None:
        return {"updated": False, "reason": "project_not_found"}

    profiles = load_expert_profiles()
    profile, created = _ensure_project_expert_profile(project, profiles)
    if created:
        save_expert_profiles(profiles)
        save_projects(projects)

    feedback_records = _build_feedback_records_for_project(project_id)
    if feedback_records:
        current_weights_norm = dict(profile.get("weights_norm") or {})
        calibrated = calibrate_weights(
            current_weights_norm,
            feedback_records,
            half_life_days=30.0,
            lr_tag=0.08,
            lr_global=0.004,
            ridge_lambda=0.06,
            min_weight=0.005,
        )
        new_weights_norm = dict(calibrated.get("weights_norm") or {})
        if new_weights_norm:
            new_weights_raw = _weights_raw_from_norm(new_weights_norm)
            auto_name = f"{project.get('name', '项目')}_tag_guided_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            new_profile = _new_expert_profile(auto_name, new_weights_raw)
            # 保留算法输出的精确归一化权重，避免反推整数造成信息损失
            new_profile["weights_norm"] = {
                dim_id: float(new_weights_norm.get(dim_id, 1.0 / len(DIMENSION_IDS)))
                for dim_id in DIMENSION_IDS
            }
            profiles.append(new_profile)
            save_expert_profiles(profiles)

            project["expert_profile_id"] = new_profile["id"]
            project["updated_at"] = _now_iso()
            save_projects(projects)

            return {
                "updated": True,
                "strategy": "tag_guided_calibration",
                "sample_count": len(feedback_records),
                "new_profile_id": new_profile["id"],
                "calibration_stats": calibrated.get("stats") or {},
                "new_weights_norm": dict(new_profile.get("weights_norm") or {}),
                "new_dimension_multipliers": _weights_norm_to_dimension_multipliers(
                    dict(new_profile.get("weights_norm") or {})
                ),
            }

    # 回退：历史 DELTA_CASE 方案
    return _auto_update_from_delta_cases(project_id)


def _sync_feedback_weights_to_evolution(
    project_id: str,
    weight_update: Dict[str, object],
) -> Dict[str, object]:
    if not bool(weight_update.get("updated")):
        return {"synced": False, "reason": "weight_not_updated"}
    multipliers = weight_update.get("new_dimension_multipliers") or {}
    if not isinstance(multipliers, dict) or not multipliers:
        return {"synced": False, "reason": "missing_multipliers"}
    reports = load_evolution_reports()
    evo = reports.get(project_id) or {}
    scoring_evolution = (
        evo.get("scoring_evolution") if isinstance(evo.get("scoring_evolution"), dict) else {}
    )
    scoring_evolution = dict(scoring_evolution or {})
    scoring_evolution["dimension_multipliers"] = {
        dim_id: float(multipliers.get(dim_id, 1.0)) for dim_id in DIMENSION_IDS
    }
    scoring_evolution.setdefault("rationale", {})
    scoring_evolution["updated_by_feedback"] = True
    scoring_evolution["updated_by_feedback_at"] = _now_iso()
    evo["scoring_evolution"] = scoring_evolution
    evo.setdefault("project_id", project_id)
    evo.setdefault("sample_count", 0)
    evo["updated_at"] = _now_iso()
    reports[project_id] = evo
    save_evolution_reports(reports)
    return {
        "synced": True,
        "dimension_multipliers_count": len(scoring_evolution.get("dimension_multipliers") or {}),
    }


def _refresh_evolution_report_from_ground_truth(project_id: str) -> Dict[str, object]:
    projects = load_projects()
    project = next((p for p in projects if str(p.get("id")) == project_id), None)
    if project is None:
        return {"refreshed": False, "reason": "project_not_found"}
    project_score_scale = _resolve_project_score_scale_max(project)
    records_raw = [r for r in load_ground_truth() if str(r.get("project_id")) == project_id]
    records = [
        _ground_truth_record_for_learning(
            r if isinstance(r, dict) else {},
            default_score_scale_max=project_score_scale,
        )
        for r in records_raw
    ]
    ctx_data = load_project_context().get(project_id) or {}
    project_context = str(ctx_data.get("text") or "").strip()
    materials_text = _merge_materials_text(project_id)
    if materials_text:
        project_context = (
            (project_context + "\n\n" + materials_text) if project_context else materials_text
        )

    report = build_evolution_report(project_id, records, project_context)
    reports = load_evolution_reports()
    prev = reports.get(project_id) or {}
    # 自动闭环刷新时仅更新规则进化结果与编制指导；保留已有 LLM 增强来源标记。
    if isinstance(prev.get("enhanced_by"), str):
        report["enhanced_by"] = prev.get("enhanced_by")
    reports[project_id] = report
    save_evolution_reports(reports)
    return {
        "refreshed": True,
        "sample_count": int(report.get("sample_count", 0) or 0),
    }


def _run_feedback_closed_loop(project_id: str, *, locale: str, trigger: str) -> Dict[str, object]:
    """
    反馈信号闭环：刷新样本 -> 自动调权重 -> 自动反演校准。
    所有步骤为 best-effort，不影响主流程返回。
    """
    result: Dict[str, object] = {
        "ok": True,
        "project_id": project_id,
        "trigger": trigger,
        "weight_update": {"updated": False},
        "weight_sync_to_evolution": {"synced": False},
        "auto_run": None,
        "evolution_refresh": {"refreshed": False},
    }
    try:
        _refresh_project_reflection_objects(project_id)
    except Exception as exc:
        result["ok"] = False
        result["refresh_error"] = str(exc)
        return result

    try:
        result["weight_update"] = _auto_update_project_weights_from_delta_cases(project_id)
    except Exception as exc:
        result["weight_update"] = {"updated": False, "error": str(exc)}
    try:
        result["weight_sync_to_evolution"] = _sync_feedback_weights_to_evolution(
            project_id, result["weight_update"]
        )
    except Exception as exc:
        result["weight_sync_to_evolution"] = {"synced": False, "error": str(exc)}

    try:
        auto_resp = auto_run_reflection_pipeline(project_id=project_id, api_key=None, locale=locale)
        if hasattr(auto_resp, "model_dump"):
            result["auto_run"] = auto_resp.model_dump()
        else:
            result["auto_run"] = dict(auto_resp)
    except Exception as exc:
        result["auto_run"] = {"ok": False, "error": str(exc)}
        result["ok"] = False
    try:
        result["evolution_refresh"] = _refresh_evolution_report_from_ground_truth(project_id)
    except Exception as exc:
        result["evolution_refresh"] = {"refreshed": False, "error": str(exc)}
    return result


def _run_feedback_closed_loop_safe(
    project_id: str,
    *,
    locale: str,
    trigger: str,
) -> Dict[str, object]:
    """
    闭环执行保护层：不抛错中断主流程，但必须显式返回失败信息并记录日志。
    """
    try:
        raw_result = _run_feedback_closed_loop(project_id, locale=locale, trigger=trigger)
        if isinstance(raw_result, dict):
            result = dict(raw_result)
        elif hasattr(raw_result, "model_dump"):
            dumped = raw_result.model_dump()
            if isinstance(dumped, dict):
                result = dict(dumped)
            else:
                result = {
                    "ok": bool(getattr(raw_result, "ok", False)),
                    "project_id": project_id,
                    "trigger": trigger,
                    "raw": str(raw_result),
                }
        else:
            result = {
                "ok": bool(getattr(raw_result, "ok", False)),
                "project_id": project_id,
                "trigger": trigger,
                "raw": str(raw_result),
            }
        if not bool(result.get("ok", True)):
            logger.warning(
                "feedback_closed_loop_non_ok project_id=%s trigger=%s result=%s",
                project_id,
                trigger,
                result,
            )
        return result
    except Exception as exc:
        logger.exception(
            "feedback_closed_loop_exception project_id=%s trigger=%s error=%s",
            project_id,
            trigger,
            exc,
        )
        return {
            "ok": False,
            "project_id": project_id,
            "trigger": trigger,
            "error": f"{type(exc).__name__}: {exc}",
        }


def _build_evolution_health_report(
    project_id: str, project: Dict[str, object]
) -> Dict[str, object]:
    """
    构建项目进化健康度报告：
    - 统计系统预测分与真实分误差（全量/30天/90天）
    - 用时间衰减评估样本新鲜度
    - 给出概念漂移风险等级和建议动作
    """
    project_score_scale = _resolve_project_score_scale_max(project)
    submissions = [s for s in load_submissions() if str(s.get("project_id")) == project_id]
    submissions_by_id = {str(s.get("id") or ""): s for s in submissions if str(s.get("id") or "")}
    ground_truth_rows = [r for r in load_ground_truth() if str(r.get("project_id")) == project_id]
    now_utc = datetime.now(timezone.utc)

    matched_rows: List[Dict[str, object]] = []
    unmatched_ground_truth = 0
    for row in ground_truth_rows:
        if not isinstance(row, dict):
            continue
        normalized = _ground_truth_record_for_learning(
            row,
            default_score_scale_max=project_score_scale,
        )
        final_score = _to_float_or_none(normalized.get("final_score"))
        if final_score is None:
            continue

        source_submission_id = str(row.get("source_submission_id") or "").strip()
        submission = submissions_by_id.get(source_submission_id) if source_submission_id else None
        if submission is None:
            gt_text = str(row.get("shigong_text") or "").strip()
            if gt_text:
                submission = next(
                    (s for s in submissions if str(s.get("text") or "").strip() == gt_text),
                    None,
                )
        if submission is None:
            unmatched_ground_truth += 1
            continue

        report = submission.get("report") if isinstance(submission.get("report"), dict) else {}
        pred_score_raw = _to_float_or_none(report.get("pred_total_score"))
        if pred_score_raw is None:
            pred_score_raw = _to_float_or_none(report.get("rule_total_score"))
        if pred_score_raw is None:
            pred_score_raw = _to_float_or_none(report.get("total_score"))
        if pred_score_raw is None:
            pred_score_raw = _to_float_or_none(submission.get("total_score"))
        if pred_score_raw is None:
            unmatched_ground_truth += 1
            continue

        report_scale_max = _normalize_score_scale_max(
            report.get("score_scale_max"),
            default=project_score_scale,
        )
        pred_score_100 = _convert_score_to_100(pred_score_raw, report_scale_max)
        if pred_score_100 is None:
            unmatched_ground_truth += 1
            continue

        created_at_dt = _parse_iso_datetime_utc(row.get("created_at")) or now_utc
        age_days = max(0.0, (now_utc - created_at_dt).total_seconds() / 86400.0)
        abs_error = abs(float(pred_score_100) - float(final_score))
        matched_rows.append(
            {
                "ground_truth_id": str(row.get("id") or ""),
                "submission_id": str(submission.get("id") or ""),
                "predicted_score": float(pred_score_100),
                "actual_score": float(final_score),
                "abs_error": float(abs_error),
                "age_days": float(age_days),
                "time_decay": float(
                    compute_time_decay_weight(
                        record_time=created_at_dt,
                        now=now_utc,
                        half_life_days=30.0,
                    )
                ),
            }
        )

    def _window_metrics(
        rows: List[Dict[str, object]],
        *,
        min_age_days: Optional[float] = None,
        max_age_days: Optional[float] = None,
    ) -> Dict[str, object]:
        scoped: List[Dict[str, object]] = []
        for item in rows:
            age_days = float(_to_float_or_none(item.get("age_days")) or 0.0)
            if min_age_days is not None and age_days < float(min_age_days):
                continue
            if max_age_days is not None and age_days > float(max_age_days):
                continue
            scoped.append(item)
        count = len(scoped)
        if count <= 0:
            return {
                "count": 0,
                "mae": None,
                "rmse": None,
                "avg_time_decay": None,
                "max_abs_error": None,
            }
        err_sq_sum = 0.0
        err_abs_sum = 0.0
        max_abs_error = 0.0
        decay_sum = 0.0
        for item in scoped:
            abs_error = float(_to_float_or_none(item.get("abs_error")) or 0.0)
            err_abs_sum += abs_error
            err_sq_sum += abs_error * abs_error
            max_abs_error = max(max_abs_error, abs_error)
            decay_sum += float(_to_float_or_none(item.get("time_decay")) or 0.0)
        return {
            "count": count,
            "mae": round(err_abs_sum / float(count), 4),
            "rmse": round((err_sq_sum / float(count)) ** 0.5, 4),
            "avg_time_decay": round(decay_sum / float(count), 6),
            "max_abs_error": round(max_abs_error, 4),
        }

    metrics_all = _window_metrics(matched_rows)
    metrics_recent_30 = _window_metrics(matched_rows, max_age_days=30.0)
    metrics_recent_90 = _window_metrics(matched_rows, max_age_days=90.0)
    metrics_prev_30_90 = _window_metrics(matched_rows, min_age_days=30.0, max_age_days=90.0)

    recent_mae = _to_float_or_none(metrics_recent_30.get("mae"))
    prev_mae = _to_float_or_none(metrics_prev_30_90.get("mae"))
    mae_delta_recent_vs_prev = None
    drift_level = "insufficient_data"
    if recent_mae is not None and prev_mae is not None:
        mae_delta_recent_vs_prev = round(float(recent_mae) - float(prev_mae), 4)
        if mae_delta_recent_vs_prev >= 2.5:
            drift_level = "high"
        elif mae_delta_recent_vs_prev >= 1.0:
            drift_level = "medium"
        else:
            drift_level = "low"
    elif int(metrics_recent_30.get("count") or 0) >= 3:
        drift_level = "watch"

    multipliers, profile_snapshot, _ = _resolve_project_scoring_context(project_id)
    evo = load_evolution_reports().get(project_id) or {}
    scoring_evolution = evo.get("scoring_evolution") or {}
    has_evolved_multipliers = bool(
        isinstance(scoring_evolution, dict)
        and isinstance(scoring_evolution.get("dimension_multipliers"), dict)
        and scoring_evolution.get("dimension_multipliers")
    )
    recommendations: List[str] = []
    if len(ground_truth_rows) < 3:
        recommendations.append("真实评标样本不足，建议至少录入 3 条以上再观察进化稳定性。")
    if int(metrics_recent_30.get("count") or 0) <= 0:
        recommendations.append("近30天无真实反馈，建议补录最新项目评分以避免概念漂移。")
    if drift_level in {"high", "medium"}:
        recommendations.append("近期误差上升，建议立即执行「学习进化」并触发 V2 一键闭环。")
    if int(unmatched_ground_truth) > 0:
        recommendations.append(
            f"有 {unmatched_ground_truth} 条真实评分未关联到施组预测记录，建议使用“从步骤4施组下拉选择”录入。"
        )
    if not has_evolved_multipliers:
        recommendations.append("尚未形成进化维度权重，建议在录入真实评分后执行一次学习进化。")

    return {
        "project_id": project_id,
        "generated_at": _now_iso(),
        "summary": {
            "ground_truth_count": len(ground_truth_rows),
            "matched_prediction_count": len(matched_rows),
            "unmatched_ground_truth_count": unmatched_ground_truth,
            "current_weights_source": _infer_weights_source(project_id, profile_snapshot),
            "current_multiplier_count": len(multipliers or {}),
            "has_evolved_multipliers": has_evolved_multipliers,
            "last_evolution_updated_at": str(evo.get("updated_at") or evo.get("created_at") or ""),
        },
        "windows": {
            "all": metrics_all,
            "recent_30d": metrics_recent_30,
            "recent_90d": metrics_recent_90,
            "prev_30_90d": metrics_prev_30_90,
        },
        "drift": {
            "level": drift_level,
            "mae_delta_recent_vs_prev_30_90": mae_delta_recent_vs_prev,
            "half_life_days": 30.0,
        },
        "recommendations": recommendations[:12],
    }


def _collect_applied_feature_ids_from_report(
    report: Dict[str, object],
    *,
    top_k_per_probe: int = 2,
) -> List[str]:
    """
    从评分报告中提取“本轮建议所采用”的高分骨架特征 ID。
    优先使用 suggestions[*].applied_feature_ids；
    若历史报告缺字段，则按探针维度回填 top-k 活跃特征。
    """
    feature_ids: set[str] = set()
    probe_ids: set[str] = set()

    suggestions = report.get("suggestions")
    if isinstance(suggestions, list):
        for item in suggestions:
            if not isinstance(item, dict):
                continue
            raw_feature_ids = item.get("applied_feature_ids")
            if isinstance(raw_feature_ids, list):
                for fid in raw_feature_ids:
                    s = str(fid or "").strip()
                    if s:
                        feature_ids.add(s)
            dim_id = str(item.get("dimension_id") or "").strip().upper()
            if dim_id.startswith("P"):
                probe_ids.add(dim_id)

    # 回填：旧报告无 applied_feature_ids 时，按低分探针补全
    if not probe_ids:
        probes = report.get("probe_dimensions")
        if isinstance(probes, list):
            for probe in probes:
                if not isinstance(probe, dict):
                    continue
                probe_id = str(probe.get("id") or "").strip().upper()
                score_rate = _to_float_or_none(probe.get("score_rate"))
                if probe_id.startswith("P") and (score_rate is None or score_rate < 0.8):
                    probe_ids.add(probe_id)

    for probe_id in sorted(probe_ids):
        for feature in select_top_logic_skeletons(
            dimension_ids=[probe_id],
            top_k=max(1, int(top_k_per_probe)),
        ):
            fid = str(feature.feature_id or "").strip()
            if fid:
                feature_ids.add(fid)

    return sorted(feature_ids)


def _auto_update_feature_confidence_on_ground_truth(
    *,
    report: Dict[str, object],
    gt_record: Dict[str, object],
    project_score_scale_max: int,
) -> Dict[str, object]:
    """
    真实评标录入后，立即执行一次 feature confidence 闭环更新。
    """
    applied_feature_ids = _collect_applied_feature_ids_from_report(report)
    if not applied_feature_ids:
        return {"updated": 0, "retired": 0, "reason": "no_applied_feature_ids"}

    gt_for_learning = _ground_truth_record_for_learning(
        gt_record,
        default_score_scale_max=project_score_scale_max,
    )
    actual_score_100 = _to_float_or_none(gt_for_learning.get("final_score"))
    if actual_score_100 is None:
        return {"updated": 0, "retired": 0, "reason": "missing_actual_score"}

    pred_score_100 = _to_float_or_none(report.get("pred_total_score"))
    if pred_score_100 is None:
        pred_score_100 = _to_float_or_none(report.get("total_score"))
    if pred_score_100 is None:
        pred_score_100 = _to_float_or_none(report.get("rule_total_score"))
    if pred_score_100 is None:
        return {"updated": 0, "retired": 0, "reason": "missing_predicted_score"}

    # 兜底兼容：若项目为5分制且报告字段偶发为5分口径，则转回100分口径。
    if int(project_score_scale_max) == 5 and pred_score_100 <= 5.0:
        pred_score_100 = float(_convert_score_to_100(pred_score_100, 5) or 0.0)

    update_result = update_feature_confidence(
        applied_feature_ids=applied_feature_ids,
        actual_score=float(actual_score_100),
        predicted_score=float(pred_score_100),
    )
    update_result["applied_feature_ids"] = applied_feature_ids
    update_result["actual_score_100"] = round(float(actual_score_100), 2)
    update_result["predicted_score_100"] = round(float(pred_score_100), 2)
    return update_result


def _sync_ground_truth_record_to_qingtian(project_id: str, gt_record: Dict[str, object]) -> None:
    projects = load_projects()
    project = _find_project(project_id, projects)
    config = load_config()
    multipliers, profile_snapshot, _ = _resolve_project_scoring_context(project_id)
    scoring_engine_version = str(project.get("scoring_engine_version_locked") or "v1")
    source_gt_id = str(gt_record.get("id") or "")
    gt_text = str(gt_record.get("shigong_text") or "")

    submissions = load_submissions()
    matched_submission = None
    for s in submissions:
        if str(s.get("project_id")) != project_id:
            continue
        if str(s.get("source_ground_truth_id") or "") == source_gt_id:
            matched_submission = s
            break
        if str(s.get("text") or "").strip() == gt_text.strip() and gt_text.strip():
            matched_submission = s
            break

    scored_submission = False
    submission_changed = False
    evidence_units_new: List[Dict[str, object]] = []
    now_iso = _now_iso()
    if matched_submission is None:
        matched_submission = {
            "id": str(uuid4()),
            "project_id": project_id,
            "filename": f"ground_truth_{source_gt_id[:8]}.txt",
            "total_score": 0.0,
            "report": _build_pending_submission_report(
                project=project,
                scoring_engine_version=scoring_engine_version,
            ),
            "text": gt_text,
            "created_at": now_iso,
            "updated_at": now_iso,
            "expert_profile_id_used": profile_snapshot.get("id") if profile_snapshot else None,
            "source_ground_truth_id": source_gt_id,
            "bidder_name": f"GT_{source_gt_id[:8]}",
        }
        submissions.append(matched_submission)
        submission_changed = True

    if str(matched_submission.get("source_ground_truth_id") or "") != source_gt_id:
        matched_submission["source_ground_truth_id"] = source_gt_id
        submission_changed = True
    if gt_text.strip() and str(matched_submission.get("text") or "").strip() != gt_text.strip():
        matched_submission["text"] = gt_text
        submission_changed = True

    if not _submission_is_scored(matched_submission):
        report, evidence_units_new = _score_submission_for_project(
            submission_id=str(matched_submission.get("id")),
            text=gt_text,
            project_id=project_id,
            project=project,
            config=config,
            multipliers=multipliers,
            profile_snapshot=profile_snapshot,
            scoring_engine_version=scoring_engine_version,
        )
        if not _report_is_blocked(report):
            _mark_report_scored(report, trigger="ground_truth_sync")
        matched_submission["report"] = report
        matched_submission["total_score"] = float(
            report.get("total_score", report.get("rule_total_score", 0.0))
        )
        matched_submission["expert_profile_id_used"] = (
            profile_snapshot.get("id") if profile_snapshot else None
        )
        matched_submission["updated_at"] = _now_iso()
        scored_submission = True
        submission_changed = True

    if submission_changed:
        save_submissions(submissions)

    if scored_submission:
        snapshots = load_score_reports()
        snapshots.append(
            _build_score_report_snapshot(
                submission_id=str(matched_submission.get("id")),
                project=project,
                report=matched_submission.get("report") or {},
                profile_snapshot=profile_snapshot,
                scoring_engine_version=scoring_engine_version,
            )
        )
        save_score_reports(snapshots)
        if evidence_units_new:
            all_units = load_evidence_units()
            all_units = _replace_submission_evidence_units(
                all_units,
                submission_id=str(matched_submission.get("id")),
                new_units=evidence_units_new,
            )
            save_evidence_units(all_units)

        report = matched_submission.get("report") or {}
        dimension_scores = {
            dim_id: (dim.get("score", 0.0) if isinstance(dim, dict) else 0.0)
            for dim_id, dim in (report.get("dimension_scores") or {}).items()
        }
        penalty_count = len(report.get("penalties", []))
        if not _report_is_blocked(report):
            record_history_score(
                project_id=project_id,
                submission_id=str(matched_submission.get("id")),
                filename=str(matched_submission.get("filename", "")),
                total_score=float(report.get("total_score", report.get("rule_total_score", 0.0))),
                dimension_scores=dimension_scores,
                penalty_count=penalty_count,
            )

    qt_results = load_qingtian_results()
    matched_qt = next(
        (
            r
            for r in qt_results
            if str((r.get("raw_payload") or {}).get("ground_truth_record_id") or "") == source_gt_id
        ),
        None,
    )
    project_score_scale = _resolve_project_score_scale_max(project)
    gt_for_learning = _ground_truth_record_for_learning(
        gt_record,
        default_score_scale_max=project_score_scale,
    )
    feature_confidence_update: Dict[str, object] = {
        "updated": 0,
        "retired": 0,
        "reason": "not_executed",
    }
    report_for_feedback = matched_submission.get("report")
    if isinstance(report_for_feedback, dict):
        try:
            feature_confidence_update = _auto_update_feature_confidence_on_ground_truth(
                report=report_for_feedback,
                gt_record=gt_record,
                project_score_scale_max=project_score_scale,
            )
        except Exception as exc:
            feature_confidence_update = {
                "updated": 0,
                "retired": 0,
                "reason": "feature_confidence_update_error",
                "error": str(exc),
            }

    if source_gt_id:
        all_gt_records = load_ground_truth()
        changed_gt = False
        for row in all_gt_records:
            if str(row.get("id") or "") != source_gt_id:
                continue
            row["feature_confidence_update"] = feature_confidence_update
            row["updated_at"] = _now_iso()
            changed_gt = True
            break
        if changed_gt:
            save_ground_truth(all_gt_records)

    if matched_qt is None:
        qt_results.append(
            {
                "id": str(uuid4()),
                "submission_id": str(matched_submission.get("id")),
                "qingtian_model_version": str(
                    project.get("qingtian_model_version") or DEFAULT_QINGTIAN_MODEL_VERSION
                ),
                "qt_total_score": float(gt_for_learning.get("final_score", 0.0)),
                "qt_dim_scores": None,
                "qt_reasons": [
                    {
                        "kind": "ground_truth",
                        "text": f"评委分: {gt_record.get('judge_scores')}",
                    }
                ],
                "raw_payload": {
                    "ground_truth_record_id": source_gt_id,
                    "source": gt_record.get("source"),
                    "judge_scores": gt_record.get("judge_scores"),
                    "final_score": gt_record.get("final_score"),
                    "final_score_raw": gt_for_learning.get("final_score_raw"),
                    "final_score_100": gt_for_learning.get("final_score"),
                    "score_scale_max": gt_for_learning.get("score_scale_max"),
                    "feature_confidence_update": feature_confidence_update,
                },
                "created_at": _now_iso(),
            }
        )
        save_qingtian_results(qt_results)
    else:
        raw_payload = matched_qt.get("raw_payload")
        if not isinstance(raw_payload, dict):
            raw_payload = {}
        raw_payload["feature_confidence_update"] = feature_confidence_update
        matched_qt["raw_payload"] = raw_payload
        save_qingtian_results(qt_results)

    if str(project.get("status") or "") == "scoring_preparation":
        project["status"] = "submitted_to_qingtian"
        project["updated_at"] = _now_iso()
        save_projects(projects)

    _refresh_project_reflection_objects(project_id)


def _rebuild_project_anchors_and_requirements(
    project_id: str,
) -> tuple[List[Dict[str, object]], List[Dict[str, object]]]:
    merged_text = _build_constraints_source_text(project_id)
    project = next((p for p in load_projects() if str(p.get("id")) == project_id), {})
    region = str(project.get("region") or DEFAULT_REGION)
    scoring_engine_version = str(
        project.get("scoring_engine_version_locked") or DEFAULT_SCORING_ENGINE_LOCKED
    )
    anchors = extract_project_anchors_from_text(project_id, merged_text)
    requirements = build_project_requirements_from_anchors(
        project_id,
        anchors,
        region=region,
        scoring_engine_version=scoring_engine_version,
    )

    all_anchors = [a for a in load_project_anchors() if str(a.get("project_id")) != project_id]
    all_requirements = [
        r for r in load_project_requirements() if str(r.get("project_id")) != project_id
    ]
    all_anchors.extend(anchors)
    all_requirements.extend(requirements)
    save_project_anchors(all_anchors)
    save_project_requirements(all_requirements)
    return anchors, requirements


def _build_constraint_pack(project_id: str) -> Dict[str, object]:
    multipliers, profile, _ = _resolve_project_scoring_context(project_id)
    anchors = [a for a in load_project_anchors() if str(a.get("project_id")) == project_id]
    requirements = [
        r for r in load_project_requirements() if str(r.get("project_id")) == project_id
    ]
    pack_versions = sorted(
        {
            str(r.get("source_pack_version") or "").strip()
            for r in requirements
            if str(r.get("source_pack_version") or "").strip()
        }
    )
    mandatory = [r for r in requirements if bool(r.get("mandatory"))]
    priority_order = sorted(
        DIMENSION_IDS, key=lambda d: float(multipliers.get(d, 1.0)), reverse=True
    )

    # 基础阈值：高优先维度略提高闭环/落地阈值
    thresholds: Dict[str, Dict[str, float]] = {}
    for dim_id in DIMENSION_IDS:
        weight = float(multipliers.get(dim_id, 1.0))
        bump = 0.2 if weight >= 1.15 else 0.0
        thresholds[dim_id] = {
            "coverage_min": round(1.2 + bump, 2),
            "closure_min": round(1.2 + bump, 2),
            "landing_min": round(1.2 + bump, 2),
            "specificity_min": round(1.0 + bump, 2),
        }

    return {
        "project_id": project_id,
        "expert_profile_snapshot": profile,
        "anchors_required": [
            {
                "anchor_key": a.get("anchor_key"),
                "expectation": "must_appear_same_value"
                if a.get("value_num") is not None
                else "must_not_conflict",
            }
            for a in anchors
        ],
        "requirements_mandatory": [
            {
                "requirement_id": r.get("id"),
                "dimension_id": r.get("dimension_id"),
                "req_label": r.get("req_label"),
                "req_type": r.get("req_type"),
            }
            for r in mandatory
        ],
        "dimension_thresholds": thresholds,
        "priority_order": priority_order,
        "requirement_pack_versions": pack_versions,
        "generated_at": _now_iso(),
    }


app = FastAPI(
    title="青天评标系统 API",
    version="1.0.0",
    description="""
## 施工组织设计智能评审系统

本系统提供施工组织设计文档的自动化评分、对比分析和自适应优化功能。

### 核心功能

- **智能评分**: 基于多维度评分规则，自动分析施组文档质量
- **批量处理**: 支持多文档并行处理，提升效率
- **对比分析**: 多版本施组横向对比，发现改进点
- **自适应学习**: 根据历史数据自动优化评分参数
- **DOCX 导出**: 生成专业格式的评审报告

### 认证方式

部分端点需要 API Key 认证，通过 `X-API-Key` 请求头传递。

### 快速开始

1. 创建项目: `POST /api/v1/projects`
2. 上传施组: `POST /api/v1/projects/{project_id}/shigong`
3. 查看结果: `GET /api/v1/projects/{project_id}/submissions`
4. 对比分析: `GET /api/v1/projects/{project_id}/compare`

### API 版本

当前 API 版本: v1 (路径前缀: `/api/v1`)
""",
    openapi_tags=OPENAPI_TAGS,
    contact={
        "name": "智飞文档生成系统",
        "email": "support@zhifei.example.com",
    },
    license_info={
        "name": "MIT License",
        "url": "https://opensource.org/licenses/MIT",
    },
)

# Setup rate limiting (infrastructure ready, decorators disabled due to compatibility)
setup_rate_limiting(app)


@app.exception_handler(StarletteHTTPException)
async def _web_405_fallback_handler(request: Request, exc: StarletteHTTPException):
    """
    将网页上传入口的 405 统一回退为首页提示，避免浏览器直接展示 JSON 报错页。
    """
    if exc.status_code == 405:
        path = (request.url.path or "").rstrip("/")
        if path in ("/web/upload_materials", "/web/upload_shigong"):
            project_id = request.query_params.get("project_id", "")
            message = "请在主页选择文件后点击“上传资料”提交。"
            anchor = "#section-materials"
            if path == "/web/upload_shigong":
                message = "请在主页选择文件后点击“上传施组”提交。"
                anchor = "#section-shigong"
            return RedirectResponse(
                url=_web_upload_redirect_url(project_id, message, anchor),
                status_code=303,
            )
    return await http_exception_handler(request, exc)


def parse_accept_language(accept_language: str | None) -> str:
    """
    解析 Accept-Language header，返回最佳匹配的语言代码。

    支持的格式：
    - "zh"
    - "zh-CN"
    - "en-US,en;q=0.9,zh;q=0.8"

    Args:
        accept_language: Accept-Language header 值

    Returns:
        匹配的语言代码 (zh/en)，默认返回 zh
    """
    if not accept_language:
        return DEFAULT_LOCALE

    # 解析语言优先级列表
    languages = []
    for part in accept_language.split(","):
        part = part.strip()
        if not part:
            continue

        # 解析 q 值
        if ";q=" in part:
            lang, q_str = part.split(";q=", 1)
            try:
                q = float(q_str)
            except ValueError:
                q = 1.0
        else:
            lang = part
            q = 1.0

        # 提取主语言代码 (zh-CN -> zh)
        lang = lang.split("-")[0].lower()
        languages.append((lang, q))

    # 按优先级排序
    languages.sort(key=lambda x: x[1], reverse=True)

    # 查找第一个支持的语言
    for lang, _ in languages:
        if lang in SUPPORTED_LOCALES:
            return lang

    return DEFAULT_LOCALE


def get_locale(
    accept_language: str | None = Header(None, alias="Accept-Language"),
) -> str:
    """
    FastAPI 依赖：从 Accept-Language header 获取语言代码。

    用法:
        @router.get("/endpoint")
        def endpoint(locale: str = Depends(get_locale)):
            message = t("api.some_key", locale=locale)
    """
    return parse_accept_language(accept_language)


def _build_data_hygiene_report(*, apply: bool) -> Dict[str, object]:
    """
    数据卫生巡检/修复：
    - 清理 project_id 不存在的孤儿记录
    - 清理 submission_id 不存在的孤儿记录
    - 清理 project_id 维度的 dict 型映射残留键
    """
    ensure_data_dirs()
    projects = load_projects()
    valid_project_ids = {str(p.get("id") or "").strip() for p in projects if str(p.get("id") or "")}
    datasets: List[Dict[str, object]] = []
    orphan_records_total = 0
    cleaned_records_total = 0

    def _append_dataset(
        *,
        name: str,
        total: int,
        orphan_count: int,
        cleaned_count: int = 0,
        mode: str = "project_id",
    ) -> None:
        nonlocal orphan_records_total, cleaned_records_total
        orphan_records_total += int(orphan_count)
        cleaned_records_total += int(cleaned_count)
        datasets.append(
            {
                "name": name,
                "total": int(total),
                "orphan_count": int(orphan_count),
                "cleaned_count": int(cleaned_count),
                "mode": mode,
            }
        )

    def _scan_project_scoped_rows(
        *,
        name: str,
        rows: List[Dict[str, object]],
        save_fn,
    ) -> List[Dict[str, object]]:
        kept: List[Dict[str, object]] = []
        orphan_count = 0
        for row in rows:
            if not isinstance(row, dict):
                kept.append(row)
                continue
            pid = str(row.get("project_id") or "").strip()
            if pid and pid not in valid_project_ids:
                orphan_count += 1
                continue
            kept.append(row)
        cleaned_count = orphan_count if apply else 0
        if apply and orphan_count > 0:
            save_fn(kept)
        _append_dataset(
            name=name,
            total=len(rows),
            orphan_count=orphan_count,
            cleaned_count=cleaned_count,
            mode="project_id",
        )
        return kept

    submissions_rows = load_submissions()
    submissions_kept = _scan_project_scoped_rows(
        name="submissions",
        rows=submissions_rows,
        save_fn=save_submissions,
    )
    valid_submission_ids = {
        str(row.get("id") or "").strip()
        for row in submissions_kept
        if isinstance(row, dict) and str(row.get("id") or "").strip()
    }

    def _scan_submission_linked_rows(
        *,
        name: str,
        rows: List[Dict[str, object]],
        save_fn,
        submission_key: str = "submission_id",
    ) -> None:
        kept: List[Dict[str, object]] = []
        orphan_count = 0
        for row in rows:
            if not isinstance(row, dict):
                kept.append(row)
                continue
            pid = str(row.get("project_id") or "").strip()
            sid = str(row.get(submission_key) or "").strip()
            orphan_by_project = bool(pid) and pid not in valid_project_ids
            orphan_by_submission = bool(sid) and sid not in valid_submission_ids
            if orphan_by_project or orphan_by_submission:
                orphan_count += 1
                continue
            kept.append(row)
        cleaned_count = orphan_count if apply else 0
        if apply and orphan_count > 0:
            save_fn(kept)
        _append_dataset(
            name=name,
            total=len(rows),
            orphan_count=orphan_count,
            cleaned_count=cleaned_count,
            mode=f"project_id|{submission_key}",
        )

    _scan_project_scoped_rows(
        name="materials",
        rows=load_materials(),
        save_fn=save_materials,
    )
    _scan_project_scoped_rows(
        name="learning_profiles",
        rows=load_learning_profiles(),
        save_fn=save_learning_profiles,
    )
    _scan_project_scoped_rows(
        name="score_history",
        rows=load_score_history(),
        save_fn=save_score_history,
    )
    _scan_project_scoped_rows(
        name="ground_truth_scores",
        rows=load_ground_truth(),
        save_fn=save_ground_truth,
    )
    _scan_project_scoped_rows(
        name="project_anchors",
        rows=load_project_anchors(),
        save_fn=save_project_anchors,
    )
    _scan_project_scoped_rows(
        name="project_requirements",
        rows=load_project_requirements(),
        save_fn=save_project_requirements,
    )
    _scan_project_scoped_rows(
        name="delta_cases",
        rows=load_delta_cases(),
        save_fn=save_delta_cases,
    )
    _scan_project_scoped_rows(
        name="calibration_samples",
        rows=load_calibration_samples(),
        save_fn=save_calibration_samples,
    )

    patch_packages_rows = _scan_project_scoped_rows(
        name="patch_packages",
        rows=load_patch_packages(),
        save_fn=save_patch_packages,
    )
    valid_patch_ids = {
        str(p.get("id") or "").strip()
        for p in patch_packages_rows
        if isinstance(p, dict) and str(p.get("id") or "").strip()
    }

    # patch_deployments 额外校验 patch_id
    patch_deployments_rows = load_patch_deployments()
    patch_deployments_kept: List[Dict[str, object]] = []
    patch_deployments_orphan = 0
    for row in patch_deployments_rows:
        if not isinstance(row, dict):
            patch_deployments_kept.append(row)
            continue
        pid = str(row.get("project_id") or "").strip()
        patch_id = str(row.get("patch_id") or "").strip()
        orphan_by_project = bool(pid) and pid not in valid_project_ids
        orphan_by_patch = bool(patch_id) and patch_id not in valid_patch_ids
        if orphan_by_project or orphan_by_patch:
            patch_deployments_orphan += 1
            continue
        patch_deployments_kept.append(row)
    if apply and patch_deployments_orphan > 0:
        save_patch_deployments(patch_deployments_kept)
    _append_dataset(
        name="patch_deployments",
        total=len(patch_deployments_rows),
        orphan_count=patch_deployments_orphan,
        cleaned_count=(patch_deployments_orphan if apply else 0),
        mode="project_id|patch_id",
    )

    _scan_submission_linked_rows(
        name="score_reports",
        rows=load_score_reports(),
        save_fn=save_score_reports,
        submission_key="submission_id",
    )
    _scan_submission_linked_rows(
        name="evidence_units",
        rows=load_evidence_units(),
        save_fn=save_evidence_units,
        submission_key="submission_id",
    )
    _scan_submission_linked_rows(
        name="qingtian_results",
        rows=load_qingtian_results(),
        save_fn=save_qingtian_results,
        submission_key="submission_id",
    )

    def _scan_project_map(*, name: str, data: Dict[str, object], save_fn) -> None:
        if not isinstance(data, dict):
            _append_dataset(name=name, total=0, orphan_count=0, cleaned_count=0, mode="project_map")
            return
        orphan_keys = [str(k) for k in data.keys() if str(k) not in valid_project_ids]
        cleaned_count = len(orphan_keys) if apply else 0
        if apply and orphan_keys:
            new_data = {k: v for k, v in data.items() if str(k) in valid_project_ids}
            save_fn(new_data)
        _append_dataset(
            name=name,
            total=len(data),
            orphan_count=len(orphan_keys),
            cleaned_count=cleaned_count,
            mode="project_map",
        )

    _scan_project_map(
        name="project_context",
        data=load_project_context(),
        save_fn=save_project_context,
    )
    _scan_project_map(
        name="evolution_reports",
        data=load_evolution_reports(),
        save_fn=save_evolution_reports,
    )

    recommendations: List[str] = []
    if orphan_records_total <= 0:
        recommendations.append("数据卫生良好：未发现跨项目孤儿记录。")
    elif apply:
        recommendations.append(
            f"已清理孤儿记录 {cleaned_records_total} 条，建议执行一次 doctor/acceptance 回归。"
        )
    else:
        recommendations.append(
            f"发现孤儿记录 {orphan_records_total} 条，建议调用 /api/v1/system/data_hygiene/repair 进行修复。"
        )
    if orphan_records_total > 0:
        recommendations.append(
            "建议在批量删除项目后执行数据卫生巡检，避免历史孤儿记录影响统计与审计。"
        )

    return {
        "generated_at": _now_iso(),
        "apply_mode": bool(apply),
        "valid_project_count": len(valid_project_ids),
        "orphan_records_total": int(orphan_records_total),
        "cleaned_records_total": int(cleaned_records_total),
        "datasets": datasets,
        "recommendations": recommendations,
    }


def _run_system_self_check(project_id: Optional[str]) -> Dict[str, object]:
    items: List[Dict[str, object]] = []

    def add(name: str, ok: bool, detail: str = "") -> None:
        items.append({"name": name, "ok": bool(ok), "detail": detail or None})

    # health (always true if request reaches server)
    add("health", True, "service reachable")

    # config
    try:
        load_config()
        add("config", True, "rubric/lexicon loaded")
    except Exception as e:
        add("config", False, str(e))

    # data dirs + writable test
    try:
        ensure_data_dirs()
        with tempfile.NamedTemporaryFile(
            prefix="selfcheck_", suffix=".tmp", dir="data", delete=True
        ) as _:
            pass
        add("data_dirs_writable", True, "data directory writable")
    except Exception as e:
        add("data_dirs_writable", False, str(e))

    # status providers
    try:
        s = get_auth_status()
        enabled = bool(s.get("auth_enabled", s.get("enabled", False)))
        add("auth_status", True, f"enabled={enabled}")
    except Exception as e:
        add("auth_status", False, str(e))
    try:
        s = get_rate_limit_status()
        add("rate_limit_status", True, f"enabled={bool(s.get('enabled'))}")
    except Exception as e:
        add("rate_limit_status", False, str(e))

    # parser/runtime capability checks
    pdf_backend = _pdf_backend_name()
    add(
        "parser_pdf",
        pdf_backend != "none",
        (f"backend={pdf_backend}" if pdf_backend != "none" else "PyMuPDF/pypdf missing"),
    )
    add(
        "parser_docx",
        Document is not None,
        "python-docx available" if Document is not None else "python-docx missing",
    )
    ocr_available = bool(pytesseract is not None and Image is not None)
    if ocr_available:
        try:
            version = str(pytesseract.get_tesseract_version()) if pytesseract is not None else ""
            add("parser_ocr", True, f"tesseract={version}")
        except Exception:
            add("parser_ocr", True, "pytesseract available")
    else:
        add("parser_ocr", False, "pytesseract or PIL missing")
    try:
        dwg_bins = _resolve_dwg_converter_binaries()
        add(
            "parser_dwg_converter",
            bool(dwg_bins),
            f"found={','.join(Path(p).name for p in dwg_bins)}" if dwg_bins else "not_found",
        )
    except Exception as e:
        add("parser_dwg_converter", False, str(e))

    # data hygiene (non-blocking): 用于识别孤儿项目数据，避免统计/审计偏差
    try:
        hygiene = _build_data_hygiene_report(apply=False)
        orphan_count = int(_to_float_or_none(hygiene.get("orphan_records_total")) or 0)
        impacted = sum(
            1
            for row in (hygiene.get("datasets") or [])
            if int(_to_float_or_none((row or {}).get("orphan_count")) or 0) > 0
        )
        add(
            "data_hygiene",
            orphan_count == 0,
            f"orphan_records={orphan_count}, impacted_datasets={impacted}",
        )
    except Exception as e:
        add("data_hygiene", False, str(e))

    # project-specific checks
    if project_id:
        try:
            projects = load_projects()
            target = next((p for p in projects if str(p.get("id")) == project_id), None)
            if target is None:
                add("project_exists", False, f"project not found: {project_id}")
            else:
                add("project_exists", True, str(target.get("name") or project_id))
                try:
                    materials_count = len(
                        [m for m in load_materials() if str(m.get("project_id")) == project_id]
                    )
                    add("project_materials_listable", True, f"count={materials_count}")
                except Exception as e:
                    add("project_materials_listable", False, str(e))
                try:
                    submissions_count = len(
                        [s for s in load_submissions() if str(s.get("project_id")) == project_id]
                    )
                    add("project_submissions_listable", True, f"count={submissions_count}")
                except Exception as e:
                    add("project_submissions_listable", False, str(e))
                try:
                    readiness = _build_scoring_readiness(project_id, target)
                    ready = bool(readiness.get("ready"))
                    issues = (
                        readiness.get("issues") if isinstance(readiness.get("issues"), list) else []
                    )
                    issues_preview = "；".join(str(x) for x in issues[:2]) if issues else "-"
                    add(
                        "project_scoring_readiness",
                        True,
                        f"ready={ready}, issues={issues_preview}",
                    )
                except Exception as e:
                    add("project_scoring_readiness", False, str(e))
        except Exception as e:
            add("project_exists", False, str(e))

    # `parser_ocr` / `parser_dwg_converter` 属于增强能力，不应阻断系统基础可用性判断。
    required_item_names = {
        "health",
        "config",
        "data_dirs_writable",
        "auth_status",
        "rate_limit_status",
        "parser_pdf",
        "parser_docx",
    }
    required_items = [x for x in items if str(x.get("name")) in required_item_names]
    all_ok = bool(required_items) and all(bool(x.get("ok")) for x in required_items)
    return {
        "ok": all_ok,
        "checked_at": _now_iso(),
        "items": items,
    }


# ==================== 健康检查端点（根路径，便于容器编排系统访问） ====================


@app.get("/health", response_model=HealthResponse, tags=["健康检查"])
def health_check() -> HealthResponse:
    """
    健康检查（Liveness Probe）。

    返回服务存活状态。只要服务进程正常运行即返回 healthy。
    用于 Kubernetes liveness probe 或负载均衡健康检查。

    - 不检查外部依赖
    - 响应时间应小于 100ms
    """
    return HealthResponse(status="healthy", version="1.0.0")


@app.get("/metrics", tags=["监控指标"], include_in_schema=True)
def prometheus_metrics():
    """
    Prometheus 指标端点。

    返回 Prometheus 文本格式的运行时指标，包括：
    - HTTP 请求计数和延迟
    - 评分请求统计和分数分布
    - 项目和提交数量
    - 配置状态

    用于 Prometheus 服务器抓取（scrape）或 Grafana 可视化。
    """
    from fastapi.responses import Response

    # 更新项目统计指标
    try:
        ensure_data_dirs()
        projects = load_projects()
        submissions = load_submissions()
        update_project_stats(len(projects), len(submissions))
    except Exception:
        pass  # 指标端点不应因统计失败而报错

    content = get_metrics()
    return Response(content=content, media_type="text/plain; charset=utf-8")


@app.get("/ready", response_model=ReadyResponse, tags=["健康检查"])
def readiness_check() -> ReadyResponse:
    """
    就绪检查（Readiness Probe）。

    返回服务是否准备好处理请求。检查配置和数据目录是否可用。
    用于 Kubernetes readiness probe，决定是否将流量路由到此实例。

    检查项目：
    - config: 配置文件是否可加载
    - data_dirs: 数据目录是否存在且可访问
    """
    checks = {}

    # 检查配置是否可加载
    try:
        load_config()
        checks["config"] = True
    except Exception:
        checks["config"] = False

    # 检查数据目录是否可用
    try:
        ensure_data_dirs()
        checks["data_dirs"] = True
    except Exception:
        checks["data_dirs"] = False

    # 所有检查通过则就绪
    all_ready = all(checks.values())
    status = "ready" if all_ready else "not_ready"

    return ReadyResponse(status=status, checks=checks)


@app.get("/__ping__", include_in_schema=False)
def ui_click_ping(btn: str = "") -> dict:
    return {"ok": True, "btn": btn}


# API v1 路由
router = APIRouter(prefix="/api/v1")


@router.get("/auth/status", tags=["系统状态"])
def auth_status() -> dict:
    """
    获取 API 认证状态。

    返回当前系统的认证配置状态，包括是否启用认证、认证方式等信息。
    """
    return get_auth_status()


@router.get("/rate_limit/status", tags=["系统状态"])
def rate_limit_status() -> dict:
    """
    获取限流状态。

    返回当前系统的请求限流配置，包括限流策略、配额等信息。
    """
    return get_rate_limit_status()


@router.get("/cache/stats", response_model=CacheStatsResponse, tags=["系统状态"])
def cache_stats() -> CacheStatsResponse:
    """
    获取评分缓存统计。

    返回缓存的运行时统计信息，包括：
    - 总请求数
    - 缓存命中数与未命中数
    - 缓存命中率
    - 当前缓存条目数
    - 驱逐数量

    用于监控缓存效率和容量规划。
    """
    stats = get_cache_stats()
    return CacheStatsResponse(**stats)


@router.post(
    "/cache/clear",
    response_model=CacheClearResponse,
    tags=["系统状态"],
    responses=RESPONSES_401,
)
def cache_clear(
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> CacheClearResponse:
    """
    清空评分缓存。

    清除所有缓存的评分结果。在配置变更后或需要强制重新评分时使用。

    **需要 API Key 认证**

    ⚠️ 此操作会清除所有缓存，不可恢复

    支持 Accept-Language header 进行多语言响应。
    """
    count = clear_score_cache()
    return CacheClearResponse(
        cleared=True,
        count=count,
        message=t("api.cache_cleared", locale=locale)
        if count > 0
        else t("api.cache_empty", locale=locale),
    )


@router.get("/config/status", response_model=ConfigStatusResponse, tags=["系统状态"])
def config_status() -> ConfigStatusResponse:
    """
    获取配置加载状态。

    返回配置文件的缓存状态、文件路径和修改时间信息。
    用于监控配置热加载机制是否正常工作。
    """
    status = get_config_status()
    return ConfigStatusResponse(
        cached=status["cached"],
        rubric_path=status["rubric_path"],
        lexicon_path=status["lexicon_path"],
        needs_reload=status["needs_reload"],
        rubric_mtime=status["rubric_mtime"],
        lexicon_mtime=status["lexicon_mtime"],
    )


@router.get("/config/llm_status", response_model=LLMBackendStatus, tags=["系统状态"])
def llm_status() -> LLMBackendStatus:
    """
    获取进化 LLM 后端配置状态（不暴露密钥）。
    便于确认当前生效的后端及哪些 API 已配置。
    """
    s = get_llm_backend_status()
    return LLMBackendStatus(
        evolution_backend=s["evolution_backend"],
        spark_configured=s["spark_configured"],
        openai_configured=s["openai_configured"],
        gemini_configured=s["gemini_configured"],
        ollama_configured=s["ollama_configured"],
    )


@router.get(
    "/scoring/factors",
    response_model=ScoringFactorsResponse,
    tags=["系统状态"],
    responses=RESPONSES_404,
)
def scoring_factors(
    project_id: Optional[str] = Query(None, description="可选：指定项目ID，返回该项目的编制要求"),
    locale: str = Depends(get_locale),
) -> ScoringFactorsResponse:
    """
    获取评分体系总览。

    返回当前系统已启用的维度评分因子、扣分规则、Lint问题码，以及章节完整性/图文/组织架构等编制要求。
    便于对外分析与审阅当前评分标准。
    """
    ensure_data_dirs()
    if project_id:
        projects = load_projects()
        if not any(str(p.get("id")) == project_id for p in projects):
            raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    return ScoringFactorsResponse(**_build_scoring_factors_overview(project_id))


@router.get(
    "/scoring/factors/markdown",
    response_model=ScoringFactorsMarkdownResponse,
    tags=["系统状态"],
    responses=RESPONSES_404,
)
def scoring_factors_markdown(
    project_id: Optional[str] = Query(None, description="可选：指定项目ID，返回该项目的编制要求"),
    locale: str = Depends(get_locale),
) -> ScoringFactorsMarkdownResponse:
    """导出评分体系总览 Markdown 文本，便于外部模型或文档系统直接使用。"""
    ensure_data_dirs()
    if project_id:
        projects = load_projects()
        if not any(str(p.get("id")) == project_id for p in projects):
            raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    payload = _build_scoring_factors_overview(project_id)
    return ScoringFactorsMarkdownResponse(
        project_id=project_id,
        markdown=_render_scoring_factors_markdown(payload),
    )


@router.get(
    "/projects/{project_id}/analysis_bundle",
    response_model=AnalysisBundleResponse,
    tags=["洞察与学习"],
    responses={**RESPONSES_404},
)
def project_analysis_bundle(
    project_id: str,
    locale: str = Depends(get_locale),
) -> AnalysisBundleResponse:
    """
    导出项目分析包（Markdown），包含：
    - 项目级 V1/V2/V2+Calib 指标
    - 当前评分体系与章节要求
    """
    ensure_data_dirs()
    projects = load_projects()
    project = next((p for p in projects if str(p.get("id")) == project_id), None)
    if project is None:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))

    factors_payload = _build_scoring_factors_overview(project_id)
    eval_payload = evaluate_project_variants(
        project_id=project_id,
        submissions=load_submissions(),
        score_reports=load_score_reports(),
        qingtian_results=load_qingtian_results(),
    )
    markdown = _render_project_analysis_bundle_markdown(
        project=project,
        factors_payload=factors_payload,
        evaluation_payload=eval_payload,
    )
    return AnalysisBundleResponse(
        project_id=project_id,
        markdown=markdown,
        generated_at=_now_iso(),
    )


@router.get(
    "/projects/{project_id}/analysis_bundle.md",
    tags=["洞察与学习"],
    responses={**RESPONSES_404},
)
def project_analysis_bundle_markdown_file(
    project_id: str,
    locale: str = Depends(get_locale),
) -> Response:
    """下载项目分析包 Markdown 文件。"""
    bundle = project_analysis_bundle(project_id=project_id, locale=locale)
    if isinstance(bundle, dict):
        markdown = str(bundle.get("markdown") or "")
    else:
        markdown = str(bundle.markdown or "")
    filename = f"analysis_bundle_{project_id}.md"
    return Response(
        content=markdown,
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post(
    "/config/reload",
    response_model=ConfigReloadResponse,
    tags=["系统状态"],
    responses=RESPONSES_401,
)
def config_reload_endpoint(
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> ConfigReloadResponse:
    """
    强制重新加载配置文件。

    立即从磁盘重新加载 rubric.yaml 和 lexicon.yaml 配置文件。
    用于在修改配置文件后立即生效，无需重启服务。

    **需要 API Key 认证**

    支持 Accept-Language header 进行多语言响应。
    """
    try:
        reload_config()
        return ConfigReloadResponse(reloaded=True, message=t("api.config_reloaded", locale=locale))
    except Exception as e:
        return ConfigReloadResponse(reloaded=False, message=f"重载失败: {str(e)}")


@router.get(
    "/system/self_check",
    response_model=SelfCheckResponse,
    tags=["系统状态"],
)
def system_self_check(
    project_id: Optional[str] = Query(None, description="可选：附带检查某项目读取能力"),
) -> SelfCheckResponse:
    """运行系统自检并返回结构化结果。"""
    ensure_data_dirs()
    return SelfCheckResponse(**_run_system_self_check(project_id))


@router.get(
    "/system/data_hygiene",
    response_model=DataHygieneResponse,
    tags=["系统状态"],
)
def system_data_hygiene() -> DataHygieneResponse:
    """
    数据卫生巡检（只读）。

    用于检查项目删除/迁移后是否残留孤儿记录，避免影响评分统计与审计结论。
    """
    ensure_data_dirs()
    payload = _build_data_hygiene_report(apply=False)
    return DataHygieneResponse(**payload)


@router.post(
    "/system/data_hygiene/repair",
    response_model=DataHygieneResponse,
    tags=["系统状态"],
    responses={**RESPONSES_401},
)
def repair_system_data_hygiene(
    api_key: Optional[str] = Depends(verify_api_key),
) -> DataHygieneResponse:
    """
    数据卫生修复（写操作）。

    清理 project_id / submission_id / patch_id 失联导致的孤儿记录。
    """
    ensure_data_dirs()
    payload = _build_data_hygiene_report(apply=True)
    return DataHygieneResponse(**payload)


@router.post(
    "/score",
    response_model=ScoreReport,
    tags=["评分"],
    responses={**RESPONSES_401, **RESPONSES_422},
)
def score_endpoint(
    payload: ScoreRequest,
    api_key: Optional[str] = Depends(verify_api_key),
) -> ScoreReport:
    """
    对施组文本进行评分。

    接收施工组织设计文本内容，返回完整的评分报告，包括：
    - 总分及各维度得分
    - 命中的关键词证据
    - 扣分项及原因
    - 改进建议

    支持评分结果缓存，相同文本重复评分时直接返回缓存结果。

    **需要 API Key 认证**
    """
    config = load_config()

    # 尝试从缓存获取结果
    cached_result = get_cached_score(payload.text)
    if cached_result is not None:
        # 记录评分指标（命中缓存）
        record_score(cached_result.get("total_score", 0.0))
        return ScoreReport(**cached_result)

    # 缓存未命中，执行评分
    result = score_text(payload.text, config.rubric, config.lexicon)
    result_dict = result.model_dump()

    # 缓存评分结果
    cache_score_result(payload.text, result_dict)

    # 记录评分指标
    record_score(result.total_score)
    return result


@router.post(
    "/projects",
    response_model=ProjectRecord,
    tags=["项目管理"],
    responses={**RESPONSES_401, **RESPONSES_422},
)
def create_project(
    payload: ProjectCreate,
    api_key: Optional[str] = Depends(verify_api_key),
) -> ProjectRecord:
    """
    创建新项目。

    项目是施组评分的容器，可以包含多个施组文档和材料文件。
    创建项目后，可以上传材料和施组进行评分。

    **需要 API Key 认证**
    """
    ensure_data_dirs()
    projects = load_projects()
    if any(p["name"] == payload.name for p in projects):
        raise HTTPException(status_code=422, detail="项目名称已存在，请更换名称")
    project_id = str(uuid4())
    record = {
        "id": project_id,
        "name": payload.name,
        "meta": payload.meta or {},
        "region": DEFAULT_REGION,
        "expert_profile_id": None,
        "qingtian_model_version": DEFAULT_QINGTIAN_MODEL_VERSION,
        "scoring_engine_version_locked": DEFAULT_SCORING_ENGINE_LOCKED,
        "calibrator_version_locked": DEFAULT_CALIBRATOR_LOCKED,
        "status": "scoring_preparation",
        "created_at": _now_iso(),
        "updated_at": _now_iso(),
    }
    _ensure_project_v2_fields(record)
    projects.append(record)
    save_projects(projects)
    return ProjectRecord(**record)


@router.get("/projects", response_model=list[ProjectRecord], tags=["项目管理"])
def list_projects() -> list[ProjectRecord]:
    """
    获取所有项目列表。

    返回系统中已创建的所有项目记录。
    """
    ensure_data_dirs()
    projects = load_projects()
    if not os.environ.get("PYTEST_CURRENT_TEST"):
        active_projects = [
            p
            for p in projects
            if str(p.get("id") or "") != "p1" and not str(p.get("name") or "").startswith("E2E_")
        ]
        if not active_projects:
            recovered = _recover_latest_orphan_project(projects)
            if recovered is not None:
                projects = load_projects()
    changed = False
    for p in projects:
        changed = _ensure_project_v2_fields(p) or changed
    if changed:
        save_projects(projects)
    return [ProjectRecord(**p) for p in projects]


@router.get(
    "/projects/{project_id}/expert-profile",
    response_model=ProjectExpertProfileResponse,
    tags=["项目管理"],
    responses={**RESPONSES_404},
)
def get_project_expert_profile(
    project_id: str, locale: str = Depends(get_locale)
) -> ProjectExpertProfileResponse:
    """获取项目当前生效的专家16维关注度配置。"""
    ensure_data_dirs()
    projects = load_projects()
    try:
        project = _find_project(project_id, projects)
    except HTTPException:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))

    project_changed = _ensure_project_v2_fields(project)
    profiles = load_expert_profiles()
    profile, created = _ensure_project_expert_profile(project, profiles)
    if project_changed or created:
        save_projects(projects)
    if created:
        save_expert_profiles(profiles)
    return ProjectExpertProfileResponse(
        project=ProjectRecord(**project),
        expert_profile=ExpertProfileRecord(**profile),
    )


@router.put(
    "/projects/{project_id}/expert-profile",
    response_model=ProjectExpertProfileResponse,
    tags=["项目管理"],
    responses={**RESPONSES_401, **RESPONSES_404, **RESPONSES_409},
)
def update_project_expert_profile(
    project_id: str,
    payload: ExpertProfileUpdate,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> ProjectExpertProfileResponse:
    """保存新的专家关注度配置并绑定到项目。"""
    ensure_data_dirs()
    projects = load_projects()
    try:
        project = _find_project(project_id, projects)
    except HTTPException:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))

    _ensure_project_v2_fields(project)
    _assert_project_profile_operation_unlocked(project, bool(payload.force_unlock))
    weights_raw = _coerce_weights_raw(payload.weights_raw)
    profile_name = (
        payload.name or ""
    ).strip() or f"{project.get('name', '项目')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    profile = _new_expert_profile(profile_name, weights_raw)

    profiles = load_expert_profiles()
    profiles.append(profile)
    save_expert_profiles(profiles)

    project["expert_profile_id"] = profile["id"]
    project["updated_at"] = _now_iso()
    save_projects(projects)

    return ProjectExpertProfileResponse(
        project=ProjectRecord(**project),
        expert_profile=ExpertProfileRecord(**profile),
    )


@router.post(
    "/projects/{project_id}/rescore",
    response_model=RescoreResponse,
    tags=["项目管理"],
    responses={**RESPONSES_401, **RESPONSES_404, **RESPONSES_409},
)
def rescore_project_submissions(
    project_id: str,
    payload: RescoreRequest,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> RescoreResponse:
    """按当前生效专家配置重算项目施组评分。"""
    ensure_data_dirs()
    started_at = _now_iso()
    projects = load_projects()
    try:
        project = _find_project(project_id, projects)
    except HTTPException:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))

    project_changed = _ensure_project_v2_fields(project)
    _assert_project_profile_operation_unlocked(project, bool(payload.force_unlock))
    score_scale_max = _normalize_score_scale_max(
        payload.score_scale_max,
        default=_resolve_project_score_scale_max(project),
    )
    project_meta = project.get("meta") if isinstance(project.get("meta"), dict) else {}
    project_meta = dict(project_meta or {})
    if int(project_meta.get("score_scale_max", DEFAULT_SCORE_SCALE_MAX)) != score_scale_max:
        project_meta["score_scale_max"] = score_scale_max
        project["meta"] = project_meta
        project_changed = True
    profiles = load_expert_profiles()
    profile, created = _ensure_project_expert_profile(project, profiles)
    if created:
        save_expert_profiles(profiles)
    if project_changed or created:
        save_projects(projects)

    if payload.scope not in {"project", "submission"}:
        raise HTTPException(status_code=422, detail="scope 仅支持 project 或 submission")
    if payload.scope == "submission" and not payload.submission_id:
        raise HTTPException(status_code=422, detail="scope=submission 时必须传 submission_id")

    if str(payload.scoring_engine_version or "").strip():
        project["scoring_engine_version_locked"] = payload.scoring_engine_version
    # 使用 _resolve_project_scoring_context 获取 multipliers，使进化产出的 dimension_multipliers 在评分时生效
    multipliers, profile_snapshot, _ = _resolve_project_scoring_context(project_id)
    if not multipliers and profile:
        multipliers = _weights_norm_to_dimension_multipliers(profile.get("weights_norm", {}))
    profile_for_meta = profile_snapshot if profile_snapshot else profile
    config = load_config()
    submissions = load_submissions()

    anchors: Optional[List[Dict[str, object]]] = None
    requirements: Optional[List[Dict[str, object]]] = None
    if payload.rebuild_anchors or payload.rebuild_requirements:
        anchors, requirements = _rebuild_project_anchors_and_requirements(project_id)

    if payload.scope == "submission":
        targets = [
            s
            for s in submissions
            if s.get("project_id") == project_id and s.get("id") == payload.submission_id
        ]
    else:
        targets = [s for s in submissions if s.get("project_id") == project_id]

    if not targets:
        raise HTTPException(status_code=404, detail=t("api.no_submissions", locale=locale))

    material_quality_snapshot, _ = _validate_material_gate_for_scoring(
        project_id,
        project,
        raise_on_fail=True,
    )

    score_reports = load_score_reports()
    all_evidence_units = load_evidence_units()
    generated = 0
    material_utilization_summaries: List[Dict[str, object]] = []
    material_utilization_by_submission: List[Dict[str, object]] = []
    material_utilization_gates: List[Dict[str, object]] = []
    failed_gate_filenames: List[str] = []
    now = _now_iso()
    for submission in targets:
        text = submission.get("text") or ""
        if not text.strip():
            continue
        report, evidence_units = _score_submission_for_project(
            submission_id=str(submission.get("id")),
            text=text,
            project_id=project_id,
            project=project,
            config=config,
            multipliers=multipliers,
            profile_snapshot=profile_snapshot,
            scoring_engine_version=payload.scoring_engine_version,
            anchors=anchors,
            requirements=requirements,
            material_quality_snapshot=material_quality_snapshot,
        )
        _apply_evolution_total_scale(project_id, report)
        all_evidence_units = _replace_submission_evidence_units(
            all_evidence_units,
            submission_id=str(submission.get("id")),
            new_units=evidence_units,
        )
        if not _report_is_blocked(report):
            _mark_report_scored(report, trigger="manual_rescore")
        report_meta = report.get("meta") if isinstance(report.get("meta"), dict) else {}
        report_meta = dict(report_meta or {})
        report_meta["score_scale_max"] = score_scale_max
        report_meta["score_scale_label"] = _score_scale_label(score_scale_max)
        report["meta"] = report_meta
        material_utilization = report_meta.get("material_utilization")
        if isinstance(material_utilization, dict):
            material_utilization_summaries.append(material_utilization)
            material_utilization_gate = (
                report_meta.get("material_utilization_gate")
                if isinstance(report_meta.get("material_utilization_gate"), dict)
                else {}
            )
            detail_item: Dict[str, object] = {
                "submission_id": str(submission.get("id") or ""),
                "filename": str(submission.get("filename") or ""),
                "summary": material_utilization,
            }
            if material_utilization_gate:
                detail_item["gate"] = material_utilization_gate
                material_utilization_gates.append(material_utilization_gate)
                if not bool(material_utilization_gate.get("passed", True)):
                    filename_text = str(submission.get("filename") or "")
                    if filename_text and filename_text not in failed_gate_filenames:
                        failed_gate_filenames.append(filename_text)
            alerts = report_meta.get("material_utilization_alerts")
            if isinstance(alerts, list):
                detail_item["alerts"] = [str(x) for x in alerts[:6] if str(x).strip()]
            material_utilization_by_submission.append(detail_item)

        submission["report"] = report
        submission["total_score"] = float(
            report.get("total_score", report.get("rule_total_score", 0.0))
        )
        submission["updated_at"] = now
        submission["expert_profile_id_used"] = (
            profile_for_meta.get("id") if profile_for_meta else None
        )

        snapshot = _build_score_report_snapshot(
            submission_id=str(submission.get("id")),
            project=project,
            report=report,
            profile_snapshot=profile_for_meta,
            scoring_engine_version=payload.scoring_engine_version,
        )
        score_reports.append(snapshot)
        generated += 1

        dimension_scores = {
            dim_id: dim.get("score", 0.0)
            for dim_id, dim in report.get("dimension_scores", {}).items()
        }
        penalty_count = len(report.get("penalties", []))
        if not _report_is_blocked(report):
            record_history_score(
                project_id=project_id,
                submission_id=str(submission.get("id")),
                filename=str(submission.get("filename", "")),
                total_score=report.get("total_score", 0.0),
                dimension_scores=dimension_scores,
                penalty_count=penalty_count,
            )

    save_submissions(submissions)
    save_score_reports(score_reports)
    save_evidence_units(all_evidence_units)
    project["updated_at"] = _now_iso()
    save_projects(projects)
    # 重评分属于有效反馈信号：自动刷新样本并触发校准/调权重闭环。
    feedback_closed_loop = _run_feedback_closed_loop_safe(
        project_id,
        locale=locale,
        trigger="rescore",
    )
    material_utilization = _aggregate_material_utilization_summaries(material_utilization_summaries)
    material_gate = (
        material_quality_snapshot.get("gate")
        if isinstance(material_quality_snapshot, dict)
        and isinstance(material_quality_snapshot.get("gate"), dict)
        else {}
    )
    material_utilization_alerts = _build_material_utilization_alerts(
        material_utilization,
        material_gate if isinstance(material_gate, dict) else {},
    )
    material_utilization_gate = _aggregate_material_utilization_gates(material_utilization_gates)
    material_utilization_gate["failed_filenames"] = failed_gate_filenames

    return RescoreResponse(
        ok=True,
        project_id=project_id,
        scoring_engine_version=payload.scoring_engine_version,
        expert_profile_id_used=str(profile.get("id")),
        submission_count=len(targets),
        reports_generated=generated,
        score_scale_max=score_scale_max,
        score_scale_label=_score_scale_label(score_scale_max),
        material_utilization=material_utilization,
        material_utilization_alerts=material_utilization_alerts,
        material_utilization_gate=material_utilization_gate,
        material_utilization_by_submission=material_utilization_by_submission[:20],
        feedback_closed_loop=feedback_closed_loop,
        started_at=started_at,
        finished_at=_now_iso(),
    )


def _delete_project_cascade(project_id: str, *, locale: str = "zh") -> Dict[str, object]:
    """
    删除项目及其关联数据。

    会删除该项目下的：
    - 项目记录
    - 资料记录与资料文件
    - 施组提交记录
    - 学习画像
    - 历史评分记录
    - 项目背景上下文
    - 真实评标记录
    - 进化报告

    **需要 API Key 认证**（未配置 API_KEYS 时无需）
    """
    ensure_data_dirs()
    projects = load_projects()
    target = next((p for p in projects if str(p.get("id")) == project_id), None)
    if target is None:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    target_name = str(target.get("name") or project_id)
    target_profile_id = str(target.get("expert_profile_id") or "")

    removed_counts = {
        "materials": 0,
        "submissions": 0,
        "score_reports": 0,
        "ground_truth": 0,
        "delta_cases": 0,
        "calibration_samples": 0,
        "patch_packages": 0,
    }

    save_projects([p for p in projects if p.get("id") != project_id])

    materials = load_materials()
    project_materials = [m for m in materials if m.get("project_id") == project_id]
    removed_counts["materials"] = len(project_materials)
    for m in project_materials:
        path = Path(str(m.get("path") or ""))
        if path.exists() and path.is_file():
            try:
                path.unlink()
            except Exception:
                pass
    save_materials([m for m in materials if m.get("project_id") != project_id])
    _invalidate_material_index_cache(project_id)

    project_dir = MATERIALS_DIR / project_id
    if project_dir.exists():
        shutil.rmtree(project_dir, ignore_errors=True)

    submissions = load_submissions()
    project_submission_ids = {
        str(s.get("id")) for s in submissions if s.get("project_id") == project_id
    }
    removed_counts["submissions"] = len(project_submission_ids)
    save_submissions([s for s in submissions if s.get("project_id") != project_id])

    score_reports = load_score_reports()
    removed_counts["score_reports"] = sum(
        1 for r in score_reports if r.get("project_id") == project_id
    )
    save_score_reports([r for r in score_reports if r.get("project_id") != project_id])
    evidence_units = load_evidence_units()
    save_evidence_units(
        [u for u in evidence_units if str(u.get("submission_id")) not in project_submission_ids]
    )
    qingtian_results = load_qingtian_results()
    save_qingtian_results(
        [q for q in qingtian_results if str(q.get("submission_id")) not in project_submission_ids]
    )
    delta_cases = load_delta_cases()
    removed_counts["delta_cases"] = sum(
        1
        for d in delta_cases
        if str(d.get("project_id")) == project_id
        or str(d.get("submission_id")) in project_submission_ids
    )
    save_delta_cases(
        [
            d
            for d in delta_cases
            if str(d.get("project_id")) != project_id
            and str(d.get("submission_id")) not in project_submission_ids
        ]
    )
    calibration_samples = load_calibration_samples()
    removed_counts["calibration_samples"] = sum(
        1
        for s in calibration_samples
        if str(s.get("project_id")) == project_id
        or str(s.get("submission_id")) in project_submission_ids
    )
    save_calibration_samples(
        [
            s
            for s in calibration_samples
            if str(s.get("project_id")) != project_id
            and str(s.get("submission_id")) not in project_submission_ids
        ]
    )
    patch_packages = load_patch_packages()
    removed_patch_ids = {
        str(p.get("id")) for p in patch_packages if str(p.get("project_id")) == project_id
    }
    removed_counts["patch_packages"] = len(removed_patch_ids)
    save_patch_packages([p for p in patch_packages if str(p.get("project_id")) != project_id])
    patch_deployments = load_patch_deployments()
    save_patch_deployments(
        [
            d
            for d in patch_deployments
            if str(d.get("project_id")) != project_id
            and str(d.get("patch_id")) not in removed_patch_ids
        ]
    )

    anchors = load_project_anchors()
    save_project_anchors([a for a in anchors if a.get("project_id") != project_id])
    requirements = load_project_requirements()
    save_project_requirements([r for r in requirements if r.get("project_id") != project_id])

    learning_profiles = load_learning_profiles()
    save_learning_profiles([p for p in learning_profiles if p.get("project_id") != project_id])

    score_history = load_score_history()
    save_score_history([h for h in score_history if h.get("project_id") != project_id])

    context = load_project_context()
    if project_id in context:
        context.pop(project_id, None)
        save_project_context(context)

    ground_truth = load_ground_truth()
    removed_counts["ground_truth"] = sum(
        1 for r in ground_truth if r.get("project_id") == project_id
    )
    save_ground_truth([r for r in ground_truth if r.get("project_id") != project_id])

    reports = load_evolution_reports()
    if project_id in reports:
        reports.pop(project_id, None)
        save_evolution_reports(reports)

    # 清理未被其它项目引用的专家配置
    if target_profile_id:
        remaining_projects = load_projects()
        in_use = any(
            str(p.get("expert_profile_id") or "") == target_profile_id for p in remaining_projects
        )
        if not in_use:
            profiles = load_expert_profiles()
            profiles = [ep for ep in profiles if str(ep.get("id") or "") != target_profile_id]
            save_expert_profiles(profiles)

    return {
        "project_id": project_id,
        "project_name": target_name,
        "removed_counts": removed_counts,
    }


@router.delete(
    "/projects/{project_id}",
    status_code=204,
    tags=["项目管理"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def delete_project(
    project_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> None:
    _delete_project_cascade(project_id, locale=locale)


@router.post(
    "/projects/cleanup_e2e",
    tags=["项目管理"],
    responses={**RESPONSES_401},
)
def cleanup_e2e_projects(
    prefix: str = Query("E2E_"),
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> dict:
    ensure_data_dirs()
    projects = load_projects()
    targets = [
        {
            "id": str(p.get("id") or ""),
            "name": str(p.get("name") or ""),
        }
        for p in projects
        if str(p.get("name") or "").startswith(prefix)
    ]
    removed: List[Dict[str, str]] = []
    failed: List[Dict[str, str]] = []
    for item in targets:
        pid = item["id"]
        try:
            _delete_project_cascade(pid, locale=locale)
            removed.append(item)
        except Exception as exc:  # noqa: BLE001 - bulk cleanup should continue
            failed.append({"id": pid, "name": item["name"], "detail": str(exc)})

    return {
        "ok": len(failed) == 0,
        "prefix": prefix,
        "matched": len(targets),
        "removed_count": len(removed),
        "removed": removed,
        "failed_count": len(failed),
        "failed": failed,
    }


MATERIAL_TYPE_DEFAULT = "tender_qa"
MATERIAL_TYPE_LABELS = {
    "tender_qa": "招标文件和答疑",
    "boq": "清单",
    "drawing": "图纸",
    "site_photo": "现场照片",
}
MATERIAL_TYPE_ALIASES = {
    "": MATERIAL_TYPE_DEFAULT,
    "material": MATERIAL_TYPE_DEFAULT,
    "materials": MATERIAL_TYPE_DEFAULT,
    "tender": "tender_qa",
    "bid": "tender_qa",
    "qa": "tender_qa",
    "qa_reply": "tender_qa",
    "list": "boq",
    "bill_of_quantities": "boq",
    "drawing_file": "drawing",
    "drawings": "drawing",
    "photo": "site_photo",
    "photos": "site_photo",
    "site_images": "site_photo",
}
MATERIAL_TYPE_ALLOWED_EXTS: Dict[str, tuple[str, ...]] = {
    # tender_qa 兼容旧版“单一资料上传”入口，保留宽松后缀集合。
    "tender_qa": (
        ".txt",
        ".md",
        ".pdf",
        ".doc",
        ".docx",
        ".docm",
        ".json",
        ".xlsx",
        ".xls",
        ".xlsm",
        ".csv",
        ".dxf",
        ".dwg",
        ".png",
        ".jpg",
        ".jpeg",
        ".webp",
        ".bmp",
        ".tif",
        ".tiff",
    ),
    "boq": (".xlsx", ".xls", ".xlsm", ".csv", ".pdf", ".doc", ".docx", ".txt", ".json"),
    "drawing": (
        ".pdf",
        ".doc",
        ".docx",
        ".xlsx",
        ".xls",
        ".dxf",
        ".dwg",
        ".png",
        ".jpg",
        ".jpeg",
        ".webp",
        ".bmp",
        ".tif",
        ".tiff",
        ".json",
        ".txt",
    ),
    "site_photo": (".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff"),
}
MATERIAL_ALLOWED_EXTS = tuple(
    sorted({ext for exts in MATERIAL_TYPE_ALLOWED_EXTS.values() for ext in exts})
)
MATERIAL_TYPE_ALLOWED_MIME_TOKENS: Dict[str, tuple[str, ...]] = {
    "tender_qa": (
        "text/plain",
        "text/markdown",
        "text/csv",
        "application/pdf",
        "application/json",
        "application/msword",
        "wordprocessingml",
        "spreadsheetml",
        "ms-excel",
        "application/dxf",
        "image/vnd.dxf",
        "application/acad",
        "application/x-autocad",
        "drawing/x-dxf",
        "image/",
    ),
    "boq": (
        "text/plain",
        "text/csv",
        "application/pdf",
        "application/json",
        "application/msword",
        "wordprocessingml",
        "spreadsheetml",
        "ms-excel",
    ),
    "drawing": (
        "text/plain",
        "application/pdf",
        "application/json",
        "application/msword",
        "wordprocessingml",
        "spreadsheetml",
        "ms-excel",
        "application/dxf",
        "image/vnd.dxf",
        "application/acad",
        "application/x-autocad",
        "drawing/x-dxf",
        "image/",
    ),
    "site_photo": ("image/",),
}
MATERIAL_ALLOWED_MIME_TOKENS = tuple(
    sorted({token for tokens in MATERIAL_TYPE_ALLOWED_MIME_TOKENS.values() for token in tokens})
)
MATERIAL_TYPE_DIMENSION_PRIORITY: Dict[str, List[str]] = {
    "tender_qa": ["01", "08", "09", "07"],
    "boq": ["13", "15", "11", "04"],
    "drawing": ["14", "16", "07", "12"],
    "site_photo": ["02", "03", "07", "08"],
}
MATERIAL_TYPE_KEYWORDS: Dict[str, List[str]] = {
    "tender_qa": ["答疑", "澄清", "变更", "条款", "工期", "质量标准", "计价规则"],
    "boq": ["清单", "工程量", "综合单价", "措施费", "暂估价", "设备", "甲供材"],
    "drawing": ["图纸", "节点", "平面", "剖面", "BIM", "碰撞", "深化"],
    "site_photo": ["现场", "照片", "临边", "扬尘", "消防", "高处", "塔吊"],
}
DIMENSION_RAG_KEYWORDS: Dict[str, List[str]] = {
    "01": ["工程范围", "项目理解", "招标", "答疑", "澄清"],
    "02": ["安全", "应急", "隐患", "临边", "高处", "消防"],
    "03": ["文明施工", "扬尘", "噪声", "围挡", "环保"],
    "04": ["材料", "部品", "进场", "验收", "周转"],
    "05": ["新技术", "数字化", "智能建造", "BIM", "IoT"],
    "06": ["关键工序", "工艺", "工法", "流程"],
    "07": ["危大工程", "专项方案", "监测", "应急预案"],
    "08": ["质量", "验收", "样板", "复检", "旁站"],
    "09": ["进度", "工期", "里程碑", "关键线路", "节点"],
    "10": ["专项施工", "模板", "吊装", "脚手架"],
    "11": ["人力", "班组", "劳动力", "组织"],
    "12": ["施工工艺", "总平面", "部署", "流水"],
    "13": ["物资", "设备", "清单", "机械", "采购"],
    "14": ["设计协调", "深化", "图纸", "碰撞", "会审"],
    "15": ["配置计划", "资源计划", "工程量", "措施费"],
    "16": ["技术措施", "可行性", "参数", "节点做法"],
}


def _normalize_uploaded_filename(filename: str) -> str:
    raw = unicodedata.normalize("NFKC", str(filename or "")).replace("\u3000", " ").strip()
    base = Path(raw).name.strip()
    while base.endswith("."):
        base = base[:-1].rstrip()
    return base


def _infer_material_type_from_filename(filename: object) -> str:
    normalized = _normalize_uploaded_filename(str(filename or "")).lower()
    ext = Path(normalized).suffix.lower()
    name = normalized

    if ext in {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff"}:
        if any(k in name for k in ("现场", "实景", "照片", "photo", "image", "img")):
            return "site_photo"
        return "drawing"
    if ext in {".dxf", ".dwg"}:
        return "drawing"
    if ext in {".xlsx", ".xls", ".xlsm", ".csv"}:
        return "boq"
    if any(k in name for k in ("清单", "boq", "bill_of_quantities", "工程量")):
        return "boq"
    if any(k in name for k in ("图纸", "总图", "平面", "立面", "剖面", "cad", "详图", "节点图")):
        return "drawing"
    if any(k in name for k in ("现场", "实景", "照片", "photo", "image", "img")):
        return "site_photo"
    return MATERIAL_TYPE_DEFAULT


def _normalize_material_type(material_type: object, *, filename: object = "") -> str:
    raw = str(material_type or "").strip().lower().replace("-", "_")
    if not raw:
        return _infer_material_type_from_filename(filename)
    normalized = MATERIAL_TYPE_ALIASES.get(raw, raw)
    if normalized in MATERIAL_TYPE_LABELS:
        return normalized
    return _infer_material_type_from_filename(filename)


def _parse_material_type_or_422(material_type: object, *, filename: object = "") -> str:
    raw = str(material_type or "").strip()
    if not raw:
        return _normalize_material_type("", filename=filename)
    normalized = _normalize_material_type(raw, filename=filename)
    raw_key = raw.lower().replace("-", "_")
    if raw_key in MATERIAL_TYPE_ALIASES or raw_key in MATERIAL_TYPE_LABELS:
        return normalized
    supported = "、".join(MATERIAL_TYPE_LABELS.keys())
    raise HTTPException(status_code=422, detail=f"material_type 不支持：{raw}（支持：{supported}）")


def _material_type_label(material_type: object, *, filename: object = "") -> str:
    return MATERIAL_TYPE_LABELS.get(
        _normalize_material_type(material_type, filename=filename), "项目资料"
    )


def _material_type_ext_hint(material_type: object, *, filename: object = "") -> str:
    normalized = _normalize_material_type(material_type, filename=filename)
    exts = MATERIAL_TYPE_ALLOWED_EXTS.get(normalized) or MATERIAL_ALLOWED_EXTS
    return "、".join(exts)


def _is_allowed_material_upload(filename: str, content_type: str, material_type: str) -> bool:
    normalized = _normalize_uploaded_filename(filename).lower()
    allowed_exts = MATERIAL_TYPE_ALLOWED_EXTS.get(material_type) or MATERIAL_ALLOWED_EXTS
    if normalized and any(normalized.endswith(ext) for ext in allowed_exts):
        return True
    ctype = str(content_type or "").lower().strip()
    allowed_tokens = (
        MATERIAL_TYPE_ALLOWED_MIME_TOKENS.get(material_type) or MATERIAL_ALLOWED_MIME_TOKENS
    )
    return any(token in ctype for token in allowed_tokens)


SUBMISSION_DUPLICATE_WINDOW_SECONDS = 15


def _parse_iso_datetime_utc(value: object) -> Optional[datetime]:
    raw = str(value or "").strip()
    if not raw:
        return None
    if raw.endswith("Z"):
        raw = raw[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(raw)
    except ValueError:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _find_recent_duplicate_submission(
    submissions: List[dict],
    *,
    project_id: str,
    filename: str,
    text: str,
    now_utc: datetime,
) -> Optional[dict]:
    normalized_filename = _normalize_uploaded_filename(filename)
    for submission in reversed(submissions):
        if submission.get("project_id") != project_id:
            continue
        if _normalize_uploaded_filename(submission.get("filename", "")) != normalized_filename:
            continue
        if str(submission.get("text") or "") != text:
            continue
        created_at = _parse_iso_datetime_utc(submission.get("created_at"))
        if created_at is None:
            continue
        age_seconds = (now_utc - created_at).total_seconds()
        if 0 <= age_seconds <= SUBMISSION_DUPLICATE_WINDOW_SECONDS:
            return submission
    return None


@router.post(
    "/projects/{project_id}/materials",
    tags=["项目管理"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def upload_material(
    project_id: str,
    file: UploadFile = File(...),
    material_type: str = Form(MATERIAL_TYPE_DEFAULT),
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> dict:
    """
    上传项目材料文件（支持按资料分类上传）。

    material_type:
    - tender_qa: 招标文件和答疑
    - boq: 清单
    - drawing: 图纸
    - site_photo: 现场照片

    **需要 API Key 认证**（未配置 API_KEYS 时无需）

    支持 Accept-Language header 进行多语言响应。
    """
    ensure_data_dirs()
    raw_name = file.filename or ""
    normalized_name = _normalize_uploaded_filename(raw_name)
    normalized_material_type = _parse_material_type_or_422(material_type, filename=normalized_name)
    if not normalized_name:
        raise HTTPException(status_code=422, detail="资料文件名为空，请重试或重命名后上传。")
    if not _is_allowed_material_upload(
        normalized_name, file.content_type or "", normalized_material_type
    ):
        raise HTTPException(
            status_code=422,
            detail=(
                _material_type_label(normalized_material_type)
                + "支持："
                + _material_type_ext_hint(normalized_material_type)
            ),
        )
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    project_dir = MATERIALS_DIR / project_id / normalized_material_type
    project_dir.mkdir(parents=True, exist_ok=True)
    target = project_dir / normalized_name
    content = file.file.read()
    target.write_bytes(content)

    materials = load_materials()
    existing_ids = [
        str(m.get("id"))
        for m in materials
        if m.get("project_id") == project_id
        and _normalize_material_type(m.get("material_type"), filename=m.get("filename"))
        == normalized_material_type
        and _normalize_uploaded_filename(m.get("filename", "")) == normalized_name
        and m.get("id")
    ]
    materials = [
        m
        for m in materials
        if not (
            m.get("project_id") == project_id
            and _normalize_material_type(m.get("material_type"), filename=m.get("filename"))
            == normalized_material_type
            and _normalize_uploaded_filename(m.get("filename", "")) == normalized_name
        )
    ]
    record = {
        "id": existing_ids[0] if existing_ids else str(uuid4()),
        "project_id": project_id,
        "material_type": normalized_material_type,
        "filename": normalized_name,
        "path": str(target),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    materials.append(record)
    save_materials(materials)
    _invalidate_material_index_cache(project_id)
    # 材料更新后立即重建锚点/要求，避免后续评分继续使用旧约束。
    constraint_sync: Dict[str, object] = {"rebuilt": False}
    try:
        anchors, requirements = _rebuild_project_anchors_and_requirements(project_id)
        constraint_sync = {
            "rebuilt": True,
            "anchors": len(anchors),
            "requirements": len(requirements),
        }
    except Exception as exc:
        constraint_sync = {"rebuilt": False, "error": f"{type(exc).__name__}: {exc}"}
    return {"status": "ok", "material": record, "constraint_sync": constraint_sync}


@router.get(
    "/projects/{project_id}/materials",
    response_model=list[MaterialRecord],
    tags=["项目管理"],
    responses={**RESPONSES_404},
)
def list_materials(project_id: str, locale: str = Depends(get_locale)) -> list[MaterialRecord]:
    """获取指定项目下已上传的资料列表。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    materials = [m for m in load_materials() if m.get("project_id") == project_id]
    materials.sort(key=lambda x: str(x.get("created_at", "")), reverse=True)
    normalized_rows: List[dict] = []
    for material in materials:
        row = dict(material)
        row["material_type"] = _normalize_material_type(
            row.get("material_type"), filename=row.get("filename")
        )
        normalized_rows.append(row)
    return [MaterialRecord(**m) for m in normalized_rows]


@router.get(
    "/projects/{project_id}/materials/health",
    tags=["项目管理"],
    responses={**RESPONSES_404},
)
def get_materials_health(project_id: str, locale: str = Depends(get_locale)) -> dict:
    """返回项目资料解析质量与门禁状态（不触发评分）。"""
    ensure_data_dirs()
    projects = load_projects()
    try:
        project = _find_project(project_id, projects)
    except HTTPException:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    snapshot, _ = _validate_material_gate_for_scoring(project_id, project, raise_on_fail=False)
    return snapshot


@router.get(
    "/projects/{project_id}/scoring_readiness",
    response_model=ScoringReadinessResponse,
    tags=["项目管理"],
    responses={**RESPONSES_404},
)
def get_scoring_readiness(
    project_id: str, locale: str = Depends(get_locale)
) -> ScoringReadinessResponse:
    """评分前置检查：门禁 + 施组可评分状态。"""
    ensure_data_dirs()
    projects = load_projects()
    try:
        project = _find_project(project_id, projects)
    except HTTPException:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    payload = _build_scoring_readiness(project_id, project)
    return ScoringReadinessResponse(**payload)


@router.get(
    "/projects/{project_id}/mece_audit",
    response_model=ProjectMeceAuditResponse,
    tags=["系统诊断"],
    responses={**RESPONSES_404},
)
def get_project_mece_audit(
    project_id: str, locale: str = Depends(get_locale)
) -> ProjectMeceAuditResponse:
    """项目级 MECE 审计：输入链路、评分有效性、进化闭环、运行稳定性。"""
    ensure_data_dirs()
    projects = load_projects()
    try:
        project = _find_project(project_id, projects)
    except HTTPException:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    payload = _build_project_mece_audit(project_id, project)
    return ProjectMeceAuditResponse(**payload)


@router.get(
    "/projects/{project_id}/evolution/health",
    response_model=EvolutionHealthResponse,
    tags=["自我学习与进化"],
    responses={**RESPONSES_404},
)
def get_project_evolution_health(
    project_id: str, locale: str = Depends(get_locale)
) -> EvolutionHealthResponse:
    """项目进化健康度：误差趋势、样本时效、概念漂移风险。"""
    ensure_data_dirs()
    projects = load_projects()
    try:
        project = _find_project(project_id, projects)
    except HTTPException:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    payload = _build_evolution_health_report(project_id, project)
    return EvolutionHealthResponse(**payload)


@router.get(
    "/projects/{project_id}/materials/depth_report",
    response_model=MaterialDepthReportResponse,
    tags=["项目管理"],
    responses={**RESPONSES_404},
)
def get_material_depth_report(
    project_id: str, locale: str = Depends(get_locale)
) -> MaterialDepthReportResponse:
    """项目资料深读体检报告（JSON），用于评分前体检。"""
    ensure_data_dirs()
    projects = load_projects()
    try:
        project = _find_project(project_id, projects)
    except HTTPException:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    payload = _build_material_depth_report(project_id, project)
    return MaterialDepthReportResponse(**payload)


@router.get(
    "/projects/{project_id}/materials/depth_report/markdown",
    response_model=MaterialDepthReportMarkdownResponse,
    tags=["项目管理"],
    responses={**RESPONSES_404},
)
def get_material_depth_report_markdown(
    project_id: str, locale: str = Depends(get_locale)
) -> MaterialDepthReportMarkdownResponse:
    """项目资料深读体检报告（Markdown 文本）。"""
    ensure_data_dirs()
    projects = load_projects()
    try:
        project = _find_project(project_id, projects)
    except HTTPException:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    payload = _build_material_depth_report(project_id, project)
    markdown = _render_material_depth_report_markdown(payload)
    return MaterialDepthReportMarkdownResponse(
        project_id=project_id,
        markdown=markdown,
        generated_at=str(payload.get("generated_at") or _now_iso()),
    )


@router.get(
    "/projects/{project_id}/materials/depth_report.md",
    tags=["项目管理"],
    responses={**RESPONSES_404},
)
def download_material_depth_report_markdown(
    project_id: str, locale: str = Depends(get_locale)
) -> Response:
    """下载项目资料深读体检 Markdown 文件。"""
    result = get_material_depth_report_markdown(project_id=project_id, locale=locale)
    markdown = str(result.markdown or "")
    filename = f"material_depth_report_{project_id}.md"
    return Response(
        content=markdown,
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/projects/{project_id}/materials/knowledge_profile",
    response_model=MaterialKnowledgeProfileResponse,
    tags=["项目管理"],
    responses={**RESPONSES_404},
)
def get_material_knowledge_profile(
    project_id: str, locale: str = Depends(get_locale)
) -> MaterialKnowledgeProfileResponse:
    """项目资料知识画像（JSON），展示资料对16维评分的覆盖强度。"""
    ensure_data_dirs()
    projects = load_projects()
    try:
        _find_project(project_id, projects)
    except HTTPException:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    payload = _build_material_knowledge_profile(project_id)
    return MaterialKnowledgeProfileResponse(**payload)


@router.get(
    "/projects/{project_id}/materials/knowledge_profile/markdown",
    response_model=MaterialKnowledgeProfileMarkdownResponse,
    tags=["项目管理"],
    responses={**RESPONSES_404},
)
def get_material_knowledge_profile_markdown(
    project_id: str, locale: str = Depends(get_locale)
) -> MaterialKnowledgeProfileMarkdownResponse:
    """项目资料知识画像（Markdown 文本）。"""
    ensure_data_dirs()
    projects = load_projects()
    try:
        _find_project(project_id, projects)
    except HTTPException:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    payload = _build_material_knowledge_profile(project_id)
    markdown = _render_material_knowledge_profile_markdown(payload)
    return MaterialKnowledgeProfileMarkdownResponse(
        project_id=project_id,
        markdown=markdown,
        generated_at=str(payload.get("generated_at") or _now_iso()),
    )


@router.get(
    "/projects/{project_id}/materials/knowledge_profile.md",
    tags=["项目管理"],
    responses={**RESPONSES_404},
)
def download_material_knowledge_profile_markdown(
    project_id: str, locale: str = Depends(get_locale)
) -> Response:
    """下载项目资料知识画像 Markdown 文件。"""
    result = get_material_knowledge_profile_markdown(project_id=project_id, locale=locale)
    markdown = str(result.markdown or "")
    filename = f"material_knowledge_profile_{project_id}.md"
    return Response(
        content=markdown,
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/projects/{project_id}/anchors",
    response_model=list[ProjectAnchorRecord],
    tags=["项目管理"],
    responses={**RESPONSES_404},
)
def list_project_anchors(
    project_id: str, locale: str = Depends(get_locale)
) -> list[ProjectAnchorRecord]:
    """获取项目锚点列表。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    anchors = [a for a in load_project_anchors() if a.get("project_id") == project_id]
    return [ProjectAnchorRecord(**a) for a in anchors]


@router.post(
    "/projects/{project_id}/anchors/rebuild",
    response_model=list[ProjectAnchorRecord],
    tags=["项目管理"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def rebuild_project_anchors(
    project_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> list[ProjectAnchorRecord]:
    """基于项目资料重建锚点。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    anchors, _ = _rebuild_project_anchors_and_requirements(project_id)
    return [ProjectAnchorRecord(**a) for a in anchors]


@router.get(
    "/projects/{project_id}/requirements",
    response_model=list[ProjectRequirementRecord],
    tags=["项目管理"],
    responses={**RESPONSES_404},
)
def list_project_requirements(
    project_id: str,
    locale: str = Depends(get_locale),
) -> list[ProjectRequirementRecord]:
    """获取项目要求矩阵。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    requirements = [r for r in load_project_requirements() if r.get("project_id") == project_id]
    return [ProjectRequirementRecord(**r) for r in requirements]


@router.post(
    "/projects/{project_id}/requirements/rebuild",
    response_model=list[ProjectRequirementRecord],
    tags=["项目管理"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def rebuild_project_requirements(
    project_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> list[ProjectRequirementRecord]:
    """基于项目资料重建要求矩阵。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    _, requirements = _rebuild_project_anchors_and_requirements(project_id)
    return [ProjectRequirementRecord(**r) for r in requirements]


@router.get(
    "/projects/{project_id}/constraint_pack",
    response_model=ConstraintPack,
    tags=["项目管理"],
    responses={**RESPONSES_404},
)
def get_project_constraint_pack(
    project_id: str, locale: str = Depends(get_locale)
) -> ConstraintPack:
    """生成项目级编制约束包（Constraint Pack）。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    anchors = [a for a in load_project_anchors() if a.get("project_id") == project_id]
    requirements = [r for r in load_project_requirements() if r.get("project_id") == project_id]
    if not anchors or not requirements:
        _rebuild_project_anchors_and_requirements(project_id)
    pack = _build_constraint_pack(project_id)
    return ConstraintPack(**pack)


@router.delete(
    "/projects/{project_id}/materials/{material_id}",
    tags=["项目管理"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def delete_material(
    project_id: str,
    material_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> dict:
    """删除指定项目下的一条资料记录及对应文件。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    materials = load_materials()
    found = None
    for m in materials:
        if m.get("id") == material_id and m.get("project_id") == project_id:
            found = m
            break
    if not found:
        raise HTTPException(status_code=404, detail="资料记录不存在")
    path = Path(found["path"])
    if path.exists():
        path.unlink()
    materials = [m for m in materials if m.get("id") != material_id]
    save_materials(materials)
    _invalidate_material_index_cache(project_id)
    return {"ok": True, "id": material_id}


def _decode_dxf_text(content: bytes) -> str:
    if b"\x00" in content[:4096]:
        raise ValueError("DXF 解析失败：检测到二进制 DXF，请先另存为 ASCII DXF。")
    for encoding in ("utf-8-sig", "gb18030", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="ignore")


def _iter_dxf_group_pairs(text: str) -> List[tuple[int, str]]:
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    pairs: List[tuple[int, str]] = []
    idx = 0
    while idx + 1 < len(lines):
        code_raw = lines[idx].strip()
        value = lines[idx + 1].strip()
        idx += 2
        if not code_raw:
            continue
        try:
            code = int(code_raw)
        except ValueError:
            continue
        pairs.append((code, value))
    return pairs


def _extract_dxf_text(content: bytes) -> str:
    raw_text = _decode_dxf_text(content)
    pairs = _iter_dxf_group_pairs(raw_text)
    if not pairs:
        return "[DXF解析摘要]\n未读取到有效 DXF 组码。"

    acadver = ""
    codepage = ""
    insunits: Optional[int] = None
    unit_map = {
        0: "未指定",
        1: "英寸",
        2: "英尺",
        4: "毫米",
        5: "厘米",
        6: "米",
        20: "秒",
        21: "分",
        22: "时",
    }
    for i, (code, value) in enumerate(pairs[:-1]):
        if code != 9:
            continue
        key = value.upper()
        next_code, next_value = pairs[i + 1]
        if key == "$ACADVER" and next_code in (1, 3):
            acadver = next_value.strip()
        elif key == "$DWGCODEPAGE" and next_code in (1, 3):
            codepage = next_value.strip()
        elif key == "$INSUNITS" and next_code in (70, 280):
            try:
                insunits = int(float(next_value.strip()))
            except Exception:
                insunits = None

    text_entity_types = {"TEXT", "MTEXT", "ATTDEF", "ATTRIB"}
    entity_counts: Dict[str, int] = {}
    extracted_texts: List[str] = []
    layers: set[str] = set()
    blocks: set[str] = set()

    in_entities = False
    waiting_section_name = False
    current_entity_type = ""
    current_entity_texts: List[str] = []
    current_layer = ""
    current_block = ""

    def _flush_entity() -> None:
        nonlocal current_entity_type, current_entity_texts, current_layer, current_block
        if not current_entity_type:
            return
        entity_counts[current_entity_type] = entity_counts.get(current_entity_type, 0) + 1
        if current_layer:
            layers.add(current_layer)
        if current_block:
            blocks.add(current_block)
        for item in current_entity_texts:
            if not item:
                continue
            normalized = (
                item.replace("\\P", "\n")
                .replace("\\~", " ")
                .replace("{", "")
                .replace("}", "")
                .strip()
            )
            if normalized and normalized not in extracted_texts:
                extracted_texts.append(normalized)
        current_entity_type = ""
        current_entity_texts = []
        current_layer = ""
        current_block = ""

    for code, value in pairs:
        token = value.upper().strip()
        if waiting_section_name and code == 2:
            in_entities = token == "ENTITIES"
            waiting_section_name = False
            continue

        if code == 0:
            if token == "SECTION":
                _flush_entity()
                waiting_section_name = True
                continue
            if token in {"ENDSEC", "EOF"}:
                _flush_entity()
                in_entities = False
                continue
            if in_entities:
                _flush_entity()
                current_entity_type = token
                continue

        if not in_entities or not current_entity_type:
            continue
        if code == 8:
            current_layer = value.strip()
            continue
        if code == 2 and current_entity_type == "INSERT":
            current_block = value.strip()
            continue
        if code in (1, 3) and current_entity_type in text_entity_types:
            current_entity_texts.append(value)

    _flush_entity()

    summary_lines = ["[DXF解析摘要]"]
    if acadver:
        summary_lines.append(f"ACAD版本: {acadver}")
    if codepage:
        summary_lines.append(f"编码页: {codepage}")
    if insunits is not None:
        summary_lines.append(f"插入单位: {insunits}({unit_map.get(insunits, '未知')})")
    if entity_counts:
        top_entities = sorted(entity_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        summary_lines.append(
            "实体统计: " + "、".join(f"{etype}:{count}" for etype, count in top_entities)
        )
    if layers:
        summary_lines.append("图层: " + "、".join(sorted(layers)[:20]))
    if blocks:
        summary_lines.append("块参照: " + "、".join(sorted(blocks)[:20]))

    if extracted_texts:
        summary_lines.append("")
        summary_lines.append("[DXF文本实体提取]")
        summary_lines.extend(extracted_texts[:160])
    return "\n".join(summary_lines).strip()


def _looks_like_ascii_dxf(content: bytes) -> bool:
    sample = content[:4096]
    if not sample or b"\x00" in sample:
        return False
    text = sample.decode("latin-1", errors="ignore").replace("\r", "\n").upper()
    return "SECTION" in text and ("ENTITIES" in text or "HEADER" in text) and "\n0\n" in text


def _extract_dwg_binary_markers(content: bytes, *, max_tokens: int = 30) -> Dict[str, object]:
    sample = content[: min(len(content), 1_500_000)]
    versions = sorted(
        {
            item.decode("ascii", errors="ignore")
            for item in re.findall(rb"AC10\d{2}", sample)
            if item
        }
    )
    raw_tokens = re.findall(rb"[A-Za-z_][A-Za-z0-9_./:-]{2,48}", sample)
    blocklist_prefix = ("http", "https", "xmlns", "schema", "version", "content")
    token_counter: Counter[str] = Counter()
    for token_bytes in raw_tokens:
        token = token_bytes.decode("latin-1", errors="ignore").strip()
        lower = token.lower()
        if len(token) < 3:
            continue
        if any(lower.startswith(prefix) for prefix in blocklist_prefix):
            continue
        if lower in {"acdb", "objectdbx", "autocad", "dwg"}:
            continue
        if re.fullmatch(r"[0-9a-f]{8,}", lower):
            continue
        token_counter[token] += 1
    top_tokens = [tok for tok, _ in token_counter.most_common(max(1, int(max_tokens)))]
    return {"versions": versions, "tokens": top_tokens}


def _dwg_converter_command_candidates(
    binary: str,
    *,
    in_path: Path,
    input_dir: Path,
    output_dir: Path,
) -> List[List[str]]:
    name = Path(binary).name.lower()
    out_file = output_dir / f"{in_path.stem}.dxf"
    candidates: List[List[str]] = []
    if "dwg2dxf" in name:
        candidates.append([binary, str(in_path), str(out_file)])
    elif any(mark in name for mark in ("odafileconverter", "oda_file_converter", "teigha")):
        # ODA/Teigha 不同版本参数略有差异，按候选命令依次尝试。
        candidates.append([binary, str(input_dir), str(output_dir), "ACAD2018", "DXF", "0", "1"])
        candidates.append(
            [binary, str(input_dir), str(output_dir), "ACAD2018", "DXF", "0", "1", "*.DWG"]
        )
        candidates.append([binary, str(in_path), str(out_file)])
    else:
        candidates.append([binary, str(in_path), str(out_file)])
        candidates.append([binary, str(input_dir), str(output_dir)])
    return candidates


def _resolve_dwg_converter_binaries() -> List[str]:
    converter_names: List[str] = []
    converter_paths: List[str] = []
    env_converters_raw = str(os.getenv("DWG_CONVERTER_BIN") or "").strip()
    if env_converters_raw:
        for item in re.split(r"[;,]", env_converters_raw):
            s = item.strip()
            if not s:
                continue
            p = Path(s)
            if p.exists() and p.is_file():
                converter_paths.append(str(p))
                continue
            if p.exists() and p.is_dir():
                for bin_name in (
                    "dwg2dxf",
                    "ODAFileConverter",
                    "oda_file_converter",
                    "TeighaFileConverter",
                ):
                    candidate = p / bin_name
                    if candidate.exists() and candidate.is_file():
                        converter_paths.append(str(candidate))
                continue
            if s not in converter_names:
                converter_names.append(s)
    converter_names.extend(
        [
            "dwg2dxf",
            "ODAFileConverter",
            "oda_file_converter",
            "TeighaFileConverter",
            "dwgread",
        ]
    )
    converter_names = list(dict.fromkeys(converter_names))
    binaries: List[str] = [p for p in converter_paths if p]
    common_paths = [
        "/Applications/ODAFileConverter.app/Contents/MacOS/ODAFileConverter",
        "/Applications/Teigha File Converter.app/Contents/MacOS/TeighaFileConverter",
        "/opt/homebrew/bin/dwg2dxf",
        "/usr/local/bin/dwg2dxf",
        "/opt/homebrew/bin/dwgread",
        "/usr/local/bin/dwgread",
    ]
    for raw_path in common_paths:
        p = Path(raw_path)
        if p.exists() and p.is_file():
            binaries.append(str(p))
    for name in converter_names:
        if not name:
            continue
        resolved = name if Path(name).exists() else shutil.which(name)
        if resolved:
            binaries.append(str(resolved))
    return list(dict.fromkeys(binaries))


def _extract_dwg_text(content: bytes, filename: str) -> str:
    """
    DWG 预处理链：
    1) 优先尝试系统级转换器将 DWG 转 DXF（若已安装）
    2) 转换成功后复用 DXF 解析
    3) 无转换器或转换失败时保留元信息并给出明确提示
    """
    if _looks_like_ascii_dxf(content):
        try:
            dxf_text = _extract_dxf_text(content)
            return f"[DWG预处理] 文件: {filename}\n检测到ASCII DXF内容，按DXF解析。\n\n{dxf_text}"
        except Exception:
            pass
    # 对明显异常的小体量 DWG 直接走标识兜底，避免调用外部转换器造成长时间阻塞。
    if len(content) < 256:
        markers = _extract_dwg_binary_markers(content, max_tokens=26)
        versions = [str(x) for x in (markers.get("versions") or []) if str(x).strip()]
        token_preview = [str(x) for x in (markers.get("tokens") or []) if str(x).strip()]
        marker_text = "、".join(versions[:4]) if versions else "未识别"
        tokens_text = "、".join(token_preview[:16]) if token_preview else "未提取到有效标识"
        return (
            f"[DWG图纸] 文件: {filename}，字节数: {len(content)}\n"
            "DWG预处理: 文件体积过小，已跳过外部转换器尝试\n"
            f"版本标记: {marker_text}\n"
            f"二进制标识提取: {tokens_text}\n"
            "当前未完成稳定结构化解析，建议同时上传 PDF 或 ASCII DXF 以提升评分准确性。"
        )

    binaries = _resolve_dwg_converter_binaries()

    converter_display = ["dwg2dxf", "ODAFileConverter", "oda_file_converter", "TeighaFileConverter"]
    notes: List[str] = []
    with tempfile.TemporaryDirectory(prefix="dwg_bridge_") as tmpdir:
        tmp_root = Path(tmpdir)
        input_dir = tmp_root / "in"
        output_dir = tmp_root / "out"
        input_dir.mkdir(parents=True, exist_ok=True)
        output_dir.mkdir(parents=True, exist_ok=True)
        in_path = input_dir / _normalize_uploaded_filename(filename)
        in_path.write_bytes(content)

        for name in converter_display:
            if not any(Path(b).name.lower() == name.lower() for b in binaries):
                notes.append(f"{name}: not_found")
        for binary in binaries:
            cmd_candidates = _dwg_converter_command_candidates(
                binary,
                in_path=in_path,
                input_dir=input_dir,
                output_dir=output_dir,
            )
            for cmd in cmd_candidates:
                cmd_signature = " ".join(cmd[:3])
                try:
                    completed = subprocess.run(
                        cmd,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        timeout=45,
                        check=False,
                        text=True,
                    )
                    if completed.returncode != 0:
                        err = (completed.stderr or completed.stdout or "").strip().splitlines()
                        notes.append(
                            f"{Path(binary).name}: rc={completed.returncode} {err[0] if err else ''}"
                        )
                        continue
                    dxf_candidates = sorted(
                        {
                            *output_dir.rglob("*.dxf"),
                            *tmp_root.rglob(f"{in_path.stem}.dxf"),
                        }
                    )
                    if not dxf_candidates:
                        notes.append(f"{Path(binary).name}: no_dxf_output ({cmd_signature})")
                        continue
                    for dxf_candidate in dxf_candidates:
                        try:
                            dxf_text = _extract_dxf_text(dxf_candidate.read_bytes())
                        except Exception as exc:  # noqa: BLE001 - converter output might be malformed
                            notes.append(
                                f"{Path(binary).name}: dxf_parse_failed {type(exc).__name__}: {exc}"
                            )
                            continue
                        head = [
                            f"[DWG预处理] 文件: {filename}",
                            f"转换器: {Path(binary).name}",
                            f"命令: {cmd_signature}",
                            f"输出DXF: {dxf_candidate.name}",
                        ]
                        return "\n".join(head + ["", dxf_text]).strip()
                except subprocess.TimeoutExpired:
                    notes.append(f"{Path(binary).name}: timeout ({cmd_signature})")
                except Exception as exc:  # noqa: BLE001 - continue next converter
                    notes.append(f"{Path(binary).name}: exception {type(exc).__name__}: {exc}")

    markers = _extract_dwg_binary_markers(content, max_tokens=26)
    versions = [str(x) for x in (markers.get("versions") or []) if str(x).strip()]
    token_preview = [str(x) for x in (markers.get("tokens") or []) if str(x).strip()]
    notes_text = "；".join(notes[:6]) if notes else "未检测到可用转换器"
    marker_text = "、".join(versions[:4]) if versions else "未识别"
    tokens_text = "、".join(token_preview[:16]) if token_preview else "未提取到有效标识"
    return (
        f"[DWG图纸] 文件: {filename}，字节数: {len(content)}\n"
        f"DWG预处理: {notes_text}\n"
        f"版本标记: {marker_text}\n"
        f"二进制标识提取: {tokens_text}\n"
        "当前未完成稳定结构化解析，建议同时上传 PDF 或 ASCII DXF 以提升评分准确性。"
    )


def _safe_float_from_cell(value: object) -> Optional[float]:
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return None
    raw = str(value or "").strip()
    if not raw:
        return None
    raw = raw.replace(",", "").replace("，", "")
    if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", raw):
        return None
    try:
        return float(raw)
    except Exception:
        return None


def _extract_terms(text: str, *, max_terms: int = 40) -> List[str]:
    tokens = re.findall(r"[A-Za-z]{3,}|[\u4e00-\u9fff]{2,8}", str(text or ""))
    counter: Counter[str] = Counter()
    stop_words = {"施工", "工程", "项目", "内容", "要求", "进行", "以及", "方案"}
    for token in tokens:
        key = token.strip().lower()
        if not key:
            continue
        if key in stop_words:
            continue
        counter[key] += 1
    return [k for k, _ in counter.most_common(max(1, int(max_terms)))]


def _normalize_numeric_token(token: object) -> str:
    raw = str(token or "").strip()
    if not raw:
        return ""
    if not re.fullmatch(r"\d+(?:\.\d+)?", raw):
        return ""
    if "." in raw:
        raw = raw.rstrip("0").rstrip(".")
    if not raw:
        return ""
    if "." not in raw:
        try:
            raw = str(int(raw))
        except Exception:
            pass
    return raw


def _extract_numeric_terms(text: str, *, max_terms: int = 24) -> List[str]:
    tokens = re.findall(r"\d+(?:\.\d+)?", str(text or ""))
    counter: Counter[str] = Counter()
    for token in tokens:
        normalized = _normalize_numeric_token(token)
        if not normalized:
            continue
        # 过滤噪声：单字符整数通常无业务含义（页码/编号），保留小数与2位以上整数。
        if len(normalized) < 2 and "." not in normalized:
            continue
        counter[normalized] += 1
    return [k for k, _ in counter.most_common(max(1, int(max_terms)))]


def _build_boq_structured_summary(
    content: bytes,
    filename: str,
    *,
    parsed_text: str = "",
) -> Dict[str, object]:
    summary: Dict[str, object] = {
        "filename": filename,
        "detected_format": Path(filename).suffix.lower().lstrip("."),
    }
    ext = Path(filename).suffix.lower()
    header_alias = {
        "code": ["项目编码", "清单编码", "编码", "子目号"],
        "name": ["项目名称", "项目特征", "工程内容", "名称", "描述"],
        "unit": ["单位"],
        "quantity": ["工程量", "数量"],
        "unit_price": ["综合单价", "单价"],
        "amount": ["合价", "金额", "总价", "小计"],
    }

    def _find_col_map(header_row: List[str]) -> Dict[str, int]:
        out: Dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            s = str(cell or "").strip()
            if not s:
                continue
            for field, aliases in header_alias.items():
                if field in out:
                    continue
                if any(a in s for a in aliases):
                    out[field] = idx
        return out

    def _sheet_struct_from_rows(rows: List[List[object]], sheet_name: str) -> Dict[str, object]:
        header_idx = -1
        col_map: Dict[str, int] = {}
        for i, row in enumerate(rows[:40]):
            str_row = [str(x or "").strip() for x in row]
            maybe = _find_col_map(str_row)
            if len(maybe) >= 2:
                header_idx = i
                col_map = maybe
                break
        if header_idx < 0:
            return {
                "sheet": sheet_name,
                "parsed_items": 0,
                "detected_columns": {},
                "quantity_sum": 0.0,
                "amount_sum": 0.0,
                "top_items_by_amount": [],
            }

        parsed_items = 0
        quantity_sum = 0.0
        amount_sum = 0.0
        unit_set: set[str] = set()
        top_items: List[Dict[str, object]] = []
        for row in rows[header_idx + 1 :]:
            if not row:
                continue

            def _cell(name: str) -> object:
                idx = col_map.get(name)
                if idx is None or idx >= len(row):
                    return ""
                return row[idx]

            code = str(_cell("code") or "").strip()
            name = str(_cell("name") or "").strip()
            if not code and not name:
                continue
            parsed_items += 1
            unit = str(_cell("unit") or "").strip()
            if unit:
                unit_set.add(unit)
            qty = _safe_float_from_cell(_cell("quantity"))
            amt = _safe_float_from_cell(_cell("amount"))
            if qty is not None:
                quantity_sum += qty
            if amt is not None:
                amount_sum += amt
            if name or code:
                top_items.append(
                    {
                        "code": code,
                        "name": name,
                        "unit": unit,
                        "quantity": round(float(qty), 4) if qty is not None else None,
                        "amount": round(float(amt), 2) if amt is not None else None,
                    }
                )
        top_items.sort(key=lambda x: float(x.get("amount") or 0.0), reverse=True)
        return {
            "sheet": sheet_name,
            "parsed_items": parsed_items,
            "detected_columns": col_map,
            "quantity_sum": round(quantity_sum, 4),
            "amount_sum": round(amount_sum, 2),
            "units": sorted(unit_set)[:20],
            "top_items_by_amount": top_items[:10],
        }

    if ext in {".xlsx", ".xls", ".xlsm"}:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet_summaries: List[Dict[str, object]] = []
            total_items = 0
            total_qty = 0.0
            total_amt = 0.0
            for sheet in wb.worksheets:
                rows: List[List[object]] = []
                for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    rows.append(list(row))
                    if idx >= 3000:
                        break
                sheet_summary = _sheet_struct_from_rows(rows, sheet.title)
                sheet_summaries.append(sheet_summary)
                total_items += int(sheet_summary.get("parsed_items", 0))
                total_qty += float(sheet_summary.get("quantity_sum", 0.0))
                total_amt += float(sheet_summary.get("amount_sum", 0.0))
            wb.close()
            summary["sheets"] = sheet_summaries
            summary["total_parsed_items"] = total_items
            summary["total_quantity"] = round(total_qty, 4)
            summary["total_amount"] = round(total_amt, 2)
            return summary
        except Exception as exc:
            summary["parse_error"] = f"excel_parse_failed: {type(exc).__name__}: {exc}"

    if ext == ".csv":
        try:
            decoded = content.decode("utf-8", errors="ignore")
            reader = csv.reader(io.StringIO(decoded))
            rows = [list(r) for _, r in zip(range(3000), reader)]
            csv_summary = _sheet_struct_from_rows(rows, "csv")
            summary["sheets"] = [csv_summary]
            summary["total_parsed_items"] = int(csv_summary.get("parsed_items", 0))
            summary["total_quantity"] = float(csv_summary.get("quantity_sum", 0.0))
            summary["total_amount"] = float(csv_summary.get("amount_sum", 0.0))
            return summary
        except Exception as exc:
            summary["parse_error"] = f"csv_parse_failed: {type(exc).__name__}: {exc}"

    # 兜底：从已解析文本里识别关键信号，至少提供结构化可见性。
    text = str(parsed_text or "")
    summary["keyword_hits"] = {
        "工程量": len(re.findall(r"工程量|数量", text)),
        "综合单价": len(re.findall(r"综合单价|单价", text)),
        "合价金额": len(re.findall(r"合价|金额|总价|小计", text)),
    }
    summary["text_chars"] = len(text)
    summary["signal_density"] = round(
        (float(sum(summary["keyword_hits"].values())) / max(1.0, float(len(text) / 100.0))),
        4,
    )
    return summary


def _extract_binary_text_snippet(content: bytes, *, max_chars: int = 4000) -> str:
    decoded = content.decode("utf-8", errors="ignore")
    cleaned = "".join(ch if ch.isprintable() else " " for ch in decoded)
    compact = " ".join(cleaned.split())
    if not compact:
        return ""
    return compact[: max(256, int(max_chars))]


def _extract_image_content(content: bytes, filename: str) -> str:
    lines = [f"[图像资料] 文件: {filename}", f"字节数: {len(content)}"]
    if Image is None:
        lines.append("图像解析: 当前环境未安装 Pillow，已纳入文件元信息。")
        return "\n".join(lines)
    try:
        with Image.open(io.BytesIO(content)) as img:
            lines.append(f"格式: {img.format or '未知'}")
            lines.append(f"尺寸: {img.width}x{img.height}")
            lines.append(f"模式: {img.mode}")
            if pytesseract is not None:
                try:
                    ocr_text = str(
                        pytesseract.image_to_string(img, lang="chi_sim+eng") or ""
                    ).strip()
                except Exception:
                    ocr_text = ""
                if ocr_text:
                    lines.append("[OCR文本提取]")
                    lines.append(ocr_text[:4000])
                else:
                    lines.append("OCR文本提取: 未识别到有效文本。")
            else:
                lines.append("OCR文本提取: 当前环境未安装 pytesseract，已纳入图像元信息。")
    except Exception as exc:
        lines.append(f"图像解析失败: {exc}")
    return "\n".join(lines)


def _pdf_backend_name() -> str:
    if pymupdf is not None:
        return "pymupdf"
    if PdfReader is not None:
        return "pypdf"
    return "none"


def _extract_pdf_text_with_pypdf(content: bytes) -> str:
    if PdfReader is None:
        return ""
    if not bytes(content or b"").lstrip().startswith(b"%PDF"):
        return ""
    try:
        reader = PdfReader(io.BytesIO(content))
    except Exception:
        return ""
    parts: List[str] = []
    for idx, page in enumerate(getattr(reader, "pages", []) or [], start=1):
        try:
            page_text = str(page.extract_text() or "")
        except Exception:
            page_text = ""
        page_text = page_text.strip()
        if page_text:
            parts.append(f"[PAGE:{idx}]\n{page_text}")
    return "\n\n".join(parts).strip()


def _extract_pdf_text(content: bytes, filename: str) -> str:
    if pymupdf is not None:
        doc = pymupdf.open(stream=content, filetype="pdf")
        try:
            parts: List[str] = []
            for idx, page in enumerate(doc, start=1):
                # Embed stable page markers so downstream diagnostics can map evidence to pages.
                page_text = page.get_text() or ""
                parts.append(f"[PAGE:{idx}]\n{page_text}")
            merged_pdf_text = "\n\n".join(parts).strip()
            text_chars = len(merged_pdf_text.replace("\n", "").strip())
            need_ocr = text_chars < DEFAULT_PDF_TEXT_MIN_CHARS_FOR_OCR
            if need_ocr and pytesseract is not None and Image is not None:
                ocr_chunks: List[str] = []
                for idx, page in enumerate(doc, start=1):
                    if idx > DEFAULT_PDF_OCR_MAX_PAGES:
                        break
                    try:
                        pix = page.get_pixmap(matrix=pymupdf.Matrix(2.0, 2.0), alpha=False)
                        with Image.open(io.BytesIO(pix.tobytes("png"))) as img:
                            ocr_text = str(
                                pytesseract.image_to_string(img, lang="chi_sim+eng") or ""
                            ).strip()
                    except Exception:
                        ocr_text = ""
                    if ocr_text:
                        ocr_chunks.append(f"[PAGE_OCR:{idx}]\n{ocr_text[:5000]}")
                if ocr_chunks:
                    merged_pdf_text = (
                        merged_pdf_text + "\n\n[PDF_OCR_FALLBACK]\n" + "\n\n".join(ocr_chunks)
                    ).strip()
            if merged_pdf_text:
                return f"[PDF_BACKEND:pymupdf]\n{merged_pdf_text}"
        finally:
            doc.close()

    pypdf_text = _extract_pdf_text_with_pypdf(content)
    if pypdf_text:
        return f"[PDF_BACKEND:pypdf]\n{pypdf_text}"

    if _pdf_backend_name() == "none":
        raise ValueError(
            "PDF 解析不可用：请安装与当前系统架构兼容的 PyMuPDF，或安装 pypdf 作为兼容解析后端。"
        )
    return f"[PDF资料] 文件: {filename}（未提取到有效文本）"


def _read_uploaded_file_content(content: bytes, filename: str) -> str:
    """根据文件名解析上传文件为文本，覆盖招标/清单/图纸/现场照片常见格式。"""
    name = filename.lower()
    if name.endswith(".txt") or name.endswith(".md") or name.endswith(".csv"):
        return content.decode("utf-8", errors="ignore")
    if name.endswith(".docx"):
        if Document is None:
            raise ValueError("DOCX 解析不可用：请安装与当前系统架构兼容的 python-docx/lxml。")
        doc = Document(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs)
    if name.endswith(".doc") or name.endswith(".docm"):
        snippet = _extract_binary_text_snippet(content)
        if snippet:
            return snippet
        return f"[DOC资料] 文件: {filename}（当前环境未启用结构化解析，已纳入文件元信息）"
    if name.endswith(".pdf"):
        return _extract_pdf_text(content, filename)
    if name.endswith(".json"):
        return content.decode("utf-8", errors="ignore")
    if name.endswith(".xlsx") or name.endswith(".xls") or name.endswith(".xlsm"):
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            parts = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    parts.append("\t".join(str(c) if c is not None else "" for c in row))
            wb.close()
            return "\n".join(parts)
        except Exception as e:
            raise ValueError(f"Excel 解析失败: {e}") from e
    if name.endswith(".dxf"):
        try:
            return _extract_dxf_text(content)
        except ValueError:
            raise
        except Exception as e:
            raise ValueError(f"DXF 解析失败: {e}") from e
    if name.endswith(".dwg"):
        return _extract_dwg_text(content, filename)
    if name.endswith((".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff")):
        return _extract_image_content(content, filename)
    snippet = _extract_binary_text_snippet(content, max_chars=2000)
    if snippet:
        return snippet
    raise ValueError(
        "仅支持 .txt、.md、.csv、.doc/.docx/.docm、.pdf、.json、.xlsx/.xls/.xlsm、.dxf/.dwg、图片格式"
    )


def _merge_materials_text(project_id: str) -> str:
    """将本项目已上传资料按分类合并为文本，作为评分与学习进化的依据。"""
    material_index = _build_project_material_index(project_id)
    files = material_index.get("files") if isinstance(material_index.get("files"), list) else []
    if not files:
        return ""
    groups: Dict[str, List[Dict[str, object]]] = {}
    for file_entry in files:
        if not isinstance(file_entry, dict):
            continue
        mat_type = _normalize_material_type(
            file_entry.get("material_type"), filename=file_entry.get("filename")
        )
        groups.setdefault(mat_type, []).append(file_entry)
    merge_order = ["tender_qa", "boq", "drawing", "site_photo"]
    for key in sorted(groups.keys()):
        if key not in merge_order:
            merge_order.append(key)

    sections: List[str] = []
    for mat_type in merge_order:
        rows = groups.get(mat_type) or []
        if not rows:
            continue
        rows.sort(key=lambda x: str(x.get("created_at") or ""))
        section_lines: List[str] = [f"=== {_material_type_label(mat_type)} ==="]
        for row in rows:
            filename = str(row.get("filename") or "")
            created_at = str(row.get("created_at") or "")[:19]
            if not bool(row.get("parsed_ok")):
                reason = str(row.get("parse_error") or "解析失败")
                if reason == "missing_path":
                    section_lines.append(f"--- {filename} ---\n[资料缺失] 未记录路径。")
                elif reason == "path_not_exists":
                    section_lines.append(f"--- {filename} ---\n[资料缺失] 文件已不存在。")
                else:
                    section_lines.append(
                        f"--- {filename} ({created_at}) ---\n[资料解析失败] {reason}"
                    )
                continue
            text = str(row.get("text") or "").strip()
            if not text:
                section_lines.append(
                    f"--- {filename} ({created_at}) ---\n[资料为空] 文件解析后为空内容。"
                )
                continue
            section_block = f"--- {filename} ({created_at}) ---\n{text}"
            if mat_type == "boq":
                boq_struct = (
                    row.get("boq_structured_summary")
                    if isinstance(row.get("boq_structured_summary"), dict)
                    else {}
                )
                if boq_struct:
                    section_block = (
                        section_block
                        + "\n\n[BOQ结构化摘要]\n"
                        + json.dumps(boq_struct, ensure_ascii=False)
                    )
            section_lines.append(section_block)
        sections.append("\n\n".join(section_lines))
    return "\n\n".join(sections).strip()


def _split_material_text_chunks(text: str, *, max_chars: int = 900) -> List[str]:
    raw_parts = re.split(r"\n{2,}|[。！？]\s*", str(text or ""))
    chunks: List[str] = []
    buf: List[str] = []
    size = 0
    for part in raw_parts:
        s = str(part or "").strip()
        if not s:
            continue
        if len(s) > max_chars:
            s = s[:max_chars]
        if size + len(s) > max_chars and buf:
            chunks.append("；".join(buf))
            buf = [s]
            size = len(s)
            continue
        buf.append(s)
        size += len(s)
    if buf:
        chunks.append("；".join(buf))
    return chunks[:80]


def _infer_chunk_dimension_id(material_type: str, chunk_text: str) -> str:
    candidates = MATERIAL_TYPE_DIMENSION_PRIORITY.get(material_type) or ["01"]
    lower = str(chunk_text or "").lower()
    best_dim = candidates[0]
    best_score = -1
    for dim_id in candidates:
        score = 0
        for kw in DIMENSION_RAG_KEYWORDS.get(dim_id, []):
            if kw.lower() in lower:
                score += 1
        if score > best_score:
            best_score = score
            best_dim = dim_id
    return best_dim


def _select_material_retrieval_chunks(
    project_id: str,
    submission_text: str,
    *,
    top_k: int = 12,
    per_type_quota: int = 1,
    per_file_quota: int = 3,
    query_terms_extra: Optional[List[str]] = None,
    query_numeric_terms: Optional[List[str]] = None,
    material_index: Optional[Dict[str, object]] = None,
) -> List[Dict[str, object]]:
    query_terms = set(_extract_terms(submission_text, max_terms=60))
    if isinstance(query_terms_extra, list):
        query_terms.update(str(x).strip().lower() for x in query_terms_extra if str(x).strip())
    # 确保不同资料类型关键词也参与检索，避免“只用施组词命中”。
    for kws in MATERIAL_TYPE_KEYWORDS.values():
        query_terms.update(k.lower() for k in kws)
    numeric_terms = set(_extract_numeric_terms(submission_text, max_terms=24))
    if isinstance(query_numeric_terms, list):
        numeric_terms.update(
            _normalize_numeric_token(x) for x in query_numeric_terms if _normalize_numeric_token(x)
        )

    index = (
        material_index
        if isinstance(material_index, dict)
        else _build_project_material_index(project_id)
    )
    available_types = [str(x) for x in (index.get("available_types") or []) if str(x).strip()]
    file_entries = index.get("files") if isinstance(index.get("files"), list) else []
    candidates: List[Dict[str, object]] = []
    backfill_candidates: List[Dict[str, object]] = []
    for file_entry in file_entries:
        if not isinstance(file_entry, dict):
            continue
        if not bool(file_entry.get("parsed_ok")):
            continue
        filename = str(file_entry.get("filename") or "").strip()
        mat_type = str(file_entry.get("material_type") or "").strip()
        chunks = file_entry.get("chunks") if isinstance(file_entry.get("chunks"), list) else []
        if not chunks:
            text_raw = str(file_entry.get("text") or "")
            if text_raw:
                chunks = _split_material_text_chunks(text_raw, max_chars=900)
        for idx, chunk in enumerate(chunks, start=1):
            chunk_text = str(chunk or "")
            lower = chunk_text.lower()
            matched_terms = [t for t in query_terms if t and t in lower][:8]
            chunk_numeric_terms = _extract_numeric_terms(chunk_text, max_terms=30)
            chunk_numeric_set = {str(x) for x in chunk_numeric_terms}
            matched_numeric_terms = [
                token for token in numeric_terms if token and token in chunk_numeric_set
            ][:8]
            type_keyword_hits = [
                kw.lower() for kw in MATERIAL_TYPE_KEYWORDS.get(mat_type, []) if kw.lower() in lower
            ][:6]
            if not matched_terms:
                # 兜底：若命中该类型关键词也纳入候选
                matched_terms = list(type_keyword_hits)
            if not matched_terms and not matched_numeric_terms:
                # 二级兜底：即使没有直接命中查询词，也为每种资料类型保留“高信息密度”候选，
                # 后续用于类型补全，避免某类资料长期无法进入评分证据链。
                lexical_terms = _extract_terms(chunk_text, max_terms=20)
                chunk_numeric_norm = [
                    token
                    for token in (
                        _normalize_numeric_token(item)
                        for item in _extract_numeric_terms(chunk_text, max_terms=18)
                    )
                    if token
                ]
                fallback_score = 0
                fallback_score += min(4, len(lexical_terms))
                fallback_score += min(4, len(chunk_numeric_norm))
                fallback_score += min(3, len(type_keyword_hits))
                if mat_type in {"boq", "drawing"}:
                    fallback_score += 1
                if fallback_score <= 0 or len(chunk_text.strip()) < 24:
                    continue
                backfill_candidates.append(
                    {
                        "material_type": mat_type,
                        "filename": filename,
                        "chunk_id": f"{filename}#c{idx:03d}",
                        "dimension_id": _infer_chunk_dimension_id(mat_type, chunk),
                        "score": fallback_score,
                        "matched_terms": type_keyword_hits or lexical_terms[:6],
                        "matched_numeric_terms": chunk_numeric_norm[:6],
                        "chunk_preview": chunk_text[:220],
                        "is_backfill": True,
                    }
                )
                continue
            dimension_id = _infer_chunk_dimension_id(mat_type, chunk)
            score = len(matched_terms) + (2 if mat_type in {"boq", "drawing"} else 1)
            if matched_numeric_terms:
                score += min(4, len(matched_numeric_terms))
                if mat_type in {"tender_qa", "boq", "drawing"}:
                    score += 1
            candidates.append(
                {
                    "material_type": mat_type,
                    "filename": filename,
                    "chunk_id": f"{filename}#c{idx:03d}",
                    "dimension_id": dimension_id,
                    "score": score,
                    "matched_terms": matched_terms,
                    "matched_numeric_terms": matched_numeric_terms,
                    "chunk_preview": chunk_text[:220],
                }
            )

    if not candidates and not backfill_candidates:
        return []
    candidates.sort(
        key=lambda x: (
            -int(x.get("score", 0)),
            str(x.get("material_type")),
            str(x.get("filename")),
            str(x.get("chunk_id")),
        )
    )
    backfill_candidates.sort(
        key=lambda x: (
            -int(x.get("score", 0)),
            str(x.get("material_type")),
            str(x.get("filename")),
            str(x.get("chunk_id")),
        )
    )
    target_k = max(1, int(top_k))
    selected: List[Dict[str, object]] = []
    seen_chunk_ids: set[str] = set()
    selected_per_file: Dict[str, int] = {}

    type_quota = max(1, int(per_type_quota))
    file_quota = max(1, int(per_file_quota))

    def _try_pick(candidate: Dict[str, object], *, selected_via: str) -> bool:
        chunk_id = str(candidate.get("chunk_id") or "")
        if not chunk_id or chunk_id in seen_chunk_ids:
            return False
        filename_text = str(candidate.get("filename") or "").strip()
        if filename_text and int(selected_per_file.get(filename_text, 0)) >= file_quota:
            return False
        row = dict(candidate)
        row["selected_via"] = selected_via
        selected.append(row)
        seen_chunk_ids.add(chunk_id)
        if filename_text:
            selected_per_file[filename_text] = int(selected_per_file.get(filename_text, 0)) + 1
        return True

    # 先保证“每个已上传资料类型至少取 N 个块”，避免单一资料类型垄断检索。
    for mat_type in available_types:
        type_selected = 0
        for candidate in candidates:
            if str(candidate.get("material_type") or "") != mat_type:
                continue
            if not _try_pick(candidate, selected_via="type_quota"):
                continue
            type_selected += 1
            if len(selected) >= target_k:
                return selected
            if type_selected >= type_quota:
                break

    selected_types = {
        str(item.get("material_type") or "").strip()
        for item in selected
        if str(item.get("material_type") or "").strip()
    }
    for mat_type in available_types:
        if mat_type in selected_types:
            continue
        for candidate in backfill_candidates:
            if str(candidate.get("material_type") or "") != mat_type:
                continue
            if not _try_pick(candidate, selected_via="type_backfill"):
                continue
            selected_types.add(mat_type)
            if len(selected) >= target_k:
                return selected
            break

    # 再按全局相关性补齐 top_k。
    for candidate in candidates:
        if not _try_pick(candidate, selected_via="global_rank"):
            continue
        if len(selected) >= target_k:
            break
    if len(selected) < target_k:
        for candidate in backfill_candidates:
            if not _try_pick(candidate, selected_via="global_backfill"):
                continue
            if len(selected) >= target_k:
                break
    return selected


def _build_material_retrieval_requirements(
    project_id: str,
    retrieval_chunks: List[Dict[str, object]],
) -> List[Dict[str, object]]:
    now = _now_iso()
    reqs: List[Dict[str, object]] = []
    for idx, chunk in enumerate(retrieval_chunks, start=1):
        dim_id = str(chunk.get("dimension_id") or "01")
        chunk_id = str(chunk.get("chunk_id") or f"chunk-{idx}")
        mat_type = str(chunk.get("material_type") or MATERIAL_TYPE_DEFAULT)
        filename = str(chunk.get("filename") or "")
        hints = [str(x) for x in (chunk.get("matched_terms") or []) if str(x).strip()]
        preview = str(chunk.get("chunk_preview") or "").strip()
        if preview:
            hints.append(preview[:120])
        hints = _to_text_items(hints, max_items=8)
        if not hints:
            continue
        minimum_hint_hits = 2 if mat_type in {"tender_qa", "boq", "drawing"} else 1
        reqs.append(
            {
                "id": f"runtime-rag-{project_id[:8]}-{idx}",
                "project_id": project_id,
                "dimension_id": dim_id,
                "req_label": f"资料检索证据：{_material_type_label(mat_type)} / {filename} / {chunk_id}",
                "req_type": "semantic",
                "patterns": {
                    "hints": hints,
                    "minimum_hint_hits": minimum_hint_hits,
                    "material_type": mat_type,
                    "source_filename": filename,
                    "chunk_id": chunk_id,
                    "source_mode": str(chunk.get("selected_via") or "retrieval_chunks"),
                },
                "mandatory": False,
                "weight": 0.7,
                "material_type": mat_type,
                "source_anchor_id": None,
                "source_pack_id": "runtime_material_rag",
                "source_pack_version": "v2-rag-1",
                "priority": 88.0,
                "override_key": f"runtime::material_rag::{chunk_id}",
                "lint": {},
                "created_at": now,
            }
        )
    return reqs


def _build_material_consistency_requirements(
    project_id: str,
    retrieval_chunks: List[Dict[str, object]],
    *,
    available_material_types: Optional[List[str]] = None,
) -> List[Dict[str, object]]:
    now = _now_iso()
    grouped: Dict[str, List[Dict[str, object]]] = {}
    for chunk in retrieval_chunks:
        mat_type = str(chunk.get("material_type") or MATERIAL_TYPE_DEFAULT)
        grouped.setdefault(mat_type, []).append(chunk)

    normalized_available_types: List[str] = []
    for item in available_material_types or []:
        key = _normalize_material_type(item)
        if key and key not in normalized_available_types:
            normalized_available_types.append(key)
    for mat_type in normalized_available_types:
        grouped.setdefault(mat_type, [])

    reqs: List[Dict[str, object]] = []
    req_index = 0
    ordered_types = normalized_available_types + [
        mat_type for mat_type in grouped.keys() if mat_type not in normalized_available_types
    ]
    for mat_type in ordered_types:
        chunks = grouped.get(mat_type) or []
        keyword_counter: Counter[str] = Counter()
        numeric_counter: Counter[str] = Counter()
        preview_terms: List[str] = []
        filenames: List[str] = []
        for chunk in chunks[:12]:
            filename = str(chunk.get("filename") or "")
            if filename:
                filenames.append(filename)
            for term in chunk.get("matched_terms") or []:
                token = str(term or "").strip()
                if token:
                    keyword_counter[token] += 1
            for token in chunk.get("matched_numeric_terms") or []:
                numeric = _normalize_numeric_token(token)
                if numeric:
                    numeric_counter[numeric] += 1
            preview_terms.extend(
                _extract_terms(str(chunk.get("chunk_preview") or ""), max_terms=10)
            )
            for token in _extract_numeric_terms(str(chunk.get("chunk_preview") or ""), max_terms=8):
                numeric = _normalize_numeric_token(token)
                if numeric:
                    numeric_counter[numeric] += 1
        must_terms = [item for item, _ in keyword_counter.most_common(8)]
        if len(must_terms) < 4:
            must_terms.extend([item for item in preview_terms if item not in must_terms])
        if len(must_terms) < 3:
            must_terms.extend(
                [
                    item
                    for item in MATERIAL_TYPE_KEYWORDS.get(mat_type, [])
                    if item not in must_terms
                ]
            )
        must_terms = _to_text_items(must_terms, max_items=8)
        must_numbers = [item for item, _ in numeric_counter.most_common(4)]
        if not must_terms and not must_numbers:
            continue

        req_index += 1
        minimum_terms = (
            2
            if (must_terms and mat_type in {"tender_qa", "boq", "drawing"})
            else (1 if must_terms else 0)
        )
        minimum_numbers = 1 if must_numbers else 0
        known_required_types = {"tender_qa", "boq", "drawing", "site_photo"}
        mandatory = mat_type in known_required_types and (
            (not normalized_available_types) or (mat_type in normalized_available_types)
        )
        source_mode = "retrieval_chunks" if chunks else "fallback_keywords"
        dim_id = (MATERIAL_TYPE_DIMENSION_PRIORITY.get(mat_type) or ["01"])[0]
        reqs.append(
            {
                "id": f"runtime-consistency-{project_id[:8]}-{req_index}",
                "project_id": project_id,
                "dimension_id": dim_id,
                "req_label": f"跨资料一致性：施组需体现{_material_type_label(mat_type)}关键约束",
                "req_type": "material_consistency",
                "patterns": {
                    "must_hit_terms": must_terms,
                    "minimum_terms": minimum_terms,
                    "must_hit_numbers": must_numbers,
                    "minimum_numbers": minimum_numbers,
                    "material_type": mat_type,
                    "sample_filenames": _to_text_items(filenames, max_items=3),
                    "within_dimension_scope": False,
                    "source_mode": source_mode,
                },
                "mandatory": mandatory,
                "weight": 1.1
                if (mandatory and mat_type in {"tender_qa", "boq", "drawing"})
                else 0.9,
                "material_type": mat_type,
                "source_anchor_id": None,
                "source_pack_id": "runtime_material_consistency",
                "source_pack_version": "v2-consistency-1",
                "priority": 92.0,
                "override_key": f"runtime::material_consistency::{mat_type}",
                "lint": {},
                "created_at": now,
            }
        )
    return reqs


def _build_material_quality_snapshot(project_id: str) -> Dict[str, object]:
    index = _build_project_material_index(project_id)
    snapshot = (
        index.get("quality_snapshot") if isinstance(index.get("quality_snapshot"), dict) else {}
    )
    return copy.deepcopy(snapshot)


def _resolve_material_gate_config(project: Dict[str, object]) -> Dict[str, object]:
    meta = project.get("meta") if isinstance(project.get("meta"), dict) else {}
    enforce = bool(meta.get("enforce_material_gate", DEFAULT_ENFORCE_MATERIAL_GATE))
    enforce_depth_gate = bool(
        meta.get("enforce_material_depth_gate", DEFAULT_ENFORCE_MATERIAL_DEPTH_GATE)
    )
    required_raw = meta.get("required_material_types", DEFAULT_REQUIRED_MATERIAL_TYPES)
    required_types: List[str] = []
    if isinstance(required_raw, list):
        for item in required_raw:
            key = _normalize_material_type(item)
            if key not in required_types:
                required_types.append(key)
    if not required_types:
        required_types = list(DEFAULT_REQUIRED_MATERIAL_TYPES)

    min_chars_map = dict(DEFAULT_MIN_PARSED_CHARS_BY_TYPE)
    raw_map = meta.get("min_parsed_chars_by_type")
    if isinstance(raw_map, dict):
        for k, v in raw_map.items():
            key = _normalize_material_type(k)
            numeric = _to_float_or_none(v)
            if numeric is None:
                continue
            min_chars_map[key] = max(0, int(round(numeric)))

    min_chunks_map = dict(DEFAULT_MIN_PARSED_CHUNKS_BY_TYPE)
    raw_chunks_map = meta.get("min_parsed_chunks_by_type")
    if isinstance(raw_chunks_map, dict):
        for k, v in raw_chunks_map.items():
            key = _normalize_material_type(k)
            numeric = _to_float_or_none(v)
            if numeric is None:
                continue
            min_chunks_map[key] = max(0, int(round(numeric)))

    min_numeric_terms_map = dict(DEFAULT_MIN_NUMERIC_TERMS_BY_TYPE)
    raw_numeric_map = meta.get("min_numeric_terms_by_type")
    if isinstance(raw_numeric_map, dict):
        for k, v in raw_numeric_map.items():
            key = _normalize_material_type(k)
            numeric = _to_float_or_none(v)
            if numeric is None:
                continue
            min_numeric_terms_map[key] = max(0, int(round(numeric)))

    total_chunks_raw = _to_float_or_none(meta.get("min_total_parsed_chunks"))
    min_total_chunks = (
        max(0, int(round(total_chunks_raw)))
        if total_chunks_raw is not None
        else int(DEFAULT_MIN_TOTAL_PARSED_CHUNKS)
    )

    total_chars = _to_float_or_none(meta.get("min_total_parsed_chars"))
    min_total_chars = (
        max(0, int(round(total_chars)))
        if total_chars is not None
        else int(DEFAULT_MIN_TOTAL_PARSED_CHARS)
    )
    fail_ratio_raw = _to_float_or_none(meta.get("max_material_parse_fail_ratio"))
    max_fail_ratio = (
        min(1.0, max(0.0, float(fail_ratio_raw)))
        if fail_ratio_raw is not None
        else float(DEFAULT_MAX_MATERIAL_PARSE_FAIL_RATIO)
    )
    block_on_parse_failure = bool(
        meta.get(
            "block_on_any_material_parse_failure",
            DEFAULT_BLOCK_ON_ANY_MATERIAL_PARSE_FAILURE,
        )
    )
    return {
        "enforce": enforce,
        "enforce_depth_gate": enforce_depth_gate,
        "required_types": required_types,
        "min_chars_by_type": min_chars_map,
        "min_chunks_by_type": min_chunks_map,
        "min_numeric_terms_by_type": min_numeric_terms_map,
        "min_total_chunks": min_total_chunks,
        "min_total_chars": min_total_chars,
        "max_fail_ratio": max_fail_ratio,
        "block_on_any_parse_failure": block_on_parse_failure,
    }


def _validate_material_gate_for_scoring(
    project_id: str,
    project: Dict[str, object],
    *,
    raise_on_fail: bool = True,
) -> tuple[Dict[str, object], List[str]]:
    snapshot = _build_material_quality_snapshot(project_id)
    cfg = _resolve_material_gate_config(project)
    counts_by_type = snapshot.get("counts_by_type") if isinstance(snapshot, dict) else {}
    chars_by_type = snapshot.get("chars_by_type") if isinstance(snapshot, dict) else {}
    chunks_by_type = snapshot.get("chunks_by_type") if isinstance(snapshot, dict) else {}
    numeric_terms_by_type = (
        snapshot.get("numeric_terms_by_type") if isinstance(snapshot, dict) else {}
    )
    counts_by_type = counts_by_type if isinstance(counts_by_type, dict) else {}
    chars_by_type = chars_by_type if isinstance(chars_by_type, dict) else {}
    chunks_by_type = chunks_by_type if isinstance(chunks_by_type, dict) else {}
    numeric_terms_by_type = numeric_terms_by_type if isinstance(numeric_terms_by_type, dict) else {}
    issues: List[str] = []
    depth_issues: List[str] = []
    required_types = cfg.get("required_types") if isinstance(cfg, dict) else []
    required_types = required_types if isinstance(required_types, list) else []
    for tpe in required_types:
        label = _material_type_label(tpe)
        count = int(counts_by_type.get(tpe, 0))
        if count <= 0:
            issues.append(f"缺少必需资料类型：{label}")
            continue
        min_chars = int((cfg.get("min_chars_by_type") or {}).get(tpe, 0))
        parsed_chars = int(chars_by_type.get(tpe, 0))
        if parsed_chars < min_chars:
            issues.append(f"{label}解析文本不足：{parsed_chars} 字（要求至少 {min_chars} 字）")
        min_chunks = int((cfg.get("min_chunks_by_type") or {}).get(tpe, 0))
        parsed_chunks = int(chunks_by_type.get(tpe, 0))
        if parsed_chunks < min_chunks:
            depth_issues.append(
                f"{label}解析分块不足：{parsed_chunks} 段（建议至少 {min_chunks} 段）"
            )
        min_numeric_terms = int((cfg.get("min_numeric_terms_by_type") or {}).get(tpe, 0))
        parsed_numeric_terms = int(numeric_terms_by_type.get(tpe, 0))
        if min_numeric_terms > 0 and parsed_numeric_terms < min_numeric_terms:
            depth_issues.append(
                f"{label}数字约束提取不足：{parsed_numeric_terms} 项（建议至少 {min_numeric_terms} 项）"
            )

    total_parsed_chars = int(snapshot.get("total_parsed_chars", 0))
    min_total_chars = int(cfg.get("min_total_chars", DEFAULT_MIN_TOTAL_PARSED_CHARS))
    if total_parsed_chars < min_total_chars:
        issues.append(
            f"项目资料总解析文本不足：{total_parsed_chars} 字（要求至少 {min_total_chars} 字）"
        )
    total_parsed_chunks = int(snapshot.get("total_parsed_chunks", 0))
    min_total_chunks = int(cfg.get("min_total_chunks", DEFAULT_MIN_TOTAL_PARSED_CHUNKS))
    if total_parsed_chunks < min_total_chunks:
        depth_issues.append(
            f"项目资料总解析分块不足：{total_parsed_chunks} 段（建议至少 {min_total_chunks} 段）"
        )

    parse_fail_ratio = _to_float_or_none(snapshot.get("parse_fail_ratio")) or 0.0
    parsed_failed_files = int(_to_float_or_none(snapshot.get("parsed_failed_files")) or 0)
    max_fail_ratio = float(cfg.get("max_fail_ratio", DEFAULT_MAX_MATERIAL_PARSE_FAIL_RATIO))
    block_on_parse_failure = bool(
        cfg.get(
            "block_on_any_parse_failure",
            DEFAULT_BLOCK_ON_ANY_MATERIAL_PARSE_FAILURE,
        )
    )
    if block_on_parse_failure and parsed_failed_files > 0:
        parsed_fail_details = (
            snapshot.get("parsed_fail_details")
            if isinstance(snapshot.get("parsed_fail_details"), list)
            else []
        )
        preview = "；".join(
            str(item.get("filename") or "")
            for item in parsed_fail_details[:3]
            if isinstance(item, dict) and str(item.get("filename") or "").strip()
        )
        detail_suffix = f"（示例：{preview}）" if preview else ""
        issues.append(f"存在 {parsed_failed_files} 份资料解析失败，硬闸门已阻断评分{detail_suffix}")
    if parse_fail_ratio > max_fail_ratio:
        issues.append(f"资料解析失败比例过高：{parse_fail_ratio:.1%}（阈值 {max_fail_ratio:.1%}）")

    snapshot["gate"] = {
        "enforce": bool(cfg.get("enforce", False)),
        "required_types": required_types,
        "min_chars_by_type": cfg.get("min_chars_by_type", {}),
        "min_total_chars": min_total_chars,
        "max_fail_ratio": max_fail_ratio,
        "block_on_any_parse_failure": block_on_parse_failure,
        "issues": issues,
        "passed": len(issues) == 0,
    }
    enforce_depth_gate = bool(cfg.get("enforce_depth_gate", DEFAULT_ENFORCE_MATERIAL_DEPTH_GATE))
    snapshot["depth_gate"] = {
        "enforce": enforce_depth_gate,
        "required_types": required_types,
        "min_chunks_by_type": cfg.get("min_chunks_by_type", {}),
        "min_numeric_terms_by_type": cfg.get("min_numeric_terms_by_type", {}),
        "min_total_chunks": min_total_chunks,
        "issues": depth_issues,
        "passed": len(depth_issues) == 0,
    }

    blocking_reasons: List[str] = []
    if bool(cfg.get("enforce", False)):
        blocking_reasons.extend(issues)
    if enforce_depth_gate:
        blocking_reasons.extend(depth_issues)

    if raise_on_fail and blocking_reasons:
        detail = "；".join(blocking_reasons)
        raise HTTPException(status_code=422, detail=f"资料门禁未通过：{detail}")
    return snapshot, issues


def _build_scoring_readiness(project_id: str, project: Dict[str, object]) -> Dict[str, object]:
    snapshot, issues = _validate_material_gate_for_scoring(
        project_id,
        project,
        raise_on_fail=False,
    )
    gate = snapshot.get("gate") if isinstance(snapshot.get("gate"), dict) else {}
    depth_gate = snapshot.get("depth_gate") if isinstance(snapshot.get("depth_gate"), dict) else {}
    depth_enforced = bool(depth_gate.get("enforce", False))
    depth_passed = bool(depth_gate.get("passed", True))
    gate_passed = bool(gate.get("passed", True)) and (depth_passed if depth_enforced else True)

    submissions = [s for s in load_submissions() if str(s.get("project_id")) == project_id]
    submission_count = len(submissions)
    non_empty_submission_count = sum(1 for s in submissions if str(s.get("text") or "").strip())
    scored_submission_count = sum(1 for s in submissions if _submission_is_scored(s))

    warnings: List[str] = []
    if non_empty_submission_count <= 0:
        warnings.append("尚未上传施组文件，上传后才能执行评分。")
    elif scored_submission_count <= 0:
        warnings.append("当前施组尚未完成评分，可点击“评分施组”触发。")

    parsed_ok_files = int(_to_float_or_none(snapshot.get("parsed_ok_files")) or 0)
    total_files = int(_to_float_or_none(snapshot.get("total_files")) or 0)
    if total_files > 0 and parsed_ok_files < total_files:
        warnings.append(f"存在 {total_files - parsed_ok_files} 份资料解析失败，建议修复后再评分。")
    depth_issues = depth_gate.get("issues") if isinstance(depth_gate.get("issues"), list) else []
    depth_issue_texts = [str(x).strip() for x in depth_issues if str(x).strip()]
    if depth_issue_texts:
        if depth_enforced:
            warnings.append("资料深读门禁已开启，需补齐深读质量后才能评分。")
        else:
            warnings.append("资料深读预警：建议补充关键约束资料，提高评分稳定性。")

    ready = gate_passed and non_empty_submission_count > 0
    issue_texts = [str(x).strip() for x in issues if str(x).strip()]
    if depth_enforced and depth_issue_texts:
        issue_texts.extend([f"资料深读门禁：{text}" for text in depth_issue_texts if text])

    return {
        "project_id": project_id,
        "ready": ready,
        "score_button_enabled": ready,
        "gate_passed": gate_passed,
        "issues": issue_texts,
        "warnings": warnings[:8],
        "material_quality": snapshot,
        "material_gate": gate,
        "material_depth_gate": depth_gate,
        "submissions": {
            "total": submission_count,
            "non_empty": non_empty_submission_count,
            "scored": scored_submission_count,
        },
        "retrieval_policy": _resolve_material_utilization_policy(project),
        "generated_at": _now_iso(),
    }


def _build_project_mece_audit(project_id: str, project: Dict[str, object]) -> Dict[str, object]:
    """
    项目级 MECE 审计（Mutually Exclusive, Collectively Exhaustive）。

    目标：对“资料输入 -> 评分输出 -> 进化闭环 -> 运行稳定性”给出可执行诊断，
    用于自动化验收与持续回归。
    """
    readiness = _build_scoring_readiness(project_id, project)
    depth = _build_material_depth_report(project_id, project)
    evo_health = _build_evolution_health_report(project_id, project)
    self_check = _run_system_self_check(project_id)

    submissions_all = [s for s in load_submissions() if str(s.get("project_id")) == project_id]
    scored_submissions = [s for s in submissions_all if _submission_is_scored(s)]
    scored_totals: List[float] = []
    for row in scored_submissions:
        score_fields = _resolve_submission_score_fields(
            row,
            allow_pred_score=True,
            score_scale_max=100,
        )
        scored_totals.append(float(_to_float_or_none(score_fields.get("total_score")) or 0.0))
    scored_totals = [max(0.0, min(100.0, x)) for x in scored_totals]
    if scored_totals:
        mean_score = sum(scored_totals) / len(scored_totals)
        variance = sum((x - mean_score) ** 2 for x in scored_totals) / len(scored_totals)
        std_score = variance**0.5
    else:
        std_score = 0.0
        mean_score = 0.0

    latest_scored: Optional[Dict[str, object]] = None
    if scored_submissions:
        latest_scored = max(
            scored_submissions,
            key=lambda x: str(x.get("updated_at") or x.get("created_at") or ""),
        )
    scoring_basis_ok = False
    evidence_trace_ok = False
    if isinstance(latest_scored, dict):
        try:
            basis_payload = _build_submission_scoring_basis_report(
                project_id=project_id,
                submission=latest_scored,
            )
            mece_inputs = (
                basis_payload.get("mece_inputs")
                if isinstance(basis_payload.get("mece_inputs"), dict)
                else {}
            )
            required_keys = (
                "project_materials_extracted",
                "shigong_parsed",
                "bid_requirements_loaded",
                "attention_16d_weights_injected",
                "custom_instructions_injected",
            )
            scoring_basis_ok = all(bool(mece_inputs.get(k)) for k in required_keys)
            evidence_trace = (
                basis_payload.get("evidence_trace")
                if isinstance(basis_payload.get("evidence_trace"), dict)
                else {}
            )
            total_req = int(_to_float_or_none(evidence_trace.get("total_requirements")) or 0)
            hit_req = int(_to_float_or_none(evidence_trace.get("total_hits")) or 0)
            evidence_trace_ok = total_req > 0 and hit_req > 0
        except Exception:
            scoring_basis_ok = False
            evidence_trace_ok = False

    dim_rows: List[Dict[str, object]] = []

    def _append_dim(
        *,
        key: str,
        title: str,
        checks: Dict[str, object],
        pass_condition: bool,
        warnings: Optional[List[str]] = None,
        fail_reasons: Optional[List[str]] = None,
    ) -> None:
        warn_rows = [str(x).strip() for x in (warnings or []) if str(x).strip()]
        fail_rows = [str(x).strip() for x in (fail_reasons or []) if str(x).strip()]
        status = "pass" if pass_condition else ("warn" if warn_rows else "fail")
        if fail_rows:
            status = "fail"
        dim_rows.append(
            {
                "key": key,
                "title": title,
                "status": status,
                "checks": checks,
                "warnings": warn_rows,
                "issues": fail_rows,
            }
        )

    _append_dim(
        key="input_chain",
        title="输入链路完整性",
        checks={
            "gate_passed": bool(readiness.get("gate_passed")),
            "ready": bool(readiness.get("ready")),
            "material_files": int(
                _to_float_or_none(((depth.get("quality_summary") or {}).get("total_files"))) or 0
            ),
            "materials_total_parsed_chars": int(
                _to_float_or_none(
                    ((readiness.get("material_quality") or {}).get("total_parsed_chars"))
                )
                or 0
            ),
            "required_types": list(
                ((readiness.get("material_gate") or {}).get("required_types")) or []
            ),
        },
        pass_condition=bool(readiness.get("gate_passed")) and bool(readiness.get("ready")),
        warnings=[str(x) for x in (readiness.get("warnings") or [])],
        fail_reasons=[str(x) for x in (readiness.get("issues") or [])],
    )

    score_variance_warn = []
    if len(scored_totals) >= 3 and std_score < 1.0:
        score_variance_warn.append(
            f"已评分样本 {len(scored_totals)} 份，但总分标准差仅 {std_score:.2f}，建议检查区分度。"
        )
    _append_dim(
        key="scoring_validity",
        title="评分有效性与可解释性",
        checks={
            "scored_submissions": len(scored_submissions),
            "mean_total_score": round(mean_score, 2),
            "std_total_score": round(std_score, 4),
            "scoring_basis_ok": bool(scoring_basis_ok),
            "evidence_trace_ok": bool(evidence_trace_ok),
        },
        pass_condition=len(scored_submissions) > 0 and scoring_basis_ok and evidence_trace_ok,
        warnings=score_variance_warn,
        fail_reasons=[],
    )

    gt_summary = evo_health.get("summary") if isinstance(evo_health.get("summary"), dict) else {}
    gt_count = int(_to_float_or_none(gt_summary.get("ground_truth_count")) or 0)
    matched_pred_count = int(_to_float_or_none(gt_summary.get("matched_prediction_count")) or 0)
    has_evolved_multipliers = bool(gt_summary.get("has_evolved_multipliers"))
    evo_fail_reasons: List[str] = []
    evo_warnings: List[str] = []
    if gt_count <= 0:
        evo_fail_reasons.append("未录入真实评标，进化闭环未激活。")
    elif gt_count < 3:
        evo_warnings.append(f"真实评标样本仅 {gt_count} 条，建议至少 3 条以上。")
    if gt_count > 0 and matched_pred_count <= 0:
        evo_fail_reasons.append("真实评标与系统预测未形成有效匹配，闭环训练不可用。")
    if gt_count >= 3 and not has_evolved_multipliers:
        evo_warnings.append("已具备样本但尚未产出进化权重，建议执行“学习进化”。")
    _append_dim(
        key="self_evolution_loop",
        title="自我进化闭环",
        checks={
            "ground_truth_count": gt_count,
            "matched_prediction_count": matched_pred_count,
            "has_evolved_multipliers": has_evolved_multipliers,
            "weights_source": str(gt_summary.get("current_weights_source") or "-"),
            "drift_level": str((evo_health.get("drift") or {}).get("level") or "unknown"),
        },
        pass_condition=gt_count >= 1 and matched_pred_count >= 1,
        warnings=evo_warnings,
        fail_reasons=evo_fail_reasons,
    )

    self_check_items = self_check.get("items") if isinstance(self_check.get("items"), list) else []
    failed_items = [str(x.get("name") or "") for x in self_check_items if not bool(x.get("ok"))]
    _append_dim(
        key="runtime_stability",
        title="运行稳定性",
        checks={
            "self_check_ok": bool(self_check.get("ok")),
            "failed_items": failed_items,
            "service_health": bool(
                next(
                    (bool(i.get("ok")) for i in self_check_items if str(i.get("name")) == "health"),
                    False,
                )
            ),
        },
        pass_condition=bool(self_check.get("ok")),
        warnings=[],
        fail_reasons=[],
    )

    pass_count = sum(1 for row in dim_rows if row.get("status") == "pass")
    warn_count = sum(1 for row in dim_rows if row.get("status") == "warn")
    fail_count = sum(1 for row in dim_rows if row.get("status") == "fail")
    total_dims = max(1, len(dim_rows))
    pass_rate = round(pass_count / total_dims, 4)
    health_score = round((pass_count + 0.5 * warn_count) / total_dims * 100, 2)
    level = "good"
    if fail_count > 0:
        level = "critical" if fail_count >= 2 else "watch"
    elif warn_count > 0:
        level = "watch"

    recommendations: List[str] = []
    for row in dim_rows:
        for text in (row.get("issues") or []) + (row.get("warnings") or []):
            s = str(text).strip()
            if s and s not in recommendations:
                recommendations.append(s)

    return {
        "project_id": project_id,
        "generated_at": _now_iso(),
        "overall": {
            "health_score": health_score,
            "level": level,
            "pass_rate": pass_rate,
            "pass_count": pass_count,
            "warn_count": warn_count,
            "fail_count": fail_count,
            "total_dimensions": total_dims,
        },
        "dimensions": dim_rows,
        "summary": {
            "submission_total": len(submissions_all),
            "submission_scored": len(scored_submissions),
            "score_mean_100": round(mean_score, 2),
            "score_std_100": round(std_score, 4),
            "ground_truth_count": gt_count,
            "matched_prediction_count": matched_pred_count,
            "self_check_ok": bool(self_check.get("ok")),
        },
        "recommendations": recommendations[:20],
    }


def _build_material_depth_report(project_id: str, project: Dict[str, object]) -> Dict[str, object]:
    readiness = _build_scoring_readiness(project_id, project)
    quality = (
        readiness.get("material_quality")
        if isinstance(readiness.get("material_quality"), dict)
        else {}
    )
    gate = (
        readiness.get("material_gate") if isinstance(readiness.get("material_gate"), dict) else {}
    )
    depth_gate = (
        readiness.get("material_depth_gate")
        if isinstance(readiness.get("material_depth_gate"), dict)
        else {}
    )
    counts_by_type = (
        quality.get("counts_by_type") if isinstance(quality.get("counts_by_type"), dict) else {}
    )
    chars_by_type = (
        quality.get("chars_by_type") if isinstance(quality.get("chars_by_type"), dict) else {}
    )
    chunks_by_type = (
        quality.get("chunks_by_type") if isinstance(quality.get("chunks_by_type"), dict) else {}
    )
    numeric_terms_by_type = (
        quality.get("numeric_terms_by_type")
        if isinstance(quality.get("numeric_terms_by_type"), dict)
        else {}
    )
    parsed_ok_by_type = (
        quality.get("parsed_ok_by_type")
        if isinstance(quality.get("parsed_ok_by_type"), dict)
        else {}
    )
    parsed_fail_by_type = (
        quality.get("parsed_fail_by_type")
        if isinstance(quality.get("parsed_fail_by_type"), dict)
        else {}
    )
    required_types = (
        gate.get("required_types") if isinstance(gate.get("required_types"), list) else []
    )
    all_types: List[str] = []
    for src in (required_types, list(counts_by_type.keys())):
        for item in src:
            key = _normalize_material_type(item)
            if key and key not in all_types:
                all_types.append(key)

    per_type_rows: List[Dict[str, object]] = []
    min_chars_by_type = (
        gate.get("min_chars_by_type") if isinstance(gate.get("min_chars_by_type"), dict) else {}
    )
    min_chunks_by_type = (
        depth_gate.get("min_chunks_by_type")
        if isinstance(depth_gate.get("min_chunks_by_type"), dict)
        else {}
    )
    min_numeric_terms_by_type = (
        depth_gate.get("min_numeric_terms_by_type")
        if isinstance(depth_gate.get("min_numeric_terms_by_type"), dict)
        else {}
    )
    for mat_type in all_types:
        files = int(_to_float_or_none(counts_by_type.get(mat_type)) or 0)
        chars = int(_to_float_or_none(chars_by_type.get(mat_type)) or 0)
        chunks = int(_to_float_or_none(chunks_by_type.get(mat_type)) or 0)
        numeric_terms = int(_to_float_or_none(numeric_terms_by_type.get(mat_type)) or 0)
        parsed_ok = int(_to_float_or_none(parsed_ok_by_type.get(mat_type)) or 0)
        parsed_fail = int(_to_float_or_none(parsed_fail_by_type.get(mat_type)) or 0)
        min_chars = int(_to_float_or_none(min_chars_by_type.get(mat_type)) or 0)
        min_chunks = int(_to_float_or_none(min_chunks_by_type.get(mat_type)) or 0)
        min_numeric_terms = int(_to_float_or_none(min_numeric_terms_by_type.get(mat_type)) or 0)
        per_type_rows.append(
            {
                "material_type": mat_type,
                "material_type_label": _material_type_label(mat_type),
                "files": files,
                "parsed_ok_files": parsed_ok,
                "parsed_failed_files": parsed_fail,
                "parsed_chars": chars,
                "parsed_chunks": chunks,
                "numeric_terms": numeric_terms,
                "targets": {
                    "min_chars": min_chars,
                    "min_chunks": min_chunks,
                    "min_numeric_terms": min_numeric_terms,
                },
                "meets_chars": chars >= min_chars,
                "meets_chunks": chunks >= min_chunks,
                "meets_numeric_terms": numeric_terms >= min_numeric_terms,
            }
        )

    recommendations: List[str] = []
    for issue in readiness.get("issues") or []:
        text = str(issue).strip()
        if text and text not in recommendations:
            recommendations.append(text)
    for warning in readiness.get("warnings") or []:
        text = str(warning).strip()
        if text and text not in recommendations:
            recommendations.append(text)
    for detail in (quality.get("parsed_fail_details") or [])[:8]:
        if not isinstance(detail, dict):
            continue
        fn = str(detail.get("filename") or "").strip()
        rs = str(detail.get("reason") or "").strip()
        tip = f"解析失败文件：{fn or '未命名文件'}（{rs or '未知原因'}），建议重新上传可解析版本。"
        if tip not in recommendations:
            recommendations.append(tip)

    material_rows = [m for m in load_materials() if str(m.get("project_id")) == project_id]
    dwg_files = 0
    site_photo_files = 0
    pdf_files = 0
    for row in material_rows:
        filename = str(row.get("filename") or "").strip()
        ext = Path(filename).suffix.lower()
        material_type = _normalize_material_type(row.get("material_type"), filename=filename)
        if ext == ".dwg" or (material_type == "drawing" and ext == ".dwg"):
            dwg_files += 1
        if ext == ".pdf":
            pdf_files += 1
        if material_type == "site_photo" and ext in {
            ".png",
            ".jpg",
            ".jpeg",
            ".webp",
            ".bmp",
            ".tif",
            ".tiff",
        }:
            site_photo_files += 1
    ocr_available = bool(pytesseract is not None and Image is not None)
    dwg_converter_bins = _resolve_dwg_converter_binaries()
    dwg_converter_available = bool(dwg_converter_bins)
    pdf_backend = _pdf_backend_name()
    pdf_parser_available = pdf_backend != "none"
    capabilities = {
        "pdf_parser_available": pdf_parser_available,
        "pdf_backend": pdf_backend,
        "pdf_file_count": pdf_files,
        "ocr_available": ocr_available,
        "dwg_converter_available": dwg_converter_available,
        "dwg_converter_bins": [Path(p).name for p in dwg_converter_bins][:8],
        "dwg_file_count": dwg_files,
        "site_photo_file_count": site_photo_files,
    }
    if site_photo_files > 0 and not ocr_available:
        tip = "已上传现场照片，但当前环境未启用 OCR（pytesseract/PIL），建议安装后重评分以提升图片信息利用率。"
        if tip not in recommendations:
            recommendations.append(tip)
    if dwg_files > 0 and not dwg_converter_available:
        tip = "已上传 DWG 图纸，但未检测到 DWG 转换器（如 dwg2dxf/ODAFileConverter），建议安装后重评分以提升图纸解析深度。"
        if tip not in recommendations:
            recommendations.append(tip)
    if pdf_files > 0 and not pdf_parser_available:
        tip = "已上传 PDF 资料，但当前环境未启用 PDF 解析后端（PyMuPDF/pypdf），建议安装后重评分。"
        if tip not in recommendations:
            recommendations.append(tip)

    quality_summary = {
        "total_files": int(_to_float_or_none(quality.get("total_files")) or 0),
        "parsed_ok_files": int(_to_float_or_none(quality.get("parsed_ok_files")) or 0),
        "parsed_failed_files": int(_to_float_or_none(quality.get("parsed_failed_files")) or 0),
        "total_parsed_chars": int(_to_float_or_none(quality.get("total_parsed_chars")) or 0),
        "total_parsed_chunks": int(_to_float_or_none(quality.get("total_parsed_chunks")) or 0),
        "total_numeric_terms": int(_to_float_or_none(quality.get("total_numeric_terms")) or 0),
        "parse_fail_ratio": float(_to_float_or_none(quality.get("parse_fail_ratio")) or 0.0),
    }

    return {
        "project_id": project_id,
        "generated_at": _now_iso(),
        "ready_to_score": bool(readiness.get("ready")),
        "capabilities": capabilities,
        "gate": gate,
        "depth_gate": depth_gate,
        "quality_summary": quality_summary,
        "by_type": per_type_rows,
        "recommendations": recommendations[:20],
    }


def _render_material_depth_report_markdown(payload: Dict[str, object]) -> str:
    by_type = payload.get("by_type") if isinstance(payload.get("by_type"), list) else []
    quality_summary = (
        payload.get("quality_summary") if isinstance(payload.get("quality_summary"), dict) else {}
    )
    capabilities = (
        payload.get("capabilities") if isinstance(payload.get("capabilities"), dict) else {}
    )
    gate = payload.get("gate") if isinstance(payload.get("gate"), dict) else {}
    depth_gate = payload.get("depth_gate") if isinstance(payload.get("depth_gate"), dict) else {}
    recommendations = (
        payload.get("recommendations") if isinstance(payload.get("recommendations"), list) else []
    )
    lines = [
        "# 项目资料深读体检报告",
        "",
        f"- 项目ID：`{payload.get('project_id') or '-'}`",
        f"- 生成时间：`{payload.get('generated_at') or '-'}`",
        f"- 评分就绪：`{bool(payload.get('ready_to_score'))}`",
        "",
        "## 总览",
        "",
        f"- 资料总数：`{quality_summary.get('total_files', 0)}`",
        f"- 解析成功：`{quality_summary.get('parsed_ok_files', 0)}`",
        f"- 解析失败：`{quality_summary.get('parsed_failed_files', 0)}`",
        f"- 解析字数：`{quality_summary.get('total_parsed_chars', 0)}`",
        f"- 解析分块：`{quality_summary.get('total_parsed_chunks', 0)}`",
        f"- 数字约束项：`{quality_summary.get('total_numeric_terms', 0)}`",
        f"- 失败率：`{quality_summary.get('parse_fail_ratio', 0.0):.2%}`",
        "",
        "## 解析能力",
        "",
        f"- PDF 解析可用：`{bool(capabilities.get('pdf_parser_available'))}`（后端：`{capabilities.get('pdf_backend', '-')}`，PDF 文件数：`{capabilities.get('pdf_file_count', 0)}`）",
        f"- OCR 可用：`{bool(capabilities.get('ocr_available'))}`（现场照片文件数：`{capabilities.get('site_photo_file_count', 0)}`）",
        f"- DWG 转换器可用：`{bool(capabilities.get('dwg_converter_available'))}`（DWG 文件数：`{capabilities.get('dwg_file_count', 0)}`）",
        f"- 已检测转换器：`{', '.join(str(x) for x in (capabilities.get('dwg_converter_bins') or [])) or '-'}`",
        "",
        "## 门禁状态",
        "",
        f"- 基础资料门禁：`{bool(gate.get('passed', True))}`（enforce={bool(gate.get('enforce', False))}）",
        f"- 深读门禁：`{bool(depth_gate.get('passed', True))}`（enforce={bool(depth_gate.get('enforce', False))}）",
        "",
        "## 分类型体检",
        "",
        "| 资料类型 | 文件数 | 成功/失败 | 解析字数 | 分块数 | 数字约束项 | 目标(字数/分块/数字) |",
        "|---|---:|---:|---:|---:|---:|---|",
    ]
    for row in by_type:
        if not isinstance(row, dict):
            continue
        targets = row.get("targets") if isinstance(row.get("targets"), dict) else {}
        lines.append(
            f"| {row.get('material_type_label', row.get('material_type', '-'))} "
            f"| {row.get('files', 0)} "
            f"| {row.get('parsed_ok_files', 0)}/{row.get('parsed_failed_files', 0)} "
            f"| {row.get('parsed_chars', 0)} "
            f"| {row.get('parsed_chunks', 0)} "
            f"| {row.get('numeric_terms', 0)} "
            f"| {targets.get('min_chars', 0)}/{targets.get('min_chunks', 0)}/{targets.get('min_numeric_terms', 0)} |"
        )
    lines.extend(["", "## 建议动作", ""])
    if recommendations:
        for item in recommendations:
            text = str(item or "").strip()
            if text:
                lines.append(f"- {text}")
    else:
        lines.append("- 当前资料深读质量满足要求，无需额外整改。")
    return "\n".join(lines).strip()


def _build_material_knowledge_profile(project_id: str) -> Dict[str, object]:
    """
    构建项目资料知识画像：
    - 按资料类型聚合：提取词项、数字约束、覆盖维度
    - 按16维聚合：评估每个维度在资料中的证据覆盖强度
    """
    material_index = _build_project_material_index(project_id)
    rows = material_index.get("files") if isinstance(material_index.get("files"), list) else []
    rows = sorted(
        [r for r in rows if isinstance(r, dict)],
        key=lambda x: str(x.get("created_at") or ""),
    )

    by_type_term_counter: Dict[str, Counter[str]] = {}
    by_type_numeric_counter: Dict[str, Counter[str]] = {}
    by_type_dim_counter: Dict[str, Counter[str]] = {}
    by_type_file_count: Dict[str, int] = {}
    by_type_ok_files: Dict[str, int] = {}
    by_type_chars: Dict[str, int] = {}
    by_type_chunks: Dict[str, int] = {}

    by_dim_stats: Dict[str, Dict[str, object]] = {
        dim_id: {
            "keyword_hits": 0,
            "numeric_signal_hits": 0,
            "source_types": set(),
            "source_files": set(),
        }
        for dim_id in DIMENSION_IDS
    }

    parsed_ok_files = 0
    parsed_failed_files = 0
    total_chars = 0
    total_chunks = 0
    total_numeric_terms = 0
    parse_fail_details: List[Dict[str, str]] = []

    for row in rows:
        filename = str(row.get("filename") or "").strip()
        mat_type = _normalize_material_type(row.get("material_type"), filename=filename)
        by_type_file_count[mat_type] = int(by_type_file_count.get(mat_type, 0)) + 1
        by_type_term_counter.setdefault(mat_type, Counter())
        by_type_numeric_counter.setdefault(mat_type, Counter())
        by_type_dim_counter.setdefault(mat_type, Counter())
        by_type_ok_files.setdefault(mat_type, 0)
        by_type_chars.setdefault(mat_type, 0)
        by_type_chunks.setdefault(mat_type, 0)

        if not bool(row.get("parsed_ok")):
            parsed_failed_files += 1
            parse_fail_details.append(
                {
                    "filename": filename,
                    "material_type": mat_type,
                    "reason": str(row.get("parse_error") or "parse_failed"),
                }
            )
            continue
        text = str(row.get("text") or "").strip()
        if not text:
            parsed_failed_files += 1
            parse_fail_details.append(
                {"filename": filename, "material_type": mat_type, "reason": "empty_text"}
            )
            continue

        parsed_ok_files += 1
        by_type_ok_files[mat_type] = int(by_type_ok_files.get(mat_type, 0)) + 1

        chars = len(text)
        chunks = len(row.get("chunks") if isinstance(row.get("chunks"), list) else [])
        if chunks <= 0:
            chunks = len(_split_material_text_chunks(text, max_chars=900))
        total_chars += chars
        total_chunks += chunks
        by_type_chars[mat_type] = int(by_type_chars.get(mat_type, 0)) + chars
        by_type_chunks[mat_type] = int(by_type_chunks.get(mat_type, 0)) + chunks

        lexical_terms = (
            row.get("lexical_terms")
            if isinstance(row.get("lexical_terms"), list)
            else _extract_terms(text, max_terms=220)
        )
        by_type_term_counter[mat_type].update(lexical_terms)

        numeric_terms = (
            row.get("numeric_terms_norm") if isinstance(row.get("numeric_terms_norm"), list) else []
        )
        if not numeric_terms:
            numeric_terms = [
                token
                for token in (
                    _normalize_numeric_token(item)
                    for item in _extract_numeric_terms(text, max_terms=260)
                )
                if token
            ]
        total_numeric_terms += len(set(numeric_terms))
        by_type_numeric_counter[mat_type].update(numeric_terms)

        lower = text.lower()
        file_numeric_strength = min(8, len(set(numeric_terms)))
        file_name = filename
        for dim_id in DIMENSION_IDS:
            hit_count = 0
            for kw in DIMENSION_RAG_KEYWORDS.get(dim_id, [])[:10]:
                token = str(kw or "").strip().lower()
                if token:
                    hit_count += lower.count(token)
            if hit_count <= 0:
                continue
            dim_row = by_dim_stats[dim_id]
            dim_row["keyword_hits"] = int(dim_row.get("keyword_hits", 0)) + int(hit_count)
            dim_row["numeric_signal_hits"] = int(dim_row.get("numeric_signal_hits", 0)) + int(
                file_numeric_strength
            )
            (dim_row.get("source_types") or set()).add(mat_type)
            (dim_row.get("source_files") or set()).add(file_name)
            by_type_dim_counter[mat_type][dim_id] += int(hit_count)

    dwg_files = 0
    site_photo_files = 0
    pdf_files = 0
    for row in rows:
        filename = str(row.get("filename") or "").strip()
        ext = Path(filename).suffix.lower()
        mat_type = _normalize_material_type(row.get("material_type"), filename=filename)
        if ext == ".dwg" or (mat_type == "drawing" and ext == ".dwg"):
            dwg_files += 1
        if ext == ".pdf":
            pdf_files += 1
        if mat_type == "site_photo" and ext in {
            ".png",
            ".jpg",
            ".jpeg",
            ".webp",
            ".bmp",
            ".tif",
            ".tiff",
        }:
            site_photo_files += 1

    pdf_backend = _pdf_backend_name()
    capabilities = {
        "pdf_parser_available": pdf_backend != "none",
        "pdf_backend": pdf_backend,
        "pdf_file_count": pdf_files,
        "ocr_available": bool(pytesseract is not None and Image is not None),
        "dwg_converter_available": bool(_resolve_dwg_converter_binaries()),
        "dwg_file_count": dwg_files,
        "site_photo_file_count": site_photo_files,
    }

    all_types: List[str] = []
    for mat_type in ["tender_qa", "boq", "drawing", "site_photo"]:
        if int(by_type_file_count.get(mat_type, 0)) > 0:
            all_types.append(mat_type)
    for mat_type in sorted(by_type_file_count.keys()):
        if mat_type not in all_types:
            all_types.append(mat_type)

    by_type_rows: List[Dict[str, object]] = []
    for mat_type in all_types:
        dim_hits = by_type_dim_counter.get(mat_type, Counter())
        top_dims = [
            {
                "dimension_id": dim_id,
                "dimension_name": str((DIMENSIONS.get(dim_id) or {}).get("name") or dim_id),
                "keyword_hits": int(score),
            }
            for dim_id, score in dim_hits.most_common(4)
        ]
        by_type_rows.append(
            {
                "material_type": mat_type,
                "material_type_label": _material_type_label(mat_type),
                "files": int(by_type_file_count.get(mat_type, 0)),
                "parsed_ok_files": int(by_type_ok_files.get(mat_type, 0)),
                "parsed_chars": int(by_type_chars.get(mat_type, 0)),
                "parsed_chunks": int(by_type_chunks.get(mat_type, 0)),
                "unique_terms": int(len(by_type_term_counter.get(mat_type, Counter()))),
                "numeric_terms": int(
                    sum(by_type_numeric_counter.get(mat_type, Counter()).values())
                ),
                "top_terms": [
                    term
                    for term, _ in by_type_term_counter.get(mat_type, Counter()).most_common(10)
                ],
                "top_numeric_terms": [
                    term
                    for term, _ in by_type_numeric_counter.get(mat_type, Counter()).most_common(8)
                ],
                "top_dimensions": top_dims,
            }
        )

    by_dimension_rows: List[Dict[str, object]] = []
    low_dims: List[Dict[str, object]] = []
    for dim_id in DIMENSION_IDS:
        row = by_dim_stats.get(dim_id) or {}
        keyword_hits = int(row.get("keyword_hits", 0))
        numeric_signal_hits = int(row.get("numeric_signal_hits", 0))
        source_types = sorted(str(x) for x in (row.get("source_types") or set()) if str(x))
        source_files = sorted(str(x) for x in (row.get("source_files") or set()) if str(x))
        # 覆盖评分：关键词命中 + 跨类型覆盖共同决定（0..1）。
        coverage_score = min(1.0, (keyword_hits / 8.0) + (len(source_types) * 0.18))
        if numeric_signal_hits > 0:
            coverage_score = min(1.0, coverage_score + 0.06)
        coverage_level = (
            "high" if coverage_score >= 0.75 else ("medium" if coverage_score >= 0.35 else "low")
        )
        dim_row = {
            "dimension_id": dim_id,
            "dimension_name": str((DIMENSIONS.get(dim_id) or {}).get("name") or dim_id),
            "keyword_hits": keyword_hits,
            "numeric_signal_hits": numeric_signal_hits,
            "source_types": source_types,
            "source_file_count": len(source_files),
            "source_files_preview": source_files[:6],
            "coverage_score": round(float(coverage_score), 4),
            "coverage_level": coverage_level,
            "suggested_keywords": list((DIMENSION_RAG_KEYWORDS.get(dim_id) or [])[:4]),
        }
        by_dimension_rows.append(dim_row)
        if coverage_level == "low":
            low_dims.append(dim_row)

    covered_dimensions = sum(1 for item in by_dimension_rows if item.get("coverage_level") != "low")
    coverage_rate = (
        round(float(covered_dimensions) / float(len(DIMENSION_IDS)), 4) if DIMENSION_IDS else 0.0
    )

    recommendations: List[str] = []
    if not rows:
        recommendations.append("尚未上传任何项目资料，请先上传招答/清单/图纸/现场照片后再评分。")
    if parsed_failed_files > 0:
        recommendations.append(
            f"存在 {parsed_failed_files} 份资料未解析成功，建议优先修复（见失败详情）。"
        )
    if capabilities["site_photo_file_count"] > 0 and not capabilities["ocr_available"]:
        recommendations.append("已上传现场照片但 OCR 不可用，建议安装 OCR 组件后重评分。")
    if capabilities["dwg_file_count"] > 0 and not capabilities["dwg_converter_available"]:
        recommendations.append("已上传 DWG 图纸但转换器不可用，建议安装 DWG 转换器后重评分。")
    if capabilities["pdf_file_count"] > 0 and not capabilities["pdf_parser_available"]:
        recommendations.append(
            "已上传 PDF 资料但 PDF 解析后端不可用，建议安装 PyMuPDF 或 pypdf 后重评分。"
        )
    for dim_row in low_dims[:6]:
        dim_name = str(dim_row.get("dimension_name") or dim_row.get("dimension_id") or "")
        kw = "、".join(str(x) for x in (dim_row.get("suggested_keywords") or [])[:3])
        if kw:
            recommendations.append(f"维度[{dim_name}]证据薄弱，建议在资料/施组中补充：{kw}。")
        else:
            recommendations.append(
                f"维度[{dim_name}]证据薄弱，建议补充可量化的执行条款与验收口径。"
            )
    for item in parse_fail_details[:8]:
        fn = str(item.get("filename") or "").strip() or "未命名文件"
        rs = str(item.get("reason") or "").strip() or "未知原因"
        recommendations.append(f"解析失败：{fn}（{rs}）。")

    return {
        "project_id": project_id,
        "generated_at": _now_iso(),
        "capabilities": capabilities,
        "summary": {
            "total_files": len(rows),
            "parsed_ok_files": parsed_ok_files,
            "parsed_failed_files": parsed_failed_files,
            "total_parsed_chars": total_chars,
            "total_parsed_chunks": total_chunks,
            "total_numeric_terms": total_numeric_terms,
            "covered_dimensions": covered_dimensions,
            "dimension_coverage_rate": coverage_rate,
            "low_coverage_dimensions": len(low_dims),
        },
        "by_type": by_type_rows,
        "by_dimension": by_dimension_rows,
        "recommendations": recommendations[:24],
    }


def _render_material_knowledge_profile_markdown(payload: Dict[str, object]) -> str:
    summary = payload.get("summary") if isinstance(payload.get("summary"), dict) else {}
    by_type = payload.get("by_type") if isinstance(payload.get("by_type"), list) else []
    by_dimension = (
        payload.get("by_dimension") if isinstance(payload.get("by_dimension"), list) else []
    )
    recommendations = (
        payload.get("recommendations") if isinstance(payload.get("recommendations"), list) else []
    )
    capabilities = (
        payload.get("capabilities") if isinstance(payload.get("capabilities"), dict) else {}
    )
    lines = [
        "# 项目资料知识画像报告",
        "",
        f"- 项目ID：`{payload.get('project_id') or '-'}`",
        f"- 生成时间：`{payload.get('generated_at') or '-'}`",
        "",
        "## 总览",
        "",
        f"- 资料总数：`{summary.get('total_files', 0)}`",
        f"- 解析成功/失败：`{summary.get('parsed_ok_files', 0)}/{summary.get('parsed_failed_files', 0)}`",
        f"- 解析字数：`{summary.get('total_parsed_chars', 0)}`",
        f"- 分块数：`{summary.get('total_parsed_chunks', 0)}`",
        f"- 数字约束项：`{summary.get('total_numeric_terms', 0)}`",
        f"- 维度覆盖率：`{summary.get('dimension_coverage_rate', 0.0):.2%}`",
        "",
        "## 解析能力",
        "",
        f"- PDF 解析可用：`{bool(capabilities.get('pdf_parser_available'))}`（后端：`{capabilities.get('pdf_backend', '-')}`）",
        f"- OCR 可用：`{bool(capabilities.get('ocr_available'))}`",
        f"- DWG 转换器可用：`{bool(capabilities.get('dwg_converter_available'))}`",
        f"- PDF 文件数：`{capabilities.get('pdf_file_count', 0)}`",
        f"- 现场照片文件数：`{capabilities.get('site_photo_file_count', 0)}`",
        f"- DWG 文件数：`{capabilities.get('dwg_file_count', 0)}`",
        "",
        "## 按资料类型",
        "",
        "| 资料类型 | 文件数 | 解析字数 | 分块数 | 词项数 | 数字约束项 |",
        "|---|---:|---:|---:|---:|---:|",
    ]
    for row in by_type:
        if not isinstance(row, dict):
            continue
        lines.append(
            f"| {row.get('material_type_label', row.get('material_type', '-'))} "
            f"| {row.get('files', 0)} "
            f"| {row.get('parsed_chars', 0)} "
            f"| {row.get('parsed_chunks', 0)} "
            f"| {row.get('unique_terms', 0)} "
            f"| {row.get('numeric_terms', 0)} |"
        )

    lines.extend(
        [
            "",
            "## 按评分维度",
            "",
            "| 维度 | 关键词命中 | 来源类型数 | 覆盖评分 | 等级 |",
            "|---|---:|---:|---:|---|",
        ]
    )
    for row in by_dimension:
        if not isinstance(row, dict):
            continue
        lines.append(
            f"| {row.get('dimension_id', '-')} {row.get('dimension_name', '')} "
            f"| {row.get('keyword_hits', 0)} "
            f"| {len(row.get('source_types') or [])} "
            f"| {row.get('coverage_score', 0)} "
            f"| {row.get('coverage_level', '-')} |"
        )

    lines.extend(["", "## 建议动作", ""])
    if recommendations:
        for item in recommendations:
            text = str(item or "").strip()
            if text:
                lines.append(f"- {text}")
    else:
        lines.append("- 当前知识覆盖正常，无需额外动作。")
    return "\n".join(lines).strip()


def _compute_multipliers_hash(multipliers: dict) -> str:
    """计算 multipliers 的 hash，用于缓存 key"""
    import hashlib
    import json

    content = json.dumps(multipliers, sort_keys=True)
    return hashlib.sha256(content.encode("utf-8")).hexdigest()[:16]


def _get_effective_dimension_multipliers(project_id: str) -> Dict[str, float]:
    multipliers, _, _ = _resolve_project_scoring_context(project_id)
    return multipliers


def _get_evolution_total_score_scale(project_id: str) -> float | None:
    """进化报告中的总分缩放因子，用于使本系统总分贴近青天平均分。"""
    reports = load_evolution_reports()
    evo = reports.get(project_id) or {}
    se = evo.get("scoring_evolution") or {}
    scale = se.get("total_score_scale")
    if scale is not None and isinstance(scale, (int, float)):
        return float(scale)
    return None


def _apply_evolution_total_scale(project_id: str, report: Dict[str, object]) -> None:
    """若进化报告有 total_score_scale，对总分进行缩放（原地修改 report）。"""
    scale = _get_evolution_total_score_scale(project_id)
    if scale is None or abs(scale - 1.0) < 1e-6:
        return
    for key in ("total_score", "rule_total_score", "pred_total_score", "llm_total_score"):
        v = report.get(key)
        if v is not None:
            try:
                report[key] = round(min(100.0, max(0.0, float(v) * scale)), 2)
            except (TypeError, ValueError):
                pass
    # 若存在 pred_total_score，则保持 total_score 与展示/排序主分一致。
    pred_total = _to_float_or_none(report.get("pred_total_score"))
    rule_total = _to_float_or_none(report.get("rule_total_score"))
    if pred_total is not None:
        report["total_score"] = pred_total
    elif rule_total is not None:
        report["total_score"] = rule_total


@router.post(
    "/projects/{project_id}/shigong",
    response_model=SubmissionRecord,
    tags=["施组提交"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def upload_shigong(
    project_id: str,
    file: UploadFile = File(...),
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> SubmissionRecord:
    """
    上传施工组织设计文档（仅上传，不自动评分）。

    上传 TXT / DOCX / PDF / JSON / XLSX / DXF 格式施组文档并保存解析文本。
    上传后提交会进入“待评分”状态，需手动调用“评分施组”才会产生分数。

    **需要 API Key 认证**

    支持 Accept-Language header 进行多语言响应。
    """
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    raw_filename = file.filename or ""
    normalized_filename = _normalize_uploaded_filename(raw_filename)
    if not normalized_filename:
        raise HTTPException(status_code=422, detail="施组文件名为空，请重试或重命名后上传。")
    content = file.file.read()
    try:
        text = _read_uploaded_file_content(content, normalized_filename)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    submissions = load_submissions()
    now_utc = datetime.now(timezone.utc)
    duplicate = _find_recent_duplicate_submission(
        submissions,
        project_id=project_id,
        filename=normalized_filename,
        text=text,
        now_utc=now_utc,
    )
    if duplicate is not None:
        return SubmissionRecord(**duplicate)

    _, profile_snapshot, project = _resolve_project_scoring_context(project_id)
    scoring_engine_version = str(project.get("scoring_engine_version_locked") or "v1")
    submission_id = str(uuid4())

    report = _build_pending_submission_report(
        project=project,
        scoring_engine_version=scoring_engine_version,
    )
    if profile_snapshot:
        report_meta = report.get("meta")
        report_meta = report_meta if isinstance(report_meta, dict) else {}
        report_meta["expert_profile_snapshot"] = profile_snapshot
        report_meta["expert_profile_id"] = profile_snapshot.get("id")
        report["meta"] = report_meta

    record = {
        "id": submission_id,
        "project_id": project_id,
        "filename": normalized_filename,
        "total_score": 0.0,
        "report": report,
        "text": text,
        "created_at": now_utc.isoformat(),
        "updated_at": now_utc.isoformat(),
        "expert_profile_id_used": profile_snapshot.get("id") if profile_snapshot else None,
    }
    submissions.append(record)
    save_submissions(submissions)

    return SubmissionRecord(**record)


@router.post(
    "/projects/{project_id}/score",
    response_model=SubmissionRecord,
    tags=["施组提交"],
    responses={**RESPONSES_401, **RESPONSES_404, **RESPONSES_422},
)
def score_text_for_project(
    project_id: str,
    payload: ScoreRequest,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> SubmissionRecord:
    """
    对文本内容进行项目级评分。

    与直接上传文件不同，此端点接收内联文本进行评分。
    适用于已解析的文本内容或 API 集成场景。
    支持评分结果缓存，相同文本和配置重复评分时直接返回缓存结果。

    **需要 API Key 认证**

    支持 Accept-Language header 进行多语言响应。
    """
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    config = load_config()
    multipliers, profile_snapshot, project = _resolve_project_scoring_context(project_id)
    submission_id = str(uuid4())
    scoring_engine_version = str(project.get("scoring_engine_version_locked") or "v1")
    engine_version = _determine_engine_version(project, scoring_engine_version)

    if engine_version == "v1":
        config_hash = _compute_multipliers_hash(multipliers) if multipliers else None
        cached_result = get_cached_score(payload.text, config_hash)
        if cached_result is not None:
            report = dict(cached_result)
        else:
            raw_report, _ = _score_submission_for_project(
                submission_id=submission_id,
                text=payload.text,
                project_id=project_id,
                project=project,
                config=config,
                multipliers=multipliers,
                profile_snapshot=profile_snapshot,
                scoring_engine_version=scoring_engine_version,
            )
            # 缓存仅存“未缩放原始分”，避免后续读取时重复应用 total_score_scale。
            cache_score_result(payload.text, raw_report, config_hash)
            report = dict(raw_report)
        _apply_evolution_total_scale(project_id, report)
        evidence_units: List[Dict[str, object]] = []
    else:
        report, evidence_units = _score_submission_for_project(
            submission_id=submission_id,
            text=payload.text,
            project_id=project_id,
            project=project,
            config=config,
            multipliers=multipliers,
            profile_snapshot=profile_snapshot,
            scoring_engine_version=scoring_engine_version,
        )
        _apply_evolution_total_scale(project_id, report)

    record = {
        "id": submission_id,
        "project_id": project_id,
        "filename": "inline",
        "total_score": float(report.get("total_score", report.get("rule_total_score", 0.0))),
        "report": report,
        "text": payload.text,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expert_profile_id_used": profile_snapshot.get("id") if profile_snapshot else None,
    }
    submissions = load_submissions()
    submissions.append(record)
    save_submissions(submissions)

    snapshots = load_score_reports()
    snapshots.append(
        _build_score_report_snapshot(
            submission_id=submission_id,
            project=project,
            report=report,
            profile_snapshot=profile_snapshot,
            scoring_engine_version=scoring_engine_version,
        )
    )
    save_score_reports(snapshots)
    if evidence_units:
        all_units = load_evidence_units()
        all_units = _replace_submission_evidence_units(
            all_units,
            submission_id=submission_id,
            new_units=evidence_units,
        )
        save_evidence_units(all_units)

    # 记录评分历史
    dimension_scores = {
        dim_id: dim.get("score", 0.0) for dim_id, dim in report.get("dimension_scores", {}).items()
    }
    penalty_count = len(report.get("penalties", []))
    record_history_score(
        project_id=project_id,
        submission_id=submission_id,
        filename="inline",
        total_score=float(report.get("total_score", report.get("rule_total_score", 0.0))),
        dimension_scores=dimension_scores,
        penalty_count=penalty_count,
    )

    return SubmissionRecord(**record)


@router.get(
    "/projects/{project_id}/submissions",
    response_model=list[SubmissionRecord] | ProjectPreScoreListResponse,
    tags=["施组提交"],
)
def list_submissions(
    project_id: str,
    with_: Optional[str] = Query(None, alias="with"),
):
    """
    获取项目的所有施组提交记录。

    返回指定项目下的所有历史评分记录，包括评分报告详情。
    """
    ensure_data_dirs()
    projects = load_projects()
    project = next((p for p in projects if str(p.get("id")) == project_id), {"id": project_id})
    allow_pred_score = _select_calibrator_model(project) is not None
    score_scale_max = _resolve_project_score_scale_max(project)
    score_scale_max = _resolve_project_score_scale_max(project)

    submissions = [s for s in load_submissions() if s["project_id"] == project_id]

    def _view_submission(item: Dict[str, object]) -> Dict[str, object]:
        view = dict(item)
        report_obj = item.get("report")
        if not isinstance(report_obj, dict):
            total_display = _convert_score_from_100(item.get("total_score"), score_scale_max)
            if total_display is not None:
                view["total_score"] = total_display
            return view
        report = dict(report_obj)
        rule_total = _to_float_or_none(report.get("rule_total_score"))
        if rule_total is None:
            rule_total = _to_float_or_none(report.get("total_score"))
        if rule_total is None:
            rule_total = _to_float_or_none(item.get("total_score"))
        if rule_total is None:
            rule_total = 0.0
        if not allow_pred_score:
            report["pred_total_score"] = None
            report["llm_total_score"] = None
            report["pred_confidence"] = None
            report["score_blend"] = None
            report["total_score"] = round(float(rule_total), 2)
            view["total_score"] = round(float(rule_total), 2)
        raw_total = _to_float_or_none(report.get("total_score"))
        raw_rule = _to_float_or_none(report.get("rule_total_score"))
        raw_pred = _to_float_or_none(report.get("pred_total_score"))
        raw_llm = _to_float_or_none(report.get("llm_total_score"))
        report["raw_total_score_100"] = raw_total
        report["raw_rule_total_score_100"] = raw_rule
        report["raw_pred_total_score_100"] = raw_pred
        report["raw_llm_total_score_100"] = raw_llm
        report["score_scale_max"] = score_scale_max
        report["score_scale_label"] = _score_scale_label(score_scale_max)
        display_pred = _convert_score_from_100(raw_pred, score_scale_max)
        display_rule = _convert_score_from_100(raw_rule, score_scale_max)
        display_llm = _convert_score_from_100(raw_llm, score_scale_max)
        display_total = _convert_score_from_100(raw_total, score_scale_max)
        if display_total is None:
            display_total = _convert_score_from_100(item.get("total_score"), score_scale_max)
        report["pred_total_score"] = display_pred
        report["rule_total_score"] = display_rule
        report["llm_total_score"] = display_llm
        report["total_score"] = display_total
        if display_total is not None:
            view["total_score"] = display_total
        view["report"] = report
        return view

    submissions_view = [_view_submission(s) for s in submissions]
    if with_ != "latest_report":
        return [SubmissionRecord(**s) for s in submissions_view]

    latest_reports = _latest_records_by_submission(
        [r for r in load_score_reports() if str(r.get("project_id")) == project_id]
    )

    rows: List[Dict[str, object]] = []
    for s in submissions_view:
        sid = str(s.get("id"))
        latest = latest_reports.get(sid, {})
        suggestions = latest.get("suggestions") or []
        top_gain = 0.0
        if suggestions and isinstance(suggestions[0], dict):
            top_gain = float(suggestions[0].get("expected_gain", 0.0))
        pred_total_raw = latest.get("pred_total_score")
        if not allow_pred_score:
            pred_total_raw = None
        rule_total_raw = float(latest.get("rule_total_score", s.get("total_score", 0.0)))
        pred_total = _convert_score_from_100(pred_total_raw, score_scale_max)
        rule_total = _convert_score_from_100(rule_total_raw, score_scale_max)
        top_gain_display = _convert_score_from_100(top_gain, score_scale_max)
        rows.append(
            {
                "submission_id": sid,
                "bidder_name": s.get("bidder_name") or s.get("filename") or sid,
                "latest_report": {
                    "report_id": latest.get("id"),
                    "rule_total_score": float(rule_total if rule_total is not None else 0.0),
                    "pred_total_score": pred_total,
                    "rank_by_pred": None,
                    "rank_by_rule": None,
                    "top_expected_gain": round(
                        float(top_gain_display if top_gain_display is not None else 0.0), 2
                    ),
                    "updated_at": latest.get("created_at") or s.get("created_at"),
                },
            }
        )

    pred_sorted = sorted(
        rows,
        key=lambda x: (
            -float(x["latest_report"]["pred_total_score"])
            if x["latest_report"]["pred_total_score"] is not None
            else float("inf")
        ),
    )
    for idx, row in enumerate(pred_sorted, start=1):
        if row["latest_report"]["pred_total_score"] is not None:
            row["latest_report"]["rank_by_pred"] = idx

    rule_sorted = sorted(rows, key=lambda x: -float(x["latest_report"]["rule_total_score"]))
    for idx, row in enumerate(rule_sorted, start=1):
        row["latest_report"]["rank_by_rule"] = idx

    return ProjectPreScoreListResponse(
        project_id=project_id,
        expert_profile_id=project.get("expert_profile_id"),
        submissions=rows,
    )


def _delete_submission_record(project_id: str, submission_id: str, locale: str) -> None:
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    submissions = load_submissions()
    found = next(
        (
            s
            for s in submissions
            if s.get("id") == submission_id and s.get("project_id") == project_id
        ),
        None,
    )
    if not found:
        raise HTTPException(status_code=404, detail="施组提交记录不存在")
    raw_path = str(found.get("path") or "").strip()
    if raw_path:
        p = Path(raw_path)
        if p.exists():
            p.unlink()
    submissions = [
        s
        for s in submissions
        if not (s.get("id") == submission_id and s.get("project_id") == project_id)
    ]
    save_submissions(submissions)
    snapshots = load_score_reports()
    snapshots = [
        r
        for r in snapshots
        if not (r.get("submission_id") == submission_id and r.get("project_id") == project_id)
    ]
    save_score_reports(snapshots)
    evidence_units = load_evidence_units()
    evidence_units = [u for u in evidence_units if str(u.get("submission_id")) != submission_id]
    save_evidence_units(evidence_units)
    qingtian_results = load_qingtian_results()
    qingtian_results = [q for q in qingtian_results if str(q.get("submission_id")) != submission_id]
    save_qingtian_results(qingtian_results)
    delta_cases = load_delta_cases()
    delta_cases = [d for d in delta_cases if str(d.get("submission_id")) != submission_id]
    save_delta_cases(delta_cases)
    calibration_samples = load_calibration_samples()
    calibration_samples = [
        s for s in calibration_samples if str(s.get("submission_id")) != submission_id
    ]
    save_calibration_samples(calibration_samples)
    # 删除属于显式反馈信号：自动刷新样本并触发校准/调权重闭环。
    _run_feedback_closed_loop_safe(project_id, locale=locale, trigger="delete_submission")


@router.delete(
    "/projects/{project_id}/submissions/{submission_id}",
    tags=["施组提交"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def delete_submission(
    project_id: str,
    submission_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> dict:
    """删除指定项目下的一条施组提交记录。"""
    _delete_submission_record(project_id=project_id, submission_id=submission_id, locale=locale)
    return {"ok": True, "id": submission_id}


@router.delete(
    "/projects/{project_id}/shigong/{file_id}",
    tags=["施组提交"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def delete_shigong_file(
    project_id: str,
    file_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> dict:
    """删除指定项目下的一条施组记录（shigong 别名路径）。"""
    _delete_submission_record(project_id=project_id, submission_id=file_id, locale=locale)
    return {"ok": True, "id": file_id}


@router.get(
    "/submissions/{submission_id}/reports/latest",
    response_model=LatestReportResponse,
    tags=["施组提交"],
    responses={**RESPONSES_404},
)
def get_latest_submission_report(submission_id: str) -> LatestReportResponse:
    """获取某个提交的最新评分报告（含UI摘要）。"""
    ensure_data_dirs()
    reports = [r for r in load_score_reports() if str(r.get("submission_id")) == submission_id]
    reports.sort(key=lambda x: str(x.get("created_at", "")), reverse=True)

    report_obj: Dict[str, object]
    if reports:
        latest = reports[0]
        report_obj = {
            "id": latest.get("id"),
            "submission_id": latest.get("submission_id"),
            "scoring_engine_version": latest.get("scoring_engine_version"),
            "rule_total_score": latest.get("rule_total_score"),
            "pred_total_score": latest.get("pred_total_score"),
            "llm_total_score": latest.get("llm_total_score"),
            "pred_confidence": latest.get("pred_confidence"),
            "score_blend": latest.get("score_blend"),
            "rule_dim_scores": latest.get("rule_dim_scores", {}),
            "pred_dim_scores": latest.get("pred_dim_scores"),
            "penalties": latest.get("penalties", []),
            "lint_findings": latest.get("lint_findings", []),
            "suggestions": latest.get("suggestions", []),
            "expert_profile_snapshot": latest.get("expert_profile_snapshot", {}),
            "created_at": latest.get("created_at"),
        }
    else:
        submissions = load_submissions()
        try:
            submission = _find_submission(submission_id, submissions)
        except HTTPException:
            raise HTTPException(status_code=404, detail="评分报告不存在")
        report_obj = dict(submission.get("report") or {})
        if not report_obj:
            raise HTTPException(status_code=404, detail="评分报告不存在")
        report_obj.setdefault("submission_id", submission_id)
        report_obj.setdefault("rule_total_score", report_obj.get("total_score", 0.0))
        report_obj.setdefault("pred_total_score", report_obj.get("pred_total_score"))
        report_obj.setdefault("llm_total_score", report_obj.get("llm_total_score"))
        report_obj.setdefault("pred_confidence", report_obj.get("pred_confidence"))
        report_obj.setdefault("score_blend", report_obj.get("score_blend"))
        report_obj.setdefault("rule_dim_scores", report_obj.get("rule_dim_scores", {}))
        report_obj.setdefault("penalties", report_obj.get("penalties", []))
        report_obj.setdefault("lint_findings", report_obj.get("lint_findings", []))
        report_obj.setdefault("suggestions", report_obj.get("suggestions", []))

    penalties = report_obj.get("penalties") or []
    lint_findings = report_obj.get("lint_findings") or []
    suggestions = report_obj.get("suggestions") or []

    top_conflicts = [p for p in penalties if str(p.get("code") or "") == "P-CONSIST-001"][:10]
    top_missing_requirements = [
        f for f in lint_findings if str(f.get("issue_code") or "") == "MissingRequirement"
    ][:10]

    ui_summary = {
        "pred_total_score": report_obj.get("pred_total_score"),
        "llm_total_score": report_obj.get("llm_total_score"),
        "pred_confidence": report_obj.get("pred_confidence"),
        "score_blend": report_obj.get("score_blend"),
        "rule_total_score": report_obj.get("rule_total_score", report_obj.get("total_score")),
        "top10_suggestions": suggestions[:10],
        "top_conflicts": top_conflicts,
        "top_missing_requirements": top_missing_requirements,
    }
    return LatestReportResponse(report=report_obj, ui_summary=ui_summary)


@router.get(
    "/projects/{project_id}/submissions/{submission_id}/evidence_trace",
    response_model=EvidenceTraceResponse,
    tags=["施组提交"],
    responses={**RESPONSES_404},
)
def get_submission_evidence_trace(
    project_id: str,
    submission_id: str,
    locale: str = Depends(get_locale),
) -> EvidenceTraceResponse:
    """获取单份施组的证据追溯报告（结构化 JSON）。"""
    ensure_data_dirs()
    projects = load_projects()
    _find_project(project_id, projects)
    submissions = load_submissions()
    submission = next(
        (
            s
            for s in submissions
            if str(s.get("id") or "") == submission_id
            and str(s.get("project_id") or "") == project_id
        ),
        None,
    )
    if submission is None:
        raise HTTPException(status_code=404, detail=t("api.no_submissions", locale=locale))
    payload = _build_submission_evidence_trace_report(project_id=project_id, submission=submission)
    return EvidenceTraceResponse(**payload)


@router.get(
    "/projects/{project_id}/submissions/{submission_id}/evidence_trace/markdown",
    response_model=EvidenceTraceMarkdownResponse,
    tags=["施组提交"],
    responses={**RESPONSES_404},
)
def get_submission_evidence_trace_markdown(
    project_id: str,
    submission_id: str,
    locale: str = Depends(get_locale),
) -> EvidenceTraceMarkdownResponse:
    """获取单份施组证据追溯的 Markdown 文本。"""
    payload = get_submission_evidence_trace(
        project_id=project_id,
        submission_id=submission_id,
        locale=locale,
    ).model_dump()
    markdown = _render_evidence_trace_markdown(payload)
    return EvidenceTraceMarkdownResponse(
        project_id=project_id,
        submission_id=submission_id,
        markdown=markdown,
        generated_at=str(payload.get("generated_at") or _now_iso()),
    )


@router.get(
    "/projects/{project_id}/submissions/{submission_id}/evidence_trace.md",
    tags=["施组提交"],
    responses={**RESPONSES_404},
)
def download_submission_evidence_trace_markdown(
    project_id: str,
    submission_id: str,
    locale: str = Depends(get_locale),
) -> Response:
    """下载单份施组证据追溯 Markdown 文件。"""
    payload = get_submission_evidence_trace_markdown(
        project_id=project_id,
        submission_id=submission_id,
        locale=locale,
    )
    filename = f"evidence_trace_{project_id}_{submission_id}.md"
    return Response(
        content=payload.markdown.encode("utf-8"),
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/projects/{project_id}/evidence_trace/latest",
    response_model=EvidenceTraceResponse,
    tags=["施组提交"],
    responses={**RESPONSES_404},
)
def get_latest_submission_evidence_trace(
    project_id: str,
    locale: str = Depends(get_locale),
) -> EvidenceTraceResponse:
    """获取本项目最新施组（优先已评分）的证据追溯报告。"""
    ensure_data_dirs()
    projects = load_projects()
    _find_project(project_id, projects)
    submissions = load_submissions()
    latest = _latest_project_submission(project_id, submissions, prefer_scored=True)
    if latest is None:
        raise HTTPException(status_code=404, detail=t("api.no_submissions", locale=locale))
    submission_id = str(latest.get("id") or "")
    return get_submission_evidence_trace(
        project_id=project_id,
        submission_id=submission_id,
        locale=locale,
    )


@router.get(
    "/projects/{project_id}/submissions/{submission_id}/scoring_basis",
    response_model=ScoringBasisResponse,
    tags=["施组提交"],
    responses={**RESPONSES_404},
)
def get_submission_scoring_basis(
    project_id: str,
    submission_id: str,
    locale: str = Depends(get_locale),
) -> ScoringBasisResponse:
    """获取单份施组评分依据审计（结构化 JSON）。"""
    ensure_data_dirs()
    projects = load_projects()
    _find_project(project_id, projects)
    submissions = load_submissions()
    submission = next(
        (
            s
            for s in submissions
            if str(s.get("id") or "") == submission_id
            and str(s.get("project_id") or "") == project_id
        ),
        None,
    )
    if submission is None:
        raise HTTPException(status_code=404, detail=t("api.no_submissions", locale=locale))
    payload = _build_submission_scoring_basis_report(project_id=project_id, submission=submission)
    return ScoringBasisResponse(**payload)


@router.get(
    "/projects/{project_id}/scoring_basis/latest",
    response_model=ScoringBasisResponse,
    tags=["施组提交"],
    responses={**RESPONSES_404},
)
def get_latest_submission_scoring_basis(
    project_id: str,
    locale: str = Depends(get_locale),
) -> ScoringBasisResponse:
    """获取本项目最新施组（优先已评分）的评分依据审计。"""
    ensure_data_dirs()
    projects = load_projects()
    _find_project(project_id, projects)
    submissions = load_submissions()
    latest = _latest_project_submission(project_id, submissions, prefer_scored=True)
    if latest is None:
        raise HTTPException(status_code=404, detail=t("api.no_submissions", locale=locale))
    submission_id = str(latest.get("id") or "")
    return get_submission_scoring_basis(
        project_id=project_id,
        submission_id=submission_id,
        locale=locale,
    )


@router.post(
    "/submissions/{submission_id}/qingtian-results",
    response_model=QingTianResultRecord,
    tags=["施组提交"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def ingest_qingtian_result(
    submission_id: str,
    payload: QingTianResultCreate,
    api_key: Optional[str] = Depends(verify_api_key),
) -> QingTianResultRecord:
    """写入青天真实评标结果。"""
    ensure_data_dirs()
    submissions = load_submissions()
    submission = _find_submission(submission_id, submissions)
    project_id = str(submission.get("project_id") or "")
    projects = load_projects()
    project = _find_project(project_id, projects)

    model_version = str(
        payload.qingtian_model_version
        or project.get("qingtian_model_version")
        or DEFAULT_QINGTIAN_MODEL_VERSION
    )
    record = {
        "id": str(uuid4()),
        "submission_id": submission_id,
        "qingtian_model_version": model_version,
        "qt_total_score": float(payload.qt_total_score),
        "qt_dim_scores": payload.qt_dim_scores,
        "qt_reasons": payload.qt_reasons,
        "raw_payload": payload.raw_payload,
        "created_at": _now_iso(),
    }
    results = load_qingtian_results()
    results.append(record)
    save_qingtian_results(results)

    if str(project.get("status") or "") == "scoring_preparation":
        project["status"] = "submitted_to_qingtian"
        project["updated_at"] = _now_iso()
        save_projects(projects)

    return QingTianResultRecord(**record)


@router.get(
    "/submissions/{submission_id}/qingtian-results/latest",
    response_model=QingTianResultRecord,
    tags=["施组提交"],
    responses={**RESPONSES_404},
)
def get_latest_qingtian_result(submission_id: str) -> QingTianResultRecord:
    """获取某个提交最新的青天真实评标结果。"""
    ensure_data_dirs()
    results = [r for r in load_qingtian_results() if str(r.get("submission_id")) == submission_id]
    if not results:
        raise HTTPException(status_code=404, detail="暂无青天评标结果")
    latest = sorted(results, key=lambda x: str(x.get("created_at", "")), reverse=True)[0]
    return QingTianResultRecord(**latest)


@router.post(
    "/calibration/train",
    response_model=CalibratorModelRecord,
    tags=["洞察与学习"],
    responses={**RESPONSES_401, **RESPONSES_422},
)
def train_calibrator(
    payload: CalibratorTrainRequest,
    api_key: Optional[str] = Depends(verify_api_key),
) -> CalibratorModelRecord:
    """训练校准器（支持 auto / ridge / offset / linear1d / isotonic1d）。"""
    ensure_data_dirs()
    model_type = str(payload.model_type or "ridge").lower().strip()
    if model_type not in {"auto", "ridge", "offset", "linear1d", "isotonic1d"}:
        raise HTTPException(
            status_code=422, detail="model_type 仅支持 auto/ridge/offset/linear1d/isotonic1d"
        )

    # 优先使用已落库的 FEATURE_ROW 样本
    stored_samples = load_calibration_samples()
    if payload.project_id:
        stored_samples = [
            s for s in stored_samples if str(s.get("project_id")) == payload.project_id
        ]

    feature_rows: List[Dict[str, object]] = []
    for sample in stored_samples:
        feature_rows.append(
            {
                "feature_schema_version": sample.get("feature_schema_version", "v2"),
                "x_features": sample.get("x_features") or {},
                "y_label": sample.get("y_label"),
                "submission_id": sample.get("submission_id"),
            }
        )

    # 若样本不足，在线拼接一次并反写样本表
    if len(feature_rows) < 3:
        submissions = load_submissions()
        if payload.project_id:
            submissions = [s for s in submissions if str(s.get("project_id")) == payload.project_id]
        submission_map = {str(s.get("id")): s for s in submissions}

        reports = load_score_reports()
        if payload.project_id:
            reports = [r for r in reports if str(r.get("project_id")) == payload.project_id]
        latest_reports = _latest_records_by_submission(reports)

        qt_results = load_qingtian_results()
        latest_qt = _latest_records_by_submission(qt_results)

        rebuilt_samples = build_calibration_samples(
            project_id=str(payload.project_id or "__all__"),
            latest_reports_by_submission=latest_reports,
            latest_qingtian_by_submission=latest_qt,
            submissions_by_id=submission_map,
        )
        if rebuilt_samples:
            saved = load_calibration_samples()
            for row in rebuilt_samples:
                sid = str(row.get("submission_id"))
                saved = [x for x in saved if str(x.get("submission_id")) != sid]
                saved.append(row)
            save_calibration_samples(saved)
            feature_rows = [
                {
                    "feature_schema_version": s.get("feature_schema_version", "v2"),
                    "x_features": s.get("x_features") or {},
                    "y_label": s.get("y_label"),
                    "submission_id": s.get("submission_id"),
                }
                for s in rebuilt_samples
            ]

    try:
        if model_type == "auto":
            model_artifact = train_best_calibrator_auto(feature_rows, alpha=float(payload.alpha))
        elif model_type == "offset":
            model_artifact = train_offset_calibrator(feature_rows)
        elif model_type == "linear1d":
            model_artifact = train_linear1d_calibrator(feature_rows, alpha=float(payload.alpha))
        elif model_type == "isotonic1d":
            model_artifact = train_isotonic1d_calibrator(feature_rows)
        else:
            model_artifact = train_ridge_calibrator(feature_rows, alpha=float(payload.alpha))
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    selected_type = str(model_artifact.get("model_type") or model_type or "ridge")
    # 为审计保留 auto 来源
    version_prefix = f"calib_{'auto_' if model_type == 'auto' else ''}{selected_type}"

    # 统一用 CV 口径做上线闸门（避免 in-sample 过拟合）
    cv = cross_validate_calibrator(
        model_type=selected_type,
        feature_rows=feature_rows,
        alpha=float(payload.alpha),
        seed=42,
    )
    # baseline: raw rule_total_score
    y_true = [float(r.get("y_label")) for r in feature_rows if r.get("y_label") is not None]
    baseline_pred = [
        max(0.0, min(100.0, float(((r.get("x_features") or {}).get("rule_total_score") or 0.0))))
        for r in feature_rows
        if r.get("y_label") is not None
    ]
    baseline_metrics = calc_metrics(y_true, baseline_pred)
    cv_metrics = (
        (cv.get("metrics") or {})
        if bool(cv.get("ok"))
        else {"mae": 0.0, "rmse": 0.0, "spearman": 0.0}
    )
    improve_threshold = max(0.2, float(baseline_metrics.get("mae") or 0.0) * 0.01)
    spearman_tolerance = 0.02
    gate_passed = (
        bool(cv.get("ok"))
        and float(cv_metrics.get("mae") or 0.0)
        <= float(baseline_metrics.get("mae") or 0.0) - improve_threshold
        and float(cv_metrics.get("spearman") or 0.0)
        >= float(baseline_metrics.get("spearman") or 0.0) - spearman_tolerance
    )
    model_artifact.setdefault("metrics", {})
    model_artifact["metrics"]["cv_mae"] = cv_metrics.get("mae")
    model_artifact["metrics"]["cv_rmse"] = cv_metrics.get("rmse")
    model_artifact["metrics"]["cv_spearman"] = cv_metrics.get("spearman")
    model_artifact["metrics"]["cv_mode"] = cv.get("mode")
    model_artifact["metrics"]["cv_pred_count"] = cv.get("pred_count")
    model_artifact["metrics"]["baseline_mae"] = baseline_metrics.get("mae")
    model_artifact["metrics"]["baseline_rmse"] = baseline_metrics.get("rmse")
    model_artifact["metrics"]["baseline_spearman"] = baseline_metrics.get("spearman")
    model_artifact["metrics"]["gate_improve_threshold"] = round(improve_threshold, 4)
    model_artifact["metrics"]["gate_spearman_tolerance"] = spearman_tolerance
    model_artifact["gate_passed"] = gate_passed

    auto_candidates = _extract_auto_candidates(model_artifact)
    version = f"{version_prefix}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"
    calibrator_summary = _build_calibrator_summary(
        model_type=selected_type,
        calibrator_version=version,
        gate_passed=bool(gate_passed),
        cv_metrics={
            "mae": cv_metrics.get("mae"),
            "rmse": cv_metrics.get("rmse"),
            "spearman": cv_metrics.get("spearman"),
            "mode": cv.get("mode"),
            "pred_count": cv.get("pred_count"),
        },
        baseline_metrics={
            "mae": baseline_metrics.get("mae"),
            "rmse": baseline_metrics.get("rmse"),
            "spearman": baseline_metrics.get("spearman"),
        },
        improve_threshold=improve_threshold,
        spearman_tolerance=spearman_tolerance,
        auto_candidates=auto_candidates,
        sample_count=len(feature_rows),
    )
    record = {
        "calibrator_version": version,
        "model_type": selected_type,
        "feature_schema_version": str(model_artifact.get("feature_schema_version", "v2")),
        "train_filter": {"project_id": payload.project_id},
        "metrics": {
            **(model_artifact.get("metrics") or {}),
            "gate_passed": bool(gate_passed),
        },
        "calibrator_summary": calibrator_summary,
        "artifact_uri": f"json://calibration_models/{version}",
        "model_artifact": model_artifact,
        "deployed": False,
        "created_at": _now_iso(),
    }

    models = load_calibration_models()
    train_scope_project_id = str(payload.project_id or "").strip()
    if payload.auto_deploy and bool(gate_passed) and train_scope_project_id:
        for m in models:
            if (
                str(((m.get("train_filter") or {}).get("project_id") or ""))
                == train_scope_project_id
            ):
                m["deployed"] = False
        record["deployed"] = True
        projects = load_projects()
        for p in projects:
            if str(p.get("id")) == train_scope_project_id:
                p["calibrator_version_locked"] = version
                p["updated_at"] = _now_iso()
        save_projects(projects)
    models.append(record)
    save_calibration_models(models)
    return CalibratorModelRecord(**record)


@router.get(
    "/calibration/models",
    response_model=list[CalibratorModelRecord],
    tags=["洞察与学习"],
)
def list_calibration_models() -> list[CalibratorModelRecord]:
    """获取校准器版本列表。"""
    ensure_data_dirs()
    models = sorted(
        load_calibration_models(), key=lambda x: str(x.get("created_at", "")), reverse=True
    )
    return [CalibratorModelRecord(**m) for m in models]


@router.post(
    "/calibration/deploy",
    response_model=CalibratorModelRecord,
    tags=["洞察与学习"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def deploy_calibrator(
    payload: CalibratorDeployRequest,
    api_key: Optional[str] = Depends(verify_api_key),
) -> CalibratorModelRecord:
    """部署某个校准器版本。"""
    ensure_data_dirs()
    models = load_calibration_models()
    target = None
    for model in models:
        if str(model.get("calibrator_version")) == payload.calibrator_version:
            target = model
            break
    if target is None:
        raise HTTPException(status_code=404, detail="校准器版本不存在")

    target_scope = str(((target.get("train_filter") or {}).get("project_id") or "")).strip()
    bind_project_id = str(payload.project_id or "").strip()
    if bind_project_id:
        if target_scope and target_scope != bind_project_id:
            raise HTTPException(status_code=422, detail="校准器与目标项目不匹配，禁止跨项目部署")
        if not target_scope:
            target.setdefault("train_filter", {})
            target["train_filter"]["project_id"] = bind_project_id
            target_scope = bind_project_id

    for model in models:
        model_scope = str(((model.get("train_filter") or {}).get("project_id") or "")).strip()
        if target_scope and model_scope == target_scope:
            model["deployed"] = False
    target["deployed"] = True
    save_calibration_models(models)

    if payload.project_id:
        projects = load_projects()
        for project in projects:
            if str(project.get("id")) == payload.project_id:
                project["calibrator_version_locked"] = payload.calibrator_version
                project["updated_at"] = _now_iso()
        save_projects(projects)

    return CalibratorModelRecord(**target)


@router.post(
    "/projects/{project_id}/calibration/predict",
    response_model=CalibratorPredictResponse,
    tags=["洞察与学习"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def apply_calibration_prediction(
    project_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> CalibratorPredictResponse:
    """将已部署校准器应用到项目已有评分报告。"""
    ensure_data_dirs()
    projects = load_projects()
    try:
        project = _find_project(project_id, projects)
    except HTTPException:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))

    model = _select_calibrator_model(project)
    if not model:
        return CalibratorPredictResponse(
            ok=True,
            project_id=project_id,
            model_version=None,
            updated_reports=0,
            updated_submissions=0,
        )

    submissions = load_submissions()
    submission_map = {
        str(s.get("id")): s for s in submissions if str(s.get("project_id")) == project_id
    }

    reports = load_score_reports()
    updated_reports = 0
    for report in reports:
        if str(report.get("project_id")) != project_id:
            continue
        sid = str(report.get("submission_id") or "")
        sub = submission_map.get(sid)
        if not sid or not sub:
            continue
        _apply_prediction_to_report(report, submission_like=sub, project=project)
        updated_reports += 1
    save_score_reports(reports)

    updated_submissions = 0
    for submission in submissions:
        if str(submission.get("project_id")) != project_id:
            continue
        report = submission.get("report")
        if not isinstance(report, dict):
            continue
        _apply_prediction_to_report(report, submission_like=submission, project=project)
        updated_submissions += 1
    save_submissions(submissions)

    return CalibratorPredictResponse(
        ok=True,
        project_id=project_id,
        model_version=str(model.get("calibrator_version") or ""),
        updated_reports=updated_reports,
        updated_submissions=updated_submissions,
    )


@router.post(
    "/projects/{project_id}/delta_cases/rebuild",
    response_model=list[DeltaCaseRecord],
    tags=["洞察与学习"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def rebuild_delta_cases(
    project_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> list[DeltaCaseRecord]:
    """重建项目 DELTA_CASE（反演误差案例）。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(str(p.get("id")) == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))

    reports = [r for r in load_score_reports() if str(r.get("project_id")) == project_id]
    latest_reports = _latest_records_by_submission(reports)
    qtrs = load_qingtian_results()
    latest_qt = _latest_records_by_submission(
        [q for q in qtrs if str(q.get("submission_id")) in latest_reports]
    )

    new_cases = build_delta_cases(
        project_id=project_id,
        latest_reports_by_submission=latest_reports,
        latest_qingtian_by_submission=latest_qt,
    )
    all_cases = [d for d in load_delta_cases() if str(d.get("project_id")) != project_id]
    all_cases.extend(new_cases)
    save_delta_cases(all_cases)
    return [DeltaCaseRecord(**d) for d in new_cases]


@router.get(
    "/projects/{project_id}/delta_cases",
    response_model=list[DeltaCaseRecord],
    tags=["洞察与学习"],
    responses={**RESPONSES_404},
)
def list_delta_cases(
    project_id: str,
    locale: str = Depends(get_locale),
) -> list[DeltaCaseRecord]:
    """查询项目 DELTA_CASE。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(str(p.get("id")) == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    rows = [d for d in load_delta_cases() if str(d.get("project_id")) == project_id]
    rows = sorted(rows, key=lambda x: str(x.get("created_at", "")), reverse=True)
    return [DeltaCaseRecord(**d) for d in rows]


@router.post(
    "/projects/{project_id}/calibration_samples/rebuild",
    response_model=list[CalibrationSampleRecord],
    tags=["洞察与学习"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def rebuild_calibration_samples(
    project_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> list[CalibrationSampleRecord]:
    """重建项目 FEATURE_ROW 校准样本。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(str(p.get("id")) == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))

    submissions = [s for s in load_submissions() if str(s.get("project_id")) == project_id]
    submissions_by_id = {str(s.get("id")): s for s in submissions}
    latest_reports = _latest_records_by_submission(
        [r for r in load_score_reports() if str(r.get("project_id")) == project_id]
    )
    latest_qt = _latest_records_by_submission(
        [q for q in load_qingtian_results() if str(q.get("submission_id")) in submissions_by_id]
    )

    samples = build_calibration_samples(
        project_id=project_id,
        latest_reports_by_submission=latest_reports,
        latest_qingtian_by_submission=latest_qt,
        submissions_by_id=submissions_by_id,
    )
    all_samples = [s for s in load_calibration_samples() if str(s.get("project_id")) != project_id]
    all_samples.extend(samples)
    save_calibration_samples(all_samples)
    return [CalibrationSampleRecord(**s) for s in samples]


@router.get(
    "/projects/{project_id}/calibration_samples",
    response_model=list[CalibrationSampleRecord],
    tags=["洞察与学习"],
    responses={**RESPONSES_404},
)
def list_calibration_samples(
    project_id: str,
    locale: str = Depends(get_locale),
) -> list[CalibrationSampleRecord]:
    """查询项目 FEATURE_ROW 样本。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(str(p.get("id")) == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    rows = [s for s in load_calibration_samples() if str(s.get("project_id")) == project_id]
    rows = sorted(rows, key=lambda x: str(x.get("created_at", "")), reverse=True)
    return [CalibrationSampleRecord(**s) for s in rows]


@router.post(
    "/projects/{project_id}/patches/mine",
    response_model=PatchPackageRecord,
    tags=["洞察与学习"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def mine_patch(
    project_id: str,
    payload: PatchMineRequest,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> PatchPackageRecord:
    """基于 DELTA_CASE 挖掘候选补丁包（PATCH_PACKAGE）。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(str(p.get("id")) == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    delta_cases = [d for d in load_delta_cases() if str(d.get("project_id")) == project_id]
    if not delta_cases:
        raise HTTPException(status_code=404, detail="暂无 DELTA_CASE，请先重建")

    packages = load_patch_packages()
    rollback_pointer = None
    deployed = [
        p
        for p in packages
        if str(p.get("project_id")) == project_id and str(p.get("status")) == "deployed"
    ]
    if deployed:
        deployed = sorted(deployed, key=lambda x: str(x.get("updated_at", "")), reverse=True)
        rollback_pointer = str(deployed[0].get("id") or "")

    package = mine_patch_package(
        project_id=project_id,
        delta_cases=delta_cases,
        patch_type=payload.patch_type,
        top_k=int(payload.top_k),
        rollback_pointer=rollback_pointer,
    )
    packages.append(package)
    save_patch_packages(packages)
    return PatchPackageRecord(**package)


@router.get(
    "/projects/{project_id}/patches",
    response_model=list[PatchPackageRecord],
    tags=["洞察与学习"],
    responses={**RESPONSES_404},
)
def list_patches(
    project_id: str,
    locale: str = Depends(get_locale),
) -> list[PatchPackageRecord]:
    """查询项目补丁包列表。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(str(p.get("id")) == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    rows = [p for p in load_patch_packages() if str(p.get("project_id")) == project_id]
    rows = sorted(rows, key=lambda x: str(x.get("updated_at", "")), reverse=True)
    return [PatchPackageRecord(**p) for p in rows]


@router.post(
    "/patches/{patch_id}/shadow_eval",
    response_model=PatchShadowEvalResponse,
    tags=["洞察与学习"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def shadow_eval_patch(
    patch_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
) -> PatchShadowEvalResponse:
    """对候选补丁做 shadow 评估并更新状态。"""
    ensure_data_dirs()
    packages = load_patch_packages()
    patch = next((p for p in packages if str(p.get("id")) == patch_id), None)
    if patch is None:
        raise HTTPException(status_code=404, detail="补丁包不存在")
    project_id = str(patch.get("project_id") or "")
    delta_cases = [d for d in load_delta_cases() if str(d.get("project_id")) == project_id]

    result = evaluate_patch_shadow(patch=patch, delta_cases=delta_cases)
    patch["shadow_metrics"] = result.get("metrics_before_after", {})
    patch["status"] = "shadow_pass" if bool(result.get("gate_passed")) else "candidate"
    patch["updated_at"] = _now_iso()
    save_patch_packages(packages)
    return PatchShadowEvalResponse(**result)


@router.post(
    "/patches/{patch_id}/deploy",
    response_model=PatchDeploymentRecord,
    tags=["洞察与学习"],
    responses={**RESPONSES_401, **RESPONSES_404, **RESPONSES_422},
)
def deploy_or_rollback_patch(
    patch_id: str,
    payload: PatchDeployRequest,
    api_key: Optional[str] = Depends(verify_api_key),
) -> PatchDeploymentRecord:
    """发布或回滚补丁。"""
    ensure_data_dirs()
    packages = load_patch_packages()
    patch = next((p for p in packages if str(p.get("id")) == patch_id), None)
    if patch is None:
        raise HTTPException(status_code=404, detail="补丁包不存在")

    action = str(payload.action or "deploy").lower()
    if action not in {"deploy", "rollback"}:
        raise HTTPException(status_code=422, detail="action 仅支持 deploy 或 rollback")

    project_id = str(patch.get("project_id") or "")
    deployed = action == "deploy"
    if deployed:
        for p in packages:
            if str(p.get("project_id")) == project_id and str(p.get("status")) == "deployed":
                p["status"] = "shadow_pass"
                p["updated_at"] = _now_iso()
        patch["status"] = "deployed"
    else:
        patch["status"] = "rolled_back"
    patch["updated_at"] = _now_iso()
    save_patch_packages(packages)

    rec = {
        "id": str(uuid4()),
        "patch_id": patch_id,
        "project_id": project_id,
        "action": action,
        "deployed": deployed,
        "metrics_before_after": patch.get("shadow_metrics") or {},
        "rollback_to_version": payload.rollback_to_version or patch.get("rollback_pointer"),
        "created_at": _now_iso(),
    }
    deploys = load_patch_deployments()
    deploys.append(rec)
    save_patch_deployments(deploys)
    return PatchDeploymentRecord(**rec)


@router.post(
    "/projects/{project_id}/reflection/auto_run",
    response_model=ReflectionAutoRunResponse,
    tags=["洞察与学习"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def auto_run_reflection_pipeline(
    project_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> ReflectionAutoRunResponse:
    """
    一键执行反演闭环：
    1) 刷新 DELTA_CASE + FEATURE_ROW
    2) 训练并自动部署校准器（闸门通过）
    3) 回填预测分
    4) 挖掘/影子评估/发布补丁（闸门通过）
    """
    ensure_data_dirs()
    projects = load_projects()
    project = None
    for p in projects:
        if str(p.get("id")) == project_id:
            project = p
            break
    if project is None:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))

    _refresh_project_reflection_objects(project_id)
    delta_cases = [d for d in load_delta_cases() if str(d.get("project_id")) == project_id]
    samples = [s for s in load_calibration_samples() if str(s.get("project_id")) == project_id]

    calibrator_version = None
    calibrator_deployed = False
    calibrator_summary = _build_calibrator_summary(
        model_type=None,
        calibrator_version=None,
        gate_passed=None,
        sample_count=len(samples),
        skipped_reason="insufficient_samples" if len(samples) < 3 else None,
    )
    calibrator_model_type = calibrator_summary.get("model_type")
    calibrator_gate_passed = calibrator_summary.get("gate_passed")
    calibrator_cv_metrics: Dict[str, Any] = calibrator_summary.get("cv_metrics") or {}
    calibrator_baseline_metrics: Dict[str, Any] = calibrator_summary.get("baseline_metrics") or {}
    calibrator_gate: Dict[str, Any] = calibrator_summary.get("gate") or {}
    calibrator_auto_candidates: List[Dict[str, Any]] = (
        calibrator_summary.get("auto_candidates") or []
    )
    if len(samples) >= 3:
        feature_rows = [
            {
                "feature_schema_version": s.get("feature_schema_version", "v2"),
                "x_features": s.get("x_features") or {},
                "y_label": s.get("y_label"),
                "submission_id": s.get("submission_id"),
            }
            for s in samples
        ]
        # “最强自动校准”：多候选 + CV 闸门 + 自动选择最佳模型
        model_artifact = train_best_calibrator_auto(feature_rows, alpha=1.0)
        selected_type = str(model_artifact.get("model_type") or "ridge")
        calibrator_model_type = selected_type

        # 统一用 CV 口径做上线闸门（避免 in-sample 过拟合）
        cv = cross_validate_calibrator(
            model_type=selected_type,
            feature_rows=feature_rows,
            alpha=1.0,
            seed=42,
        )
        # baseline: raw rule_total_score
        y_true = [float(r.get("y_label")) for r in feature_rows if r.get("y_label") is not None]
        baseline_pred = [
            float(((r.get("x_features") or {}).get("rule_total_score") or 0.0))
            for r in feature_rows
            if r.get("y_label") is not None
        ]
        baseline_metrics = calc_metrics(y_true, baseline_pred)
        cv_metrics = (
            (cv.get("metrics") or {})
            if bool(cv.get("ok"))
            else {"mae": 0.0, "rmse": 0.0, "spearman": 0.0}
        )
        improve_threshold = max(0.2, float(baseline_metrics.get("mae") or 0.0) * 0.01)
        spearman_tolerance = 0.02
        gate_passed = (
            bool(cv.get("ok"))
            and float(cv_metrics.get("mae") or 0.0)
            <= float(baseline_metrics.get("mae") or 0.0) - improve_threshold
            and float(cv_metrics.get("spearman") or 0.0)
            >= float(baseline_metrics.get("spearman") or 0.0) - spearman_tolerance
        )
        model_artifact.setdefault("metrics", {})
        model_artifact["metrics"]["cv_mae"] = cv_metrics.get("mae")
        model_artifact["metrics"]["cv_rmse"] = cv_metrics.get("rmse")
        model_artifact["metrics"]["cv_spearman"] = cv_metrics.get("spearman")
        model_artifact["metrics"]["cv_mode"] = cv.get("mode")
        model_artifact["metrics"]["cv_pred_count"] = cv.get("pred_count")
        model_artifact["metrics"]["baseline_mae"] = baseline_metrics.get("mae")
        model_artifact["metrics"]["baseline_rmse"] = baseline_metrics.get("rmse")
        model_artifact["metrics"]["baseline_spearman"] = baseline_metrics.get("spearman")
        model_artifact["metrics"]["gate_improve_threshold"] = round(improve_threshold, 4)
        model_artifact["metrics"]["gate_spearman_tolerance"] = spearman_tolerance
        model_artifact["gate_passed"] = gate_passed

        auto_candidates = _extract_auto_candidates(model_artifact)
        calibrator_version = (
            f"calib_auto_{selected_type}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"
        )
        calibrator_summary = _build_calibrator_summary(
            model_type=selected_type,
            calibrator_version=calibrator_version,
            gate_passed=bool(gate_passed),
            cv_metrics={
                "mae": cv_metrics.get("mae"),
                "rmse": cv_metrics.get("rmse"),
                "spearman": cv_metrics.get("spearman"),
                "mode": cv.get("mode"),
                "pred_count": cv.get("pred_count"),
            },
            baseline_metrics={
                "mae": baseline_metrics.get("mae"),
                "rmse": baseline_metrics.get("rmse"),
                "spearman": baseline_metrics.get("spearman"),
            },
            improve_threshold=improve_threshold,
            spearman_tolerance=spearman_tolerance,
            auto_candidates=auto_candidates,
            sample_count=len(feature_rows),
        )
        calibrator_model_type = calibrator_summary.get("model_type")
        calibrator_gate_passed = calibrator_summary.get("gate_passed")
        calibrator_cv_metrics = calibrator_summary.get("cv_metrics") or {}
        calibrator_baseline_metrics = calibrator_summary.get("baseline_metrics") or {}
        calibrator_gate = calibrator_summary.get("gate") or {}
        calibrator_auto_candidates = calibrator_summary.get("auto_candidates") or []
        record = {
            "calibrator_version": calibrator_version,
            "model_type": selected_type,
            "feature_schema_version": str(model_artifact.get("feature_schema_version", "v2")),
            "train_filter": {"project_id": project_id, "mode": "auto_run"},
            "metrics": {
                **(model_artifact.get("metrics") or {}),
                "gate_passed": bool(gate_passed),
            },
            "calibrator_summary": calibrator_summary,
            "artifact_uri": f"json://calibration_models/{calibrator_version}",
            "model_artifact": model_artifact,
            "deployed": bool(gate_passed),
            "created_at": _now_iso(),
        }
        models = load_calibration_models()
        if record["deployed"]:
            for m in models:
                if str(((m.get("train_filter") or {}).get("project_id") or "")) == project_id:
                    m["deployed"] = False
            project["calibrator_version_locked"] = calibrator_version
            project["updated_at"] = _now_iso()
            save_projects(projects)
            calibrator_deployed = True
        models.append(record)
        save_calibration_models(models)

    updated_reports = 0
    updated_submissions = 0
    if calibrator_deployed:
        submissions = load_submissions()
        submission_map = {
            str(s.get("id")): s for s in submissions if str(s.get("project_id")) == project_id
        }
        reports = load_score_reports()
        for report in reports:
            if str(report.get("project_id")) != project_id:
                continue
            sid = str(report.get("submission_id") or "")
            sub = submission_map.get(sid)
            if not sub:
                continue
            _apply_prediction_to_report(report, submission_like=sub, project=project)
            updated_reports += 1
        save_score_reports(reports)

        for sub in submissions:
            if str(sub.get("project_id")) != project_id:
                continue
            rep = sub.get("report")
            if not isinstance(rep, dict):
                continue
            _apply_prediction_to_report(rep, submission_like=sub, project=project)
            updated_submissions += 1
        save_submissions(submissions)

    patch_id = None
    patch_gate_passed = None
    patch_deployed = False
    patch_auto_govern: Dict[str, object] = {
        "checked": False,
        "reason": "not_run",
        "action": "skip",
    }
    if delta_cases:
        patch_auto_govern = _auto_govern_deployed_patch(
            project_id=project_id,
            delta_cases=delta_cases,
        )
        packages = load_patch_packages()
        deployed = [
            p
            for p in packages
            if str(p.get("project_id")) == project_id and str(p.get("status")) == "deployed"
        ]
        rollback_pointer = str(deployed[0].get("id")) if deployed else None
        patch = mine_patch_package(
            project_id=project_id,
            delta_cases=delta_cases,
            patch_type="threshold",
            top_k=5,
            rollback_pointer=rollback_pointer,
        )
        patch_id = str(patch.get("id"))
        shadow = evaluate_patch_shadow(patch=patch, delta_cases=delta_cases)
        patch_gate_passed = bool(shadow.get("gate_passed"))
        patch["shadow_metrics"] = shadow.get("metrics_before_after", {})
        patch["status"] = "shadow_pass" if patch_gate_passed else "candidate"
        patch["updated_at"] = _now_iso()

        if patch_gate_passed:
            for p in packages:
                if str(p.get("project_id")) == project_id and str(p.get("status")) == "deployed":
                    p["status"] = "shadow_pass"
            patch["status"] = "deployed"
            patch_deployed = True
            deploy_rec = {
                "id": str(uuid4()),
                "patch_id": patch_id,
                "project_id": project_id,
                "action": "deploy",
                "deployed": True,
                "metrics_before_after": patch.get("shadow_metrics") or {},
                "rollback_to_version": patch.get("rollback_pointer"),
                "created_at": _now_iso(),
            }
            deploys = load_patch_deployments()
            deploys.append(deploy_rec)
            save_patch_deployments(deploys)

        packages.append(patch)
        save_patch_packages(packages)

    return ReflectionAutoRunResponse(
        ok=True,
        project_id=project_id,
        delta_cases=len(delta_cases),
        calibration_samples=len(samples),
        calibrator_version=calibrator_version,
        calibrator_deployed=calibrator_deployed,
        calibrator_summary=calibrator_summary,
        calibrator_model_type=calibrator_model_type,
        calibrator_gate_passed=calibrator_gate_passed,
        calibrator_cv_metrics=calibrator_cv_metrics,
        calibrator_baseline_metrics=calibrator_baseline_metrics,
        calibrator_gate=calibrator_gate,
        calibrator_auto_candidates=calibrator_auto_candidates,
        prediction_updated_reports=updated_reports,
        prediction_updated_submissions=updated_submissions,
        patch_id=patch_id,
        patch_gate_passed=patch_gate_passed,
        patch_deployed=patch_deployed,
        patch_auto_govern=patch_auto_govern,
    )


@router.get(
    "/projects/{project_id}/evaluation",
    response_model=ProjectEvaluationResponse,
    tags=["洞察与学习"],
    responses={**RESPONSES_404},
)
def get_project_evaluation(
    project_id: str,
    locale: str = Depends(get_locale),
) -> ProjectEvaluationResponse:
    """输出项目级 V1/V2/V2+Calib 对比指标。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(str(p.get("id")) == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    data = evaluate_project_variants(
        project_id=project_id,
        submissions=load_submissions(),
        score_reports=load_score_reports(),
        qingtian_results=load_qingtian_results(),
    )
    return ProjectEvaluationResponse(**data)


@router.get(
    "/evaluation/summary",
    response_model=EvaluationSummaryResponse,
    tags=["洞察与学习"],
)
def get_evaluation_summary() -> EvaluationSummaryResponse:
    """跨项目汇总 V1/V2/V2+Calib 验收指标。"""
    ensure_data_dirs()
    projects = load_projects()
    submissions = load_submissions()
    reports = load_score_reports()
    qts = load_qingtian_results()

    project_ids = [str(p.get("id")) for p in projects if str(p.get("id") or "")]
    project_metrics = []
    for pid in project_ids:
        project_metrics.append(
            evaluate_project_variants(
                project_id=pid,
                submissions=submissions,
                score_reports=reports,
                qingtian_results=qts,
            )
        )

    def _avg(values: List[float]) -> float:
        if not values:
            return 0.0
        return round(sum(values) / len(values), 4)

    agg: Dict[str, Dict[str, Any]] = {}
    for variant in ("v1", "v2", "v2_calib"):
        maes, rmses, cors = [], [], []
        profile_sims, hit_rates, sample_counts = [], [], []
        for item in project_metrics:
            v = (item.get("variants") or {}).get(variant) or {}
            if int(v.get("sample_count") or 0) <= 0:
                continue
            maes.append(float(v.get("mae") or 0.0))
            rmses.append(float(v.get("rmse") or 0.0))
            cors.append(float(v.get("spearman") or 0.0))
            sample_counts.append(int(v.get("sample_count") or 0))
            if v.get("profile_similarity") is not None:
                profile_sims.append(float(v.get("profile_similarity")))
            if v.get("penalty_hit_rate") is not None:
                hit_rates.append(float(v.get("penalty_hit_rate")))
        agg[variant] = {
            "project_count": len(sample_counts),
            "sample_count_total": sum(sample_counts),
            "mae_avg": _avg(maes),
            "rmse_avg": _avg(rmses),
            "spearman_avg": _avg(cors),
            "profile_similarity_avg": _avg(profile_sims) if profile_sims else None,
            "penalty_hit_rate_avg": _avg(hit_rates) if hit_rates else None,
        }

    pass_count = {
        "mae_rmse_improved_vs_v1": 0,
        "rank_corr_not_worse_vs_v1": 0,
        "profile_similarity_improved_v2_vs_v1": 0,
        "penalty_hit_rate_improved_v2_vs_v1": 0,
    }
    for item in project_metrics:
        acc = item.get("acceptance") or {}
        for key in pass_count:
            if bool(acc.get(key)):
                pass_count[key] += 1

    return EvaluationSummaryResponse(
        project_count=len(project_ids),
        project_ids=project_ids,
        aggregate=agg,
        acceptance_pass_count=pass_count,
        computed_at=_now_iso(),
    )


@router.get(
    "/projects/{project_id}/compare",
    response_model=CompareReport,
    tags=["对比分析"],
    responses=RESPONSES_NO_SUBMISSIONS,
)
def compare_submissions(
    project_id: str,
    locale: str = Depends(get_locale),
) -> CompareReport:
    """
    对比项目的多次施组提交。

    返回排名、各维度平均分、常见扣分项统计等对比分析数据。

    支持 Accept-Language header 进行多语言响应。
    """
    ensure_data_dirs()
    projects = load_projects()
    project = next((p for p in projects if str(p.get("id")) == project_id), {"id": project_id})
    allow_pred_score = _select_calibrator_model(project) is not None
    score_scale_max = _resolve_project_score_scale_max(project)
    submissions_all = [s for s in load_submissions() if s["project_id"] == project_id]
    if not submissions_all:
        raise HTTPException(status_code=404, detail=t("api.no_submissions", locale=locale))
    submissions = [s for s in submissions_all if _submission_is_scored(s)]
    if not submissions:
        raise HTTPException(status_code=404, detail="暂无已评分施组，请先点击“评分施组”。")
    rankings = []
    for s in submissions:
        score_fields = _resolve_submission_score_fields(
            s,
            allow_pred_score=allow_pred_score,
            score_scale_max=score_scale_max,
        )
        rankings.append(
            {
                "submission_id": s["id"],
                "filename": s["filename"],
                "total_score": score_fields["total_score"],
                "pred_total_score": score_fields["pred_total_score"],
                "rule_total_score": score_fields["rule_total_score"],
                "score_source": score_fields["score_source"],
                "created_at": s["created_at"],
            }
        )
    rankings = sorted(rankings, key=lambda x: float(x["total_score"]), reverse=True)
    dimension_totals: dict[str, float] = {}
    dimension_counts: dict[str, int] = {}
    penalty_stats: dict[str, int] = {}
    for s in submissions:
        report = s.get("report") if isinstance(s.get("report"), dict) else {}
        for dim_id, dim in (report.get("dimension_scores") or {}).items():
            dimension_totals[dim_id] = dimension_totals.get(dim_id, 0.0) + float(
                dim.get("score", 0.0)
            )
            dimension_counts[dim_id] = dimension_counts.get(dim_id, 0) + 1
        for p in report.get("penalties") or []:
            if not isinstance(p, dict):
                continue
            code = p.get("code", "UNKNOWN")
            penalty_stats[code] = penalty_stats.get(code, 0) + 1

    dimension_avg = {
        dim_id: round(dimension_totals[dim_id] / dimension_counts[dim_id], 2)
        for dim_id in dimension_totals
    }
    return CompareReport(
        project_id=project_id,
        rankings=rankings,
        dimension_avg=dimension_avg,
        penalty_stats=penalty_stats,
    )


@router.get(
    "/projects/{project_id}/compare_report",
    response_model=CompareNarrative,
    tags=["对比分析"],
    responses=RESPONSES_NO_SUBMISSIONS,
)
def compare_report(
    project_id: str,
    locale: str = Depends(get_locale),
) -> CompareNarrative:
    """
    生成对比分析叙述报告。

    返回自然语言格式的对比分析报告，包含趋势分析和改进建议。

    支持 Accept-Language header 进行多语言响应。
    """
    ensure_data_dirs()
    projects = load_projects()
    project = next((p for p in projects if str(p.get("id")) == project_id), {"id": project_id})
    allow_pred_score = _select_calibrator_model(project) is not None
    score_scale_max = _resolve_project_score_scale_max(project)
    submissions_all = [s for s in load_submissions() if s["project_id"] == project_id]
    if not submissions_all:
        raise HTTPException(status_code=404, detail=t("api.no_submissions", locale=locale))
    submissions = [s for s in submissions_all if _submission_is_scored(s)]
    if not submissions:
        raise HTTPException(status_code=404, detail="暂无已评分施组，请先点击“评分施组”。")
    submissions_for_compare = []
    by_id: Dict[str, Dict[str, object]] = {}
    for s in submissions:
        score_fields_display = _resolve_submission_score_fields(
            s,
            allow_pred_score=allow_pred_score,
            score_scale_max=score_scale_max,
        )
        score_fields_raw = _resolve_submission_score_fields(
            s,
            allow_pred_score=allow_pred_score,
            score_scale_max=100,
        )
        item = dict(s)
        item["total_score"] = float(score_fields_raw["total_score"])
        report = item.get("report")
        report = dict(report) if isinstance(report, dict) else {}
        report["pred_total_score"] = score_fields_raw["pred_total_score"]
        report["rule_total_score"] = score_fields_raw["rule_total_score"]
        item["report"] = report
        item["score_source"] = score_fields_display["score_source"]
        submissions_for_compare.append(item)
        by_id[str(item.get("id") or "")] = item

    narrative = build_compare_narrative(submissions_for_compare)
    for key in ("top_submission", "bottom_submission"):
        row = narrative.get(key)
        if not isinstance(row, dict):
            continue
        sid = str(row.get("id") or "")
        source_submission = by_id.get(sid)
        if not source_submission:
            continue
        score_fields = _resolve_submission_score_fields(
            source_submission,
            allow_pred_score=allow_pred_score,
            score_scale_max=score_scale_max,
        )
        row["pred_total_score"] = score_fields["pred_total_score"]
        row["rule_total_score"] = score_fields["rule_total_score"]
        row["score_source"] = score_fields["score_source"]
    return CompareNarrative(project_id=project_id, **narrative)


@router.get(
    "/projects/{project_id}/adaptive",
    response_model=AdaptiveSuggestions,
    tags=["自适应优化"],
    responses=RESPONSES_NO_SUBMISSIONS,
)
def adaptive_suggestions(
    project_id: str,
    locale: str = Depends(get_locale),
) -> AdaptiveSuggestions:
    """
    获取自适应优化建议。

    基于项目历史评分数据，分析常见扣分模式，提供评分规则优化建议。

    支持 Accept-Language header 进行多语言响应。
    """
    ensure_data_dirs()
    submissions_all = [s for s in load_submissions() if s["project_id"] == project_id]
    if not submissions_all:
        raise HTTPException(status_code=404, detail=t("api.no_submissions", locale=locale))
    submissions = [s for s in submissions_all if _submission_is_scored(s)]
    if not submissions:
        raise HTTPException(status_code=404, detail="暂无已评分施组，请先点击“评分施组”。")
    config = load_config()
    result = build_adaptive_suggestions(submissions, config.lexicon)
    return AdaptiveSuggestions(project_id=project_id, **result)


@router.get(
    "/projects/{project_id}/adaptive_patch",
    response_model=AdaptivePatch,
    tags=["自适应优化"],
    responses=RESPONSES_NO_SUBMISSIONS,
)
def adaptive_patch(
    project_id: str,
    locale: str = Depends(get_locale),
) -> AdaptivePatch:
    """
    获取自适应优化补丁。

    根据扣分统计生成词库调整补丁，包含建议的阈值修改。

    支持 Accept-Language header 进行多语言响应。
    """
    ensure_data_dirs()
    submissions = [s for s in load_submissions() if s["project_id"] == project_id]
    if not submissions:
        raise HTTPException(status_code=404, detail=t("api.no_submissions", locale=locale))
    config = load_config()
    stats_result = build_adaptive_suggestions(submissions, config.lexicon)
    stats = stats_result["penalty_stats"]
    patch = build_adaptive_patch(config.lexicon, stats)
    return AdaptivePatch(project_id=project_id, source=stats_result.get("source") or {}, **patch)


@router.post(
    "/projects/{project_id}/adaptive_apply",
    response_model=AdaptiveApplyResult,
    tags=["自适应优化"],
    responses={**RESPONSES_401, **RESPONSES_NO_SUBMISSIONS},
)
def adaptive_apply(
    project_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> AdaptiveApplyResult:
    """
    应用自适应优化补丁。

    将优化补丁应用到词库与评分规则配置，自动备份原配置文件。

    **需要 API Key 认证**

    ⚠️ 此操作会修改系统配置文件（lexicon.yaml、rubric.yaml）

    支持 Accept-Language header 进行多语言响应。
    """
    ensure_data_dirs()
    submissions = [s for s in load_submissions() if s["project_id"] == project_id]
    if not submissions:
        raise HTTPException(status_code=404, detail=t("api.no_submissions", locale=locale))
    config = load_config()
    stats_result = build_adaptive_suggestions(submissions, config.lexicon)
    stats = stats_result["penalty_stats"]
    patch = build_adaptive_patch(config.lexicon, stats)

    from pathlib import Path

    import yaml

    res_dir = Path(__file__).resolve().parent / "resources"
    ts = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    all_changes: list[str] = []

    # 词库补丁：备份并写回
    lexicon_path = res_dir / "lexicon.yaml"
    lex_backup = lexicon_path.with_name(f"lexicon.yaml.bak_{ts}")
    lex_backup.write_text(lexicon_path.read_text(encoding="utf-8"), encoding="utf-8")
    updated_lexicon, lex_changes = apply_adaptive_patch(config.lexicon, patch)
    lexicon_path.write_text(yaml.safe_dump(updated_lexicon, allow_unicode=True), encoding="utf-8")
    all_changes.extend(lex_changes)

    # 规则补丁：备份并写回
    rubric_path = res_dir / "rubric.yaml"
    rubric_backup = rubric_path.with_name(f"rubric.yaml.bak_{ts}")
    rubric_backup.write_text(rubric_path.read_text(encoding="utf-8"), encoding="utf-8")
    updated_rubric, rubric_changes = apply_rubric_patch(
        config.rubric, patch.get("rubric_adjustments", {})
    )
    rubric_path.write_text(yaml.safe_dump(updated_rubric, allow_unicode=True), encoding="utf-8")
    all_changes.extend(rubric_changes)

    # 使内存中的配置失效，下次 load_config 会重新读盘
    reload_config()

    return AdaptiveApplyResult(
        project_id=project_id,
        applied=True,
        changes=all_changes,
        backup_path=str(lex_backup),
        source=stats_result.get("source") or {},
    )


@router.get(
    "/projects/{project_id}/adaptive_validate",
    response_model=AdaptiveValidation,
    tags=["自适应优化"],
    responses=RESPONSES_NO_SUBMISSIONS,
)
def adaptive_validate(
    project_id: str,
    locale: str = Depends(get_locale),
) -> AdaptiveValidation:
    """
    验证自适应优化效果。

    使用当前配置重新评分历史提交，对比新旧分数差异，验证优化效果。

    支持 Accept-Language header 进行多语言响应。
    """
    ensure_data_dirs()
    submissions = [s for s in load_submissions() if s["project_id"] == project_id]
    if not submissions:
        raise HTTPException(status_code=404, detail=t("api.no_submissions", locale=locale))
    config = load_config()
    try:
        multipliers, _, _ = _resolve_project_scoring_context(project_id)
    except HTTPException as exc:
        # 兼容历史/清理场景：项目记录缺失时，仍允许按默认权重验证当前施组数据。
        if exc.status_code == 404:
            multipliers = {}
        else:
            raise
    comparisons = []
    deltas = []
    for s in submissions:
        text = s.get("text")
        if not text:
            continue
        new_report = score_text(
            text,
            config.rubric,
            config.lexicon,
            dimension_multipliers=multipliers,
        ).model_dump()
        old_score = float(s.get("total_score", 0.0))
        new_score = float(new_report.get("total_score", 0.0))
        delta = round(new_score - old_score, 2)
        comparisons.append(
            {
                "submission_id": s.get("id"),
                "filename": s.get("filename"),
                "old_score": old_score,
                "new_score": new_score,
                "delta": delta,
            }
        )
        deltas.append(delta)
    avg_delta = round(sum(deltas) / len(deltas), 2) if deltas else 0.0
    return AdaptiveValidation(project_id=project_id, avg_delta=avg_delta, comparisons=comparisons)


@router.get(
    "/projects/{project_id}/insights",
    response_model=InsightsReport,
    tags=["洞察与学习"],
    responses=RESPONSES_NO_SUBMISSIONS,
)
def project_insights(
    project_id: str,
    locale: str = Depends(get_locale),
) -> InsightsReport:
    """
    获取项目洞察报告。

    分析项目历史评分数据，识别强项、弱项和改进机会。

    支持 Accept-Language header 进行多语言响应。
    """
    ensure_data_dirs()
    submissions_all = [s for s in load_submissions() if s["project_id"] == project_id]
    if not submissions_all:
        raise HTTPException(status_code=404, detail=t("api.no_submissions", locale=locale))
    submissions = [s for s in submissions_all if _submission_is_scored(s)]
    if not submissions:
        raise HTTPException(status_code=404, detail="暂无已评分施组，请先点击“评分施组”。")
    insights = build_project_insights(submissions)
    return InsightsReport(project_id=project_id, **insights)


@router.post(
    "/projects/{project_id}/learning",
    response_model=LearningProfile,
    tags=["洞察与学习"],
    responses={**RESPONSES_401, **RESPONSES_NO_SUBMISSIONS},
)
def update_learning_profile(
    project_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> LearningProfile:
    """
    生成/更新项目学习画像。

    基于历史评分模式，生成项目特定的维度权重调整系数。
    后续评分将应用这些调整以提供更精准的评估。

    **需要 API Key 认证**

    支持 Accept-Language header 进行多语言响应。
    """
    ensure_data_dirs()
    submissions_all = [s for s in load_submissions() if s["project_id"] == project_id]
    if not submissions_all:
        raise HTTPException(status_code=404, detail=t("api.no_submissions", locale=locale))
    submissions = [s for s in submissions_all if _submission_is_scored(s)]
    if not submissions:
        raise HTTPException(status_code=404, detail="暂无已评分施组，请先点击“评分施组”。")
    profile = build_learning_profile(submissions)
    profiles = load_learning_profiles()
    record = {
        "project_id": project_id,
        "dimension_multipliers": profile["dimension_multipliers"],
        "rationale": profile["rationale"],
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    profiles = [p for p in profiles if p.get("project_id") != project_id]
    profiles.append(record)
    save_learning_profiles(profiles)
    return LearningProfile(**record)


@router.get(
    "/projects/{project_id}/learning",
    response_model=LearningProfile,
    tags=["洞察与学习"],
    responses=RESPONSES_NO_PROFILE,
)
def get_learning_profile(
    project_id: str,
    locale: str = Depends(get_locale),
) -> LearningProfile:
    """
    获取项目学习画像。

    返回项目已保存的学习画像，包含维度权重调整系数和分析依据。

    支持 Accept-Language header 进行多语言响应。
    """
    ensure_data_dirs()
    profiles = load_learning_profiles()
    for p in profiles:
        if p.get("project_id") == project_id:
            return LearningProfile(**p)
    raise HTTPException(status_code=404, detail=t("api.no_profile", locale=locale))


# ==================== 历史记录与趋势分析端点 ====================


@router.get(
    "/projects/{project_id}/history",
    response_model=ProjectScoreHistory,
    tags=["历史与趋势"],
    responses=RESPONSES_404,
)
def get_project_history(
    project_id: str,
    locale: str = Depends(get_locale),
) -> ProjectScoreHistory:
    """
    获取项目评分历史记录。

    返回项目的所有评分历史，按时间顺序排列。
    历史记录包含每次评分的总分、各维度得分和扣分项数量。

    支持 Accept-Language header 进行多语言响应。
    """
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))

    history = get_history(project_id)
    return ProjectScoreHistory(**history)


@router.get(
    "/projects/{project_id}/trend",
    response_model=TrendAnalysis,
    tags=["历史与趋势"],
    responses=RESPONSES_404,
)
def get_project_trend(
    project_id: str,
    locale: str = Depends(get_locale),
) -> TrendAnalysis:
    """
    获取项目评分趋势分析。

    基于历史评分数据进行趋势分析，包括：
    - 整体评分趋势（上升/下降/稳定）
    - 各维度得分趋势
    - 扣分项数量变化
    - 基于趋势的改进建议

    支持 Accept-Language header 进行多语言响应。
    """
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))

    # 加载配置获取维度名称
    config = load_config()
    dimension_names = {}
    for dim_id, dim_config in config.rubric.get("dimensions", {}).items():
        dimension_names[dim_id] = dim_config.get("name", dim_id)

    trend = analyze_trend(project_id, dimension_names)
    return TrendAnalysis(**trend)


# ==================== 自我学习与进化 ====================


@router.put(
    "/projects/{project_id}/context",
    response_model=ProjectContextOut,
    tags=["自我学习与进化"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def set_project_context(
    project_id: str,
    payload: ProjectContextIn,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> ProjectContextOut:
    """
    设置项目投喂包/项目背景文本（招标文件、清单、图纸、设计等合并内容）。
    用于自我学习时结合项目信息分析高分逻辑。
    """
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    ctx = load_project_context()
    ctx[project_id] = {
        "text": payload.text,
        "filename": payload.filename,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    save_project_context(ctx)
    return ProjectContextOut(
        project_id=project_id,
        text=ctx[project_id]["text"],
        filename=ctx[project_id].get("filename"),
        updated_at=ctx[project_id].get("updated_at"),
    )


@router.get(
    "/projects/{project_id}/context",
    response_model=ProjectContextOut,
    tags=["自我学习与进化"],
    responses=RESPONSES_404,
)
def get_project_context_endpoint(
    project_id: str,
    locale: str = Depends(get_locale),
) -> ProjectContextOut:
    """获取项目投喂包/项目背景文本。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    ctx = load_project_context()
    data = ctx.get(project_id)
    if not data:
        return ProjectContextOut(project_id=project_id, text="", filename=None, updated_at=None)
    return ProjectContextOut(
        project_id=project_id,
        text=data.get("text", ""),
        filename=data.get("filename"),
        updated_at=data.get("updated_at"),
    )


def _normalize_judge_scores_or_422(
    judge_scores: object,
    *,
    field_name: str = "judge_scores",
) -> List[float]:
    if not isinstance(judge_scores, list):
        raise HTTPException(status_code=422, detail=f"{field_name} 必须为数组。")
    judge_count = len(judge_scores)
    if judge_count not in (5, 7):
        raise HTTPException(status_code=422, detail=f"{field_name} 必须为 5 或 7 个评委得分。")
    normalized: List[float] = []
    for idx, value in enumerate(judge_scores, start=1):
        try:
            normalized.append(float(value))
        except Exception as exc:
            raise HTTPException(
                status_code=422,
                detail=f"{field_name}[{idx}] 不是有效数字：{exc}",
            )
    return normalized


def _normalize_judge_weights_or_422(
    judge_weights: object,
    *,
    expected_count: int,
) -> Optional[List[float]]:
    if judge_weights is None:
        return None
    if not isinstance(judge_weights, list):
        raise HTTPException(status_code=422, detail="judge_weights 必须为数组。")
    if len(judge_weights) != expected_count:
        raise HTTPException(
            status_code=422,
            detail=f"judge_weights 长度需与 judge_scores 一致（当前应为 {expected_count}）。",
        )
    normalized: List[float] = []
    for idx, value in enumerate(judge_weights, start=1):
        try:
            normalized.append(float(value))
        except Exception as exc:
            raise HTTPException(
                status_code=422,
                detail=f"judge_weights[{idx}] 不是有效数字：{exc}",
            )
    return normalized


def _normalize_qualitative_tags_or_422(
    tags_by_judge: object,
    *,
    expected_count: int,
) -> Optional[List[List[str]]]:
    if tags_by_judge is None:
        return None
    if not isinstance(tags_by_judge, list):
        raise HTTPException(status_code=422, detail="qualitative_tags_by_judge 必须为数组。")
    if len(tags_by_judge) != expected_count:
        raise HTTPException(
            status_code=422,
            detail=f"qualitative_tags_by_judge 长度需与 judge_scores 一致（当前应为 {expected_count}）。",
        )
    normalized: List[List[str]] = []
    for idx, tags in enumerate(tags_by_judge, start=1):
        if tags is None:
            normalized.append([])
            continue
        if not isinstance(tags, list):
            raise HTTPException(
                status_code=422,
                detail=f"qualitative_tags_by_judge[{idx}] 必须为字符串数组。",
            )
        clean_tags = [str(x).strip() for x in tags if str(x).strip()]
        normalized.append(clean_tags)
    return normalized


def _parse_judge_scores_form(judge_scores: str) -> List[float]:
    try:
        scores = json.loads(judge_scores)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"评委得分格式错误：{e}")
    return _normalize_judge_scores_or_422(scores)


def _assert_valid_final_score(final_score: float, *, score_scale_max: int = 100) -> None:
    scale = _normalize_score_scale_max(score_scale_max, default=100)
    if not (0 <= float(final_score) <= float(scale)):
        raise HTTPException(status_code=422, detail=f"最终得分应在 0～{scale} 之间。")


def _ground_truth_record_for_learning(
    record: Dict[str, object],
    *,
    default_score_scale_max: int,
) -> Dict[str, object]:
    score_scale_max = _normalize_score_scale_max(
        record.get("score_scale_max"),
        default=default_score_scale_max,
    )
    final_raw = _to_float_or_none(record.get("final_score_raw"))
    if final_raw is None:
        final_raw = _to_float_or_none(record.get("final_score"))
    if final_raw is None:
        final_raw = 0.0
    final_100 = _to_float_or_none(record.get("final_score_100"))
    if final_100 is None:
        final_100 = _convert_score_to_100(final_raw, score_scale_max)
    final_100 = float(final_100 if final_100 is not None else 0.0)
    judge_scores = record.get("judge_scores") or []
    judge_count = len(judge_scores) if isinstance(judge_scores, list) else 0
    normalized = dict(record)
    normalized["score_scale_max"] = score_scale_max
    normalized["final_score_raw"] = round(float(final_raw), 2)
    normalized["final_score_100"] = round(final_100, 2)
    normalized["final_score"] = round(final_100, 2)
    normalized["judge_count"] = judge_count
    return normalized


def _new_ground_truth_record(
    project_id: str,
    shigong_text: str,
    judge_scores: List[float],
    final_score: float,
    source: str,
    score_scale_max: int,
    judge_weights: Optional[List[float]] = None,
    qualitative_tags_by_judge: Optional[List[List[str]]] = None,
) -> Dict[str, object]:
    score_scale = _normalize_score_scale_max(score_scale_max, default=100)
    normalized_judge_scores = _normalize_judge_scores_or_422(judge_scores)
    normalized_judge_weights = _normalize_judge_weights_or_422(
        judge_weights,
        expected_count=len(normalized_judge_scores),
    )
    normalized_tags = _normalize_qualitative_tags_or_422(
        qualitative_tags_by_judge,
        expected_count=len(normalized_judge_scores),
    )
    final_raw = float(final_score)
    final_100 = _convert_score_to_100(final_raw, score_scale)
    final_100 = float(final_100 if final_100 is not None else 0.0)
    return {
        "id": str(uuid4()),
        "project_id": project_id,
        "shigong_text": shigong_text,
        "judge_scores": normalized_judge_scores,
        "judge_count": len(normalized_judge_scores),
        "score_scale_max": score_scale,
        "final_score": round(final_raw, 2),
        "final_score_raw": round(final_raw, 2),
        "final_score_100": round(final_100, 2),
        "judge_weights": normalized_judge_weights,
        "qualitative_tags_by_judge": normalized_tags,
        "source": source,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }


@router.post(
    "/projects/{project_id}/ground_truth",
    response_model=GroundTruthRecord,
    tags=["自我学习与进化"],
    responses={**RESPONSES_401, **RESPONSES_404, **RESPONSES_422},
)
def add_ground_truth(
    project_id: str,
    payload: GroundTruthCreate,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> GroundTruthRecord:
    """
    录入真实评标结果（如青天大模型在交易中心评标后的施组+5/7评委得分+最终得分）。
    用于系统学习高分逻辑并进化。
    """
    ensure_data_dirs()
    if len((payload.shigong_text or "").strip()) < 50:
        raise HTTPException(
            status_code=422,
            detail="施组全文过短，至少 50 字以便学习分析。",
        )
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    if project is None:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    score_scale_max = _resolve_project_score_scale_max(project)
    _assert_valid_final_score(payload.final_score, score_scale_max=score_scale_max)
    records = load_ground_truth()
    record = _new_ground_truth_record(
        project_id=project_id,
        shigong_text=payload.shigong_text,
        judge_scores=payload.judge_scores,
        final_score=payload.final_score,
        source=payload.source,
        score_scale_max=score_scale_max,
        judge_weights=payload.judge_weights,
        qualitative_tags_by_judge=payload.qualitative_tags_by_judge,
    )
    records.append(record)
    save_ground_truth(records)
    _sync_ground_truth_record_to_qingtian(project_id, record)
    record["feedback_closed_loop"] = _run_feedback_closed_loop_safe(
        project_id,
        locale=locale,
        trigger="ground_truth_add",
    )
    return GroundTruthRecord(**record)


@router.post(
    "/projects/{project_id}/ground_truth/from_submission",
    response_model=GroundTruthRecord,
    tags=["自我学习与进化"],
    responses={**RESPONSES_401, **RESPONSES_404, **RESPONSES_422},
)
def add_ground_truth_from_submission(
    project_id: str,
    payload: GroundTruthFromSubmissionCreate,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> GroundTruthRecord:
    """从“步骤4已上传施组”中选择一份文件录入真实评标结果，避免重复上传。"""
    ensure_data_dirs()
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    if project is None:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    score_scale_max = _resolve_project_score_scale_max(project)
    _assert_valid_final_score(payload.final_score, score_scale_max=score_scale_max)

    submission_id = str(payload.submission_id or "").strip()
    submissions = load_submissions()
    submission = next(
        (
            s
            for s in submissions
            if str(s.get("id")) == submission_id and str(s.get("project_id")) == project_id
        ),
        None,
    )
    if not submission:
        raise HTTPException(status_code=404, detail="未找到对应施组，请先在步骤4上传施组。")

    shigong_text = str(submission.get("text") or "").strip()
    if len(shigong_text) < 50:
        raise HTTPException(status_code=422, detail="该施组文本过短，暂不支持录入真实评标。")

    record = _new_ground_truth_record(
        project_id=project_id,
        shigong_text=shigong_text,
        judge_scores=[float(x) for x in payload.judge_scores],
        final_score=float(payload.final_score),
        source=payload.source,
        score_scale_max=score_scale_max,
        judge_weights=None,
        qualitative_tags_by_judge=payload.qualitative_tags_by_judge,
    )
    record["source_submission_id"] = submission_id
    record["source_submission_filename"] = submission.get("filename")

    records = load_ground_truth()
    records.append(record)
    save_ground_truth(records)
    _sync_ground_truth_record_to_qingtian(project_id, record)
    record["feedback_closed_loop"] = _run_feedback_closed_loop_safe(
        project_id,
        locale=locale,
        trigger="ground_truth_add",
    )
    return GroundTruthRecord(**record)


@router.post(
    "/projects/{project_id}/ground_truth/from_file",
    response_model=GroundTruthRecord,
    tags=["自我学习与进化"],
    responses={**RESPONSES_401, **RESPONSES_404, **RESPONSES_422},
)
async def add_ground_truth_from_file(
    project_id: str,
    file: UploadFile = File(...),
    judge_scores: str = Form(..., description="5或7个评委得分，JSON 数组如 [1,2,3,4,5]"),
    final_score: float = Form(...),
    source: str = Form("青天大模型"),
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> GroundTruthRecord:
    """
    通过上传施组文件录入一条真实评标。后端解析文件为文本后保存。
    用于界面简洁录入：选文件 + 评委分 + 最终分即可。
    """
    ensure_data_dirs()
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    if project is None:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    score_scale_max = _resolve_project_score_scale_max(project)
    judge_scores_list = _parse_judge_scores_form(judge_scores)
    _assert_valid_final_score(final_score, score_scale_max=score_scale_max)
    content = await file.read()
    try:
        shigong_text = _read_uploaded_file_content(content, file.filename or "")
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    if len(shigong_text.strip()) < 50:
        raise HTTPException(status_code=422, detail="施组全文过短，至少 50 字以便学习分析。")
    records = load_ground_truth()
    record = _new_ground_truth_record(
        project_id=project_id,
        shigong_text=shigong_text,
        judge_scores=judge_scores_list,
        final_score=final_score,
        source=source,
        score_scale_max=score_scale_max,
        judge_weights=None,
        qualitative_tags_by_judge=None,
    )
    records.append(record)
    save_ground_truth(records)
    _sync_ground_truth_record_to_qingtian(project_id, record)
    record["feedback_closed_loop"] = _run_feedback_closed_loop_safe(
        project_id,
        locale=locale,
        trigger="ground_truth_add",
    )
    return GroundTruthRecord(**record)


@router.post(
    "/projects/{project_id}/ground_truth/from_files",
    response_model=GroundTruthBatchResponse,
    tags=["自我学习与进化"],
    responses={**RESPONSES_401, **RESPONSES_404, **RESPONSES_422},
)
async def add_ground_truth_from_files(
    project_id: str,
    files: List[UploadFile] = File(...),
    judge_scores: str = Form(..., description="5或7个评委得分，JSON 数组如 [1,2,3,4,5]"),
    final_score: float = Form(...),
    source: str = Form("青天大模型"),
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> GroundTruthBatchResponse:
    """
    通过上传多个施组文件批量录入真实评标。后端将逐个解析并保存。
    """
    ensure_data_dirs()
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    if project is None:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    score_scale_max = _resolve_project_score_scale_max(project)
    judge_scores_list = _parse_judge_scores_form(judge_scores)
    _assert_valid_final_score(final_score, score_scale_max=score_scale_max)

    items: List[Dict[str, object]] = []
    success_records: List[Dict[str, object]] = []
    for file in files:
        filename = file.filename or "unknown"
        content = await file.read()
        try:
            shigong_text = _read_uploaded_file_content(content, filename)
            if len(shigong_text.strip()) < 50:
                raise ValueError("施组全文过短，至少 50 字以便学习分析。")
            record = _new_ground_truth_record(
                project_id=project_id,
                shigong_text=shigong_text,
                judge_scores=judge_scores_list,
                final_score=final_score,
                source=source,
                score_scale_max=score_scale_max,
                judge_weights=None,
                qualitative_tags_by_judge=None,
            )
            success_records.append(record)
            items.append(
                {
                    "filename": filename,
                    "ok": True,
                    "record": record,
                    "detail": None,
                }
            )
        except Exception as e:
            items.append(
                {
                    "filename": filename,
                    "ok": False,
                    "record": None,
                    "detail": str(e),
                }
            )

    if success_records:
        records = load_ground_truth()
        records.extend(success_records)
        save_ground_truth(records)
        for item in items:
            record = item.get("record")
            if item.get("ok") and isinstance(record, dict):
                try:
                    _sync_ground_truth_record_to_qingtian(project_id, record)
                except Exception as e:
                    item["detail"] = f"已保存，但同步青天失败：{e}"
        closed_loop_result = _run_feedback_closed_loop_safe(
            project_id,
            locale=locale,
            trigger="ground_truth_batch_add",
        )
        for item in items:
            if item.get("ok") and isinstance(item.get("record"), dict):
                item["record"]["feedback_closed_loop"] = closed_loop_result

    success_count = sum(1 for item in items if item.get("ok"))
    failed_count = len(items) - success_count
    return GroundTruthBatchResponse(
        project_id=project_id,
        total_files=len(items),
        success_count=success_count,
        failed_count=failed_count,
        items=items,
    )


@router.get(
    "/projects/{project_id}/ground_truth",
    response_model=List[GroundTruthRecord],
    tags=["自我学习与进化"],
    responses=RESPONSES_404,
)
def list_ground_truth(
    project_id: str,
    locale: str = Depends(get_locale),
) -> List[GroundTruthRecord]:
    """列出本项目的所有真实评标记录。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    records = [r for r in load_ground_truth() if r.get("project_id") == project_id]
    return [GroundTruthRecord(**r) for r in records]


@router.delete(
    "/projects/{project_id}/ground_truth/{record_id}",
    status_code=204,
    tags=["自我学习与进化"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def delete_ground_truth(
    project_id: str,
    record_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> None:
    """删除指定项目下的一条真实评标记录。"""
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    records = load_ground_truth()
    if not any(r.get("id") == record_id and r.get("project_id") == project_id for r in records):
        raise HTTPException(status_code=404, detail="真实评标记录不存在")
    removed = next(
        (r for r in records if r.get("id") == record_id and r.get("project_id") == project_id),
        None,
    )
    records = [
        r for r in records if not (r.get("id") == record_id and r.get("project_id") == project_id)
    ]
    save_ground_truth(records)
    if removed is not None:
        gt_id = str(removed.get("id") or "")
        qtrs = load_qingtian_results()
        linked_submission_ids = {
            str(q.get("submission_id") or "")
            for q in qtrs
            if str((q.get("raw_payload") or {}).get("ground_truth_record_id") or "") == gt_id
        }
        qtrs = [
            q
            for q in qtrs
            if str((q.get("raw_payload") or {}).get("ground_truth_record_id") or "") != gt_id
        ]
        save_qingtian_results(qtrs)

        submissions = load_submissions()
        auto_submission_ids = {
            str(s.get("id") or "")
            for s in submissions
            if str(s.get("source_ground_truth_id") or "") == gt_id
            and str(s.get("project_id")) == project_id
        }
        remove_submission_ids = linked_submission_ids.union(auto_submission_ids)
        if remove_submission_ids:
            submissions = [
                s for s in submissions if str(s.get("id") or "") not in remove_submission_ids
            ]
            save_submissions(submissions)
            reports = load_score_reports()
            reports = [
                r for r in reports if str(r.get("submission_id") or "") not in remove_submission_ids
            ]
            save_score_reports(reports)
            units = load_evidence_units()
            units = [
                u for u in units if str(u.get("submission_id") or "") not in remove_submission_ids
            ]
            save_evidence_units(units)
        _refresh_project_reflection_objects(project_id)


@router.post(
    "/projects/{project_id}/evolve",
    response_model=EvolutionReport,
    tags=["自我学习与进化"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def evolve_project(
    project_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> EvolutionReport:
    """
    根据已录入的真实评标结果进行学习，生成高分逻辑与编制指导并保存。
    执行后可通过「编制指导」接口或页面查看。
    """
    ensure_data_dirs()
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    if project is None:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    project_score_scale = _resolve_project_score_scale_max(project)
    records_raw = [r for r in load_ground_truth() if r.get("project_id") == project_id]
    records = [
        _ground_truth_record_for_learning(
            r if isinstance(r, dict) else {},
            default_score_scale_max=project_score_scale,
        )
        for r in records_raw
    ]
    ctx_data = load_project_context().get(project_id) or {}
    project_context = (ctx_data.get("text") or "").strip()
    materials_text = _merge_materials_text(project_id)
    if materials_text:
        project_context = (
            (project_context + "\n\n" + materials_text) if project_context else materials_text
        )
    report = build_evolution_report(project_id, records, project_context)
    enhanced = enhance_evolution_report_with_llm(project_id, report, records, project_context)
    if enhanced is not None:
        report["high_score_logic"] = enhanced.get("high_score_logic", report["high_score_logic"])
        report["writing_guidance"] = enhanced.get("writing_guidance", report["writing_guidance"])
        report["sample_count"] = enhanced.get("sample_count", report["sample_count"])
        report["updated_at"] = enhanced.get("updated_at", report["updated_at"])
        report["enhanced_by"] = enhanced.get("enhanced_by")  # 可追溯：spark | openai | gemini
        # 保留规则版产出的 scoring_evolution、compilation_instructions（LLM 仅增强文字部分）
    reports = load_evolution_reports()
    reports[project_id] = report
    save_evolution_reports(reports)
    return EvolutionReport(**report)


@router.post(
    "/projects/{project_id}/evolve/ollama_preview",
    response_model=OllamaEvolutionPreviewResponse,
    tags=["自我学习与进化"],
    responses={**RESPONSES_401, **RESPONSES_404},
)
def preview_project_evolution_with_ollama(
    project_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> OllamaEvolutionPreviewResponse:
    """
    手动触发 Ollama 增强预览，仅返回临时结果，不写入正式 evolution_reports。
    规则版报告仍由 build_evolution_report 生成，Ollama 失败时返回 fallback 信息。
    """
    ensure_data_dirs()
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    if project is None:
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    project_score_scale = _resolve_project_score_scale_max(project)
    records_raw = [r for r in load_ground_truth() if r.get("project_id") == project_id]
    records = [
        _ground_truth_record_for_learning(
            r if isinstance(r, dict) else {},
            default_score_scale_max=project_score_scale,
        )
        for r in records_raw
    ]
    ctx_data = load_project_context().get(project_id) or {}
    project_context = (ctx_data.get("text") or "").strip()
    materials_text = _merge_materials_text(project_id)
    if materials_text:
        project_context = (
            (project_context + "\n\n" + materials_text) if project_context else materials_text
        )
    report = build_evolution_report(project_id, records, project_context)
    preview = preview_evolution_report_with_ollama(project_id, report, records, project_context)
    return OllamaEvolutionPreviewResponse(**preview)


@router.get(
    "/projects/{project_id}/writing_guidance",
    response_model=WritingGuidance,
    tags=["自我学习与进化"],
    responses=RESPONSES_404,
)
def get_writing_guidance(
    project_id: str,
    locale: str = Depends(get_locale),
) -> WritingGuidance:
    """
    获取编制指导。若已执行过「学习进化」则返回学习结果；否则返回引导说明。
    """
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    reports = load_evolution_reports()
    data = reports.get(project_id)
    if data:
        return WritingGuidance(
            project_id=project_id,
            guidance=data.get("writing_guidance", []),
            high_score_logic=data.get("high_score_logic", []),
            sample_count=data.get("sample_count", 0),
            updated_at=data.get("updated_at"),
        )
    return WritingGuidance(
        project_id=project_id,
        guidance=[
            "请先录入「真实评标结果」（施组+5/7评委得分+最终得分），再点击「学习进化」生成编制指导。"
        ],
        high_score_logic=[],
        sample_count=0,
        updated_at=None,
    )


@router.get(
    "/projects/{project_id}/scoring_context",
    tags=["自我学习与进化"],
    responses=RESPONSES_404,
)
def get_scoring_context(project_id: str) -> Dict[str, object]:
    """
    返回当前项目评分生效的维度权重与总分缩放，用于诊断进化是否生效。
    source: evolution | expert_profile | learning_profile | none
    """
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail="项目不存在")
    multipliers, profile_snapshot, _ = _resolve_project_scoring_context(project_id)
    scale = _get_evolution_total_score_scale(project_id)
    source = "none"
    if profile_snapshot:
        source = "expert_profile"
    elif (
        load_evolution_reports()
        .get(project_id, {})
        .get("scoring_evolution", {})
        .get("dimension_multipliers")
    ):
        source = "evolution"
    elif any(p.get("project_id") == project_id for p in load_learning_profiles()):
        source = "learning_profile"
    return {
        "project_id": project_id,
        "source": source,
        "dimension_multipliers": multipliers,
        "total_score_scale": scale,
        "has_non_default_multipliers": any(
            abs(float(multipliers.get(d, 1.0)) - 1.0) > 0.01 for d in DIMENSION_IDS
        ),
    }


@router.get(
    "/projects/{project_id}/compilation_instructions",
    response_model=CompilationInstructions,
    tags=["自我学习与进化"],
    responses=RESPONSES_404,
)
def get_compilation_instructions(
    project_id: str,
    locale: str = Depends(get_locale),
) -> CompilationInstructions:
    """
    获取编制系统指令。基于学习进化结果，用于约束施组编制输出（必备章节、图表、要素及禁止表述）。
    可导出为系统指令，强制要求按此输出内容和图表。
    """
    ensure_data_dirs()
    projects = load_projects()
    if not any(p["id"] == project_id for p in projects):
        raise HTTPException(status_code=404, detail=t("api.project_not_found", locale=locale))
    reports = load_evolution_reports()
    data = reports.get(project_id)
    ci = data.get("compilation_instructions") if data else None
    if ci:
        return CompilationInstructions(
            project_id=project_id,
            required_sections=ci.get("required_sections", []),
            required_charts_images=ci.get("required_charts_images", []),
            mandatory_elements=ci.get("mandatory_elements", []),
            forbidden_patterns=ci.get("forbidden_patterns", []),
            guidance_items=ci.get("guidance_items", []),
            high_score_summary=ci.get("high_score_summary", []),
        )
    return CompilationInstructions(
        project_id=project_id,
        required_sections=[],
        required_charts_images=[],
        mandatory_elements=[],
        forbidden_patterns=[],
        guidance_items=["请先录入真实评标并执行「学习进化」后，编制系统指令将基于学习结果生成。"],
        high_score_summary=[],
    )


@router.post(
    "/tools/parse_text",
    tags=["工具"],
    include_in_schema=False,
)
async def parse_file_to_text(file: UploadFile = File(...)) -> Dict[str, str]:
    """解析上传文件为纯文本。支持 .txt、.docx、.pdf、.json、.xlsx/.xls、.dxf"""
    content = await file.read()
    try:
        text = _read_uploaded_file_content(content, file.filename or "")
        return {"text": text}
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))


# API 兼容路由（与执行文档中的 /api/projects/... 路径保持一致）
compat_router = APIRouter(prefix="/api")


@compat_router.get("/scoring/factors", response_model=ScoringFactorsResponse, tags=["系统状态"])
def compat_scoring_factors(
    project_id: Optional[str] = Query(None),
    locale: str = Depends(get_locale),
) -> ScoringFactorsResponse:
    return scoring_factors(project_id=project_id, locale=locale)


@compat_router.get(
    "/scoring/factors/markdown", response_model=ScoringFactorsMarkdownResponse, tags=["系统状态"]
)
def compat_scoring_factors_markdown(
    project_id: Optional[str] = Query(None),
    locale: str = Depends(get_locale),
) -> ScoringFactorsMarkdownResponse:
    return scoring_factors_markdown(project_id=project_id, locale=locale)


@compat_router.get("/system/self_check", response_model=SelfCheckResponse, tags=["系统状态"])
def compat_system_self_check(
    project_id: Optional[str] = Query(None),
) -> SelfCheckResponse:
    return system_self_check(project_id=project_id)


@compat_router.get("/system/data_hygiene", response_model=DataHygieneResponse, tags=["系统状态"])
def compat_system_data_hygiene() -> DataHygieneResponse:
    return system_data_hygiene()


@compat_router.post(
    "/system/data_hygiene/repair",
    response_model=DataHygieneResponse,
    tags=["系统状态"],
    responses={**RESPONSES_401},
)
def compat_repair_system_data_hygiene(
    api_key: Optional[str] = Depends(verify_api_key),
) -> DataHygieneResponse:
    return repair_system_data_hygiene(api_key=api_key)


@compat_router.get(
    "/projects/{project_id}/analysis_bundle",
    response_model=AnalysisBundleResponse,
    tags=["洞察与学习"],
)
def compat_project_analysis_bundle(
    project_id: str,
    locale: str = Depends(get_locale),
) -> AnalysisBundleResponse:
    return project_analysis_bundle(project_id=project_id, locale=locale)


@compat_router.get("/projects/{project_id}/analysis_bundle.md", tags=["洞察与学习"])
def compat_project_analysis_bundle_markdown_file(
    project_id: str,
    locale: str = Depends(get_locale),
) -> Response:
    return project_analysis_bundle_markdown_file(project_id=project_id, locale=locale)


@compat_router.get(
    "/projects/{project_id}/materials/depth_report",
    response_model=MaterialDepthReportResponse,
    tags=["项目管理"],
)
def compat_material_depth_report(
    project_id: str,
    locale: str = Depends(get_locale),
) -> MaterialDepthReportResponse:
    return get_material_depth_report(project_id=project_id, locale=locale)


@compat_router.get(
    "/projects/{project_id}/materials/depth_report/markdown",
    response_model=MaterialDepthReportMarkdownResponse,
    tags=["项目管理"],
)
def compat_material_depth_report_markdown(
    project_id: str,
    locale: str = Depends(get_locale),
) -> MaterialDepthReportMarkdownResponse:
    return get_material_depth_report_markdown(project_id=project_id, locale=locale)


@compat_router.get("/projects/{project_id}/materials/depth_report.md", tags=["项目管理"])
def compat_material_depth_report_markdown_file(
    project_id: str,
    locale: str = Depends(get_locale),
) -> Response:
    return download_material_depth_report_markdown(project_id=project_id, locale=locale)


@compat_router.get(
    "/projects/{project_id}/expert-profile",
    response_model=ProjectExpertProfileResponse,
    tags=["项目管理"],
)
def compat_get_project_expert_profile(
    project_id: str, locale: str = Depends(get_locale)
) -> ProjectExpertProfileResponse:
    return get_project_expert_profile(project_id=project_id, locale=locale)


@compat_router.put(
    "/projects/{project_id}/expert-profile",
    response_model=ProjectExpertProfileResponse,
    tags=["项目管理"],
)
def compat_update_project_expert_profile(
    project_id: str,
    payload: ExpertProfileUpdate,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> ProjectExpertProfileResponse:
    return update_project_expert_profile(
        project_id=project_id,
        payload=payload,
        api_key=api_key,
        locale=locale,
    )


@compat_router.post(
    "/projects/{project_id}/rescore", response_model=RescoreResponse, tags=["项目管理"]
)
def compat_rescore_project_submissions(
    project_id: str,
    payload: RescoreRequest,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> RescoreResponse:
    return rescore_project_submissions(
        project_id=project_id,
        payload=payload,
        api_key=api_key,
        locale=locale,
    )


@compat_router.get(
    "/projects/{project_id}/submissions",
    response_model=list[SubmissionRecord] | ProjectPreScoreListResponse,
    tags=["施组提交"],
)
def compat_list_submissions(
    project_id: str,
    with_: Optional[str] = Query(None, alias="with"),
):
    return list_submissions(project_id=project_id, with_=with_)


@compat_router.get(
    "/submissions/{submission_id}/reports/latest",
    response_model=LatestReportResponse,
    tags=["施组提交"],
)
def compat_latest_submission_report(submission_id: str) -> LatestReportResponse:
    return get_latest_submission_report(submission_id=submission_id)


@compat_router.get(
    "/projects/{project_id}/submissions/{submission_id}/evidence_trace",
    response_model=EvidenceTraceResponse,
    tags=["施组提交"],
)
def compat_submission_evidence_trace(
    project_id: str,
    submission_id: str,
    locale: str = Depends(get_locale),
) -> EvidenceTraceResponse:
    return get_submission_evidence_trace(
        project_id=project_id,
        submission_id=submission_id,
        locale=locale,
    )


@compat_router.get(
    "/projects/{project_id}/submissions/{submission_id}/evidence_trace/markdown",
    response_model=EvidenceTraceMarkdownResponse,
    tags=["施组提交"],
)
def compat_submission_evidence_trace_markdown(
    project_id: str,
    submission_id: str,
    locale: str = Depends(get_locale),
) -> EvidenceTraceMarkdownResponse:
    return get_submission_evidence_trace_markdown(
        project_id=project_id,
        submission_id=submission_id,
        locale=locale,
    )


@compat_router.get(
    "/projects/{project_id}/submissions/{submission_id}/evidence_trace.md",
    tags=["施组提交"],
)
def compat_download_submission_evidence_trace_markdown(
    project_id: str,
    submission_id: str,
    locale: str = Depends(get_locale),
) -> Response:
    return download_submission_evidence_trace_markdown(
        project_id=project_id,
        submission_id=submission_id,
        locale=locale,
    )


@compat_router.get(
    "/projects/{project_id}/evidence_trace/latest",
    response_model=EvidenceTraceResponse,
    tags=["施组提交"],
)
def compat_latest_submission_evidence_trace(
    project_id: str,
    locale: str = Depends(get_locale),
) -> EvidenceTraceResponse:
    return get_latest_submission_evidence_trace(project_id=project_id, locale=locale)


@compat_router.get(
    "/projects/{project_id}/submissions/{submission_id}/scoring_basis",
    response_model=ScoringBasisResponse,
    tags=["施组提交"],
)
def compat_get_submission_scoring_basis(
    project_id: str,
    submission_id: str,
    locale: str = Depends(get_locale),
) -> ScoringBasisResponse:
    return get_submission_scoring_basis(
        project_id=project_id,
        submission_id=submission_id,
        locale=locale,
    )


@compat_router.get(
    "/projects/{project_id}/scoring_basis/latest",
    response_model=ScoringBasisResponse,
    tags=["施组提交"],
)
def compat_get_latest_submission_scoring_basis(
    project_id: str,
    locale: str = Depends(get_locale),
) -> ScoringBasisResponse:
    return get_latest_submission_scoring_basis(project_id=project_id, locale=locale)


@compat_router.post(
    "/submissions/{submission_id}/qingtian-results",
    response_model=QingTianResultRecord,
    tags=["施组提交"],
)
def compat_ingest_qingtian_result(
    submission_id: str,
    payload: QingTianResultCreate,
    api_key: Optional[str] = Depends(verify_api_key),
) -> QingTianResultRecord:
    return ingest_qingtian_result(submission_id=submission_id, payload=payload, api_key=api_key)


@compat_router.get(
    "/submissions/{submission_id}/qingtian-results/latest",
    response_model=QingTianResultRecord,
    tags=["施组提交"],
)
def compat_latest_qingtian_result(submission_id: str) -> QingTianResultRecord:
    return get_latest_qingtian_result(submission_id=submission_id)


@compat_router.post(
    "/projects/{project_id}/reflection/auto_run",
    response_model=ReflectionAutoRunResponse,
    tags=["洞察与学习"],
)
def compat_auto_run_reflection(
    project_id: str,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> ReflectionAutoRunResponse:
    return auto_run_reflection_pipeline(project_id=project_id, api_key=api_key, locale=locale)


@compat_router.post(
    "/projects/{project_id}/ground_truth/from_submission",
    response_model=GroundTruthRecord,
    tags=["自我学习与进化"],
)
def compat_add_ground_truth_from_submission(
    project_id: str,
    payload: GroundTruthFromSubmissionCreate,
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> GroundTruthRecord:
    return add_ground_truth_from_submission(
        project_id=project_id,
        payload=payload,
        api_key=api_key,
        locale=locale,
    )


@compat_router.post(
    "/projects/{project_id}/ground_truth/from_files",
    response_model=GroundTruthBatchResponse,
    tags=["自我学习与进化"],
)
async def compat_add_ground_truth_from_files(
    project_id: str,
    files: List[UploadFile] = File(...),
    judge_scores: str = Form(...),
    final_score: float = Form(...),
    source: str = Form("青天大模型"),
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
) -> GroundTruthBatchResponse:
    return await add_ground_truth_from_files(
        project_id=project_id,
        files=files,
        judge_scores=judge_scores,
        final_score=final_score,
        source=source,
        api_key=api_key,
        locale=locale,
    )


@compat_router.get(
    "/projects/{project_id}/evaluation",
    response_model=ProjectEvaluationResponse,
    tags=["洞察与学习"],
)
def compat_project_evaluation(
    project_id: str,
    locale: str = Depends(get_locale),
) -> ProjectEvaluationResponse:
    return get_project_evaluation(project_id=project_id, locale=locale)


@compat_router.get(
    "/evaluation/summary", response_model=EvaluationSummaryResponse, tags=["洞察与学习"]
)
def compat_evaluation_summary() -> EvaluationSummaryResponse:
    return get_evaluation_summary()


# 注册 API v1 路由
app.include_router(router)
app.include_router(compat_router)


@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    return Response(status_code=204)


@app.get("/apple-touch-icon-precomposed.png", include_in_schema=False)
async def apple_touch_icon():
    return Response(status_code=204)


@app.get("/apple-touch-icon.png", include_in_schema=False)
async def apple_touch_icon_alt():
    return Response(status_code=204)


@app.post("/web/create_project", include_in_schema=False)
def web_create_project(
    name: str = Form(...),
    api_key: Optional[str] = Depends(verify_api_key),
):
    clean_name = (name or "").strip()
    if not clean_name:
        return RedirectResponse(
            url="/?create_error=" + quote_plus("项目名称不能为空"), status_code=303
        )
    try:
        rec = create_project(ProjectCreate(name=clean_name), api_key=api_key)
    except HTTPException as exc:
        detail = str(getattr(exc, "detail", "创建失败"))
        return RedirectResponse(url="/?create_error=" + quote_plus(detail), status_code=303)
    project_id_value = str(getattr(rec, "id", "") or "")
    return RedirectResponse(
        url="/?create_ok=" + quote_plus(clean_name) + "&project_id=" + quote_plus(project_id_value),
        status_code=303,
    )


@app.post("/web/delete_project", include_in_schema=False)
def web_delete_project(
    project_id: str = Form(""),
    api_key: Optional[str] = Depends(verify_api_key),
):
    pid = (project_id or "").strip()
    if not pid:
        return RedirectResponse(
            url="/?msg_type=error&msg=" + quote_plus("删除失败：请先选择项目"),
            status_code=303,
        )
    try:
        result = _delete_project_cascade(pid, locale="zh")
    except HTTPException as exc:
        detail = str(getattr(exc, "detail", "删除失败"))
        return RedirectResponse(
            url="/?msg_type=error&msg=" + quote_plus("删除失败：" + detail),
            status_code=303,
        )
    return RedirectResponse(
        url="/?msg_type=success&msg="
        + quote_plus("项目已删除：" + str(result.get("project_name") or pid)),
        status_code=303,
    )


@app.post("/web/upload_materials", include_in_schema=False)
def web_upload_materials(
    project_id: str = Form(""),
    material_type: str = Form(MATERIAL_TYPE_DEFAULT),
    file: List[UploadFile] = File(default=[]),
    api_key: Optional[str] = Depends(verify_api_key),
):
    pid = (project_id or "").strip()
    files = file or []
    if not pid:
        return RedirectResponse(
            url="/?msg_type=error&msg="
            + quote_plus("上传资料失败：请先选择项目")
            + "#section-materials",
            status_code=303,
        )
    if not files:
        return RedirectResponse(
            url="/?project_id="
            + quote_plus(pid)
            + "&msg_type=error&msg="
            + quote_plus("上传资料失败：未选择文件")
            + "#section-materials",
            status_code=303,
        )
    ok_count = 0
    fail_count = 0
    first_error = ""
    for f in files:
        try:
            upload_material(
                project_id=pid,
                file=f,
                material_type=material_type,
                api_key=api_key,
                locale="zh",
            )
            ok_count += 1
        except Exception as exc:  # noqa: BLE001 - web fallback should keep processing
            fail_count += 1
            if not first_error:
                first_error = str(exc)
    msg = f"资料上传完成：成功 {ok_count}，失败 {fail_count}"
    if fail_count > 0 and first_error:
        msg += f"；首个错误：{first_error}"
    return RedirectResponse(
        url="/?project_id="
        + quote_plus(pid)
        + "&msg_type="
        + ("error" if fail_count > 0 else "success")
        + "&msg="
        + quote_plus(msg)
        + "#section-materials",
        status_code=303,
    )


def _web_upload_redirect_url(project_id: str, message: str, anchor: str = "") -> str:
    pid = (project_id or "").strip()
    base = "/?msg_type=error&msg=" + quote_plus(message)
    if pid:
        base = "/?project_id=" + quote_plus(pid) + "&msg_type=error&msg=" + quote_plus(message)
    anchor_part = str(anchor or "").strip()
    if anchor_part and not anchor_part.startswith("#"):
        anchor_part = "#" + anchor_part
    return base + anchor_part


@app.api_route(
    "/web/upload_materials",
    methods=["GET", "HEAD", "PUT", "PATCH", "DELETE", "OPTIONS"],
    include_in_schema=False,
)
def web_upload_materials_get_fallback(project_id: str = ""):
    """
    兼容误触发非 POST 的场景，避免直接 405 中断用户流程。
    """
    return RedirectResponse(
        url=_web_upload_redirect_url(
            project_id, "请在主页选择文件后点击“上传资料”提交。", "#section-materials"
        ),
        status_code=303,
    )


@app.post("/web/upload_shigong", include_in_schema=False)
def web_upload_shigong(
    project_id: str = Form(""),
    file: List[UploadFile] = File(default=[]),
    api_key: Optional[str] = Depends(verify_api_key),
):
    pid = (project_id or "").strip()
    files = file or []
    if not pid:
        return RedirectResponse(
            url="/?msg_type=error&msg="
            + quote_plus("上传施组失败：请先选择项目")
            + "#section-shigong",
            status_code=303,
        )
    if not files:
        return RedirectResponse(
            url="/?project_id="
            + quote_plus(pid)
            + "&msg_type=error&msg="
            + quote_plus("上传施组失败：未选择文件")
            + "#section-shigong",
            status_code=303,
        )
    ok_count = 0
    fail_count = 0
    first_error = ""
    for f in files:
        try:
            upload_shigong(project_id=pid, file=f, api_key=api_key, locale="zh")
            ok_count += 1
        except Exception as exc:  # noqa: BLE001 - web fallback should keep processing
            fail_count += 1
            if not first_error:
                first_error = str(exc)
    msg = f"施组上传完成：成功 {ok_count}，失败 {fail_count}"
    if fail_count > 0 and first_error:
        msg += f"；首个错误：{first_error}"
    return RedirectResponse(
        url="/?project_id="
        + quote_plus(pid)
        + "&msg_type="
        + ("error" if fail_count > 0 else "success")
        + "&msg="
        + quote_plus(msg)
        + "#section-shigong",
        status_code=303,
    )


@app.post("/web/score_shigong", include_in_schema=False)
def web_score_shigong(
    project_id: str = Form(""),
    score_scale_max: int = Form(DEFAULT_SCORE_SCALE_MAX),
    api_key: Optional[str] = Depends(verify_api_key),
    locale: str = Depends(get_locale),
):
    pid = (project_id or "").strip()
    if not pid:
        return RedirectResponse(
            url="/?msg_type=error&msg="
            + quote_plus("施组评分失败：请先选择项目")
            + "#section-shigong",
            status_code=303,
        )
    try:
        result = rescore_project_submissions(
            project_id=pid,
            payload=RescoreRequest(
                scoring_engine_version="v2",
                scope="project",
                score_scale_max=score_scale_max,
                rebuild_anchors=False,
                rebuild_requirements=False,
                retrain_calibrator=False,
                force_unlock=False,
            ),
            api_key=api_key,
            locale=locale,
        )
        updated = int(
            getattr(
                result,
                "updated_submissions",
                getattr(result, "reports_generated", getattr(result, "submission_count", 0)),
            )
        )
        msg = f"施组评分完成：已重算 {updated} 份"
        gate_obj = getattr(result, "material_utilization_gate", {}) or {}
        blocked_count = int(_to_float_or_none(gate_obj.get("blocked_submissions")) or 0)
        warn_count = int(_to_float_or_none(gate_obj.get("warn_submissions")) or 0)
        if blocked_count > 0:
            msg += f"；资料门禁阻断 {blocked_count} 份"
        elif warn_count > 0:
            msg += f"；资料门禁预警 {warn_count} 份"
        return RedirectResponse(
            url="/?project_id="
            + quote_plus(pid)
            + "&msg_type=success&msg="
            + quote_plus(msg)
            + "#section-shigong",
            status_code=303,
        )
    except Exception as exc:  # noqa: BLE001 - web fallback should keep processing
        return RedirectResponse(
            url="/?project_id="
            + quote_plus(pid)
            + "&msg_type=error&msg="
            + quote_plus("施组评分失败：" + str(exc))
            + "#section-shigong",
            status_code=303,
        )


@app.api_route(
    "/web/upload_shigong",
    methods=["GET", "HEAD", "PUT", "PATCH", "DELETE", "OPTIONS"],
    include_in_schema=False,
)
def web_upload_shigong_get_fallback(project_id: str = ""):
    return RedirectResponse(
        url=_web_upload_redirect_url(
            project_id, "请在主页选择文件后点击“上传施组”提交。", "#section-shigong"
        ),
        status_code=303,
    )


@app.api_route(
    "/web/score_shigong",
    methods=["GET", "HEAD", "PUT", "PATCH", "DELETE", "OPTIONS"],
    include_in_schema=False,
)
def web_score_shigong_get_fallback(project_id: str = ""):
    return RedirectResponse(
        url=_web_upload_redirect_url(
            project_id, "请在主页选择项目后点击“评分施组”提交。", "#section-shigong"
        ),
        status_code=303,
    )


@app.head("/", include_in_schema=False)
def index_head() -> Response:
    """
    某些浏览器会先发 HEAD 检测主页可达性，避免返回 405 干扰用户判断。
    """
    return Response(status_code=200)


@app.get("/", tags=["系统状态"], include_in_schema=False)
def index(
    create_ok: Optional[str] = Query(None),
    create_error: Optional[str] = Query(None),
    project_id: Optional[str] = Query(None),
    msg: Optional[str] = Query(None),
    msg_type: Optional[str] = Query(None),
) -> Response:
    ensure_data_dirs()
    projects = load_projects()
    if not os.environ.get("PYTEST_CURRENT_TEST"):
        active_projects = [
            p
            for p in projects
            if str(p.get("id") or "") != "p1" and not str(p.get("name") or "").startswith("E2E_")
        ]
        if not active_projects:
            recovered = _recover_latest_orphan_project(projects)
            if recovered is not None:
                projects = load_projects()
    project_ids = [str(p.get("id", "")) for p in projects]
    if (not os.environ.get("PYTEST_CURRENT_TEST")) and project_id and project_id not in project_ids:
        recovered = _recover_missing_project_from_artifacts(project_id, projects)
        if recovered is not None:
            projects = load_projects()
            project_ids = [str(p.get("id", "")) for p in projects]
    selected_project_id = ""
    if project_id and project_id in project_ids:
        selected_project_id = project_id
    elif project_ids:
        # Default to latest created project for better usability.
        selected_project_id = project_ids[-1]
    project_options = []
    for p in projects:
        pid_raw = str(p.get("id", ""))
        pid = html_lib.escape(pid_raw)
        pname = html_lib.escape(str(p.get("name", p.get("id", ""))))
        short_id = html_lib.escape(str(p.get("id", ""))[:8])
        selected_attr = " selected" if pid_raw == selected_project_id else ""
        project_options.append(
            f'<option value="{pid}"{selected_attr}>{pname} ({short_id}…)</option>'
        )
    project_options_html = "".join(project_options)
    create_notice_html = ""
    if create_ok:
        create_notice_html = (
            '<p style="margin:6px 0 0 0;font-size:13px;color:#15803d">'
            "创建成功（表单模式）：" + html_lib.escape(create_ok) + "</p>"
        )
    elif create_error:
        create_notice_html = (
            '<p style="margin:6px 0 0 0;font-size:13px;color:#b91c1c">'
            "创建失败：" + html_lib.escape(create_error) + "</p>"
        )

    global_notice_html = ""
    if msg:
        is_error = str(msg_type or "").lower() == "error"
        notice_color = "#b91c1c" if is_error else "#15803d"
        notice_bg = "#fef2f2" if is_error else "#ecfdf5"
        global_notice_html = (
            f'<div style="margin:0 0 16px 0;padding:10px 12px;border-radius:8px;background:{notice_bg};color:{notice_color};font-size:13px">'
            + html_lib.escape(msg)
            + "</div>"
        )

    ui_dimension_labels = {
        "01": "01 工程项目整体理解与实施路径",
        "02": "02 安全生产管理体系与控制措施",
        "03": "03 文明施工管理体系与实施措施",
        "04": "04 材料与部品采购及管理机制",
        "05": "05 四新技术的应用与实施方案",
        "06": "06 工程关键工序识别与控制措施",
        "07": "07 工程重难点及危险性较大工程管控",
        "08": "08 工程质量管理体系与保证措施",
        "09": "09 工期目标保障与进度控制措施",
        "10": "10 专项施工工艺与技术方案",
        "11": "11 人力资源配置与管理方案",
        "12": "12 总体施工工艺流程与组织逻辑",
        "13": "13 物资与施工设备配置方案",
        "14": "14 设计协调与深化实施能力",
        "15": "15 总体资源配置与实施计划",
        "16": "16 技术措施的可行性与落地性",
    }
    initial_weights_raw: Dict[str, int] = _default_weights_raw()
    initial_weights_norm: Dict[str, float] = _normalize_weights(initial_weights_raw)
    initial_profile_status = "请先选择项目并加载配置。"
    if selected_project_id:
        try:
            project = _find_project(selected_project_id, projects)
            profiles = load_expert_profiles()
            profile: Optional[Dict[str, object]] = None
            profile_id = str(project.get("expert_profile_id") or "")
            if profile_id:
                for item in profiles:
                    if str(item.get("id") or "") == profile_id:
                        profile = item
                        break
            if profile and isinstance(profile.get("weights_raw"), dict):
                initial_weights_raw = _coerce_weights_raw(profile.get("weights_raw", {}))
                initial_weights_norm = _normalize_weights(initial_weights_raw)
                profile_name = str(profile.get("name") or "项目默认配置")
                profile_id_text = str(profile.get("id") or "-")
                updated_at_text = (
                    str(project.get("updated_at") or profile.get("updated_at") or "")[:19] or "-"
                )
                initial_profile_status = (
                    "当前生效配置："
                    + html_lib.escape(profile_name)
                    + "（ID: "
                    + html_lib.escape(profile_id_text)
                    + "，更新时间: "
                    + html_lib.escape(updated_at_text)
                    + "）"
                )
            else:
                profile_name = str(project.get("name") or "项目") + " 默认配置"
                updated_at_text = str(project.get("updated_at") or "")[:19] or "-"
                initial_profile_status = (
                    "当前生效配置："
                    + html_lib.escape(profile_name)
                    + "（ID: 未绑定，更新时间: "
                    + html_lib.escape(updated_at_text)
                    + "）"
                )
        except Exception:
            initial_profile_status = "请先选择项目并加载配置。"

    initial_weights_rows = []
    for dim_id in DIMENSION_IDS:
        label = ui_dimension_labels.get(dim_id, f"{dim_id} {_normalize_dimension_id(dim_id)}")
        raw = int(initial_weights_raw.get(dim_id, 5))
        norm_pct = float(initial_weights_norm.get(dim_id, 0.0) * 100.0)
        initial_weights_rows.append(
            '<div class="weight-row">'
            + f'<label for="w_{dim_id}">{html_lib.escape(label)}</label>'
            + f'<input id="w_{dim_id}" data-dim="{html_lib.escape(dim_id)}" type="range" min="0" max="10" step="1" value="{raw}" />'
            + f'<span class="raw-value" id="w_raw_{dim_id}">{raw}</span>'
            + f'<span class="norm-value" id="w_norm_{dim_id}">{norm_pct:.2f}%</span>'
            + "</div>"
        )
    initial_weights_rows_html = "".join(initial_weights_rows)
    initial_weights_summary = " | ".join(
        [
            f"{dim_id}:{(float(initial_weights_norm.get(dim_id, 0.0)) * 100.0):.2f}%"
            for dim_id in DIMENSION_IDS
        ]
    )
    selected_project_for_view = (
        next((p for p in projects if str(p.get("id", "")) == selected_project_id), {})
        if selected_project_id
        else {}
    )
    score_scale_initial = (
        _resolve_project_score_scale_max(selected_project_for_view)
        if selected_project_for_view
        else DEFAULT_SCORE_SCALE_MAX
    )
    allow_pred_initial = bool(selected_project_for_view) and (
        _select_calibrator_model(selected_project_for_view) is not None
    )
    initial_material_rows: List[str] = []
    initial_submission_rows: List[str] = []
    if selected_project_id:
        try:
            materials_all = load_materials()
            selected_materials = [
                m for m in materials_all if str(m.get("project_id", "")) == selected_project_id
            ]
            selected_materials.sort(key=lambda x: str(x.get("created_at", "")), reverse=True)
            for m in selected_materials:
                material_id = html_lib.escape(str(m.get("id", "")))
                filename_raw = str(m.get("filename", ""))
                filename = html_lib.escape(filename_raw)
                material_type_label = html_lib.escape(
                    _material_type_label(m.get("material_type"), filename=m.get("filename"))
                )
                created_at = html_lib.escape(str(m.get("created_at", ""))[:19])
                initial_material_rows.append(
                    "<tr>"
                    + f"<td>{material_type_label}</td>"
                    + f"<td>{filename}</td>"
                    + f"<td>{created_at}</td>"
                    + (
                        "<td>"
                        + f'<button type="button" class="btn-danger js-delete-material" data-material-id="{material_id}" data-project-id="{html_lib.escape(str(m.get("project_id") or ""))}" data-filename="{html_lib.escape(filename_raw)}" onclick="return window.__zhifeiFallbackDelete(event, \'material\', this.getAttribute(\'data-material-id\'), this.getAttribute(\'data-filename\'), this.getAttribute(\'data-project-id\'))">删除</button>'
                        + "</td>"
                    )
                    + "</tr>"
                )
        except Exception:
            initial_material_rows = []
        try:
            submissions_all = load_submissions()
            selected_submissions = [
                s for s in submissions_all if str(s.get("project_id", "")) == selected_project_id
            ]
            selected_submissions.sort(key=lambda x: str(x.get("created_at", "")), reverse=True)
            for s in selected_submissions:
                submission_id = html_lib.escape(str(s.get("id", "")))
                filename_raw = str(s.get("filename", ""))
                filename = html_lib.escape(filename_raw)
                report_obj = s.get("report")
                report = report_obj if isinstance(report_obj, dict) else {}
                pred_total_raw = report.get("pred_total_score")
                rule_total_raw = report.get("rule_total_score")
                if not allow_pred_initial:
                    pred_total_raw = None
                llm_total_raw = report.get("llm_total_score")
                pred_total = _convert_score_from_100(pred_total_raw, score_scale_initial)
                rule_total = _convert_score_from_100(rule_total_raw, score_scale_initial)
                llm_total = _convert_score_from_100(llm_total_raw, score_scale_initial)
                scoring_status = str(report.get("scoring_status") or "").strip().lower()
                is_pending = scoring_status == "pending"
                is_blocked = scoring_status == "blocked"
                report_meta = report.get("meta") if isinstance(report.get("meta"), dict) else {}
                util_gate = (
                    report_meta.get("material_utilization_gate")
                    if isinstance(report_meta.get("material_utilization_gate"), dict)
                    else {}
                )
                util_blocked = bool(util_gate.get("blocked"))
                evidence_trace = (
                    report_meta.get("evidence_trace")
                    if isinstance(report_meta.get("evidence_trace"), dict)
                    else {}
                )
                primary_total = (
                    pred_total
                    if pred_total is not None
                    else _convert_score_from_100(s.get("total_score"), score_scale_initial)
                )
                if is_pending:
                    score_cell = '<span class="note">待评分</span>'
                elif is_blocked:
                    score_cell = '<span class="error">待补资料后重评分</span>'
                elif pred_total is not None:
                    score_cell = html_lib.escape(str(pred_total))
                    note_items: List[str] = []
                    if rule_total is not None:
                        note_items.append("规则: " + html_lib.escape(str(rule_total)))
                    if llm_total is not None:
                        note_items.append("LLM: " + html_lib.escape(str(llm_total)))
                    if note_items:
                        score_cell += '<div class="note">' + " / ".join(note_items) + "</div>"
                else:
                    score_cell = (
                        "-" if primary_total is None else html_lib.escape(str(primary_total))
                    )
                evidence_hits = int(_to_float_or_none(evidence_trace.get("total_hits")) or 0)
                evidence_file_hits = int(
                    _to_float_or_none(evidence_trace.get("source_files_hit_count")) or 0
                )
                if not is_pending and evidence_hits > 0:
                    score_cell += (
                        '<div class="note">证据命中: '
                        + html_lib.escape(str(evidence_hits))
                        + " 条 / 文件覆盖: "
                        + html_lib.escape(str(evidence_file_hits))
                        + " 份</div>"
                    )
                util_summary = (
                    report_meta.get("material_utilization")
                    if isinstance(report_meta.get("material_utilization"), dict)
                    else {}
                )
                util_by_type = (
                    util_summary.get("by_type")
                    if isinstance(util_summary.get("by_type"), dict)
                    else {}
                )
                util_available_types = (
                    util_summary.get("available_types")
                    if isinstance(util_summary.get("available_types"), list)
                    else []
                )

                def _type_short(t: str) -> str:
                    if t == "tender_qa":
                        return "招答"
                    if t == "boq":
                        return "清单"
                    if t == "drawing":
                        return "图纸"
                    if t == "site_photo":
                        return "照片"
                    return t or "-"

                coverage_tokens: List[str] = []
                for type_key in ["tender_qa", "boq", "drawing", "site_photo"]:
                    in_scope = type_key in util_available_types
                    if not in_scope:
                        coverage_tokens.append(_type_short(type_key) + "·")
                        continue
                    row = util_by_type.get(type_key) if isinstance(util_by_type, dict) else {}
                    row = row if isinstance(row, dict) else {}
                    retrieval_hit = int(_to_float_or_none(row.get("retrieval_hit")) or 0)
                    consistency_hit = int(_to_float_or_none(row.get("consistency_hit")) or 0)
                    coverage_tokens.append(
                        _type_short(type_key)
                        + ("✓" if (retrieval_hit + consistency_hit) > 0 else "×")
                    )
                if not is_pending and coverage_tokens:
                    score_cell += (
                        '<div class="note">类型覆盖: '
                        + html_lib.escape(" / ".join(coverage_tokens))
                        + "</div>"
                    )
                evidence_files = (
                    evidence_trace.get("source_files_hit")
                    if isinstance(evidence_trace.get("source_files_hit"), list)
                    else []
                )
                evidence_files = [str(x).strip() for x in evidence_files if str(x).strip()]
                if not is_pending and evidence_files:
                    preview = "；".join(evidence_files[:2])
                    suffix = " 等" if len(evidence_files) > 2 else ""
                    score_cell += (
                        '<div class="note">命中文件: '
                        + html_lib.escape(preview)
                        + suffix
                        + "</div>"
                    )
                if util_blocked:
                    score_cell += (
                        '<div class="error">资料利用门禁未达标（建议补齐资料后重评分）</div>'
                    )
                created_at = html_lib.escape(str(s.get("created_at", ""))[:19])
                initial_submission_rows.append(
                    "<tr>"
                    + f"<td>{filename}</td>"
                    + f"<td>{score_cell}</td>"
                    + f"<td>{created_at}</td>"
                    + (
                        "<td>"
                        + f'<button type="button" class="btn-danger js-delete-submission" data-submission-id="{submission_id}" data-project-id="{html_lib.escape(str(s.get("project_id") or ""))}" data-filename="{html_lib.escape(filename_raw)}" onclick="return window.__zhifeiFallbackDelete(event, \'submission\', this.getAttribute(\'data-submission-id\'), this.getAttribute(\'data-filename\'), this.getAttribute(\'data-project-id\'))">删除</button>'
                        + "</td>"
                    )
                    + "</tr>"
                )
        except Exception:
            initial_submission_rows = []
    initial_material_rows_html = "".join(initial_material_rows)
    initial_submission_rows_html = "".join(initial_submission_rows)
    initial_materials_empty_display = "none" if initial_material_rows else "block"
    initial_submissions_empty_display = "none" if initial_submission_rows else "block"
    html = """
    <html>
    <head>
      <meta charset="utf-8">
      <title>青天评标系统</title>
      <style>
        :root { --bg:#f4f6fb; --card:#fff; --border:#dbe2ef; --primary:#2563eb; --text:#1e293b; }
        body { font-family: system-ui, sans-serif; margin: 0 auto; max-width: 1680px; padding: 20px; background: var(--bg); color: var(--text); line-height: 1.5; }
        h2 { margin-top: 12px; font-size: 1.8rem; color: var(--primary); }
        .card { background: var(--card); border: 1px solid var(--border); border-radius: 14px; padding: 18px 20px; margin-bottom: 18px; box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04); }
        input[type="text"], select { padding: 8px 10px; margin-right: 8px; min-width: 200px; border:1px solid #cbd5e1; border-radius:8px; background:#f8fafc; }
        button { padding: 10px 16px; background: var(--primary); color: #fff; border: none; border-radius: 8px; cursor: pointer; font-weight: 600; }
        button:hover { opacity: 0.9; }
        button:disabled { opacity: 0.45; cursor: not-allowed; }
        button.secondary { background: #64748b; }
        button.btn-danger { background: #dc2626; color: #fff; position:relative; z-index:2; pointer-events:auto; }
        pre { white-space: pre-wrap; font-size: 12px; margin: 0; line-height: 1.45; }
        #output { font-size: 13px; }
        .section { margin-bottom: 24px; }
        .toolbar { display:flex; flex-wrap:wrap; align-items:center; gap:10px; }
        .inline-form { display:inline-flex; flex-wrap:wrap; align-items:center; gap:10px; margin:0; }
        .action-row { display:flex; flex-wrap:wrap; align-items:center; gap:10px; margin: 8px 0; }
        .action-row button { position:relative; z-index:2; pointer-events:auto; }
        .upload-box { margin-top:14px; padding:12px 12px 10px 12px; border-top:1px solid var(--border); border-radius:10px; background:#fbfdff; }
        .upload-panel-title { margin:0 0 8px 0; font-size:22px; font-weight:700; color:#1e293b; }
        .upload-zones { display:grid; grid-template-columns: 1fr; gap:12px; margin-top:12px; }
        .upload-zone { border:1px solid var(--border); border-radius:12px; background:#f8fafc; padding:12px; }
        .upload-zone h4 { margin:0 0 8px 0; font-size:18px; color:#1e293b; }
        .upload-zone .inline-form { width:100%; }
        .upload-zone input[type="file"] { flex:1 1 360px; min-width:260px; }
        .upload-zone .note { display:block; width:100%; }
        .field-group { display:flex; flex-wrap:wrap; align-items:center; gap:8px; margin-bottom:8px; }
        .field-group p { flex-basis:100%; margin:6px 0 0 0; }
        .note { font-size:12px; color:#64748b; }
        .muted { margin:4px 0 0 0; font-size:13px; color:#64748b; }
        .result-block { margin-top: 10px; padding: 12px; background: #f8fafc; border-radius: 8px; border-left: 3px solid var(--primary); overflow-x:auto; }
        table { border-collapse: collapse; width: 100%; font-size: 14px; margin-top: 6px; }
        th, td { border: 1px solid var(--border); padding: 10px 12px; text-align: left; vertical-align: top; }
        th { background: #e2e8f0; }
        tbody tr:nth-child(even) { background: #f8fafc; }
        .error { color: #b91c1c; }
        .success { color: #15803d; }
        details { margin-top: 8px; }
        summary { cursor: pointer; font-weight: 600; }
        .weight-row { display:grid; grid-template-columns: 280px 1fr 48px 78px; gap:10px; align-items:center; }
        .weight-row label { font-size:13px; color:#1e293b; }
        .weight-row input[type="range"] { width:100%; }
        .weight-row .raw-value { font-weight:600; color:#0f172a; text-align:right; }
        .weight-row .norm-value { color:#334155; text-align:right; font-size:12px; }
      </style>
    </head>
    <body>
      <h1>青天评标系统 - 上传与对比 (v2)</h1>
      <p style="margin:-8px 0 16px 0;padding:10px;background:#e0f2fe;border-radius:6px;font-size:14px;">
        <strong>首次使用：</strong>① 创建项目 → ② 刷新并选择项目 → ③ 上传施组文件 → ④ 点击“评分施组”出分。数据保存在本机，无需额外配置。
      </p>
      __GLOBAL_NOTICE_HTML__
      <script>
        (function () {
          // Early fallback shim: guarantees visible response even if later scripts crash.
          if (window.__zhifeiEarlyFallbackReady) return;
          window.__zhifeiEarlyFallbackReady = true;

          function pickProjectId() {
            const sel = document.getElementById('projectSelect');
            return (sel && sel.value) ? String(sel.value) : '';
          }
          function apiHeaders(isJson) {
            const h = {};
            try {
              const k = localStorage.getItem('api_key') || '';
              if (k) h['X-API-Key'] = k;
            } catch (_) {}
            if (isJson) h['Content-Type'] = 'application/json';
            return h;
          }
          function esc(v) {
            return String(v == null ? '' : v)
              .replace(/&/g, '&amp;')
              .replace(/</g, '&lt;')
              .replace(/>/g, '&gt;')
              .replace(/"/g, '&quot;')
              .replace(/'/g, '&#39;');
          }
          function setResult(resultId, text, isError) {
            const el = resultId ? document.getElementById(resultId) : null;
            if (!el) return;
            el.style.display = 'block';
            el.innerHTML = '<span class="' + (isError ? 'error' : 'success') + '">' + esc(text || '') + '</span>';
          }
          function setResultHtml(resultId, html) {
            const el = resultId ? document.getElementById(resultId) : null;
            if (!el) return;
            el.style.display = 'block';
            el.innerHTML = html || '';
          }
          function setOutput(text) {
            const out = document.getElementById('output');
            if (out) out.textContent = String(text || '');
          }
          function parseJson(text) {
            try { return JSON.parse(text || '{}'); } catch (_) { return {}; }
          }
          function selectedScoreScaleMax() {
            const el = document.getElementById('scoreScaleSelect');
            const raw = (el && el.value) ? String(el.value).trim() : '100';
            return raw === '5' ? 5 : 100;
          }

          const EARLY_ACTIONS = {
            btnMaterialDepthReport: { resultId: 'materialDepthReportResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/materials/depth_report', loading: '资料深读体检生成中...' },
            btnMaterialDepthReportDownload: { resultId: 'materialDepthReportResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/materials/depth_report.md', loading: '体检报告下载准备中...' },
            btnMaterialKnowledgeProfile: { resultId: 'materialKnowledgeProfileResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/materials/knowledge_profile', loading: '资料知识画像生成中...' },
            btnMaterialKnowledgeProfileDownload: { resultId: 'materialKnowledgeProfileResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/materials/knowledge_profile.md', loading: '知识画像下载准备中...' },
            btnCompare: { resultId: 'compareResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/compare', loading: '对比排名加载中...' },
            btnCompareReport: { resultId: 'compareReportResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/compare_report', loading: '对比报告生成中...' },
            btnInsights: { resultId: 'insightsResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/insights', loading: '洞察分析中...' },
            btnLearning: { resultId: 'learningResult', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/learning', loading: '学习画像生成中...' },
            btnEvidenceTrace: { resultId: 'evidenceTraceResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/evidence_trace/latest', loading: '证据追溯生成中...' },
            btnScoringBasis: { resultId: 'scoringBasisResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/scoring_basis/latest', loading: '评分依据审计生成中...' },
            btnAdaptive: { resultId: 'adaptiveResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/adaptive', loading: '自适应建议生成中...' },
            btnAdaptivePatch: { resultId: 'adaptivePatchResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/adaptive_patch', loading: '补丁生成中...' },
            btnAdaptiveValidate: { resultId: 'adaptiveValidateResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/adaptive_validate', loading: '验证效果计算中...' },
            btnAdaptiveApply: { resultId: 'adaptiveApplyResult', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/adaptive_apply', loading: '应用补丁中...' },
            btnRefreshGroundTruth: { resultId: 'evolveResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/ground_truth', loading: '真实评标列表刷新中...' },
            btnRefreshGroundTruthSubmissionOptions: { resultId: 'evolveResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/submissions', loading: '施组选项刷新中...' },
            btnEvolve: { resultId: 'evolveResult', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/evolve', loading: '学习进化执行中...' },
            btnOllamaPreview: { resultId: 'ollamaPreviewResult', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/evolve/ollama_preview', loading: 'Ollama 增强预览生成中...' },
            btnEvolutionHealth: { resultId: 'evolutionHealthResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/evolution/health', loading: '进化健康度分析中...' },
            btnWritingGuidance: { resultId: 'guidanceResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/writing_guidance', loading: '正在生成编制指导...' },
            btnCompilationInstructions: { resultId: 'compilationInstructionsResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/compilation_instructions', loading: '正在生成编制系统指令...' },
            btnScoreShigong: { resultId: 'shigongActionStatus', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/rescore', loading: '施组评分中...' },
          };
          window.__ZHIFEI_FALLBACK_ACTIONS = window.__ZHIFEI_FALLBACK_ACTIONS || {};
          Object.keys(EARLY_ACTIONS).forEach((k) => { window.__ZHIFEI_FALLBACK_ACTIONS[k] = EARLY_ACTIONS[k]; });

          async function runEarlyAction(actionId) {
            const cfg = EARLY_ACTIONS[String(actionId || '')];
            if (!cfg) return true;
            const projectId = pickProjectId();
            if (!projectId) {
              setResult(cfg.resultId, '请先在「2) 选择项目」中选择项目', true);
              setOutput('[' + actionId + '] 缺少项目ID');
              return false;
            }
            if (actionId === 'btnMaterialDepthReportDownload') {
              const dlUrl = '/api/v1/projects/' + encodeURIComponent(projectId) + '/materials/depth_report.md';
              const a = document.createElement('a');
              a.href = dlUrl;
              a.download = 'material_depth_report_' + projectId + '.md';
              document.body.appendChild(a);
              a.click();
              a.remove();
              setResult(cfg.resultId, '资料深读体检报告下载已触发。', false);
              setOutput('[' + actionId + '] download ' + dlUrl);
              return true;
            }
            if (actionId === 'btnMaterialKnowledgeProfileDownload') {
              const dlUrl = '/api/v1/projects/' + encodeURIComponent(projectId) + '/materials/knowledge_profile.md';
              const a = document.createElement('a');
              a.href = dlUrl;
              a.download = 'material_knowledge_profile_' + projectId + '.md';
              document.body.appendChild(a);
              a.click();
              a.remove();
              setResult(cfg.resultId, '资料知识画像报告下载已触发。', false);
              setOutput('[' + actionId + '] download ' + dlUrl);
              return true;
            }
            setResult(cfg.resultId, cfg.loading || '处理中...', false);
            let url = cfg.path(projectId);
            let method = cfg.method || 'GET';
            let headers = apiHeaders(false);
            let body;
            if (actionId === 'btnScoreShigong') {
              headers = apiHeaders(true);
              body = JSON.stringify({
                scope: 'project',
                scoring_engine_version: 'v2',
                score_scale_max: selectedScoreScaleMax(),
              });
            }
            let res;
            let text = '';
            try {
              res = await fetch(url, { method, headers, body });
              text = await res.text();
            } catch (err) {
              setResult(cfg.resultId, '网络异常：' + String((err && err.message) || err || 'unknown'), true);
              setOutput('[' + actionId + '] 网络异常');
              return false;
            }
            const data = parseJson(text);
            setOutput('[' + actionId + '] HTTP ' + String(res.status || 0) + '\\n' + (text || ''));
            if (!res.ok) {
              setResult(cfg.resultId, '[' + actionId + '] 请求失败：' + String((data && data.detail) || ('HTTP ' + res.status)), true);
              return false;
            }
            if (actionId === 'btnCompare') {
              const rows = Array.isArray(data.rankings) ? data.rankings : [];
              const html = '<strong>排名</strong><table><tr><th>文件名</th><th>总分</th><th>时间</th></tr>' +
                (rows.length
                  ? rows.map((r) => '<tr><td>' + esc(r.filename || '') + '</td><td>' + esc(r.total_score) + '</td><td>' + esc(r.created_at || '') + '</td></tr>').join('')
                  : '<tr><td colspan="3">暂无施组评分数据</td></tr>') +
                '</table>';
              setResultHtml(cfg.resultId, html);
              return true;
            }
            if (actionId === 'btnEvidenceTrace') {
              const summary = (data && typeof data.summary === 'object') ? data.summary : {};
              const conflicts = (
                data && data.material_conflicts && typeof data.material_conflicts === 'object'
              ) ? data.material_conflicts : {};
              const recommendations = Array.isArray(data.recommendations) ? data.recommendations : [];
              const html = '<strong>证据追溯（最新施组）</strong>'
                + '<p style="margin:6px 0">文件：' + esc(data.filename || '-') + '</p>'
                + '<table><tr><th>要求总数</th><th>命中总数</th><th>整体命中率</th><th>强制项命中率</th><th>命中文件数</th></tr>'
                + '<tr><td>' + esc(summary.total_requirements || 0) + '</td><td>' + esc(summary.total_hits || 0) + '</td><td>' + esc(summary.overall_hit_rate ?? '-') + '</td><td>' + esc(summary.mandatory_hit_rate ?? '-') + '</td><td>' + esc(summary.source_files_hit_count || 0) + '</td></tr></table>'
                + '<p style="margin:6px 0">一致性冲突：' + esc(conflicts.conflict_count || 0) + '（高风险 ' + esc(conflicts.high_severity_count || 0) + '）</p>'
                + (recommendations.length
                  ? '<ul style="margin:6px 0 0 18px;color:#92400e">' + recommendations.slice(0, 6).map((x) => '<li>' + esc(x) + '</li>').join('') + '</ul>'
                  : '');
              setResultHtml(cfg.resultId, html);
              return true;
            }
            if (actionId === 'btnScoringBasis') {
              const mece = (data && typeof data.mece_inputs === 'object') ? data.mece_inputs : {};
              const util = (data && typeof data.material_utilization === 'object') ? data.material_utilization : {};
              const gate = (data && typeof data.material_utilization_gate === 'object') ? data.material_utilization_gate : {};
              const trace = (data && typeof data.evidence_trace === 'object') ? data.evidence_trace : {};
              const recs = Array.isArray(data.recommendations) ? data.recommendations : [];
              let html = '<strong>评分依据审计（最新施组）</strong>'
                + '<p style="margin:6px 0">文件：' + esc(data.filename || '-') + '；状态：' + esc(data.scoring_status || '-') + '</p>'
                + '<table><tr><th>资料门禁</th><th>资料检索命中率</th><th>文件覆盖率</th><th>强制项命中率</th><th>命中文件数</th></tr>'
                + '<tr><td>' + (mece.materials_quality_gate_passed ? '<span class="success">通过</span>' : '<span class="error">未通过</span>') + '</td>'
                + '<td>' + esc(util.retrieval_hit_rate ?? '-') + '</td>'
                + '<td>' + esc(util.retrieval_file_coverage_rate ?? '-') + '</td>'
                + '<td>' + esc(trace.mandatory_hit_rate ?? '-') + '</td>'
                + '<td>' + esc(trace.source_files_hit_count || 0) + '</td></tr></table>';
              const hitFiles = Array.isArray(trace.source_files_hit) ? trace.source_files_hit : [];
              if (hitFiles.length) {
                html += '<p style="margin:6px 0">命中文件：' + esc(hitFiles.slice(0, 6).join('；')) + (hitFiles.length > 6 ? ' 等' : '') + '</p>';
              }
              const gateReasons = Array.isArray(gate.reasons) ? gate.reasons : [];
              if (gateReasons.length) {
                html += '<details style="margin-top:6px"><summary>门禁原因</summary><ul>' + gateReasons.slice(0, 8).map((x) => '<li>' + esc(x) + '</li>').join('') + '</ul></details>';
              }
              if (recs.length) {
                html += '<strong>建议动作</strong><ul>' + recs.slice(0, 8).map((x) => '<li>' + esc(x) + '</li>').join('') + '</ul>';
              }
              setResultHtml(cfg.resultId, html);
              return true;
            }
            if (actionId === 'btnRefreshGroundTruthSubmissionOptions') {
              const sel = document.getElementById('groundTruthSubmissionSelect');
              const subs = Array.isArray(data) ? data : [];
              if (sel) {
                sel.innerHTML = '';
                const lead = document.createElement('option');
                lead.value = '';
                lead.textContent = subs.length ? '-- 请选择步骤4已上传施组文件 --' : '-- 暂无施组，请先在步骤4上传 --';
                sel.appendChild(lead);
                subs.forEach((s) => {
                  const opt = document.createElement('option');
                  opt.value = String((s && s.id) || '');
                  opt.textContent = String((s && s.filename) || '未命名施组');
                  sel.appendChild(opt);
                });
              }
              setResult(cfg.resultId, '施组选项刷新完成：共 ' + subs.length + ' 份。', false);
              return true;
            }
            if (actionId === 'btnScoreShigong') {
              const updated = Number(
                (data && (data.updated_submissions ?? data.reports_generated ?? data.submission_count)) || 0
              );
              const scaleLabel = (data && data.score_scale_label) ? String(data.score_scale_label) : '100分制';
              const gate = (data && typeof data.material_utilization_gate === 'object')
                ? data.material_utilization_gate
                : {};
              const closedLoop = (data && typeof data.feedback_closed_loop === 'object')
                ? data.feedback_closed_loop
                : {};
              const blockedCount = Number((gate && gate.blocked_submissions) || 0);
              const warnCount = Number((gate && gate.warn_submissions) || 0);
              let msg = '评分完成（' + scaleLabel + '）：已重算 ' + updated + ' 份。';
              if (blockedCount > 0) {
                msg += ' 资料门禁阻断 ' + blockedCount + ' 份。';
              } else if (warnCount > 0) {
                msg += ' 资料门禁预警 ' + warnCount + ' 份。';
              } else if (gate && gate.enabled) {
                msg += ' 资料门禁通过。';
              }
              if (Object.keys(closedLoop).length) {
                if (closedLoop.ok === false) msg += ' 反馈闭环执行异常，请查看下方原始输出。';
                else msg += ' 反馈闭环已执行。';
              }
              setResult(cfg.resultId, msg, blockedCount > 0);
              if (window.renderMaterialUtilizationPanel && typeof window.renderMaterialUtilizationPanel === 'function') {
                window.renderMaterialUtilizationPanel(data || {});
              }
              return true;
            }
            if (actionId === 'btnMaterialDepthReport') {
              const byType = Array.isArray(data.by_type) ? data.by_type : [];
              const quality = (data && typeof data.quality_summary === 'object') ? data.quality_summary : {};
              const recs = Array.isArray(data.recommendations) ? data.recommendations : [];
              const html = ''
                + '<strong>资料深读体检</strong>'
                + '<p style="margin:6px 0">评分就绪：'
                + (data && data.ready_to_score ? '<span class="success">是</span>' : '<span class="error">否</span>')
                + '；总分块 ' + esc(quality.total_parsed_chunks || 0)
                + '；数字约束项 ' + esc(quality.total_numeric_terms || 0)
                + '</p>'
                + '<table><tr><th>资料类型</th><th>文件数</th><th>解析字数</th><th>分块数</th><th>数字约束项</th></tr>'
                + (byType.length
                  ? byType.map((row) => '<tr>'
                    + '<td>' + esc((row && row.material_type_label) || (row && row.material_type) || '-') + '</td>'
                    + '<td>' + esc((row && row.files) || 0) + '</td>'
                    + '<td>' + esc((row && row.parsed_chars) || 0) + '</td>'
                    + '<td>' + esc((row && row.parsed_chunks) || 0) + '</td>'
                    + '<td>' + esc((row && row.numeric_terms) || 0) + '</td>'
                    + '</tr>').join('')
                  : '<tr><td colspan="5">暂无资料体检数据</td></tr>')
                + '</table>'
                + (recs.length
                  ? '<ul style="margin:6px 0 0 18px;color:#92400e">' + recs.slice(0, 6).map((x) => '<li>' + esc(x) + '</li>').join('') + '</ul>'
                  : '');
              setResultHtml(cfg.resultId, html);
              return true;
            }
            if (actionId === 'btnMaterialKnowledgeProfile') {
              const summary = (data && typeof data.summary === 'object') ? data.summary : {};
              const dims = Array.isArray(data.by_dimension) ? data.by_dimension : [];
              const rows = dims.slice().sort((a, b) => Number((b && b.coverage_score) || 0) - Number((a && a.coverage_score) || 0)).slice(0, 8);
              const recs = Array.isArray(data.recommendations) ? data.recommendations : [];
              const html = ''
                + '<strong>资料知识画像</strong>'
                + '<p style="margin:6px 0">维度覆盖率：' + esc(summary.dimension_coverage_rate || 0)
                + '；低覆盖维度：' + esc(summary.low_coverage_dimensions || 0)
                + '；总解析字数：' + esc(summary.total_parsed_chars || 0) + '</p>'
                + '<table><tr><th>维度</th><th>关键词命中</th><th>来源类型数</th><th>覆盖评分</th><th>等级</th></tr>'
                + (rows.length
                  ? rows.map((row) => '<tr>'
                    + '<td>' + esc((row && row.dimension_id) || '-') + ' ' + esc((row && row.dimension_name) || '') + '</td>'
                    + '<td>' + esc((row && row.keyword_hits) || 0) + '</td>'
                    + '<td>' + esc(((row && row.source_types) || []).length) + '</td>'
                    + '<td>' + esc((row && row.coverage_score) || 0) + '</td>'
                    + '<td>' + esc((row && row.coverage_level) || '-') + '</td>'
                    + '</tr>').join('')
                  : '<tr><td colspan="5">暂无知识画像数据</td></tr>')
                + '</table>'
                + (recs.length
                  ? '<ul style="margin:6px 0 0 18px;color:#92400e">' + recs.slice(0, 6).map((x) => '<li>' + esc(x) + '</li>').join('') + '</ul>'
                  : '');
              setResultHtml(cfg.resultId, html);
              return true;
            }
            if (actionId === 'btnEvolutionHealth') {
              const summary = (data && typeof data.summary === 'object') ? data.summary : {};
              const drift = (data && typeof data.drift === 'object') ? data.drift : {};
              const w30 = (data && data.windows && typeof data.windows.recent_30d === 'object') ? data.windows.recent_30d : {};
              const wall = (data && data.windows && typeof data.windows.all === 'object') ? data.windows.all : {};
              const recs = Array.isArray(data.recommendations) ? data.recommendations : [];
              const html = ''
                + '<strong>进化健康度</strong>'
                + '<p style="margin:6px 0">漂移等级：'
                + esc((drift && drift.level) || 'insufficient_data')
                + '；近30天 MAE=' + esc((w30 && w30.mae) != null ? w30.mae : '-')
                + '；全量 MAE=' + esc((wall && wall.mae) != null ? wall.mae : '-')
                + '</p>'
                + '<table><tr><th>指标</th><th>值</th></tr>'
                + '<tr><td>真实评分样本</td><td>' + esc((summary && summary.ground_truth_count) || 0) + '</td></tr>'
                + '<tr><td>已匹配预测</td><td>' + esc((summary && summary.matched_prediction_count) || 0) + '</td></tr>'
                + '<tr><td>未匹配样本</td><td>' + esc((summary && summary.unmatched_ground_truth_count) || 0) + '</td></tr>'
                + '<tr><td>当前权重来源</td><td>' + esc((summary && summary.current_weights_source) || '-') + '</td></tr>'
                + '</table>'
                + (recs.length
                  ? '<ul style="margin:6px 0 0 18px;color:#92400e">' + recs.slice(0, 6).map((x) => '<li>' + esc(x) + '</li>').join('') + '</ul>'
                  : '');
              setResultHtml(cfg.resultId, html);
              return true;
            }
            if (actionId === 'btnOllamaPreview') {
              const preview = (data && typeof data.preview === 'object') ? data.preview : {};
              const logic = Array.isArray(preview.high_score_logic) ? preview.high_score_logic : [];
              const guidance = Array.isArray(preview.writing_guidance) ? preview.writing_guidance : [];
              const enhancedBy = (data && data.enhanced_by) || 'rules';
              const fallback = !!(data && data.fallback);
              const errorSummary = (data && data.error_summary) || '-';
              const updatedAt = String((preview && preview.updated_at) || (data && data.updated_at) || '-');
              const html = ''
                + '<strong>Ollama 增强预览</strong>'
                + '<p style="margin:6px 0">状态：'
                + (fallback ? '<span class="error">已回退</span>' : '<span class="success">增强成功</span>')
                + '</p>'
                + '<table><tr><th>enhanced_by</th><th>fallback</th><th>error_summary</th><th>更新时间</th></tr>'
                + '<tr><td>' + esc(enhancedBy) + '</td>'
                + '<td>' + esc(String(fallback)) + '</td>'
                + '<td>' + esc(errorSummary) + '</td>'
                + '<td>' + esc(updatedAt) + '</td></tr></table>'
                + (data && data.error_summary ? '<p class="error">' + esc(data.error_summary) + '</p>' : '')
                + (logic.length ? '<strong>高分逻辑预览</strong><ul>' + logic.map((x) => '<li>' + esc(x) + '</li>').join('') + '</ul>' : '')
                + (guidance.length ? '<strong>编制指导预览</strong><ul>' + guidance.map((x) => '<li>' + esc(x) + '</li>').join('') + '</ul>' : '')
                + '<p style="font-size:12px;color:#64748b;margin-top:8px">仅预览，不写入正式学习进化结果；不影响评分，不进入核心评分主链。</p>';
              setResultHtml(cfg.resultId, html || '<pre>' + esc(text || '{}') + '</pre>');
              return true;
            }
            setResultHtml(cfg.resultId, '<pre>' + esc(text || '{}') + '</pre>');
            return true;
          }

          window.__zhifeiFallbackClick = function (ev, actionId) {
            if (ev) {
              if (typeof ev.preventDefault === 'function') ev.preventDefault();
              if (typeof ev.stopPropagation === 'function') ev.stopPropagation();
              if (typeof ev.stopImmediatePropagation === 'function') ev.stopImmediatePropagation();
            }
            runEarlyAction(actionId);
            return false;
          };
        })();
      </script>

      <div class="section card">
        <h2>1) 创建项目</h2>
        <form id="createProject" method="post" action="/web/create_project">
          项目名称：<input name="name" placeholder="例如：XX标段施组评审" />
          <button type="submit" id="btnCreateProject">创建</button>
        </form>
        __CREATE_NOTICE_HTML__
        <p id="createProjectMessage" style="margin:8px 0 0 0;font-size:13px;min-height:1.2em"></p>
        <p style="margin:4px 0 0 0;font-size:13px;color:#64748b">创建后可从下方下拉选择项目，或复制返回的 id 使用。</p>
      </div>

      <div class="section card">
        <h2>2) 选择项目</h2>
        <div class="toolbar">
          <button type="button" id="refreshProjects">刷新项目列表</button>
          <span style="margin-left:4px">项目：</span>
          <select id="projectSelect">
            <option value="">-- 请先刷新并选择项目 --</option>
            __PROJECT_OPTIONS__
          </select>
          <span id="currentProjectTag" style="margin-left:4px;font-size:12px;color:#475569"></span>
          <form id="deleteProjectForm" method="post" action="/web/delete_project" class="inline-form">
            <input type="hidden" name="project_id" id="deleteProjectId" value="__SELECTED_PROJECT_ID__" />
            <button type="submit" id="deleteCurrentProject" class="secondary" style="background:#dc2626">删除当前项目</button>
          </form>
          <div class="inline-form" style="align-items:center;gap:6px">
            <span style="margin-left:4px;font-size:12px;color:#475569">批量删除：</span>
            <select id="projectDeleteSelect" multiple size="3" style="min-width:260px;max-width:420px"></select>
            <button type="button" id="deleteSelectedProjects" class="secondary" style="background:#b91c1c">删除所选项目</button>
          </div>
        </div>
        <details style="margin-top:8px">
          <summary style="cursor:pointer;color:#334155;font-size:13px">高级工具（系统诊断 / 评分体系 / 分析包）</summary>
          <div class="toolbar" style="margin-top:8px">
            <button type="button" id="btnSelfCheck" class="secondary">系统自检</button>
            <button type="button" id="btnScoringFactors" class="secondary">评分体系一览</button>
            <button type="button" id="btnScoringFactorsMd" class="secondary">评分体系Markdown</button>
            <button type="button" id="btnAnalysisBundle" class="secondary">项目分析包</button>
            <button type="button" id="btnAnalysisBundleDownload" class="secondary">下载分析包(.md)</button>
            <button type="button" id="btnCleanupE2EProjects" class="secondary" style="background:#b45309">清理E2E测试项目</button>
          </div>
        </details>
        <p id="selectProjectMessage" style="margin:8px 0 0 0;font-size:13px;min-height:1.2em"></p>
        <div id="selfCheckResult" class="result-block" style="display:none"></div>
        <div id="scoringFactorsResult" class="result-block" style="display:none"></div>
        <p class="muted">下方所有操作将使用选中的项目。选择项目后建议先上传项目资料（招标、清单等），再上传施组进行评分。删除项目会同时删除该项目全部资料与记录。</p>
      </div>

      <div class="section card" id="section-materials">
        <h2>3) 项目资料</h2>
        <p style="font-size:13px;color:#64748b;margin:-8px 0 8px 0">评分会读取本区资料并注入到锚点/要求矩阵中，作为后续施组打分依据。请按资料类型上传，避免混投。</p>
        <div style="margin-bottom:10px">
          <strong>本项目资料列表</strong>
          <button type="button" id="btnRefreshMaterials" class="secondary" style="margin-left:8px">刷新</button>
          <button type="button" id="btnMaterialDepthReport" class="secondary" style="margin-left:8px" onclick="return window.__zhifeiFallbackClick(event, 'btnMaterialDepthReport')">深读体检</button>
          <button type="button" id="btnMaterialDepthReportDownload" class="secondary" style="margin-left:8px" onclick="return window.__zhifeiFallbackClick(event, 'btnMaterialDepthReportDownload')">下载体检报告(.md)</button>
          <button type="button" id="btnMaterialKnowledgeProfile" class="secondary" style="margin-left:8px" onclick="return window.__zhifeiFallbackClick(event, 'btnMaterialKnowledgeProfile')">知识画像</button>
          <button type="button" id="btnMaterialKnowledgeProfileDownload" class="secondary" style="margin-left:8px" onclick="return window.__zhifeiFallbackClick(event, 'btnMaterialKnowledgeProfileDownload')">下载知识画像(.md)</button>
        </div>
        <table id="materialsTable"><thead><tr><th>资料类型</th><th>文件名</th><th>上传时间</th><th>操作</th></tr></thead><tbody>__MATERIAL_ROWS__</tbody></table>
        <p id="materialsEmpty" style="font-size:13px;color:#64748b;margin:6px 0 0 0;display:__MATERIALS_EMPTY_DISPLAY__">暂无资料，请下方添加。</p>
        <div id="materialDepthReportResult" class="result-block" style="display:none"></div>
        <div id="materialKnowledgeProfileResult" class="result-block" style="display:none"></div>
        <div class="upload-box">
          <h3 class="upload-panel-title">文件上传区</h3>
          <div class="upload-zones">
            <div class="upload-zone">
              <h4>招标文件和答疑（可多选）</h4>
              <form id="uploadMaterial" method="post" action="/web/upload_materials" enctype="multipart/form-data" class="inline-form">
                <input type="hidden" name="project_id" id="uploadMaterialProjectId" value="__SELECTED_PROJECT_ID__" />
                <input type="hidden" name="material_type" value="tender_qa" />
                <input type="file" name="file" accept=".txt,.md,.pdf,.doc,.docx,.docm,.json" multiple />
                <button type="submit" id="btnUploadMaterials" onclick="if (window.__zhifeiFallbackClick) { return window.__zhifeiFallbackClick(event, 'btnUploadMaterials'); } return true;">上传资料</button>
                <span class="note">支持：TXT/MD/PDF/DOC/DOCX/DOCM/JSON，支持一次选择多个文件。</span>
              </form>
              <p id="materialsActionStatus" style="margin:6px 0 0 0;font-size:12px;color:#475569;min-height:1.2em"></p>
            </div>
            <div class="upload-zone">
              <h4>清单（可多选）</h4>
              <form id="uploadMaterialBoq" method="post" action="/web/upload_materials" enctype="multipart/form-data" class="inline-form">
                <input type="hidden" name="project_id" id="uploadMaterialBoqProjectId" value="__SELECTED_PROJECT_ID__" />
                <input type="hidden" name="material_type" value="boq" />
                <input type="file" name="file" accept=".xlsx,.xls,.xlsm,.csv,.pdf,.doc,.docx,.txt,.json" multiple />
                <button type="submit" id="btnUploadBoq" onclick="if (window.__zhifeiFallbackClick) { return window.__zhifeiFallbackClick(event, 'btnUploadBoq'); } return true;">上传清单</button>
                <span class="note">支持：XLSX/XLS/XLSM/CSV/PDF/DOC/DOCX/TXT/JSON，支持一次选择多个文件。</span>
              </form>
              <p id="materialsActionStatusBoq" style="margin:6px 0 0 0;font-size:12px;color:#475569;min-height:1.2em"></p>
            </div>
            <div class="upload-zone">
              <h4>图纸（可多选，支持 DXF ASCII）</h4>
              <form id="uploadMaterialDrawing" method="post" action="/web/upload_materials" enctype="multipart/form-data" class="inline-form">
                <input type="hidden" name="project_id" id="uploadMaterialDrawingProjectId" value="__SELECTED_PROJECT_ID__" />
                <input type="hidden" name="material_type" value="drawing" />
                <input type="file" name="file" accept=".pdf,.doc,.docx,.xlsx,.xls,.dxf,.dwg,.png,.jpg,.jpeg,.webp,.bmp,.tif,.tiff,.json,.txt" multiple />
                <button type="submit" id="btnUploadDrawing" onclick="if (window.__zhifeiFallbackClick) { return window.__zhifeiFallbackClick(event, 'btnUploadDrawing'); } return true;">上传图纸</button>
                <span class="note">支持：PDF/DOC/DOCX/XLSX/XLS/DXF/DWG/图片/JSON/TXT，支持一次选择多个文件。</span>
              </form>
              <p id="materialsActionStatusDrawing" style="margin:6px 0 0 0;font-size:12px;color:#475569;min-height:1.2em"></p>
            </div>
            <div class="upload-zone">
              <h4>现场照片（可多选）</h4>
              <form id="uploadMaterialPhoto" method="post" action="/web/upload_materials" enctype="multipart/form-data" class="inline-form">
                <input type="hidden" name="project_id" id="uploadMaterialPhotoProjectId" value="__SELECTED_PROJECT_ID__" />
                <input type="hidden" name="material_type" value="site_photo" />
                <input type="file" name="file" accept=".png,.jpg,.jpeg,.webp,.bmp,.tif,.tiff" multiple />
                <button type="submit" id="btnUploadSitePhotos" onclick="if (window.__zhifeiFallbackClick) { return window.__zhifeiFallbackClick(event, 'btnUploadSitePhotos'); } return true;">上传照片</button>
                <span class="note">支持：PNG/JPG/JPEG/WEBP/BMP/TIF/TIFF，支持一次选择多个文件。</span>
              </form>
              <p id="materialsActionStatusPhoto" style="margin:6px 0 0 0;font-size:12px;color:#475569;min-height:1.2em"></p>
            </div>
          </div>
        </div>
      </div>

      <div class="section card">
        <h2>2.5) 青天评标关注度（16维）</h2>
        <p style="font-size:12px;color:#64748b;margin:-4px 0 10px 0">先设置16维关注度，再点击“应用到本项目并重算”。同一项目内所有施组将统一按该配置重算，历史快照会保留。</p>
        <div id="expertProfileStatus" style="font-size:13px;color:#334155;margin-bottom:8px">__EXPERT_PROFILE_STATUS__</div>
        <div id="expertWeightsPanel" style="display:grid;grid-template-columns:1fr;gap:8px;margin-bottom:10px">__EXPERT_WEIGHTS_ROWS__</div>
        <div style="font-size:13px;color:#0f172a;margin-bottom:8px">
          当前归一化权重（%）：<span id="expertWeightsSummary">__EXPERT_WEIGHTS_SUMMARY__</span>
        </div>
        <div>
          <button type="button" id="btnWeightsReset" class="secondary" onclick="return window.__zhifeiFallbackClick(event, 'btnWeightsReset')">重置默认(全部=5)</button>
          <button type="button" id="btnWeightsSave" class="secondary" onclick="return window.__zhifeiFallbackClick(event, 'btnWeightsSave')">保存为专家配置</button>
          <button type="button" id="btnWeightsApply" onclick="return window.__zhifeiFallbackClick(event, 'btnWeightsApply')">应用到本项目并重算所有施组</button>
        </div>
      </div>

      <div class="section card" id="section-shigong">
        <h2>4) 项目施组</h2>
        <p style="font-size:13px;color:#64748b;margin:-8px 0 8px 0">每份施组单独打分。支持 .txt、.docx、.pdf、.json、.xlsx/.xls、.dxf。基于下方列表进行对比与洞察。</p>
        <div style="margin-bottom:10px">
          <strong>本项目施组列表</strong>
          <button type="button" id="btnRefreshSubmissions" class="secondary" style="margin-left:8px">刷新</button>
        </div>
        <table id="submissionsTable"><thead><tr><th>文件名</th><th>总分</th><th>上传时间</th><th>操作</th></tr></thead><tbody>__SUBMISSION_ROWS__</tbody></table>
        <p id="submissionsEmpty" style="font-size:13px;color:#64748b;margin:6px 0 0 0;display:__SUBMISSIONS_EMPTY_DISPLAY__">暂无施组，请下方添加。</p>
        <div class="upload-box">
          <form id="uploadShigong" method="post" action="/web/upload_shigong" enctype="multipart/form-data" class="inline-form">
            <strong>添加施组：</strong>
            <input type="hidden" name="project_id" id="uploadShigongProjectId" value="__SELECTED_PROJECT_ID__" />
            <input type="file" name="file" accept=".txt,.docx,.pdf,.json,.xlsx,.xls,.dxf" multiple />
            <button type="submit" id="btnUploadShigong" name="submit_action" value="upload" onclick="if (window.__zhifeiFallbackClick) { return window.__zhifeiFallbackClick(event, 'btnUploadShigong'); } return true;">上传施组</button>
            <button type="submit" id="btnScoreShigong" class="secondary" formaction="/web/score_shigong" name="submit_action" value="score" onclick="if (window.__zhifeiFallbackClick) { return window.__zhifeiFallbackClick(event, 'btnScoreShigong'); } return true;">评分施组</button>
            <span style="margin-left:8px;color:#334155;font-size:13px">满分标准：</span>
            <select id="scoreScaleSelect" name="score_scale_max" style="margin-left:4px">
              <option value="100">100分制</option>
              <option value="5">5分制</option>
            </select>
            <span class="note">支持一次选择多个文件（Mac 按 Command，Windows 按 Ctrl）。</span>
          </form>
          <p id="shigongActionStatus" style="margin:6px 0 0 0;font-size:12px;color:#475569;min-height:1.2em"></p>
          <div id="scoringReadinessResult" class="result-block" style="display:none"></div>
          <div id="materialUtilizationResult" class="result-block" style="display:none"></div>
        </div>
      </div>

      <div class="section card">
        <h2>5) 对比与洞察</h2>
        <p style="font-size:12px;color:#64748b;margin:-4px 0 8px 0">对比排名：看多份施组分数排序；对比报告：看叙述性差异；洞察：看弱项与扣分建议；学习画像：生成维度权重供后续评分参考。</p>
        <div class="action-row">
          <button type="button" id="btnCompare" onclick="return window.__zhifeiFallbackClick(event, 'btnCompare')">对比排名</button>
          <button type="button" id="btnCompareReport" class="secondary" onclick="return window.__zhifeiFallbackClick(event, 'btnCompareReport')">对比报告（叙述）</button>
          <button type="button" id="btnInsights" class="secondary" onclick="return window.__zhifeiFallbackClick(event, 'btnInsights')">洞察</button>
          <button type="button" id="btnLearning" class="secondary" onclick="return window.__zhifeiFallbackClick(event, 'btnLearning')">生成学习画像</button>
          <button type="button" id="btnEvidenceTrace" class="secondary" onclick="return window.__zhifeiFallbackClick(event, 'btnEvidenceTrace')">证据追溯（最新施组）</button>
          <button type="button" id="btnScoringBasis" class="secondary" onclick="return window.__zhifeiFallbackClick(event, 'btnScoringBasis')">评分依据（最新施组）</button>
        </div>
        <div id="compareResult" class="result-block" style="display:none"></div>
        <div id="compareReportResult" class="result-block" style="display:none"></div>
        <div id="insightsResult" class="result-block" style="display:none"></div>
        <div id="learningResult" class="result-block" style="display:none"></div>
        <div id="evidenceTraceResult" class="result-block" style="display:none"></div>
        <div id="scoringBasisResult" class="result-block" style="display:none"></div>
      </div>

      <div class="section card" id="section-adaptive" style="display:none">
        <h2>6) 自适应优化（高级维护）</h2>
        <p style="font-size:12px;color:#64748b;margin:-4px 0 8px 0">基于本项目施组扣分统计给出词库/规则优化建议 → 生成补丁 → 验证效果 → 应用补丁（需 API Key）。</p>
        <div class="action-row" style="margin-bottom:6px">
          <button type="button" id="btnAdaptive" onclick="return window.__zhifeiFallbackClick(event, 'btnAdaptive')">自适应建议</button>
          <button type="button" id="btnAdaptivePatch" class="secondary" onclick="return window.__zhifeiFallbackClick(event, 'btnAdaptivePatch')">生成补丁</button>
        </div>
        <div class="action-row">
          <button type="button" id="btnAdaptiveValidate" class="secondary" onclick="return window.__zhifeiFallbackClick(event, 'btnAdaptiveValidate')">验证效果</button>
          <button type="button" id="btnAdaptiveApply" class="secondary" onclick="return window.__zhifeiFallbackClick(event, 'btnAdaptiveApply')">应用补丁（需 API Key）</button>
        </div>
        <div id="adaptiveResult" class="result-block" style="display:none"></div>
        <div id="adaptivePatchResult" class="result-block" style="display:none"></div>
        <div id="adaptiveValidateResult" class="result-block" style="display:none"></div>
        <div id="adaptiveApplyResult" class="result-block" style="display:none"></div>
      </div>

      <div class="section card">
        <h2>7) 自我学习与进化</h2>
        <p style="font-size:13px;color:#64748b;margin:0 0 6px 0">上传项目投喂包（招标/清单/图纸等合并文本），录入交易中心真实评标结果（5/7评委+最终得分），系统学习高分逻辑并生成编制指导。</p>
        <p style="font-size:12px;color:#475569;margin:0 0 10px 0">系统会将学习到的高分逻辑与编制指导持久保存，并用于本项目的预评分权重与编制系统指令；再次执行学习进化可基于新录入的真实评标升级这些经验。</p>
        <div style="margin-bottom:10px">
          <strong>真实评标列表（本项目 / 其它项目）</strong>
          <span style="margin-left:8px">查看范围：</span>
          <select id="groundTruthScope">
            <option value="current">本项目</option>
            <option value="other">其它项目</option>
          </select>
          <select id="groundTruthOtherProject" style="margin-left:8px;display:none">
            <option value="">-- 选择要查看的项目 --</option>
          </select>
          <button type="button" id="btnRefreshGroundTruth" class="secondary" style="margin-left:8px" onclick="return window.__zhifeiFallbackClick(event, 'btnRefreshGroundTruth')">刷新</button>
        </div>
        <table id="groundTruthTable"><thead><tr><th>序号</th><th>施组摘要</th><th>评委分（5/7）</th><th>最终分</th><th>来源</th><th>操作</th></tr></thead><tbody></tbody></table>
        <p id="groundTruthEmpty" style="font-size:13px;color:#64748b;margin:6px 0 10px 0;display:none">暂无真实评标，请下方录入。</p>
        <div style="margin-bottom:10px">
          <strong>投喂包（即本项目资料）：</strong>上传后可在下方查看文件名与上传时间。
        </div>
        <div class="field-group">
          <label>上传文件：</label>
          <input type="file" id="feedFile" accept=".txt,.pdf,.doc,.docx,.json,.xlsx,.xls,.dxf" multiple style="margin-left:8px" />
          <button type="button" id="btnUploadFeed" onclick="return window.__zhifeiFallbackClick(event, 'btnUploadFeed')">上传并保存投喂包</button>
          <span class="note">支持一次选择多个文件。</span>
          <p id="feedActionStatus" style="margin:6px 0 0 0;font-size:12px;color:#475569;min-height:1.2em"></p>
        </div>
        <div style="margin-bottom:10px">
          <strong>已上传投喂包（项目资料）</strong>
          <button type="button" id="btnRefreshFeedMaterials" class="secondary" style="margin-left:8px" onclick="return window.__zhifeiFallbackClick(event, 'btnRefreshFeedMaterials')">刷新</button>
        </div>
        <table id="feedMaterialsTable"><thead><tr><th>文件名</th><th>上传时间</th><th>操作</th></tr></thead><tbody></tbody></table>
        <p id="feedMaterialsEmpty" style="font-size:13px;color:#64748b;margin:6px 0 10px 0;display:none">暂无投喂包，请在上方或「3) 项目资料」上传。</p>
        <div style="margin-bottom:10px">
          <strong>真实评标录入：</strong>从步骤4已上传施组中选择 + 评委分 + 最终分 →
          <button type="button" id="btnAddGroundTruth" onclick="return window.__zhifeiFallbackClick(event, 'btnAddGroundTruth')">录入所选文件</button>
        </div>
        <div class="field-group">
          <label>施组文件：</label>
          <select id="groundTruthSubmissionSelect" style="margin-left:8px;min-width:360px">
            <option value="">-- 请选择步骤4已上传施组文件 --</option>
          </select>
          <button type="button" id="btnRefreshGroundTruthSubmissionOptions" class="secondary" style="margin-left:8px" onclick="return window.__zhifeiFallbackClick(event, 'btnRefreshGroundTruthSubmissionOptions')">刷新施组选项</button>
          <span class="note">无需重复上传，直接复用「4) 项目施组」已上传文件。</span>
        </div>
        <div class="field-group">
          <label style="margin-right:8px">评委人数：</label>
          <select id="gtJudgeCount">
            <option value="5" selected>5位评委</option>
            <option value="7">7位评委</option>
          </select>
          <span class="note">默认 5 位，可切换到 7 位评委录入。</span>
        </div>
        <div class="field-group">
          <span id="gtJWrap1" style="display:inline-flex;align-items:center;margin-right:8px">评委1：<input type="number" id="gtJ1" step="0.01" style="width:70px;margin-left:4px" /></span>
          <span id="gtJWrap2" style="display:inline-flex;align-items:center;margin-right:8px">评委2：<input type="number" id="gtJ2" step="0.01" style="width:70px;margin-left:4px" /></span>
          <span id="gtJWrap3" style="display:inline-flex;align-items:center;margin-right:8px">评委3：<input type="number" id="gtJ3" step="0.01" style="width:70px;margin-left:4px" /></span>
          <span id="gtJWrap4" style="display:inline-flex;align-items:center;margin-right:8px">评委4：<input type="number" id="gtJ4" step="0.01" style="width:70px;margin-left:4px" /></span>
          <span id="gtJWrap5" style="display:inline-flex;align-items:center;margin-right:8px">评委5：<input type="number" id="gtJ5" step="0.01" style="width:70px;margin-left:4px" /></span>
          <span id="gtJWrap6" style="display:none;align-items:center;margin-right:8px">评委6：<input type="number" id="gtJ6" step="0.01" style="width:70px;margin-left:4px" /></span>
          <span id="gtJWrap7" style="display:none;align-items:center;margin-right:8px">评委7：<input type="number" id="gtJ7" step="0.01" style="width:70px;margin-left:4px" /></span>
          最终得分：<input type="number" id="gtFinal" step="0.01" style="width:70px" />
        </div>
        <div class="action-row" style="margin-bottom:10px">
          <button type="button" id="btnEvolve" onclick="return window.__zhifeiFallbackClick(event, 'btnEvolve')">学习进化（根据已录入真实评标生成高分逻辑与编制指导）</button>
          <button type="button" id="btnOllamaPreview" class="secondary" onclick="return window.__zhifeiFallbackClick(event, 'btnOllamaPreview')">Ollama 增强预览</button>
          <button type="button" id="btnOllamaPreviewCopy" class="secondary" disabled>复制预览结果</button>
          <button type="button" id="btnOllamaPreviewExport" class="secondary" disabled>导出 JSON</button>
          <span id="ollamaPreviewActionStatus" class="note">生成预览后可复制或导出。</span>
          <button type="button" id="btnEvolutionHealth" class="secondary" onclick="return window.__zhifeiFallbackClick(event, 'btnEvolutionHealth')">进化健康度</button>
          <button type="button" id="btnWritingGuidance" class="secondary" onclick="return window.__zhifeiFallbackClick(event, 'btnWritingGuidance')">查看编制指导</button>
          <button type="button" id="btnCompilationInstructions" class="secondary" onclick="return window.__zhifeiFallbackClick(event, 'btnCompilationInstructions')">编制系统指令（可导出为编制约束）</button>
        </div>
        <details open style="margin:12px 0 8px 0;padding:10px;border:2px solid #f59e0b;border-radius:8px;background:#fff7ed">
          <summary style="cursor:pointer;color:#9a3412"><strong>V2 反演校准闭环（核心能力，强烈建议执行）</strong></summary>
          <p style="margin:8px 0 10px 0;color:#7c2d12;font-size:13px">
            该闭环会自动训练并部署校准器（CV闸门）、回填预测分，并联动补丁影子评估/发布。
            为使系统评分持续逼近青天标准，建议每次录入真实评标后执行一次「一键闭环执行」。
          </p>
          <div style="margin-top:8px">
            <button type="button" id="btnRebuildDelta" class="secondary">重建 DELTA_CASE</button>
            <button type="button" id="btnRebuildSamples" class="secondary">重建 FEATURE_ROW</button>
            <button type="button" id="btnTrainCalibratorV2" class="secondary">训练并部署校准器</button>
            <button type="button" id="btnApplyCalibPredict" class="secondary">回填预测分</button>
            <button type="button" id="btnAutoRunReflection">一键闭环执行</button>
            <button type="button" id="btnEvalMetricsV2" class="secondary">评估 V1/V2/校准</button>
            <button type="button" id="btnEvalSummaryV2" class="secondary">跨项目汇总评估</button>
          </div>
          <div style="margin-top:8px">
            补丁类型：
            <select id="patchType" style="margin:0 8px">
              <option value="threshold">threshold</option>
              <option value="requirement">requirement</option>
              <option value="keywords">keywords</option>
            </select>
            <button type="button" id="btnMinePatchV2" class="secondary">挖掘 PATCH_PACKAGE</button>
            <input id="patchIdInput" placeholder="补丁ID（可留空自动取最新）" style="width:260px;margin-left:8px" />
            <button type="button" id="btnShadowPatchV2" class="secondary">影子评估</button>
            <button type="button" id="btnDeployPatchV2" class="secondary">发布补丁</button>
            <button type="button" id="btnRollbackPatchV2" class="secondary">回滚补丁</button>
          </div>
        </details>
        <div id="evolveResult" class="result-block" style="display:none"></div>
        <div id="ollamaPreviewResult" class="result-block" style="display:none"></div>
        <div id="evolutionHealthResult" class="result-block" style="display:none"></div>
        <div id="compilationInstructionsResult" class="result-block" style="display:none"></div>
        <div id="guidanceResult" class="result-block" style="display:none"></div>
        <div id="deltaResult" class="result-block" style="display:none"></div>
        <div id="sampleResult" class="result-block" style="display:none"></div>
        <div id="calibTrainResult" class="result-block" style="display:none"></div>
        <div id="patchResult" class="result-block" style="display:none"></div>
        <div id="patchShadowResult" class="result-block" style="display:none"></div>
        <div id="patchDeployResult" class="result-block" style="display:none"></div>
        <div id="evalResult" class="result-block" style="display:none"></div>
      </div>

      <div class="section card">
        <h2>原始输出（最后一次请求）</h2>
        <pre id="output">（操作后这里显示原始 JSON）</pre>
      </div>

      <script>
        (function () {
          const BOOTSTRAP_SCORE_SCALE_MAX = "__PROJECT_SCORE_SCALE_MAX__";
          const bootScoreScaleEl = document.getElementById('scoreScaleSelect');
          if (bootScoreScaleEl) {
            bootScoreScaleEl.value = BOOTSTRAP_SCORE_SCALE_MAX === '5' ? '5' : '100';
          }
          const FALLBACK_ACTIONS = Object.assign(window.__ZHIFEI_FALLBACK_ACTIONS || {}, {
            btnWeightsSave: { resultId: 'output', method: 'PUT', path: (pid) => '/api/v1/projects/' + pid + '/expert-profile', loading: '专家配置保存中...' },
            btnWeightsApply: { resultId: 'output', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/rescore', loading: '按当前关注度重算中...' },
            btnUploadMaterials: { resultId: 'materialsActionStatus', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/materials', loading: '资料上传中...' },
            btnUploadBoq: { resultId: 'materialsActionStatusBoq', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/materials', loading: '清单上传中...' },
            btnUploadDrawing: { resultId: 'materialsActionStatusDrawing', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/materials', loading: '图纸上传中...' },
            btnUploadSitePhotos: { resultId: 'materialsActionStatusPhoto', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/materials', loading: '现场照片上传中...' },
            btnMaterialDepthReport: { resultId: 'materialDepthReportResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/materials/depth_report', loading: '资料深读体检生成中...' },
            btnMaterialDepthReportDownload: { resultId: 'materialDepthReportResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/materials/depth_report.md', loading: '体检报告下载准备中...' },
            btnMaterialKnowledgeProfile: { resultId: 'materialKnowledgeProfileResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/materials/knowledge_profile', loading: '资料知识画像生成中...' },
            btnMaterialKnowledgeProfileDownload: { resultId: 'materialKnowledgeProfileResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/materials/knowledge_profile.md', loading: '知识画像下载准备中...' },
            btnUploadShigong: { resultId: 'shigongActionStatus', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/shigong', loading: '施组上传中...' },
            btnScoreShigong: { resultId: 'shigongActionStatus', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/rescore', loading: '施组评分中...' },
            btnCompare: { resultId: 'compareResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/compare', loading: '对比排名加载中...' },
            btnCompareReport: { resultId: 'compareReportResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/compare_report', loading: '对比报告生成中...' },
            btnInsights: { resultId: 'insightsResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/insights', loading: '洞察分析中...' },
            btnLearning: { resultId: 'learningResult', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/learning', loading: '学习画像生成中...' },
            btnEvidenceTrace: { resultId: 'evidenceTraceResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/evidence_trace/latest', loading: '证据追溯生成中...' },
            btnScoringBasis: { resultId: 'scoringBasisResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/scoring_basis/latest', loading: '评分依据审计生成中...' },
            btnAdaptive: { resultId: 'adaptiveResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/adaptive', loading: '自适应建议生成中...' },
            btnAdaptivePatch: { resultId: 'adaptivePatchResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/adaptive_patch', loading: '补丁生成中...' },
            btnAdaptiveValidate: { resultId: 'adaptiveValidateResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/adaptive_validate', loading: '验证效果计算中...' },
            btnAdaptiveApply: { resultId: 'adaptiveApplyResult', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/adaptive_apply', loading: '应用补丁中...' },
            btnRefreshGroundTruth: { resultId: 'evolveResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/ground_truth', loading: '真实评标列表刷新中...' },
            btnRefreshGroundTruthSubmissionOptions: { resultId: 'evolveResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/submissions', loading: '施组选项刷新中...' },
            btnRefreshFeedMaterials: { resultId: 'evolveResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/materials', loading: '投喂包列表刷新中...' },
            btnUploadFeed: { resultId: 'evolveResult', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/materials', loading: '投喂包上传中...' },
            btnAddGroundTruth: { resultId: 'evolveResult', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/ground_truth/from_submission', loading: '真实评标录入中...' },
            btnEvolve: { resultId: 'evolveResult', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/evolve', loading: '学习进化执行中...' },
            btnOllamaPreview: { resultId: 'ollamaPreviewResult', method: 'POST', path: (pid) => '/api/v1/projects/' + pid + '/evolve/ollama_preview', loading: 'Ollama 增强预览生成中...' },
            btnEvolutionHealth: { resultId: 'evolutionHealthResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/evolution/health', loading: '进化健康度分析中...' },
            btnWritingGuidance: { resultId: 'guidanceResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/writing_guidance', loading: '正在生成编制指导...' },
            btnCompilationInstructions: { resultId: 'compilationInstructionsResult', method: 'GET', path: (pid) => '/api/v1/projects/' + pid + '/compilation_instructions', loading: '正在生成编制系统指令...' },
          });
          window.__ZHIFEI_FALLBACK_ACTIONS = FALLBACK_ACTIONS;
          const FALLBACK_MATERIAL_UPLOAD_ACTIONS = {
            btnUploadMaterials: {
              formId: 'uploadMaterial',
              materialType: 'tender_qa',
              resultId: 'materialsActionStatus',
              typeLabel: '招标文件和答疑',
            },
            btnUploadBoq: {
              formId: 'uploadMaterialBoq',
              materialType: 'boq',
              resultId: 'materialsActionStatusBoq',
              typeLabel: '清单',
            },
            btnUploadDrawing: {
              formId: 'uploadMaterialDrawing',
              materialType: 'drawing',
              resultId: 'materialsActionStatusDrawing',
              typeLabel: '图纸',
            },
            btnUploadSitePhotos: {
              formId: 'uploadMaterialPhoto',
              materialType: 'site_photo',
              resultId: 'materialsActionStatusPhoto',
              typeLabel: '现场照片',
            },
          };
          function fallbackMaterialUploadActionConfig(actionId) {
            return FALLBACK_MATERIAL_UPLOAD_ACTIONS[String(actionId || '').trim()] || null;
          }
          function materialTypeLabel(typeKey) {
            const t = String(typeKey || '').trim();
            if (t === 'boq') return '清单';
            if (t === 'drawing') return '图纸';
            if (t === 'site_photo') return '现场照片';
            return '招标文件和答疑';
          }

          function fallbackGetProjectId() {
            const sel = document.getElementById('projectSelect');
            return (sel && sel.value) ? sel.value : '';
          }
          function fallbackResolveProjectId(projectIdOverride='') {
            const explicit = String(projectIdOverride || '').trim();
            if (explicit) return explicit;
            const fromSelect = fallbackGetProjectId();
            if (fromSelect) return fromSelect;
            const hiddenIds = [
              'uploadShigongProjectId',
              'uploadMaterialProjectId',
              'uploadMaterialBoqProjectId',
              'uploadMaterialDrawingProjectId',
              'uploadMaterialPhotoProjectId',
            ];
            for (const hid of hiddenIds) {
              const el = document.getElementById(hid);
              const v = String((el && el.value) || '').trim();
              if (v) return v;
            }
            try {
              const qid = new URL(window.location.href).searchParams.get('project_id') || '';
              if (String(qid).trim()) return String(qid).trim();
            } catch (_) {}
            return '';
          }
          function fallbackApiKey() {
            try { return localStorage.getItem('api_key') || ''; } catch (_) { return ''; }
          }
          function fallbackCollectWeightsRaw() {
            const m = {};
            const sliders = Array.from(document.querySelectorAll('input[type="range"][data-dim]') || []);
            sliders.forEach((el) => {
              const dim = String(el.getAttribute('data-dim') || '');
              if (!dim) return;
              const v = parseInt(String(el.value || '5'), 10);
              m[dim] = Number.isFinite(v) ? Math.max(0, Math.min(10, v)) : 5;
            });
            return m;
          }
          function fallbackApplyWeightUi(rawMap) {
            const dims = Array.from({ length: 16 }, (_, i) => String(i + 1).padStart(2, '0'));
            const effective = {};
            let sum = 0;
            dims.forEach((d) => {
              const v = Number(rawMap && rawMap[d] != null ? rawMap[d] : 5);
              const vv = Number.isFinite(v) ? Math.max(0, Math.min(10, v)) : 5;
              effective[d] = vv;
              const slider = document.getElementById('w_' + d);
              if (slider) slider.value = String(vv);
              const rawEl = document.getElementById('w_raw_' + d);
              if (rawEl) rawEl.textContent = String(vv);
              sum += (0.5 + vv / 10.0);
            });
            const safeSum = sum > 0 ? sum : (dims.length * 1.0);
            dims.forEach((d) => {
              const mv = 0.5 + (effective[d] / 10.0);
              const pct = mv / safeSum * 100.0;
              const normEl = document.getElementById('w_norm_' + d);
              if (normEl) normEl.textContent = pct.toFixed(2) + '%';
            });
            const summary = dims.map((d) => {
              const mv = 0.5 + (effective[d] / 10.0);
              return d + ':' + (mv / safeSum * 100.0).toFixed(2) + '%';
            }).join(' | ');
            const summaryEl = document.getElementById('expertWeightsSummary');
            if (summaryEl) summaryEl.textContent = summary;
          }
          function fallbackSetOutput(text) {
            const out = document.getElementById('output');
            if (out) out.textContent = String(text || '');
          }
          function fallbackSetResult(resultId, msg, isError) {
            const el = resultId ? document.getElementById(resultId) : null;
            if (!el) return;
            el.style.display = 'block';
            const klass = isError ? 'error' : 'success';
            el.innerHTML = '<span class="' + klass + '">' + String(msg || '').replace(/</g, '&lt;').replace(/>/g, '&gt;') + '</span>';
          }
          function fallbackSetLoading(resultId, msg) {
            const el = resultId ? document.getElementById(resultId) : null;
            if (!el) return;
            el.style.display = 'block';
            el.innerHTML = '<span style="color:#334155">' + String(msg || '处理中...').replace(/</g, '&lt;').replace(/>/g, '&gt;') + '</span>';
          }
          function fallbackEscapeHtml(v) {
            return String(v == null ? '' : v)
              .replace(/&/g, '&amp;')
              .replace(/</g, '&lt;')
              .replace(/>/g, '&gt;')
              .replace(/"/g, '&quot;')
              .replace(/'/g, '&#39;');
          }
          function fallbackSetResultHtml(resultId, html) {
            const el = resultId ? document.getElementById(resultId) : null;
            if (!el) return;
            el.style.display = 'block';
            el.innerHTML = html || '';
          }
          function fallbackParseJson(text) {
            try { return JSON.parse(text || '{}'); } catch (_) { return {}; }
          }
          function fallbackFormatPageHint(raw) {
            const s = String(raw == null ? '' : raw).trim();
            if (!s) return '页码待核对';
            const m = s.match(/(\\d+)/);
            if (m && m[1]) return '第' + m[1] + '页';
            return s;
          }
          function fallbackRenderActionSuccess(actionId, resultId, status, text) {
            const aid = String(actionId || '').trim();
            const data = fallbackParseJson(text);
            if (aid === 'btnCompare') {
              const rows = Array.isArray(data.rankings) ? data.rankings : [];
              const html = '<strong>排名</strong><table><tr><th>文件名</th><th>总分</th><th>时间</th></tr>' +
                (rows.length
                  ? rows.map((r) => '<tr><td>' + fallbackEscapeHtml(r.filename || '') + '</td><td>' + fallbackEscapeHtml(r.total_score) + '</td><td>' + fallbackEscapeHtml(r.created_at || '') + '</td></tr>').join('')
                  : '<tr><td colspan="3">暂无施组评分数据</td></tr>') +
                '</table>';
              fallbackSetResultHtml(resultId, html);
              return true;
            }
            if (aid === 'btnCompareReport') {
              const summary = fallbackEscapeHtml(data.summary || '');
              const top = data.top_submission || {};
              const bottom = data.bottom_submission || {};
              const keyDiffs = (Array.isArray(data.key_diffs) ? data.key_diffs : []).slice(0, 5);
              const rankings = Array.isArray(data.rankings) ? data.rankings : [];
              const sourceCards = Array.isArray(data.submission_optimization_cards) ? data.submission_optimization_cards : [];
              const cards = sourceCards.slice(0, 5);
              // 若后端仅返回部分优化卡片，按排名补齐到 5 份，避免“只看到 3 份”
              if (cards.length < 5 && rankings.length > cards.length) {
                const existing = new Set(cards.map((c) => String((c && c.filename) || '')));
                rankings.forEach((r) => {
                  if (cards.length >= 5) return;
                  const fn = String((r && r.filename) || '');
                  if (!fn || existing.has(fn)) return;
                  cards.push({
                    filename: fn,
                    total_score: r && r.total_score,
                    target_score: top && top.total_score,
                    recommendations: [],
                  });
                  existing.add(fn);
                });
              }
              const cardHtml = cards.map((c) => {
                const recs = (Array.isArray(c.recommendations) ? c.recommendations : []).slice(0, 2);
                return '<div style="margin:8px 0;padding:8px;border:1px solid #e2e8f0;border-radius:6px">'
                  + '<div><strong>' + fallbackEscapeHtml(c.filename || '-') + '</strong>（当前 ' + fallbackEscapeHtml(c.total_score) + '，目标 ' + fallbackEscapeHtml(c.target_score) + '）</div>'
                  + (recs.length
                    ? '<ol style="margin:6px 0 0 18px">' + recs.map((r) => '<li>' + fallbackEscapeHtml(fallbackFormatPageHint(r.page_hint || '')) + '：'
                      + fallbackEscapeHtml(r.issue || '需补充可量化执行要素')
                      + '；优化建议：' + fallbackEscapeHtml(r.rewrite_instruction || r.action || '按“责任人-频次-阈值-验收”四段式补齐。')
                      + '</li>').join('') + '</ol>'
                    : '<div style="color:#64748b;margin-top:6px">当前未提取到该文件的具体优化动作，建议先补充证据片段后再生成。</div>')
                  + '</div>';
              }).join('');
              const html = ''
                + '<p><strong>摘要</strong>：' + (summary || '无') + '</p>'
                + '<p><strong>最高分</strong>：' + fallbackEscapeHtml(top.filename || '-') + '（' + fallbackEscapeHtml(top.total_score) + '）'
                + '，<strong>最低分</strong>：' + fallbackEscapeHtml(bottom.filename || '-') + '（' + fallbackEscapeHtml(bottom.total_score) + '）</p>'
                + (keyDiffs.length
                  ? '<strong>关键差距维度（Top5）</strong><table><tr><th>维度</th><th>分差</th></tr>'
                    + keyDiffs.map((d) => '<tr><td>' + fallbackEscapeHtml((d.dimension || d.dim_id || '-') + ' ' + (d.dimension_name || '')) + '</td><td>' + fallbackEscapeHtml(d.delta) + '</td></tr>').join('')
                    + '</table>'
                  : '')
                + '<strong>逐文件优化清单（精简执行版）</strong>'
                + (cardHtml || '<div style="color:#64748b">暂无优化清单。</div>');
              fallbackSetResultHtml(resultId, html);
              return true;
            }
            if (aid === 'btnInsights') {
              const weak = Array.isArray(data.weakest_dims) ? data.weakest_dims : [];
              const recs = Array.isArray(data.recommendations) ? data.recommendations : [];
              const html = '<strong>洞察结果</strong>'
                + (weak.length ? ('<ul>' + weak.map((d) => '<li>' + fallbackEscapeHtml((d.dimension || d.dimension_id || '') + '：' + (d.avg_score || d.avg || '-')) + '</li>').join('') + '</ul>') : '<p>暂无弱项维度。</p>')
                + (recs.length ? ('<strong>建议</strong><ul>' + recs.map((r) => '<li>' + fallbackEscapeHtml((r.reason || '') + ' — ' + (r.action || '')) + '</li>').join('') + '</ul>') : '');
              fallbackSetResultHtml(resultId, html);
              return true;
            }
            if (aid === 'btnEvolutionHealth') {
              const summary = (data && typeof data.summary === 'object') ? data.summary : {};
              const drift = (data && typeof data.drift === 'object') ? data.drift : {};
              const w30 = (data && data.windows && typeof data.windows.recent_30d === 'object') ? data.windows.recent_30d : {};
              const wall = (data && data.windows && typeof data.windows.all === 'object') ? data.windows.all : {};
              const recs = Array.isArray(data.recommendations) ? data.recommendations : [];
              const html = '<strong>进化健康度</strong>'
                + '<p style="margin:6px 0">漂移等级：'
                + fallbackEscapeHtml((drift && drift.level) || 'insufficient_data')
                + '；近30天 MAE=' + fallbackEscapeHtml((w30 && w30.mae) != null ? w30.mae : '-')
                + '；全量 MAE=' + fallbackEscapeHtml((wall && wall.mae) != null ? wall.mae : '-')
                + '</p>'
                + '<table><tr><th>指标</th><th>值</th></tr>'
                + '<tr><td>真实评分样本</td><td>' + fallbackEscapeHtml((summary && summary.ground_truth_count) || 0) + '</td></tr>'
                + '<tr><td>已匹配预测</td><td>' + fallbackEscapeHtml((summary && summary.matched_prediction_count) || 0) + '</td></tr>'
                + '<tr><td>未匹配样本</td><td>' + fallbackEscapeHtml((summary && summary.unmatched_ground_truth_count) || 0) + '</td></tr>'
                + '<tr><td>当前权重来源</td><td>' + fallbackEscapeHtml((summary && summary.current_weights_source) || '-') + '</td></tr>'
                + '</table>'
                + (recs.length
                  ? '<ul style="margin:6px 0 0 18px;color:#92400e">' + recs.slice(0, 6).map((x) => '<li>' + fallbackEscapeHtml(x) + '</li>').join('') + '</ul>'
                  : '');
              fallbackSetResultHtml(resultId, html);
              return true;
            }
            if (aid === 'btnAdaptive' || aid === 'btnAdaptivePatch' || aid === 'btnAdaptiveValidate' || aid === 'btnAdaptiveApply') {
              fallbackSetResultHtml(resultId, '<strong>自适应结果</strong><pre>' + fallbackEscapeHtml(text || '{}') + '</pre>');
              return true;
            }
            if (aid === 'btnMaterialDepthReport') {
              const byType = Array.isArray(data.by_type) ? data.by_type : [];
              const quality = (data && typeof data.quality_summary === 'object') ? data.quality_summary : {};
              const recs = Array.isArray(data.recommendations) ? data.recommendations : [];
              const html = '<strong>资料深读体检</strong>'
                + '<p style="margin:6px 0">评分就绪：'
                + (data && data.ready_to_score ? '<span class="success">是</span>' : '<span class="error">否</span>')
                + '；总分块 ' + fallbackEscapeHtml(quality.total_parsed_chunks || 0)
                + '；数字约束项 ' + fallbackEscapeHtml(quality.total_numeric_terms || 0)
                + '</p>'
                + '<table><tr><th>资料类型</th><th>文件数</th><th>解析字数</th><th>分块数</th><th>数字约束项</th></tr>'
                + (byType.length
                  ? byType.map((row) => '<tr>'
                    + '<td>' + fallbackEscapeHtml((row && row.material_type_label) || (row && row.material_type) || '-') + '</td>'
                    + '<td>' + fallbackEscapeHtml((row && row.files) || 0) + '</td>'
                    + '<td>' + fallbackEscapeHtml((row && row.parsed_chars) || 0) + '</td>'
                    + '<td>' + fallbackEscapeHtml((row && row.parsed_chunks) || 0) + '</td>'
                    + '<td>' + fallbackEscapeHtml((row && row.numeric_terms) || 0) + '</td>'
                    + '</tr>').join('')
                  : '<tr><td colspan="5">暂无资料体检数据</td></tr>')
                + '</table>'
                + (recs.length
                  ? '<ul style="margin:6px 0 0 18px;color:#92400e">' + recs.slice(0, 6).map((x) => '<li>' + fallbackEscapeHtml(x) + '</li>').join('') + '</ul>'
                  : '');
              fallbackSetResultHtml(resultId, html);
              return true;
            }
            if (aid === 'btnMaterialKnowledgeProfile') {
              const summary = (data && typeof data.summary === 'object') ? data.summary : {};
              const dims = Array.isArray(data.by_dimension) ? data.by_dimension : [];
              const rows = dims.slice().sort((a, b) => Number((b && b.coverage_score) || 0) - Number((a && a.coverage_score) || 0)).slice(0, 8);
              const recs = Array.isArray(data.recommendations) ? data.recommendations : [];
              const html = '<strong>资料知识画像</strong>'
                + '<p style="margin:6px 0">维度覆盖率：'
                + fallbackEscapeHtml(summary.dimension_coverage_rate || 0)
                + '；低覆盖维度：'
                + fallbackEscapeHtml(summary.low_coverage_dimensions || 0)
                + '；总解析字数：'
                + fallbackEscapeHtml(summary.total_parsed_chars || 0)
                + '</p>'
                + '<table><tr><th>维度</th><th>关键词命中</th><th>来源类型数</th><th>覆盖评分</th><th>等级</th></tr>'
                + (rows.length
                  ? rows.map((row) => '<tr>'
                    + '<td>' + fallbackEscapeHtml((row && row.dimension_id) || '-') + ' ' + fallbackEscapeHtml((row && row.dimension_name) || '') + '</td>'
                    + '<td>' + fallbackEscapeHtml((row && row.keyword_hits) || 0) + '</td>'
                    + '<td>' + fallbackEscapeHtml(((row && row.source_types) || []).length) + '</td>'
                    + '<td>' + fallbackEscapeHtml((row && row.coverage_score) || 0) + '</td>'
                    + '<td>' + fallbackEscapeHtml((row && row.coverage_level) || '-') + '</td>'
                    + '</tr>').join('')
                  : '<tr><td colspan="5">暂无知识画像数据</td></tr>')
                + '</table>'
                + (recs.length
                  ? '<ul style="margin:6px 0 0 18px;color:#92400e">' + recs.slice(0, 6).map((x) => '<li>' + fallbackEscapeHtml(x) + '</li>').join('') + '</ul>'
                  : '');
              fallbackSetResultHtml(resultId, html);
              return true;
            }
            if (aid === 'btnScoreShigong') {
              const updated = Number(
                (data && (data.updated_submissions ?? data.reports_generated ?? data.submission_count)) || 0
              );
              const scaleLabel = (data && data.score_scale_label) ? String(data.score_scale_label) : selectedScoreScaleLabel();
              const gate = (data && typeof data.material_utilization_gate === 'object') ? data.material_utilization_gate : {};
              const blockedCount = Number((gate && gate.blocked_submissions) || 0);
              const warnCount = Number((gate && gate.warn_submissions) || 0);
              let msg = '评分完成（' + scaleLabel + '）：已重算 ' + updated + ' 份。';
              if (blockedCount > 0) {
                msg += ' 资料门禁阻断 ' + blockedCount + ' 份。';
              } else if (warnCount > 0) {
                msg += ' 资料门禁预警 ' + warnCount + ' 份。';
              } else if (gate && gate.enabled) {
                msg += ' 资料门禁通过。';
              }
              fallbackSetResult(resultId, msg, blockedCount > 0);
              if (window.renderMaterialUtilizationPanel && typeof window.renderMaterialUtilizationPanel === 'function') {
                window.renderMaterialUtilizationPanel(data || {});
              }
              return true;
            }
            if (aid === 'btnUploadMaterials' || aid === 'btnUploadBoq' || aid === 'btnUploadDrawing' || aid === 'btnUploadSitePhotos' || aid === 'btnUploadShigong' || aid === 'btnScoreShigong' || aid === 'btnLearning' || aid === 'btnEvolve' || aid === 'btnRefreshGroundTruth' || aid === 'btnUploadFeed' || aid === 'btnAddGroundTruth' || aid === 'btnRefreshFeedMaterials' || aid === 'btnWritingGuidance' || aid === 'btnCompilationInstructions') {
              fallbackSetResultHtml(resultId, '<span class="success">[' + fallbackEscapeHtml(aid) + '] 请求成功 (HTTP ' + fallbackEscapeHtml(status) + ')</span><pre>' + fallbackEscapeHtml(text || '{}') + '</pre>');
              return true;
            }
            return false;
          }
          function fallbackStopEvent(ev) {
            if (!ev) return;
            if (typeof ev.preventDefault === 'function') ev.preventDefault();
            if (typeof ev.stopPropagation === 'function') ev.stopPropagation();
            if (typeof ev.stopImmediatePropagation === 'function') ev.stopImmediatePropagation();
          }
          function fallbackJsonBodyHeaders() {
            const headers = { 'Content-Type': 'application/json' };
            const key = fallbackApiKey();
            if (key) headers['X-API-Key'] = key;
            return headers;
          }
          function fallbackAuthHeaders() {
            const headers = {};
            const key = fallbackApiKey();
            if (key) headers['X-API-Key'] = key;
            return headers;
          }
          async function fallbackFetchScoringReadiness(projectId) {
            const pid = String(projectId || '').trim();
            if (!pid) return null;
            let res;
            let text = '';
            let data = null;
            try {
              res = await fetch('/api/v1/projects/' + encodeURIComponent(pid) + '/scoring_readiness?t=' + Date.now(), {
                method: 'GET',
                headers: fallbackAuthHeaders(),
                cache: 'no-store',
              });
              text = await res.text();
              data = fallbackParseJson(text);
            } catch (_) {
              return null;
            }
            if (!res.ok || !data || typeof data !== 'object') {
              return null;
            }
            if (window.renderScoringReadinessPanel && typeof window.renderScoringReadinessPanel === 'function') {
              window.renderScoringReadinessPanel(data);
            }
            return data;
          }
          async function fallbackRefreshMaterialsTable(projectId) {
            const pid = String(projectId || '').trim();
            if (!pid) return;
            const table = document.getElementById('materialsTable');
            const tbody = table ? table.querySelector('tbody') : null;
            const emptyEl = document.getElementById('materialsEmpty');
            if (tbody) tbody.innerHTML = '';
            let res;
            let text = '';
            let rows = [];
            try {
              res = await fetch('/api/v1/projects/' + encodeURIComponent(pid) + '/materials?t=' + Date.now(), {
                method: 'GET',
                headers: fallbackAuthHeaders(),
                cache: 'no-store',
              });
              text = await res.text();
              rows = fallbackParseJson(text);
            } catch (_) {
              if (emptyEl) {
                emptyEl.textContent = '资料列表加载失败，请稍后重试。';
                emptyEl.style.display = 'block';
              }
              return;
            }
            if (!res.ok || !Array.isArray(rows)) {
              if (emptyEl) {
                emptyEl.textContent = '资料列表加载失败（HTTP ' + String((res && res.status) || 0) + '）';
                emptyEl.style.display = 'block';
              }
              return;
            }
            if (!rows.length) {
              if (emptyEl) {
                emptyEl.textContent = '暂无资料，请下方添加。';
                emptyEl.style.display = 'block';
              }
              return;
            }
            if (emptyEl) emptyEl.style.display = 'none';
            rows.forEach((m) => {
              const tr = document.createElement('tr');
              const mid = fallbackEscapeHtml(String((m && m.id) || ''));
              const fn = fallbackEscapeHtml(String((m && m.filename) || ''));
              const mt = fallbackEscapeHtml(materialTypeLabel((m && m.material_type) || 'tender_qa'));
              const createdAt = fallbackEscapeHtml(String((m && m.created_at) || '').slice(0, 19));
              tr.innerHTML =
                '<td>' + mt + '</td>'
                + '<td>' + fn + '</td>'
                + '<td>' + createdAt + '</td>'
                + '<td><button type="button" class="btn-danger js-delete-material" data-material-id="' + mid + '" data-project-id="' + fallbackEscapeHtml(pid) + '" data-filename="' + fn + '">删除</button></td>';
              if (tbody) tbody.appendChild(tr);
            });
          }
          async function fallbackRefreshSubmissionsTable(projectId) {
            const pid = String(projectId || '').trim();
            if (!pid) return;
            const table = document.getElementById('submissionsTable');
            const tbody = table ? table.querySelector('tbody') : null;
            const emptyEl = document.getElementById('submissionsEmpty');
            if (tbody) tbody.innerHTML = '';
            let res;
            let text = '';
            let rows = [];
            try {
              res = await fetch('/api/v1/projects/' + encodeURIComponent(pid) + '/submissions?t=' + Date.now(), {
                method: 'GET',
                headers: fallbackAuthHeaders(),
                cache: 'no-store',
              });
              text = await res.text();
              rows = fallbackParseJson(text);
            } catch (_) {
              if (emptyEl) {
                emptyEl.textContent = '施组列表加载失败，请稍后重试。';
                emptyEl.style.display = 'block';
              }
              return;
            }
            if (!res.ok || !Array.isArray(rows)) {
              if (emptyEl) {
                emptyEl.textContent = '施组列表加载失败（HTTP ' + String((res && res.status) || 0) + '）';
                emptyEl.style.display = 'block';
              }
              await fallbackFetchScoringReadiness(pid);
              return;
            }
            if (!rows.length) {
              if (emptyEl) {
                emptyEl.textContent = '暂无施组，请下方添加。';
                emptyEl.style.display = 'block';
              }
              if (window.renderMaterialUtilizationPanel && typeof window.renderMaterialUtilizationPanel === 'function') {
                window.renderMaterialUtilizationPanel(null);
              }
              await fallbackFetchScoringReadiness(pid);
              return;
            }
            if (emptyEl) emptyEl.style.display = 'none';
            rows.forEach((s) => {
              const tr = document.createElement('tr');
              const sid = fallbackEscapeHtml(String((s && s.id) || ''));
              const fn = fallbackEscapeHtml(String((s && s.filename) || ''));
              const createdAt = fallbackEscapeHtml(String((s && s.created_at) || '').slice(0, 19));
              const report = (s && s.report) || {};
              const pred = report && report.pred_total_score != null ? report.pred_total_score : null;
              const rule = report && report.rule_total_score != null ? report.rule_total_score : null;
              const llm = report && report.llm_total_score != null ? report.llm_total_score : null;
              const scoringStatus = String((report && report.scoring_status) || '').toLowerCase();
              const isPending = scoringStatus === 'pending';
              const isBlocked = scoringStatus === 'blocked';
              const reportMeta = (report && typeof report.meta === 'object') ? report.meta : {};
              const utilGate = (reportMeta && typeof reportMeta.material_utilization_gate === 'object')
                ? reportMeta.material_utilization_gate
                : {};
              const evidenceTrace = (reportMeta && typeof reportMeta.evidence_trace === 'object')
                ? reportMeta.evidence_trace
                : {};
              const utilBlocked = !!(utilGate && utilGate.blocked);
              let scoreHtml = '-';
              if (isPending) {
                scoreHtml = '<span class="note">待评分</span>';
              } else if (isBlocked) {
                scoreHtml = '<span class="error">待补资料后重评分</span>';
              } else if (pred != null) {
                scoreHtml = fallbackEscapeHtml(String(pred));
                const notes = [];
                if (rule != null) notes.push('规则: ' + fallbackEscapeHtml(String(rule)));
                if (llm != null) notes.push('LLM: ' + fallbackEscapeHtml(String(llm)));
                if (notes.length) scoreHtml += '<div class="note">' + notes.join(' / ') + '</div>';
              } else if (s && s.total_score != null) {
                scoreHtml = fallbackEscapeHtml(String(s.total_score));
              }
              const evidenceCount = Number(evidenceTrace.total_hits || 0);
              const evidenceFileCount = Number(evidenceTrace.source_files_hit_count || 0);
              if (!isPending && evidenceCount > 0) {
                scoreHtml += '<div class="note">证据命中: '
                  + fallbackEscapeHtml(String(evidenceCount))
                  + ' 条 / 文件覆盖: '
                  + fallbackEscapeHtml(String(evidenceFileCount))
                  + ' 份</div>';
              }
              const utilSummary = (reportMeta && typeof reportMeta.material_utilization === 'object')
                ? reportMeta.material_utilization
                : {};
              const utilByType = (utilSummary && typeof utilSummary.by_type === 'object')
                ? utilSummary.by_type
                : {};
              const utilAvailableTypes = Array.isArray(utilSummary.available_types) ? utilSummary.available_types : [];
              const typeLabelShort = (t) => {
                const key = String(t || '').trim();
                if (key === 'tender_qa') return '招答';
                if (key === 'boq') return '清单';
                if (key === 'drawing') return '图纸';
                if (key === 'site_photo') return '照片';
                return key || '-';
              };
              const hasTypeEvidence = (t) => {
                const row = (utilByType && typeof utilByType[t] === 'object') ? utilByType[t] : {};
                const retrievalHit = Number((row && row.retrieval_hit) || 0);
                const consistencyHit = Number((row && row.consistency_hit) || 0);
                return (retrievalHit + consistencyHit) > 0;
              };
              const orderedTypes = ['tender_qa', 'boq', 'drawing', 'site_photo'];
              const coverageTokens = orderedTypes.map((t) => {
                const inScope = utilAvailableTypes.includes(t);
                if (!inScope) return typeLabelShort(t) + '·';
                return typeLabelShort(t) + (hasTypeEvidence(t) ? '✓' : '×');
              });
              if (!isPending && coverageTokens.length) {
                scoreHtml += '<div class="note">类型覆盖: ' + fallbackEscapeHtml(coverageTokens.join(' / ')) + '</div>';
              }
              const evidenceFiles = Array.isArray(evidenceTrace.source_files_hit) ? evidenceTrace.source_files_hit : [];
              if (!isPending && evidenceFiles.length) {
                scoreHtml += '<div class="note">命中文件: ' + fallbackEscapeHtml(evidenceFiles.slice(0, 2).join('；')) + (evidenceFiles.length > 2 ? ' 等' : '') + '</div>';
              }
              if (utilBlocked) {
                scoreHtml += '<div class="error">资料利用门禁未达标（建议补齐资料后重评分）</div>';
              }
              tr.innerHTML =
                '<td>' + fn + '</td>'
                + '<td>' + scoreHtml + '</td>'
                + '<td>' + createdAt + '</td>'
                + '<td><button type="button" class="btn-danger js-delete-submission" data-submission-id="' + sid + '" data-project-id="' + fallbackEscapeHtml(pid) + '" data-filename="' + fn + '">删除</button></td>';
              if (tbody) tbody.appendChild(tr);
            });
            if (window.renderMaterialUtilizationPanel && typeof window.renderMaterialUtilizationPanel === 'function') {
              let utilPayload = null;
              for (const s of rows) {
                const rep = (s && typeof s === 'object') ? (s.report || {}) : {};
                const meta = (rep && typeof rep.meta === 'object') ? rep.meta : {};
                const util = (meta && typeof meta.material_utilization === 'object') ? meta.material_utilization : null;
                if (!util) continue;
                utilPayload = {
                  material_utilization: util,
                  material_utilization_alerts: Array.isArray(meta.material_utilization_alerts)
                    ? meta.material_utilization_alerts
                    : [],
                  material_utilization_gate: (meta && typeof meta.material_utilization_gate === 'object')
                    ? meta.material_utilization_gate
                    : {},
                };
                break;
              }
              if (utilPayload) window.renderMaterialUtilizationPanel(utilPayload);
              else window.renderMaterialUtilizationPanel(null);
            }
            await fallbackFetchScoringReadiness(pid);
          }
          async function fallbackRefreshAfter(actionId) {
            const projectId = fallbackGetProjectId();
            if (fallbackMaterialUploadActionConfig(actionId)) {
              if (typeof refreshMaterials === 'function') await Promise.resolve(refreshMaterials(projectId));
              else await fallbackRefreshMaterialsTable(projectId);
              if (typeof refreshFeedMaterials === 'function') await Promise.resolve(refreshFeedMaterials(projectId));
              else await fallbackRefreshMaterialsTable(projectId);
              if (typeof refreshScoringReadiness === 'function') await Promise.resolve(refreshScoringReadiness(projectId));
              else await fallbackFetchScoringReadiness(projectId);
              return;
            }
            if (actionId === 'btnUploadShigong' || actionId === 'btnScoreShigong') {
              if (typeof refreshSubmissions === 'function') await Promise.resolve(refreshSubmissions(projectId));
              else await fallbackRefreshSubmissionsTable(projectId);
              if (typeof refreshGroundTruthSubmissionOptions === 'function') {
                await Promise.resolve(refreshGroundTruthSubmissionOptions(projectId));
              }
              if (typeof refreshScoringReadiness === 'function') await Promise.resolve(refreshScoringReadiness(projectId));
              else await fallbackFetchScoringReadiness(projectId);
              return;
            }
            if (actionId === 'btnUploadFeed') {
              if (typeof refreshMaterials === 'function') await Promise.resolve(refreshMaterials(projectId));
              else await fallbackRefreshMaterialsTable(projectId);
              if (typeof refreshFeedMaterials === 'function') await Promise.resolve(refreshFeedMaterials(projectId));
              else await fallbackRefreshMaterialsTable(projectId);
              return;
            }
            if (actionId === 'btnRefreshFeedMaterials') {
              if (typeof refreshFeedMaterials === 'function') await Promise.resolve(refreshFeedMaterials(projectId));
              else await fallbackRefreshMaterialsTable(projectId);
              return;
            }
            if (actionId === 'btnAddGroundTruth' || actionId === 'btnRefreshGroundTruth') {
              if (typeof refreshGroundTruth === 'function') await Promise.resolve(refreshGroundTruth(projectId));
              return;
            }
            if (actionId === 'btnRefreshGroundTruthSubmissionOptions') {
              if (typeof refreshGroundTruthSubmissionOptions === 'function') {
                await Promise.resolve(refreshGroundTruthSubmissionOptions(projectId));
              }
              return;
            }
            if (actionId === 'btnEvolve') {
              if (typeof refreshGroundTruth === 'function') await Promise.resolve(refreshGroundTruth(projectId));
            }
          }
          function fallbackRenderPayloadForAction(actionId, projectId) {
            if (actionId === 'btnWeightsSave') {
              const payload = { name: '', weights_raw: fallbackCollectWeightsRaw(), force_unlock: false };
              return { body: JSON.stringify(payload), headers: fallbackJsonBodyHeaders() };
            }
            if (actionId === 'btnWeightsApply') {
              const payload = {
                scoring_engine_version: 'v2',
                scope: 'project',
                score_scale_max: selectedScoreScaleMax(),
                rebuild_anchors: false,
                rebuild_requirements: false,
                retrain_calibrator: false,
                force_unlock: false,
              };
              return { body: JSON.stringify(payload), headers: fallbackJsonBodyHeaders() };
            }
            if (actionId === 'btnUploadFeed') {
              const fd = new FormData();
              const files = Array.from(((document.getElementById('feedFile') || {}).files) || []);
              files.forEach((f) => fd.append('file', f));
              return { body: fd, headers: fallbackAuthHeaders() };
            }
            const materialUploadCfg = fallbackMaterialUploadActionConfig(actionId);
            if (materialUploadCfg) {
              const form = document.getElementById(materialUploadCfg.formId);
              const fileInput = form && form.querySelector ? form.querySelector('input[name="file"]') : null;
              const files = Array.from((fileInput && fileInput.files) || []);
              const fd = new FormData();
              files.forEach((f) => fd.append('file', f));
              fd.append('material_type', materialUploadCfg.materialType);
              return { body: fd, headers: fallbackAuthHeaders() };
            }
            if (actionId === 'btnUploadShigong') {
              const form = document.getElementById('uploadShigong');
              const fileInput = form && form.querySelector ? form.querySelector('input[name="file"]') : null;
              const files = Array.from((fileInput && fileInput.files) || []);
              const fd = new FormData();
              files.forEach((f) => fd.append('file', f));
              return { body: fd, headers: fallbackAuthHeaders() };
            }
            if (actionId === 'btnScoreShigong') {
              const payload = {
                scoring_engine_version: 'v2',
                scope: 'project',
                score_scale_max: selectedScoreScaleMax(),
                rebuild_anchors: false,
                rebuild_requirements: false,
                retrain_calibrator: false,
                force_unlock: false,
              };
              return { body: JSON.stringify(payload), headers: fallbackJsonBodyHeaders() };
            }
            if (actionId === 'btnAddGroundTruth') {
              const selectedSubmissionId = String(((document.getElementById('groundTruthSubmissionSelect') || {}).value) || '').trim();
              const judgeScores = (typeof collectGroundTruthJudgeScores === 'function')
                ? collectGroundTruthJudgeScores()
                : [
                    parseFloat(((document.getElementById('gtJ1') || {}).value || '0')) || 0,
                    parseFloat(((document.getElementById('gtJ2') || {}).value || '0')) || 0,
                    parseFloat(((document.getElementById('gtJ3') || {}).value || '0')) || 0,
                    parseFloat(((document.getElementById('gtJ4') || {}).value || '0')) || 0,
                    parseFloat(((document.getElementById('gtJ5') || {}).value || '0')) || 0,
                  ];
              const finalScore = parseFloat(((document.getElementById('gtFinal') || {}).value || '0')) || 0;
              const payload = {
                submission_id: selectedSubmissionId,
                judge_scores: judgeScores,
                final_score: finalScore,
                source: '青天大模型',
              };
              return { body: JSON.stringify(payload), headers: fallbackJsonBodyHeaders() };
            }
            if (actionId === 'btnAdaptiveApply') {
              let key = fallbackApiKey();
              if (!key) {
                const prompted = prompt('应用补丁将修改 lexicon 配置，需要 API Key。请输入 X-API-Key（无则留空）：');
                key = prompted == null ? '' : prompted;
              }
              const headers = {};
              if (key) headers['X-API-Key'] = key;
              return { body: null, headers };
            }
            if (actionId === 'btnLearning' || actionId === 'btnEvolve') {
              return { body: null, headers: fallbackAuthHeaders() };
            }
            return { body: null, headers: fallbackAuthHeaders() };
          }
          async function fallbackRunAction(actionId) {
            actionId = String(actionId || '').trim();
            if (actionId === 'btnWeightsReset') {
              fallbackApplyWeightUi({});
              fallbackSetOutput('[btnWeightsReset] 已重置16维关注度到默认值(全部=5)');
              return true;
            }
            const cfg = FALLBACK_ACTIONS[actionId];
            if (!cfg) return false;
            const projectId = fallbackGetProjectId();
            if (!projectId) {
              fallbackSetResult(cfg.resultId, '请先在「2) 选择项目」中选择项目', true);
              fallbackSetOutput('[' + actionId + '] 缺少项目ID');
              return false;
            }
            fallbackSetLoading(cfg.resultId, cfg.loading);
            if (actionId === 'btnWeightsApply') {
              const saved = await fallbackRunAction('btnWeightsSave');
              if (!saved) return false;
            }
            const materialActionCfg = fallbackMaterialUploadActionConfig(actionId);
            if (materialActionCfg || actionId === 'btnUploadShigong') {
              const formId = materialActionCfg ? materialActionCfg.formId : 'uploadShigong';
              const typeLabel = materialActionCfg ? materialActionCfg.typeLabel : '施组';
              const form = document.getElementById(formId);
              const fileInput = form && form.querySelector ? form.querySelector('input[name="file"]') : null;
              const files = Array.from((fileInput && fileInput.files) || []);
              if (!files.length) {
                fallbackSetResult(cfg.resultId, '请先选择至少 1 个' + typeLabel + '文件。', true);
                fallbackSetOutput('[' + actionId + '] 未选择文件');
                return false;
              }
              let okCount = 0;
              let failCount = 0;
              const details = [];
              const headers = fallbackAuthHeaders();
              for (const f of files) {
                const fd = new FormData();
                fd.append('file', f);
                if (materialActionCfg) fd.append('material_type', materialActionCfg.materialType);
                try {
                  const res = await fetch(cfg.path(projectId), {
                    method: cfg.method || 'POST',
                    headers,
                    body: fd,
                  });
                  const text = await res.text();
                  if (res.ok) {
                    okCount += 1;
                    details.push('[成功] ' + String((f && f.name) || ''));
                  } else {
                    failCount += 1;
                    let detail = text || '';
                    try {
                      const j = JSON.parse(text || '{}');
                      detail = (j && j.detail) || detail;
                    } catch (_) {}
                    details.push('[失败] ' + String((f && f.name) || '') + ' -> HTTP ' + String(res.status || 0) + ' ' + String(detail).slice(0, 120));
                  }
                } catch (err) {
                  failCount += 1;
                  details.push('[失败] ' + String((f && f.name) || '') + ' -> ' + String((err && err.message) || err || '网络异常'));
                }
              }
              const summary = typeLabel + '上传完成：成功 ' + okCount + '，失败 ' + failCount;
              fallbackSetResult(cfg.resultId, summary, failCount > 0);
              fallbackSetOutput('[' + actionId + '] ' + summary + '\\n' + details.join('\\n'));
              if (okCount > 0) await fallbackRefreshAfter(actionId);
              if (fileInput && failCount === 0) fileInput.value = '';
              return failCount === 0;
            }
            if (actionId === 'btnAddGroundTruth') {
              const selectedSubmissionId = String(((document.getElementById('groundTruthSubmissionSelect') || {}).value) || '').trim();
              if (!selectedSubmissionId) {
                fallbackSetResult(cfg.resultId, '请先在“施组文件”下拉框选择步骤4已上传施组。', true);
                fallbackSetOutput('[' + actionId + '] 未选择施组文件');
                return false;
              }
            }
            if (actionId === 'btnMaterialDepthReportDownload') {
              const dlUrl = '/api/v1/projects/' + encodeURIComponent(projectId) + '/materials/depth_report.md';
              const a = document.createElement('a');
              a.href = dlUrl;
              a.download = 'material_depth_report_' + projectId + '.md';
              document.body.appendChild(a);
              a.click();
              a.remove();
              fallbackSetResult(cfg.resultId, '资料深读体检报告下载已触发。', false);
              fallbackSetOutput('[' + actionId + '] download ' + dlUrl);
              return true;
            }
            if (actionId === 'btnMaterialKnowledgeProfileDownload') {
              const dlUrl = '/api/v1/projects/' + encodeURIComponent(projectId) + '/materials/knowledge_profile.md';
              const a = document.createElement('a');
              a.href = dlUrl;
              a.download = 'material_knowledge_profile_' + projectId + '.md';
              document.body.appendChild(a);
              a.click();
              a.remove();
              fallbackSetResult(cfg.resultId, '资料知识画像报告下载已触发。', false);
              fallbackSetOutput('[' + actionId + '] download ' + dlUrl);
              return true;
            }
            if (actionId === 'btnScoreShigong') {
              const readiness = await fallbackFetchScoringReadiness(projectId);
              if (readiness && readiness.ready === false) {
                const issues = Array.isArray(readiness.issues) ? readiness.issues : [];
                const reason = issues.length ? String(issues[0]) : '评分前置条件未满足';
                const msg = '评分已阻止：' + reason;
                fallbackSetResult(cfg.resultId, msg, true);
                fallbackSetOutput('[' + actionId + '] ' + msg);
                return false;
              }
            }
            const req = fallbackRenderPayloadForAction(actionId, projectId);
            let res;
            let text = '';
            try {
              res = await fetch(cfg.path(projectId), {
                method: cfg.method || 'GET',
                headers: req.headers || {},
                body: req.body || undefined,
              });
              text = await res.text();
            } catch (err) {
              const msg = '[' + actionId + '] 网络异常：' + String((err && err.message) || err || 'unknown');
              fallbackSetResult(cfg.resultId, msg, true);
              fallbackSetOutput(msg);
              return false;
            }
            let detail = '';
            try {
              const j = JSON.parse(text || '{}');
              detail = String((j && j.detail) || '');
            } catch (_) {
              detail = '';
            }
            if (res.ok) {
              const rendered = fallbackRenderActionSuccess(actionId, cfg.resultId, res.status, text);
              if (!rendered) fallbackSetResult(cfg.resultId, '[' + actionId + '] 请求成功 (HTTP ' + String(res.status) + ')', false);
              await fallbackRefreshAfter(actionId);
              if (actionId === 'btnWeightsSave') {
                try {
                  const parsed = JSON.parse(text || '{}');
                  const profile = (parsed && parsed.expert_profile) || {};
                  fallbackApplyWeightUi(profile.weights_raw || {});
                  const statusEl = document.getElementById('expertProfileStatus');
                  if (statusEl) statusEl.textContent = '当前生效配置：' + String(profile.name || '项目默认配置') + '（ID: ' + String(profile.id || '-') + '）';
                } catch (_) {}
              }
            } else {
              fallbackSetResult(
                cfg.resultId,
                '[' + actionId + '] 请求失败 (HTTP ' + String(res.status) + ')' + (detail ? '：' + detail : ''),
                true
              );
            }
            if (actionId === 'btnCompareReport') {
              fallbackSetOutput('[' + actionId + '] HTTP ' + String(res.status) + '\\n已渲染精简版对比报告（详情已显示在页面，不再输出完整JSON）');
            } else {
              fallbackSetOutput('[' + actionId + '] HTTP ' + String(res.status) + '\\n' + String(text || ''));
            }
            return !!res.ok;
          }
          async function fallbackDelete(kind, fileId, filename, rowEl, projectIdOverride='') {
            const projectId = fallbackResolveProjectId(projectIdOverride);
            if (!projectId) { alert('请先选择项目'); return false; }
            if (!fileId) { alert('删除失败：记录ID为空'); return false; }
            if (!confirm('确认删除该文件？')) return false;
            const encodedPid = encodeURIComponent(String(projectId || ''));
            const encodedFid = encodeURIComponent(String(fileId || ''));
            let path = '/api/v1/projects/' + encodedPid + '/materials/' + encodedFid;
            if (kind === 'submission') path = '/api/v1/projects/' + encodedPid + '/shigong/' + encodedFid;
            if (kind === 'ground_truth') path = '/api/v1/projects/' + encodedPid + '/ground_truth/' + encodedFid;
            let res;
            let text = '';
            try {
              res = await fetch(path, { method: 'DELETE', headers: fallbackAuthHeaders() });
              text = await res.text();
            } catch (err) {
              alert('删除失败：' + String((err && err.message) || err || '网络异常'));
              return false;
            }
            if (!res.ok) {
              let detail = text || '';
              try {
                const j = JSON.parse(text || '{}');
                detail = (j && j.detail) || detail;
              } catch (_) {}
              alert('删除失败：HTTP ' + String(res.status) + ' ' + String(detail || '').slice(0, 180));
              fallbackSetOutput('删除失败：' + text);
              return false;
            }
            if (rowEl && typeof rowEl.remove === 'function') rowEl.remove();
            fallbackSetOutput(JSON.stringify({ ok: true, kind, id: fileId, filename: filename || '' }, null, 2));
            if (kind === 'material') {
              if (typeof refreshMaterials === 'function') refreshMaterials();
              if (typeof refreshFeedMaterials === 'function') refreshFeedMaterials();
            } else if (kind === 'submission' && typeof refreshSubmissions === 'function') {
              refreshSubmissions();
            } else if (kind === 'ground_truth' && typeof refreshGroundTruth === 'function') {
              refreshGroundTruth();
            }
            if (typeof refreshScoringReadiness === 'function') refreshScoringReadiness(projectId);
            else fallbackFetchScoringReadiness(projectId);
            return true;
          }
          window.__zhifeiFallbackClick = function (ev, actionId) {
            fallbackStopEvent(ev);
            fallbackRunAction(actionId);
            return false;
          };
          window.__zhifeiFallbackDelete = function (ev, kind, fileId, filename, projectIdOverride='') {
            const rowEl = ev && ev.target && ev.target.closest ? ev.target.closest('tr') : null;
            fallbackStopEvent(ev);
            fallbackDelete(kind, fileId, filename, rowEl, projectIdOverride);
            return false;
          };
          document.addEventListener('click', function (ev) {
            const btn = ev && ev.target && ev.target.closest
              ? ev.target.closest('.js-delete-material,.js-delete-submission,.js-delete-ground-truth')
              : null;
            if (!btn) return;
            let kind = 'material';
            if (btn.classList.contains('js-delete-submission')) kind = 'submission';
            if (btn.classList.contains('js-delete-ground-truth')) kind = 'ground_truth';
            let fileId = '';
            if (kind === 'submission') fileId = btn.getAttribute('data-submission-id') || '';
            else if (kind === 'ground_truth') fileId = btn.getAttribute('data-gt-id') || '';
            else fileId = btn.getAttribute('data-material-id') || '';
            const filename = btn.getAttribute('data-filename') || '';
            const projectIdOverride = btn.getAttribute('data-project-id') || '';
            window.__zhifeiFallbackDelete(ev, kind, fileId, filename, projectIdOverride);
          }, true);
        })();
      </script>
      <script>
        function reportClientError(prefix, err) {
          const msg = prefix + ': ' + String((err && (err.message || err.reason || err)) || '未知错误');
          console.error(msg, err);
          const out = document.getElementById('output');
          if (out) out.textContent = msg;
        }
        window.addEventListener('error', function (e) {
          reportClientError('前端脚本错误', e && (e.error || e.message || e));
        });
        window.addEventListener('unhandledrejection', function (e) {
          reportClientError('前端异步错误', e && (e.reason || e));
        });
        const SAFE_CLICK_HANDLERS = {};
        function safeClick(id, fn) {
          const el = document.getElementById(id);
          if (!el) return;
          const wrapped = async (ev) => {
            if (ev) ev.preventDefault();
            try {
              await fn(ev);
            } catch (err) {
              const cfg = (window.__ZHIFEI_FALLBACK_ACTIONS && window.__ZHIFEI_FALLBACK_ACTIONS[id]) || {};
              const msg = '执行失败：' + String((err && err.message) || err || '未知错误');
              if (cfg.resultId && typeof setResultError === 'function') {
                setResultError(cfg.resultId, msg);
              }
              reportClientError('按钮[' + id + ']执行失败', err);
            }
          };
          if (SAFE_CLICK_HANDLERS[id]) {
            el.removeEventListener('click', SAFE_CLICK_HANDLERS[id]);
          }
          SAFE_CLICK_HANDLERS[id] = wrapped;
          el.onclick = null;
          el.addEventListener('click', wrapped);
          // 确保按钮本身可点击，避免被潜在覆盖层吞掉点击事件
          el.style.pointerEvents = 'auto';
          if (!el.style.position) el.style.position = 'relative';
          if (!el.style.zIndex || Number(el.style.zIndex) < 2) el.style.zIndex = '2';
        }
        // 兼容旧版内联 onclick：若已绑定 safeClick，则不再走旧兜底逻辑，避免覆盖详细渲染结果。
        const LEGACY_FALLBACK_CLICK = window.__zhifeiFallbackClick;
        window.__zhifeiFallbackClick = function (ev, actionId) {
          if (SAFE_CLICK_HANDLERS[actionId]) {
            if (ev && typeof ev.preventDefault === 'function') ev.preventDefault();
            return true;
          }
          if (typeof LEGACY_FALLBACK_CLICK === 'function') return LEGACY_FALLBACK_CLICK(ev, actionId);
          return true;
        };
        function safeChange(id, fn) { const el = document.getElementById(id); if (el) el.onchange = fn; }
        function storageGet(key) {
          try { return localStorage.getItem(key) || ''; } catch (_) { return ''; }
        }
        function storageSet(key, value) {
          try { localStorage.setItem(key, value); } catch (_) {}
        }
        function storageRemove(key) {
          try { localStorage.removeItem(key); } catch (_) {}
        }
        function pickProjectFromSelect(sel) {
          if (!sel) return '';
          if (sel.value) return sel.value;
          const remembered = storageGet('selected_project_id');
          if (remembered) {
            for (let i = 0; i < sel.options.length; i += 1) {
              if (String(sel.options[i].value || '') === remembered) {
                sel.value = remembered;
                break;
              }
            }
          }
          if (!sel.value && sel.options && sel.options.length > 1) {
            sel.selectedIndex = sel.options.length - 1;
          }
          if (sel.value) storageSet('selected_project_id', sel.value);
          return sel.value || '';
        }
        function pid() {
          const sel = document.getElementById('projectSelect');
          return pickProjectFromSelect(sel);
        }
        function resolveProjectId(projectIdOverride='') {
          const explicit = String(projectIdOverride || '').trim();
          if (explicit) return explicit;
          const fromSelect = pid();
          if (fromSelect) return fromSelect;
          const hiddenIds = [
            'uploadShigongProjectId',
            'uploadMaterialProjectId',
            'uploadMaterialBoqProjectId',
            'uploadMaterialDrawingProjectId',
            'uploadMaterialPhotoProjectId',
          ];
          for (const hid of hiddenIds) {
            const el = document.getElementById(hid);
            const v = String((el && el.value) || '').trim();
            if (v) return v;
          }
          try {
            const qid = new URL(window.location.href).searchParams.get('project_id') || '';
            if (String(qid).trim()) return String(qid).trim();
          } catch (_) {}
          return '';
        }
        function selectedScoreScaleMax() {
          const el = document.getElementById('scoreScaleSelect');
          const raw = (el && el.value) ? String(el.value).trim() : '100';
          return raw === '5' ? 5 : 100;
        }
        function selectedScoreScaleLabel() {
          return selectedScoreScaleMax() === 5 ? '5分制' : '100分制';
        }
        function selectedGroundTruthJudgeCount() {
          const el = document.getElementById('gtJudgeCount');
          const raw = (el && el.value) ? String(el.value).trim() : '5';
          return raw === '7' ? 7 : 5;
        }
        function syncGroundTruthJudgeInputs() {
          const count = selectedGroundTruthJudgeCount();
          for (let i = 1; i <= 7; i += 1) {
            const wrap = document.getElementById('gtJWrap' + String(i));
            const input = document.getElementById('gtJ' + String(i));
            const enabled = i <= count;
            if (wrap) wrap.style.display = enabled ? 'inline-flex' : 'none';
            if (input) {
              input.disabled = !enabled;
              if (!enabled) input.value = '';
            }
          }
        }
        function collectGroundTruthJudgeScores() {
          const count = selectedGroundTruthJudgeCount();
          const scores = [];
          for (let i = 1; i <= count; i += 1) {
            const input = document.getElementById('gtJ' + String(i));
            const value = parseFloat(((input || {}).value || '0'));
            scores.push(Number.isFinite(value) ? value : 0);
          }
          return scores;
        }
        function applyProjectScoreScale(projectId) {
          const el = document.getElementById('scoreScaleSelect');
          if (!el) return;
          const meta = projectMetaById[String(projectId || '')] || {};
          const raw = String((meta && meta.score_scale_max) != null ? meta.score_scale_max : '100');
          el.value = (raw === '5') ? '5' : '100';
        }
        function apiHeaders(isJson=true) {
          const k = storageGet('api_key');
          const h = {};
          if (k) h['X-API-Key'] = k;
          if (isJson) h['Content-Type'] = 'application/json';
          return h;
        }
        const NL = String.fromCharCode(10);
        const DIMENSION_LABELS = {
          "01": "01 工程项目整体理解与实施路径",
          "02": "02 安全生产管理体系与控制措施",
          "03": "03 文明施工管理体系与实施措施",
          "04": "04 材料与部品采购及管理机制",
          "05": "05 四新技术的应用与实施方案",
          "06": "06 工程关键工序识别与控制措施",
          "07": "07 工程重难点及危险性较大工程管控",
          "08": "08 工程质量管理体系与保证措施",
          "09": "09 工期目标保障与进度控制措施",
          "10": "10 专项施工工艺与技术方案",
          "11": "11 人力资源配置与管理方案",
          "12": "12 总体施工工艺流程与组织逻辑",
          "13": "13 物资与施工设备配置方案",
          "14": "14 设计协调与深化实施能力",
          "15": "15 总体资源配置与实施计划",
          "16": "16 技术措施的可行性与落地性"
        };
        const DIM_IDS = Array.from({ length: 16 }, (_, i) => String(i + 1).padStart(2, '0'));
        let expertProfileLocked = false;
        let projectMetaById = {};
        let scoringReadinessState = { project_id: '', ready: false, gate_passed: false, issues: [] };
        let latestOllamaPreviewPayload = null;
        let latestOllamaPreviewProjectId = '';
        const PROJECT_REQUIRED_BUTTON_IDS = [
          'deleteCurrentProject', 'btnWeightsReset', 'btnWeightsSave', 'btnWeightsApply',
          'btnMaterialDepthReport', 'btnMaterialDepthReportDownload', 'btnMaterialKnowledgeProfile', 'btnMaterialKnowledgeProfileDownload',
          'btnRefreshGroundTruth', 'btnRefreshGroundTruthSubmissionOptions', 'btnUploadFeed', 'btnRefreshFeedMaterials', 'btnAddGroundTruth',
          'btnEvolve', 'btnOllamaPreview', 'btnEvolutionHealth', 'btnWritingGuidance', 'btnCompilationInstructions',
          'btnRebuildDelta', 'btnRebuildSamples', 'btnTrainCalibratorV2', 'btnApplyCalibPredict',
          'btnAutoRunReflection', 'btnEvalMetricsV2', 'btnEvalSummaryV2',
          'btnMinePatchV2', 'btnShadowPatchV2', 'btnDeployPatchV2', 'btnRollbackPatchV2',
        ];
        const NON_BLOCKING_ACTION_BUTTON_IDS = [
          'btnUploadMaterials', 'btnUploadBoq', 'btnUploadDrawing', 'btnUploadSitePhotos', 'btnRefreshMaterials', 'btnMaterialDepthReport', 'btnMaterialDepthReportDownload', 'btnMaterialKnowledgeProfile', 'btnMaterialKnowledgeProfileDownload', 'btnUploadShigong', 'btnScoreShigong', 'btnRefreshSubmissions',
          'btnCompare', 'btnCompareReport', 'btnInsights', 'btnLearning',
          'btnEvidenceTrace', 'btnScoringBasis',
          'btnAdaptive', 'btnAdaptivePatch', 'btnAdaptiveValidate', 'btnAdaptiveApply',
          'btnRefreshGroundTruth', 'btnRefreshGroundTruthSubmissionOptions', 'btnUploadFeed', 'btnRefreshFeedMaterials', 'btnAddGroundTruth',
          'btnEvolve', 'btnOllamaPreview', 'btnEvolutionHealth', 'btnWritingGuidance', 'btnCompilationInstructions',
          'btnRebuildDelta', 'btnRebuildSamples', 'btnTrainCalibratorV2', 'btnApplyCalibPredict',
          'btnAutoRunReflection', 'btnEvalMetricsV2', 'btnEvalSummaryV2',
          'btnMinePatchV2', 'btnShadowPatchV2', 'btnDeployPatchV2', 'btnRollbackPatchV2',
        ];
        const PROJECT_REQUIRED_INPUT_IDS = [
          'scoreScaleSelect',
          'feedFile', 'groundTruthSubmissionSelect', 'groundTruthScope', 'groundTruthOtherProject',
          'gtJudgeCount', 'gtJ1', 'gtJ2', 'gtJ3', 'gtJ4', 'gtJ5', 'gtJ6', 'gtJ7', 'gtFinal', 'patchType', 'patchIdInput',
        ];
        function setActionStatus(id, msg, isError=false) {
          const el = document.getElementById(id);
          if (!el) return;
          el.textContent = msg || '';
          el.style.color = isError ? '#b91c1c' : '#475569';
        }
        function syncProjectHiddenInputs(projectId) {
          const value = projectId || '';
          ['deleteProjectId', 'uploadMaterialProjectId', 'uploadMaterialBoqProjectId', 'uploadMaterialDrawingProjectId', 'uploadMaterialPhotoProjectId', 'uploadShigongProjectId'].forEach((id) => {
            const el = document.getElementById(id);
            if (el) el.value = value;
          });
        }
        function updateCurrentProjectTag(projectId) {
          const tag = document.getElementById('currentProjectTag');
          syncProjectHiddenInputs(projectId);
          if (!tag) return;
          if (!projectId) {
            tag.textContent = '当前项目：未选择';
            return;
          }
          const sel = document.getElementById('projectSelect');
          const opt = sel && sel.options && sel.selectedIndex >= 0 ? sel.options[sel.selectedIndex] : null;
          const label = opt && opt.textContent ? opt.textContent : projectId;
          tag.textContent = '当前项目：' + label;
        }

        function updateProjectBoundControlsState() {
          const projectId = pid();
          const hasProject = !!projectId;
          updateCurrentProjectTag(projectId);
          const title = hasProject ? '' : '请先选择项目';
          PROJECT_REQUIRED_BUTTON_IDS.forEach((id) => {
            const el = document.getElementById(id);
            if (!el) return;
            el.disabled = !hasProject;
            el.title = title;
          });
          PROJECT_REQUIRED_INPUT_IDS.forEach((id) => {
            const el = document.getElementById(id);
            if (!el) return;
            el.disabled = !hasProject;
            if (!hasProject && (el.tagName === 'INPUT' || el.tagName === 'SELECT')) {
              el.title = '请先选择项目';
            } else {
              el.title = '';
            }
          });
          // 对比/自适应按钮保持可点击：未选项目时也要有可见提示，避免“点击无反应”
          NON_BLOCKING_ACTION_BUTTON_IDS.forEach((id) => {
            const el = document.getElementById(id);
            if (!el) return;
            el.disabled = false;
            el.title = hasProject ? '' : '请先选择项目后执行（点击会显示提示）';
          });
          if (!hasProject) {
            setActionStatus('materialsActionStatus', '请先在「2) 选择项目」中选择项目。', true);
            setActionStatus('materialsActionStatusBoq', '请先在「2) 选择项目」中选择项目。', true);
            setActionStatus('materialsActionStatusDrawing', '请先在「2) 选择项目」中选择项目。', true);
            setActionStatus('materialsActionStatusPhoto', '请先在「2) 选择项目」中选择项目。', true);
            setActionStatus('shigongActionStatus', '请先在「2) 选择项目」中选择项目。', true);
            setActionStatus('feedActionStatus', '请先在「2) 选择项目」中选择项目。', true);
          }
        }

        let projectSwitchSeq = 0;
        function selectedProjectIdStrict() {
          const sel = document.getElementById('projectSelect');
          return (sel && sel.value) ? String(sel.value) : '';
        }
        function isStaleProjectResponse(expectedProjectId, seq) {
          if (!expectedProjectId) return false;
          if (typeof seq === 'number' && seq !== projectSwitchSeq) return true;
          return selectedProjectIdStrict() !== String(expectedProjectId);
        }
        function setTableStandby(tableId, emptyId, message) {
          const table = document.getElementById(tableId);
          const tbody = table ? table.querySelector('tbody') : null;
          if (tbody) tbody.innerHTML = '';
          const emptyEl = document.getElementById(emptyId);
          if (emptyEl) {
            emptyEl.textContent = message || '';
            emptyEl.style.display = 'block';
          }
        }
        function clearResultBlock(resultId) {
          const el = document.getElementById(resultId);
          if (!el) return;
          el.style.display = 'none';
          el.innerHTML = '';
        }
        function resetProjectPanelsToStandby(projectId) {
          const hasProject = !!projectId;
          setTableStandby(
            'materialsTable',
            'materialsEmpty',
            hasProject ? '待机：正在加载当前项目资料…' : '暂无资料，请先选择项目。'
          );
          setTableStandby(
            'submissionsTable',
            'submissionsEmpty',
            hasProject ? '待机：正在加载当前项目施组…' : '暂无施组，请先选择项目。'
          );
          setTableStandby(
            'feedMaterialsTable',
            'feedMaterialsEmpty',
            hasProject ? '待机：正在加载当前项目投喂包…' : '暂无投喂包，请先选择项目。'
          );
          setTableStandby(
            'groundTruthTable',
            'groundTruthEmpty',
            hasProject ? '待机：正在加载当前项目真实评标记录…' : '暂无真实评标，请先选择项目。'
          );
          [
            'scoringReadinessResult',
            'materialDepthReportResult',
            'materialUtilizationResult',
            'compareResult', 'compareReportResult', 'insightsResult', 'learningResult',
            'evidenceTraceResult', 'scoringBasisResult',
            'adaptiveResult', 'adaptivePatchResult', 'adaptiveValidateResult', 'adaptiveApplyResult',
            'evolveResult', 'ollamaPreviewResult', 'guidanceResult', 'compilationInstructionsResult',
            'deltaResult', 'sampleResult', 'calibTrainResult', 'patchResult',
            'patchShadowResult', 'patchDeployResult', 'evalResult'
          ].forEach(clearResultBlock);
          const gtScope = document.getElementById('groundTruthScope');
          if (gtScope) gtScope.value = 'current';
          const gtOther = document.getElementById('groundTruthOtherProject');
          if (gtOther) {
            gtOther.style.display = 'none';
            gtOther.value = '';
          }
          ['gtJudgeCount', 'gtJ1', 'gtJ2', 'gtJ3', 'gtJ4', 'gtJ5', 'gtJ6', 'gtJ7', 'gtFinal', 'patchIdInput'].forEach((id) => {
            const el = document.getElementById(id);
            if (el) el.value = '';
          });
          const judgeCountSel = document.getElementById('gtJudgeCount');
          if (judgeCountSel) judgeCountSel.value = '5';
          if (typeof syncGroundTruthJudgeInputs === 'function') syncGroundTruthJudgeInputs();
          if (!hasProject) {
            const scaleSel = document.getElementById('scoreScaleSelect');
            if (scaleSel) scaleSel.value = '100';
          }
          resetOllamaPreviewActions();
          const gtSubmissionSel = document.getElementById('groundTruthSubmissionSelect');
          if (gtSubmissionSel) {
            gtSubmissionSel.innerHTML = '';
            const opt = document.createElement('option');
            opt.value = '';
            opt.textContent = hasProject ? '-- 待加载步骤4施组文件 --' : '-- 请先选择项目 --';
            gtSubmissionSel.appendChild(opt);
            gtSubmissionSel.value = '';
          }
          document.querySelectorAll(
            '#uploadMaterial input[type="file"], #uploadMaterialBoq input[type="file"], #uploadMaterialDrawing input[type="file"], #uploadMaterialPhoto input[type="file"], #uploadShigong input[type="file"], #feedFile'
          ).forEach((el) => { if (el) el.value = ''; });
          setActionStatus(
            'materialsActionStatus',
            hasProject ? '待机：可上传招标文件和答疑。' : '请先在「2) 选择项目」中选择项目。',
            !hasProject
          );
          setActionStatus(
            'materialsActionStatusBoq',
            hasProject ? '待机：可上传清单。' : '请先在「2) 选择项目」中选择项目。',
            !hasProject
          );
          setActionStatus(
            'materialsActionStatusDrawing',
            hasProject ? '待机：可上传图纸。' : '请先在「2) 选择项目」中选择项目。',
            !hasProject
          );
          setActionStatus(
            'materialsActionStatusPhoto',
            hasProject ? '待机：可上传现场照片。' : '请先在「2) 选择项目」中选择项目。',
            !hasProject
          );
          setActionStatus(
            'shigongActionStatus',
            hasProject ? '待机：可上传施组或点击“评分施组”。' : '请先在「2) 选择项目」中选择项目。',
            !hasProject
          );
          setActionStatus(
            'feedActionStatus',
            hasProject ? '待机：可上传投喂包或录入真实评标。' : '请先在「2) 选择项目」中选择项目。',
            !hasProject
          );
          if (hasProject) {
            setExpertProfileStatus('待机：正在加载当前项目的16维关注度配置...');
            applyWeightsRaw({});
          } else {
            setExpertProfileStatus('请先选择项目。');
            applyWeightsRaw({});
            expertProfileLocked = false;
          }
          const out = document.getElementById('output');
          if (out) {
            out.textContent = hasProject
              ? ('已切换项目，正在自动刷新 2.5/3/4/5/6/7 区域数据…')
              : '请选择项目后开始操作。';
          }
        }

        function valueOrDefault(v, fallback) {
          return (v === undefined || v === null) ? fallback : v;
        }
        function normalizeWeightsRaw(raw) {
          const m = {};
          let sum = 0;
          DIM_IDS.forEach(id => {
            const v = Number(valueOrDefault(raw[id], 5));
            const mv = 0.5 + (v / 10);
            m[id] = mv;
            sum += mv;
          });
          const norm = {};
          DIM_IDS.forEach(id => { norm[id] = sum > 0 ? (m[id] / sum) : 0; });
          return norm;
        }
        function buildWeightsPanel() {
          const panel = document.getElementById('expertWeightsPanel');
          if (!panel) return;
          panel.innerHTML = '';
          DIM_IDS.forEach(id => {
            const row = document.createElement('div');
            row.className = 'weight-row';
            row.innerHTML =
              '<label for="w_' + id + '">' + DIMENSION_LABELS[id] + '</label>' +
              '<input id="w_' + id + '" data-dim="' + id + '" type="range" min="0" max="10" step="1" value="5" />' +
              '<span class="raw-value" id="w_raw_' + id + '">5</span>' +
              '<span class="norm-value" id="w_norm_' + id + '">6.25%</span>';
            panel.appendChild(row);
          });
          const sliders = panel.querySelectorAll('input[type="range"]');
          for (let i = 0; i < sliders.length; i += 1) {
            sliders[i].addEventListener('input', renderWeightSummary);
          }
          renderWeightSummary();
        }
        function collectWeightsRaw() {
          const raw = {};
          DIM_IDS.forEach(id => {
            const el = document.getElementById('w_' + id);
            raw[id] = Number((el && el.value) || 5);
          });
          return raw;
        }
        function applyWeightsRaw(raw) {
          DIM_IDS.forEach(id => {
            const el = document.getElementById('w_' + id);
            if (!el) return;
            el.value = String(valueOrDefault(raw[id], 5));
          });
          renderWeightSummary();
        }
        function renderWeightSummary() {
          const raw = collectWeightsRaw();
          const norm = normalizeWeightsRaw(raw);
          DIM_IDS.forEach(id => {
            const rawEl = document.getElementById('w_raw_' + id);
            if (rawEl) rawEl.textContent = String(raw[id]);
            const normEl = document.getElementById('w_norm_' + id);
            if (normEl) normEl.textContent = (norm[id] * 100).toFixed(2) + '%';
          });
          const summaryEl = document.getElementById('expertWeightsSummary');
          if (summaryEl) {
            summaryEl.textContent = DIM_IDS.map(id => id + ':' + (norm[id] * 100).toFixed(2) + '%').join(' | ');
          }
        }
        function setExpertProfileStatus(text, isError=false) {
          const el = document.getElementById('expertProfileStatus');
          if (!el) return;
          el.textContent = text || '';
          el.style.color = isError ? '#b91c1c' : '#334155';
        }
        async function loadExpertProfile(expectedProjectId=null, switchSeq=null) {
          const projectId = expectedProjectId || pid();
          if (!projectId) {
            setExpertProfileStatus('请先选择项目。');
            applyWeightsRaw({});
            expertProfileLocked = false;
            return;
          }
          setExpertProfileStatus('正在加载当前生效配置...');
          let res, data;
          try {
            res = await fetch('/api/v1/projects/' + projectId + '/expert-profile');
            data = await res.json().catch(() => ({}));
          } catch (err) {
            if (isStaleProjectResponse(projectId, switchSeq)) return;
            setExpertProfileStatus('加载失败：' + String((err && err.message) || err), true);
            return;
          }
          if (isStaleProjectResponse(projectId, switchSeq)) return;
          if (!res.ok) {
            setExpertProfileStatus('加载失败：' + (data.detail || ('HTTP ' + res.status)), true);
            return;
          }
          const profile = (data && data.expert_profile) || {};
          applyWeightsRaw(profile.weights_raw || {});
          const project = (data && data.project) || {};
          expertProfileLocked = String(project.status || '') === 'submitted_to_qingtian';
          let statusText =
            '当前生效配置：' + (profile.name || '-') +
            '（ID: ' + (profile.id || '-') + '，更新时间: ' + ((project.updated_at || profile.updated_at || '').slice(0, 19) || '-') + '）';
          if (expertProfileLocked) {
            statusText += '；当前项目已进入青天评标阶段，保存/重算需要二次确认解锁。';
          }
          setExpertProfileStatus(statusText);
        }
        async function saveExpertProfile(askName=true, forceUnlock=false) {
          const projectId = pid();
          if (!projectId) { setExpertProfileStatus('请先选择项目。', true); return false; }
          if (expertProfileLocked && !forceUnlock) {
            const unlockOk = confirm('当前项目已进入青天评标阶段。是否解锁并继续保存专家配置？');
            if (!unlockOk) return false;
            forceUnlock = true;
          }
          const customName = askName ? (prompt('请输入专家配置名称（可留空自动命名）：') || '') : '';
          const payload = { name: customName, weights_raw: collectWeightsRaw(), force_unlock: !!forceUnlock };
          setExpertProfileStatus('正在保存配置...');
          let res, data;
          try {
            res = await fetch('/api/v1/projects/' + projectId + '/expert-profile', {
              method: 'PUT',
              headers: apiHeaders(),
              body: JSON.stringify(payload),
            });
            data = await res.json().catch(() => ({}));
          } catch (err) {
            setExpertProfileStatus('保存失败：' + String((err && err.message) || err), true);
            return false;
          }
          if (!res.ok) {
            if (res.status === 409 && !forceUnlock) {
              setExpertProfileStatus('保存被锁定：项目处于青天评标阶段，请解锁后重试。', true);
              return false;
            }
            setExpertProfileStatus('保存失败：' + (data.detail || ('HTTP ' + res.status)), true);
            return false;
          }
          const profile = (data && data.expert_profile) || {};
          setExpertProfileStatus('已保存并绑定配置：' + (profile.name || '-') + '（' + (profile.id || '-') + '）');
          applyWeightsRaw(profile.weights_raw || {});
          const out = document.getElementById('output');
          if (out) out.textContent = JSON.stringify(data, null, 2);
          return true;
        }
        async function applyExpertProfileAndRescore() {
          const projectId = pid();
          if (!projectId) { setExpertProfileStatus('请先选择项目。', true); return; }
          let forceUnlock = false;
          if (expertProfileLocked) {
            const unlockOk = confirm('当前项目已进入青天评标阶段。是否解锁并执行“保存+全项目重算”？');
            if (!unlockOk) return;
            forceUnlock = true;
          }
          const ok = confirm('将保存当前16维关注度，并重算本项目全部施组。是否继续？');
          if (!ok) return;
          const saved = await saveExpertProfile(false, forceUnlock);
          if (!saved) return;
          setExpertProfileStatus('正在按当前配置重算全部施组...');
          let res, data;
          try {
            res = await fetch('/api/v1/projects/' + projectId + '/rescore', {
              method: 'POST',
              headers: apiHeaders(),
              body: JSON.stringify({
                scoring_engine_version: 'v2',
                scope: 'project',
                score_scale_max: selectedScoreScaleMax(),
                rebuild_anchors: false,
                rebuild_requirements: false,
                retrain_calibrator: false,
                force_unlock: !!forceUnlock
              }),
            });
            data = await res.json().catch(() => ({}));
          } catch (err) {
            setExpertProfileStatus('重算失败：' + String((err && err.message) || err), true);
            return;
          }
          if (!res.ok) {
            if (res.status === 409) {
              setExpertProfileStatus('重算被锁定：项目处于青天评标阶段，请解锁后重试。', true);
              return;
            }
            setExpertProfileStatus('重算失败：' + (data.detail || ('HTTP ' + res.status)), true);
            return;
          }
          const gate = (data && typeof data.material_utilization_gate === 'object') ? data.material_utilization_gate : {};
          const blockedCount = Number((gate && gate.blocked_submissions) || 0);
          const warnCount = Number((gate && gate.warn_submissions) || 0);
          let rescoreMsg =
            '重算完成（' + ((data && data.score_scale_label) ? data.score_scale_label : selectedScoreScaleLabel()) + '）：共处理 ' + (data.submission_count || 0) + ' 份，生成 ' + (data.reports_generated || 0) + ' 份报告。';
          if (blockedCount > 0) {
            rescoreMsg += ' 资料门禁阻断 ' + blockedCount + ' 份。';
          } else if (warnCount > 0) {
            rescoreMsg += ' 资料门禁预警 ' + warnCount + ' 份。';
          } else if (gate && gate.enabled) {
            rescoreMsg += ' 资料门禁通过。';
          }
          setExpertProfileStatus(rescoreMsg, blockedCount > 0);
          const out = document.getElementById('output');
          if (out) out.textContent = JSON.stringify(data, null, 2);
          if (typeof renderMaterialUtilizationPanel === 'function') {
            renderMaterialUtilizationPanel(data || {});
          }
          if (typeof refreshSubmissions === 'function') await refreshSubmissions();
          if (typeof refreshMaterials === 'function') await refreshMaterials();
        }
        function initWeightsSection() {
          try {
            buildWeightsPanel();
          } catch (err) {
            reportClientError('权重面板初始化失败', err);
          }
        }
        safeClick('btnWeightsReset', () => applyWeightsRaw({}));
        safeClick('btnWeightsSave', saveExpertProfile);
        safeClick('btnWeightsApply', applyExpertProfileAndRescore);

        function setCreateMsg(msg, isError) {
          const el = document.getElementById('createProjectMessage');
          if (!el) { alert(msg); return; }
          el.textContent = msg || '';
          el.style.color = isError ? '#b91c1c' : '#15803d';
        }
        function setSelectMsg(msg, isError) {
          const el = document.getElementById('selectProjectMessage');
          if (!el) { alert(msg); return; }
          el.textContent = msg || '';
          el.style.color = isError ? '#b91c1c' : '#15803d';
        }
        function setSelfCheckResult(summary, details, isError) {
          const el = document.getElementById('selfCheckResult');
          if (!el) return;
          el.style.display = 'block';
          el.style.borderLeftColor = isError ? '#b91c1c' : '#15803d';
          el.innerHTML = '<strong></strong><pre style="margin-top:6px"></pre>';
          const strong = el.querySelector('strong');
          const pre = el.querySelector('pre');
          if (strong) strong.textContent = summary || '';
          if (pre) pre.textContent = details || '';
        }
        function renderSelfCheckPanel(data) {
          const el = document.getElementById('selfCheckResult');
          if (!el) return;
          const payload = (data && typeof data === 'object') ? data : {};
          const items = Array.isArray(payload.items) ? payload.items : [];
          const failed = items.filter((x) => !x || !x.ok).length;
          const summary = failed === 0
            ? '系统自检通过（全部正常）'
            : ('系统自检完成：发现 ' + failed + ' 项异常');
          el.style.display = 'block';
          el.style.borderLeftColor = failed === 0 ? '#15803d' : '#b91c1c';
          const capabilityNames = ['parser_pdf', 'parser_docx', 'parser_ocr', 'parser_dwg_converter'];
          const capabilityRows = items.filter((x) => capabilityNames.includes(String((x && x.name) || '')));
          const capabilitySummary = capabilityRows.length
            ? capabilityRows.map((x) => (x && x.ok ? '✓' : '×') + String((x && x.name) || '')).join(' / ')
            : '无';
          let html = '<strong>' + escapeHtmlText(summary) + '</strong>';
          html += '<p style="margin:6px 0 0 0;font-size:12px;color:#475569">解析能力：' + escapeHtmlText(capabilitySummary) + '</p>';
          html += '<table style="margin-top:8px"><tr><th>检查项</th><th>状态</th><th>详情</th></tr>';
          html += items.length
            ? items.map((x) => {
              const ok = !!(x && x.ok);
              return '<tr>'
                + '<td>' + escapeHtmlText(String((x && x.name) || '-')) + '</td>'
                + '<td>' + (ok ? '<span class="success">OK</span>' : '<span class="error">FAIL</span>') + '</td>'
                + '<td>' + escapeHtmlText(String((x && x.detail) || '-')) + '</td>'
                + '</tr>';
            }).join('')
            : '<tr><td colspan="3">无自检数据</td></tr>';
          html += '</table>';
          el.innerHTML = html;
        }
        function setScoringFactorsResult(summary, details, isError) {
          const el = document.getElementById('scoringFactorsResult');
          if (!el) return;
          el.style.display = 'block';
          el.style.borderLeftColor = isError ? '#b91c1c' : '#2563eb';
          el.innerHTML = '<strong></strong><pre style="margin-top:6px"></pre>';
          const strong = el.querySelector('strong');
          const pre = el.querySelector('pre');
          if (strong) strong.textContent = summary || '';
          if (pre) pre.textContent = details || '';
        }
        async function probeEndpoint(url, options) {
          try {
            const res = await fetch(url, options);
            const text = await res.text();
            let data = null;
            try { data = JSON.parse(text); } catch (_) {}
            return { ok: res.ok, status: res.status, text, data };
          } catch (err) {
            return { ok: false, status: 0, error: String((err && err.message) || err || '网络异常') };
          }
        }
        async function runSystemSelfCheck() {
          setSelfCheckResult('正在执行系统自检…', '', false);
          const currentId = pid();
          const url = currentId
            ? ('/api/v1/system/self_check?project_id=' + encodeURIComponent(currentId))
            : '/api/v1/system/self_check';
          const r = await probeEndpoint(url);
          if (r.error) {
            setSelfCheckResult('系统自检失败', r.error, true);
            const out = document.getElementById('output');
            if (out) out.textContent = r.error;
            return;
          }
          if (!r.ok) {
            const detail = (r.data && r.data.detail) ? String(r.data.detail) : String(r.text || '').slice(0, 200);
            setSelfCheckResult('系统自检失败', 'HTTP ' + r.status + ' ' + detail, true);
            const out = document.getElementById('output');
            if (out) out.textContent = detail;
            return;
          }
          const data = r.data || {};
          renderSelfCheckPanel(data);
          const out = document.getElementById('output');
          if (out) out.textContent = JSON.stringify(data, null, 2);
        }
        async function loadScoringFactorsOverview() {
          const currentId = pid();
          const url = currentId
            ? ('/api/v1/scoring/factors?project_id=' + encodeURIComponent(currentId))
            : '/api/v1/scoring/factors';
          setScoringFactorsResult('正在加载评分体系总览…', '', false);
          const r = await probeEndpoint(url);
          if (r.error) {
            setScoringFactorsResult('评分体系加载失败', r.error, true);
            const out = document.getElementById('output');
            if (out) out.textContent = r.error;
            return;
          }
          if (!r.ok) {
            const detail = (r.data && r.data.detail) ? String(r.data.detail) : String(r.text || '').slice(0, 200);
            setScoringFactorsResult('评分体系加载失败', 'HTTP ' + r.status + ' ' + detail, true);
            const out = document.getElementById('output');
            if (out) out.textContent = detail;
            return;
          }
          const d = r.data || {};
          const flags = d.capability_flags || {};
          const summary =
            '评分体系已加载：维度 ' + (d.dimension_count || 0) + '，扣分规则 ' + ((d.penalty_rules || []).length) +
            '；组织机构要求=' + (flags.organization_structure_required ? '是' : '否') +
            '，章节完整性=' + (flags.chapter_content_completeness_required ? '是' : '否') +
            '，重难点要求=' + (flags.key_difficult_points_required ? '是' : '否') +
            '，图文要求=' + (flags.graphic_content_required ? '是' : '否');
          const details = JSON.stringify(d, null, 2);
          setScoringFactorsResult(summary, details, false);
          const out = document.getElementById('output');
          if (out) out.textContent = details;
        }
        async function loadScoringFactorsMarkdown() {
          const currentId = pid();
          const url = currentId
            ? ('/api/v1/scoring/factors/markdown?project_id=' + encodeURIComponent(currentId))
            : '/api/v1/scoring/factors/markdown';
          setScoringFactorsResult('正在生成评分体系 Markdown…', '', false);
          const r = await probeEndpoint(url);
          if (r.error) {
            setScoringFactorsResult('Markdown 生成失败', r.error, true);
            const out = document.getElementById('output');
            if (out) out.textContent = r.error;
            return;
          }
          if (!r.ok) {
            const detail = (r.data && r.data.detail) ? String(r.data.detail) : String(r.text || '').slice(0, 200);
            setScoringFactorsResult('Markdown 生成失败', 'HTTP ' + r.status + ' ' + detail, true);
            const out = document.getElementById('output');
            if (out) out.textContent = detail;
            return;
          }
          const markdown = String((r.data && r.data.markdown) || '');
          setScoringFactorsResult('评分体系 Markdown 已生成（可直接复制到 ChatGPT）', markdown.slice(0, 600) + (markdown.length > 600 ? '\\n...（已截断预览）' : ''), false);
          const out = document.getElementById('output');
          if (out) out.textContent = markdown;
        }
        async function loadProjectAnalysisBundle() {
          const currentId = pid();
          if (!currentId) {
            setScoringFactorsResult('请先选择项目', '项目分析包需要 project_id', true);
            return;
          }
          const url = '/api/v1/projects/' + encodeURIComponent(currentId) + '/analysis_bundle';
          setScoringFactorsResult('正在生成项目分析包…', '', false);
          const r = await probeEndpoint(url);
          if (r.error) {
            setScoringFactorsResult('项目分析包生成失败', r.error, true);
            const out = document.getElementById('output');
            if (out) out.textContent = r.error;
            return;
          }
          if (!r.ok) {
            const detail = (r.data && r.data.detail) ? String(r.data.detail) : String(r.text || '').slice(0, 200);
            setScoringFactorsResult('项目分析包生成失败', 'HTTP ' + r.status + ' ' + detail, true);
            const out = document.getElementById('output');
            if (out) out.textContent = detail;
            return;
          }
          const markdown = String((r.data && r.data.markdown) || '');
          setScoringFactorsResult('项目分析包已生成（可直接复制给 ChatGPT）', markdown.slice(0, 600) + (markdown.length > 600 ? '\\n...（已截断预览）' : ''), false);
          const out = document.getElementById('output');
          if (out) out.textContent = markdown;
        }
        async function downloadProjectAnalysisBundle() {
          const currentId = pid();
          if (!currentId) {
            setScoringFactorsResult('请先选择项目', '下载分析包需要 project_id', true);
            return;
          }
          const url = '/api/v1/projects/' + encodeURIComponent(currentId) + '/analysis_bundle.md';
          setScoringFactorsResult('正在准备下载…', '若浏览器未自动下载，请检查弹窗或下载权限设置。', false);
          const a = document.createElement('a');
          a.href = url;
          a.download = 'analysis_bundle_' + currentId + '.md';
          document.body.appendChild(a);
          a.click();
          a.remove();
        }
        async function refreshProjects() {
          setSelectMsg('正在加载…', false);
          const current = pid() || storageGet('selected_project_id') || '';
          const deleteSel = document.getElementById('projectDeleteSelect');
          const prevDeleteIds = deleteSel
            ? Array.from(deleteSel.selectedOptions || []).map((o) => String(o.value || ''))
            : [];
          let res, text;
          try {
            res = await fetch('/api/v1/projects');
            text = await res.text();
          } catch (err) {
            setSelectMsg('网络错误，请确认服务已启动（如 make run）', true);
            const out = document.getElementById('output');
            if (out) { out.textContent = '请求失败: ' + (err.message || err); out.scrollIntoView({ behavior: 'smooth' }); }
            return;
          }
          let list = [];
          try { list = JSON.parse(text); } catch (e) { list = []; }
          if (!Array.isArray(list)) list = [];
          projectMetaById = {};
          list.forEach((p) => {
            if (p && p.id) projectMetaById[String(p.id)] = (p.meta && typeof p.meta === 'object') ? p.meta : {};
          });
          list = list.slice().sort((a, b) => {
            const an = String((a && a.name) || '');
            const bn = String((b && b.name) || '');
            const ae = an.startsWith('E2E_') ? 1 : 0;
            const be = bn.startsWith('E2E_') ? 1 : 0;
            if (ae !== be) return ae - be;
            const at = String((a && (a.updated_at || a.created_at)) || '');
            const bt = String((b && (b.updated_at || b.created_at)) || '');
            return at.localeCompare(bt);
          });
          if (!res.ok) {
            const errMsg = (typeof list === 'object' && list && list.detail) ? String(list.detail) : (text || '').slice(0, 200);
            setSelectMsg('刷新失败: ' + res.status + ' ' + errMsg, true);
            const out = document.getElementById('output');
            if (out) { out.textContent = '刷新失败: ' + res.status + '\\n' + text; out.scrollIntoView({ behavior: 'smooth' }); }
            return;
          }
          setSelectMsg(list.length ? '已加载 ' + list.length + ' 个项目，请在上方下拉框选择' : '暂无项目，请先在「1) 创建项目」中创建', false);
          const sel = document.getElementById('projectSelect');
          sel.innerHTML = '<option value="">-- 选择项目 --</option>';
          list.forEach(p => {
            const o = document.createElement('option');
            o.value = p.id;
            o.textContent = (p.name || p.id) + ' (' + (p.id || '').slice(0,8) + '…)';
            sel.appendChild(o);
          });
          if (deleteSel) {
            deleteSel.innerHTML = '';
            if (!list.length) {
              const emptyOpt = document.createElement('option');
              emptyOpt.value = '';
              emptyOpt.disabled = true;
              emptyOpt.textContent = '-- 暂无项目 --';
              deleteSel.appendChild(emptyOpt);
            } else {
              list.forEach((p) => {
                const o = document.createElement('option');
                o.value = p.id;
                o.textContent = (p.name || p.id) + ' (' + (p.id || '').slice(0,8) + '…)';
                if (prevDeleteIds.includes(String(p.id))) o.selected = true;
                deleteSel.appendChild(o);
              });
            }
            deleteSel.disabled = !list.length;
          }
          const deleteSelectedBtn = document.getElementById('deleteSelectedProjects');
          if (deleteSelectedBtn) deleteSelectedBtn.disabled = !list.length;
          if (current && list.some(p => p.id === current)) {
            sel.value = current;
          } else if (list.length > 0) {
            sel.value = list[list.length - 1].id;
          }
          if (sel.value) {
            storageSet('selected_project_id', sel.value);
          } else {
            storageRemove('selected_project_id');
          }
          applyProjectScoreScale(sel.value || '');
          await onProjectChanged();
        }
        async function onProjectChanged() {
          const selectedId = selectedProjectIdStrict() || pid();
          projectSwitchSeq += 1;
          const switchSeq = projectSwitchSeq;
          if (selectedId) storageSet('selected_project_id', selectedId);
          else storageRemove('selected_project_id');
          applyProjectScoreScale(selectedId);
          updateProjectBoundControlsState();
          resetProjectPanelsToStandby(selectedId);
          if (!selectedId) {
            setSelectMsg('请先在上方选择项目。', true);
            return;
          }
          await Promise.all([
            (typeof loadExpertProfile === 'function') ? loadExpertProfile(selectedId, switchSeq) : Promise.resolve(),
            (typeof refreshSubmissions === 'function') ? refreshSubmissions(selectedId, switchSeq) : Promise.resolve(),
            (typeof refreshMaterials === 'function') ? refreshMaterials(selectedId, switchSeq) : Promise.resolve(),
            (typeof refreshScoringReadiness === 'function') ? refreshScoringReadiness(selectedId, switchSeq) : Promise.resolve(),
            (typeof refreshFeedMaterials === 'function') ? refreshFeedMaterials(selectedId, switchSeq) : Promise.resolve(),
            (typeof refreshGroundTruth === 'function') ? refreshGroundTruth(selectedId, switchSeq) : Promise.resolve(),
            (typeof refreshGroundTruthSubmissionOptions === 'function') ? refreshGroundTruthSubmissionOptions(selectedId, switchSeq) : Promise.resolve(),
          ]);
          if (isStaleProjectResponse(selectedId, switchSeq)) return;
          setSelectMsg('已切换项目并自动刷新下方所有区域。', false);
        }
        const elRefresh = document.getElementById('refreshProjects');
        if (elRefresh) elRefresh.onclick = refreshProjects;
        safeChange('projectSelect', onProjectChanged);
        const btnSelfCheck = document.getElementById('btnSelfCheck');
        if (btnSelfCheck) btnSelfCheck.onclick = runSystemSelfCheck;
        const btnScoringFactors = document.getElementById('btnScoringFactors');
        if (btnScoringFactors) btnScoringFactors.onclick = loadScoringFactorsOverview;
        const btnScoringFactorsMd = document.getElementById('btnScoringFactorsMd');
        if (btnScoringFactorsMd) btnScoringFactorsMd.onclick = loadScoringFactorsMarkdown;
        const btnAnalysisBundle = document.getElementById('btnAnalysisBundle');
        if (btnAnalysisBundle) btnAnalysisBundle.onclick = loadProjectAnalysisBundle;
        const btnAnalysisBundleDownload = document.getElementById('btnAnalysisBundleDownload');
        if (btnAnalysisBundleDownload) btnAnalysisBundleDownload.onclick = downloadProjectAnalysisBundle;
        const deleteProjectForm = document.getElementById('deleteProjectForm');
        if (deleteProjectForm) {
          deleteProjectForm.onsubmit = async (e) => {
            e.preventDefault();
            const id = pid();
            if (!id) { setSelectMsg('请先选择要删除的项目', true); return; }
            const sel = document.getElementById('projectSelect');
            const opt = sel && sel.options && sel.selectedIndex >= 0 ? sel.options[sel.selectedIndex] : null;
            const label = (opt && opt.textContent) ? opt.textContent : id;
            const ok = confirm('确认删除项目「' + label + '」？\\n此操作不可恢复，并会删除该项目全部资料和评分记录。');
            if (!ok) return;
            setSelectMsg('正在删除项目…', false);
            let res, text;
            try {
              res = await fetch('/api/v1/projects/' + id, { method: 'DELETE', headers: apiHeaders(false) });
              text = await res.text();
            } catch (err) {
              setSelectMsg('删除失败: ' + (err.message || err), true);
              return;
            }
            if (res.status === 204) {
              setSelectMsg('项目已删除', false);
              const out = document.getElementById('output');
              if (out) out.textContent = '已删除项目：' + label;
              await refreshProjects();
              if (typeof refreshSubmissions === 'function') refreshSubmissions();
              if (typeof refreshMaterials === 'function') refreshMaterials();
              if (typeof refreshFeedMaterials === 'function') refreshFeedMaterials();
              if (typeof refreshGroundTruth === 'function') refreshGroundTruth();
            } else {
              let detail = text || '';
              try { const j = JSON.parse(text || '{}'); detail = (j && j.detail) || detail; } catch (_) {}
              setSelectMsg('删除失败: ' + res.status + ' ' + String(detail || '').slice(0, 120), true);
            }
          };
        }
        const deleteSelectedProjectsBtn = document.getElementById('deleteSelectedProjects');
        if (deleteSelectedProjectsBtn) {
          deleteSelectedProjectsBtn.onclick = async () => {
            const delSel = document.getElementById('projectDeleteSelect');
            const selectedOptions = delSel ? Array.from(delSel.selectedOptions || []) : [];
            const ids = selectedOptions.map((o) => String(o.value || '')).filter(Boolean);
            if (!ids.length) {
              setSelectMsg('请先在“批量删除”列表中选择项目（可按 Command/Ctrl 多选）。', true);
              return;
            }
            const labels = selectedOptions.map((o) => String(o.textContent || o.value || ''));
            const ok = confirm(
              '确认删除所选 ' + ids.length + ' 个项目？\\n\\n'
              + labels.join('\\n')
              + '\\n\\n此操作不可恢复，并会删除对应项目全部资料与记录。'
            );
            if (!ok) return;
            setSelectMsg('正在批量删除项目…', false);
            const failed = [];
            let removedCount = 0;
            for (const id of ids) {
              try {
                const res = await fetch(
                  '/api/v1/projects/' + encodeURIComponent(id),
                  { method: 'DELETE', headers: apiHeaders(false) }
                );
                const text = await res.text();
                if (res.status === 204) {
                  removedCount += 1;
                } else {
                  let detail = text || '';
                  try { const j = JSON.parse(text || '{}'); detail = (j && j.detail) || detail; } catch (_) {}
                  failed.push({ project_id: id, detail: String(detail || ('HTTP ' + res.status)).slice(0, 160) });
                }
              } catch (err) {
                failed.push({ project_id: id, detail: String((err && err.message) || err || '网络异常') });
              }
            }
            const out = document.getElementById('output');
            if (out) {
              out.textContent = JSON.stringify(
                { action: 'batch_delete_projects', requested: ids.length, removed_count: removedCount, failed_count: failed.length, failed },
                null,
                2
              );
            }
            if (failed.length) {
              setSelectMsg('批量删除完成：成功 ' + removedCount + '，失败 ' + failed.length + '。', true);
            } else {
              setSelectMsg('批量删除完成：已删除 ' + removedCount + ' 个项目。', false);
            }
            await refreshProjects();
            if (typeof refreshSubmissions === 'function') refreshSubmissions();
            if (typeof refreshMaterials === 'function') refreshMaterials();
            if (typeof refreshFeedMaterials === 'function') refreshFeedMaterials();
            if (typeof refreshGroundTruth === 'function') refreshGroundTruth();
          };
        }
        safeClick('btnCleanupE2EProjects', async () => {
          const ok = confirm('将批量删除名称以 E2E_ 开头的测试项目及其资料/施组记录。是否继续？');
          if (!ok) return;
          const res = await fetch('/api/v1/projects/cleanup_e2e?prefix=E2E_', { method: 'POST', headers: apiHeaders(false) });
          const text = await res.text();
          let data = {};
          try { data = JSON.parse(text || '{}'); } catch (_) {}
          const out = document.getElementById('output');
          if (out) out.textContent = res.ok ? JSON.stringify(data, null, 2) : text;
          if (res.ok) {
            setSelectMsg('E2E 清理完成：删除 ' + (data.removed_count || 0) + ' 个项目。', false);
            await refreshProjects();
          } else {
            setSelectMsg('E2E 清理失败：' + (data.detail || ('HTTP ' + res.status)), true);
          }
        });

        const formCreate = document.getElementById('createProject');
        if (formCreate) {
          formCreate.onsubmit = async (e) => {
            e.preventDefault();
            const name = (formCreate.elements.name && formCreate.elements.name.value || '').trim();
            if (!name) { setCreateMsg('请填写项目名称', true); return; }
            setCreateMsg('正在创建…', false);
            let res, text;
            try {
              res = await fetch('/api/v1/projects', { method: 'POST', headers: apiHeaders(), body: JSON.stringify({ name }) });
              text = await res.text();
            } catch (err) {
              setCreateMsg('网络错误: ' + (err.message || err), true);
              const out = document.getElementById('output');
              if (out) { out.textContent = '请求失败: ' + (err.message || err); out.scrollIntoView({ behavior: 'smooth' }); }
              return;
            }
            const outEl = document.getElementById('output');
            if (outEl) outEl.textContent = text;
            if (res && res.ok) {
              try {
                const created = JSON.parse(text || '{}');
                if (created && created.id) storageSet('selected_project_id', created.id);
              } catch (_) {}
              setCreateMsg('创建成功，已刷新下方列表，请选择项目', false);
              await refreshProjects();
            } else {
              let detail = text;
              try { const j = JSON.parse(text); detail = (j && j.detail) || text; } catch (_) {}
              setCreateMsg('创建失败: ' + res.status + ' ' + (detail || '').slice(0, 100), true);
              const outSc = document.getElementById('output');
              if (outSc) outSc.scrollIntoView({ behavior: 'smooth' });
            }
          };
        } else {
          console.error('createProject form not found');
        }

        function setFormPid(form) {
          const fid = form.querySelector('input[name="project_id"]');
          if (fid) return fid.value;
          const hidden = document.createElement('input');
          hidden.name = 'project_id';
          hidden.value = pid();
          form.appendChild(hidden);
          return pid();
        }

        const MATERIAL_UPLOAD_CONFIGS = {
          tender_qa: { formId: 'uploadMaterial', statusId: 'materialsActionStatus', typeLabel: '招标文件和答疑' },
          boq: { formId: 'uploadMaterialBoq', statusId: 'materialsActionStatusBoq', typeLabel: '清单' },
          drawing: { formId: 'uploadMaterialDrawing', statusId: 'materialsActionStatusDrawing', typeLabel: '图纸' },
          site_photo: { formId: 'uploadMaterialPhoto', statusId: 'materialsActionStatusPhoto', typeLabel: '现场照片' },
        };
        const uploadMaterialsInFlightByType = {};
        function bindMaterialUploadForm(materialType) {
          const cfg = MATERIAL_UPLOAD_CONFIGS[materialType];
          if (!cfg) return;
          const form = document.getElementById(cfg.formId);
          if (!form) return;
          form.onsubmit = async (e) => {
            e.preventDefault();
            e.stopPropagation();
            await uploadMaterialsAction(materialType);
            return false;
          };
        }
        bindMaterialUploadForm('tender_qa');
        bindMaterialUploadForm('boq');
        bindMaterialUploadForm('drawing');
        bindMaterialUploadForm('site_photo');
        async function uploadMaterialsAction(materialType = 'tender_qa') {
          const cfg = MATERIAL_UPLOAD_CONFIGS[materialType] || MATERIAL_UPLOAD_CONFIGS.tender_qa;
          if (uploadMaterialsInFlightByType[materialType]) {
            setActionStatus(cfg.statusId, cfg.typeLabel + '上传进行中，请稍候…', false);
            return;
          }
          uploadMaterialsInFlightByType[materialType] = true;
          try {
            const projectId = pid();
            if (!projectId) {
              const o = document.getElementById('output');
              if (o) o.textContent = '请先选择项目';
              setActionStatus(cfg.statusId, '上传失败：请先选择项目。', true);
              updateProjectBoundControlsState();
              return;
            }
            const form = document.getElementById(cfg.formId);
            const fileInput = form && form.querySelector ? form.querySelector('input[name="file"]') : null;
            const files = Array.from((fileInput && fileInput.files) || []);
            if (!files.length) {
              const o = document.getElementById('output');
              if (o) o.textContent = '请先选择要上传的文件';
              setActionStatus(cfg.statusId, '请先选择至少 1 个' + cfg.typeLabel + '文件。', true);
              return;
            }
            const headers = {};
            const apiKey = storageGet('api_key');
            if (apiKey) headers['X-API-Key'] = apiKey;
            const out = document.getElementById('output');
            if (out) out.textContent = cfg.typeLabel + '上传中（' + files.length + ' 个）...';
            setActionStatus(cfg.statusId, cfg.typeLabel + '上传中（' + files.length + ' 个）...', false);
            let okCount = 0;
            let failCount = 0;
            const details = [];
            for (const f of files) {
              const fd = new FormData();
              fd.append('file', f);
              fd.append('material_type', materialType);
              try {
                const res = await fetch('/api/v1/projects/' + projectId + '/materials', {
                  method: 'POST',
                  headers,
                  body: fd,
                });
                const text = await res.text();
                if (res.ok) {
                  okCount += 1;
                  details.push('[成功] ' + f.name);
                } else {
                  failCount += 1;
                  let detail = text || '';
                  try { const j = JSON.parse(text || '{}'); detail = (j && j.detail) || detail; } catch (_) {}
                  details.push('[失败] ' + f.name + ' -> HTTP ' + res.status + ' ' + String(detail).slice(0, 120));
                }
              } catch (err) {
                failCount += 1;
                details.push('[失败] ' + f.name + ' -> ' + String((err && err.message) || err || '网络异常'));
              }
            }
            if (out) out.textContent = cfg.typeLabel + '上传完成：成功 ' + okCount + '，失败 ' + failCount + NL + details.join(NL);
            setActionStatus(
              cfg.statusId,
              '上传完成：成功 ' + okCount + '，失败 ' + failCount + '。',
              failCount > 0
            );
            if (okCount > 0) {
              await refreshMaterials(projectId, projectSwitchSeq);
              if (typeof refreshFeedMaterials === 'function') await refreshFeedMaterials(projectId, projectSwitchSeq);
            }
            if (fileInput && failCount === 0) fileInput.value = '';
          } finally {
            uploadMaterialsInFlightByType[materialType] = false;
          }
        }

        const formShigong = document.getElementById('uploadShigong');
        let uploadShigongInFlight = false;
        let scoreShigongInFlight = false;
        let shigongSubmitIntent = 'upload';
        const btnUploadShigong = document.getElementById('btnUploadShigong');
        const btnScoreShigong = document.getElementById('btnScoreShigong');
        if (btnUploadShigong) {
          btnUploadShigong.addEventListener('click', () => { shigongSubmitIntent = 'upload'; });
        }
        if (btnScoreShigong) {
          btnScoreShigong.addEventListener('click', () => { shigongSubmitIntent = 'score'; });
        }
        if (formShigong) {
          formShigong.onsubmit = async (e) => {
            e.preventDefault();
            e.stopPropagation();
            let sid = e && e.submitter ? String(e.submitter.id || '') : '';
            if (!sid) {
              const active = document.activeElement;
              sid = active && active.id ? String(active.id) : '';
            }
            const isScoreSubmit = sid === 'btnScoreShigong' || shigongSubmitIntent === 'score';
            shigongSubmitIntent = 'upload';
            if (isScoreSubmit) await scoreShigongAction();
            else await uploadShigongAction();
            return false;
          };
        }
        async function uploadShigongAction() {
          if (uploadShigongInFlight) {
            setActionStatus('shigongActionStatus', '施组上传进行中，请稍候…', false);
            return;
          }
          uploadShigongInFlight = true;
          try {
            const projectId = pid();
            if (!projectId) {
              const o = document.getElementById('output');
              if (o) o.textContent = '请先选择项目';
              setActionStatus('shigongActionStatus', '上传失败：请先选择项目。', true);
              updateProjectBoundControlsState();
              return;
            }
            const fileInput = formShigong && formShigong.querySelector ? formShigong.querySelector('input[name="file"]') : null;
            const files = Array.from((fileInput && fileInput.files) || []);
            if (!files.length) {
              const o = document.getElementById('output');
              if (o) o.textContent = '请先选择要上传的施组文件';
              setActionStatus('shigongActionStatus', '请先选择至少 1 个施组文件。', true);
              return;
            }
            const headers = {};
            const apiKey = storageGet('api_key');
            if (apiKey) headers['X-API-Key'] = apiKey;
            const o = document.getElementById('output');
            if (o) o.textContent = '施组上传中（' + files.length + ' 个）...';
            setActionStatus('shigongActionStatus', '施组上传中（' + files.length + ' 个）...', false);
            let okCount = 0;
            let failCount = 0;
            const details = [];
            for (const f of files) {
              const fd = new FormData();
              fd.append('file', f);
              try {
                const res = await fetch('/api/v1/projects/' + projectId + '/shigong', { method: 'POST', headers, body: fd });
                const text = await res.text();
                if (res.ok) {
                  okCount += 1;
                  details.push('[成功] ' + f.name);
                } else {
                  failCount += 1;
                  let detail = text || '';
                  try { const j = JSON.parse(text || '{}'); detail = (j && j.detail) || detail; } catch (_) {}
                  details.push('[失败] ' + f.name + ' -> HTTP ' + res.status + ' ' + String(detail).slice(0, 120));
                }
              } catch (err) {
                failCount += 1;
                details.push('[失败] ' + f.name + ' -> ' + String((err && err.message) || err || '网络异常'));
              }
            }
            if (o) o.textContent = '施组上传完成：成功 ' + okCount + '，失败 ' + failCount + '。成功文件已入库，待点击“评分施组”后出分。' + NL + details.join(NL);
            setActionStatus(
              'shigongActionStatus',
              '上传完成：成功 ' + okCount + '，失败 ' + failCount + '。成功文件待评分。',
              failCount > 0
            );
            if (okCount > 0) await refreshSubmissions(projectId, projectSwitchSeq);
            if (fileInput && failCount === 0) fileInput.value = '';
          } finally {
            uploadShigongInFlight = false;
          }
        }
        async function scoreShigongAction() {
          if (scoreShigongInFlight) {
            setActionStatus('shigongActionStatus', '施组评分进行中，请稍候…', false);
            return;
          }
          scoreShigongInFlight = true;
          try {
            const projectId = pid();
            if (!projectId) {
              const o = document.getElementById('output');
              if (o) o.textContent = '请先选择项目';
              setActionStatus('shigongActionStatus', '评分失败：请先选择项目。', true);
              return;
            }
            if (typeof refreshScoringReadiness === 'function') {
              await refreshScoringReadiness(projectId, projectSwitchSeq);
            }
            if (
              scoringReadinessState
              && String(scoringReadinessState.project_id || '') === String(projectId)
              && !scoringReadinessState.ready
            ) {
              const firstIssue = Array.isArray(scoringReadinessState.issues) && scoringReadinessState.issues.length
                ? String(scoringReadinessState.issues[0])
                : '评分前置条件未满足';
              const blockMsg = '评分已阻止：' + firstIssue;
              const o = document.getElementById('output');
              if (o) o.textContent = blockMsg;
              setActionStatus('shigongActionStatus', blockMsg, true);
              return;
            }
            const o = document.getElementById('output');
            const scaleLabel = selectedScoreScaleLabel();
            if (o) o.textContent = '施组评分中（' + scaleLabel + '）...';
            setActionStatus('shigongActionStatus', '施组评分中（' + scaleLabel + '）...', false);
            let res;
            let data = {};
            try {
              res = await fetch('/api/v1/projects/' + projectId + '/rescore', {
                method: 'POST',
                headers: apiHeaders(true),
                body: JSON.stringify({
                  scope: 'project',
                  scoring_engine_version: 'v2',
                  score_scale_max: selectedScoreScaleMax(),
                }),
              });
              data = await res.json().catch(() => ({}));
            } catch (err) {
              const msg = '评分失败：' + String((err && err.message) || err || '网络异常');
              if (o) o.textContent = msg;
              setActionStatus('shigongActionStatus', msg, true);
              return;
            }
            if (!res.ok) {
              const detail = (data && data.detail) ? String(data.detail) : ('HTTP ' + String(res.status || 0));
              if (o) o.textContent = '评分失败：' + detail;
              setActionStatus('shigongActionStatus', '评分失败：' + detail, true);
              return;
            }
            const updated = Number(
              (data && (data.updated_submissions ?? data.reports_generated ?? data.submission_count)) || 0
            );
            const doneScaleLabel = (data && data.score_scale_label) ? String(data.score_scale_label) : scaleLabel;
            const gate = (data && typeof data.material_utilization_gate === 'object') ? data.material_utilization_gate : {};
            const closedLoop = (data && typeof data.feedback_closed_loop === 'object') ? data.feedback_closed_loop : {};
            const blockedCount = Number((gate && gate.blocked_submissions) || 0);
            const warnCount = Number((gate && gate.warn_submissions) || 0);
            let doneMsg = '施组评分完成（' + doneScaleLabel + '）：已重算 ' + updated + ' 份。';
            if (blockedCount > 0) {
              doneMsg += ' 资料门禁阻断 ' + blockedCount + ' 份。';
            } else if (warnCount > 0) {
              doneMsg += ' 资料门禁预警 ' + warnCount + ' 份。';
            } else if (gate && gate.enabled) {
              doneMsg += ' 资料门禁通过。';
            }
            if (Object.keys(closedLoop).length) {
              if (closedLoop.ok === false) doneMsg += ' 反馈闭环执行异常，请查看“原始输出”。';
              else doneMsg += ' 反馈闭环已执行。';
            }
            if (o) o.textContent = doneMsg;
            setActionStatus('shigongActionStatus', doneMsg, blockedCount > 0);
            if (typeof renderMaterialUtilizationPanel === 'function') {
              renderMaterialUtilizationPanel(data || {});
            }
            await refreshSubmissions(projectId, projectSwitchSeq);
            if (typeof refreshScoringReadiness === 'function') {
              await refreshScoringReadiness(projectId, projectSwitchSeq);
            }
          } finally {
            scoreShigongInFlight = false;
          }
        }
        function updateTableEmptyState(tableId, emptyId) {
          const table = document.getElementById(tableId);
          const tbody = table ? table.querySelector('tbody') : null;
          const emptyEl = document.getElementById(emptyId);
          if (!emptyEl) return;
          emptyEl.style.display = (tbody && tbody.querySelector('tr')) ? 'none' : 'block';
        }
        function extractApiErrorMessage(status, text, data) {
          const detail = (data && data.detail) ? String(data.detail) : String(text || '').trim();
          return '删除失败：HTTP ' + String(status || 0) + (detail ? (' ' + detail) : '');
        }
        async function deleteSubmissionRow(submissionId, rowEl, filename, projectIdOverride='') {
          const id = resolveProjectId(projectIdOverride);
          if (!id) { alert('删除失败：请先选择项目'); return; }
          if (!submissionId) { alert('删除失败：记录ID为空'); return; }
          const ok = confirm('确认删除该文件？');
          if (!ok) return;
          let res;
          let text = '';
          let data = {};
          try {
            const encodedPid = encodeURIComponent(String(id || ''));
            const encodedFid = encodeURIComponent(String(submissionId || ''));
            res = await fetch('/api/v1/projects/' + encodedPid + '/shigong/' + encodedFid, { method: 'DELETE', headers: apiHeaders(false) });
            text = await res.text();
            try { data = JSON.parse(text || '{}'); } catch (_) { data = {}; }
          } catch (err) {
            alert('删除失败：' + String((err && err.message) || err || '网络异常'));
            return;
          }
          if (!res.ok) {
            alert(extractApiErrorMessage(res.status, text, data));
            return;
          }
          if (rowEl) rowEl.remove();
          updateTableEmptyState('submissionsTable', 'submissionsEmpty');
          if (typeof refreshScoringReadiness === 'function') refreshScoringReadiness();
          if (typeof refreshGroundTruthSubmissionOptions === 'function') refreshGroundTruthSubmissionOptions();
          const out = document.getElementById('output');
          if (out) out.textContent = JSON.stringify({ ok: true, id: submissionId, filename: filename || '' }, null, 2);
        }
        async function deleteMaterialRow(materialId, rowEl, filename, projectIdOverride='') {
          const id = resolveProjectId(projectIdOverride);
          if (!id) { alert('删除失败：请先选择项目'); return; }
          if (!materialId) { alert('删除失败：记录ID为空'); return; }
          const ok = confirm('确认删除该文件？');
          if (!ok) return;
          let res;
          let text = '';
          let data = {};
          try {
            const encodedPid = encodeURIComponent(String(id || ''));
            const encodedFid = encodeURIComponent(String(materialId || ''));
            res = await fetch('/api/v1/projects/' + encodedPid + '/materials/' + encodedFid, { method: 'DELETE', headers: apiHeaders(false) });
            text = await res.text();
            try { data = JSON.parse(text || '{}'); } catch (_) { data = {}; }
          } catch (err) {
            alert('删除失败：' + String((err && err.message) || err || '网络异常'));
            return;
          }
          if (!res.ok) {
            alert(extractApiErrorMessage(res.status, text, data));
            return;
          }
          if (rowEl) rowEl.remove();
          updateTableEmptyState('materialsTable', 'materialsEmpty');
          if (typeof refreshFeedMaterials === 'function') refreshFeedMaterials();
          if (typeof refreshScoringReadiness === 'function') refreshScoringReadiness();
          const out = document.getElementById('output');
          if (out) out.textContent = JSON.stringify({ ok: true, id: materialId, filename: filename || '' }, null, 2);
        }
        function bindDeleteRowHandlers() {
          const submissionsTable = document.getElementById('submissionsTable');
          if (submissionsTable && !submissionsTable.dataset.deleteBound) {
            submissionsTable.dataset.deleteBound = '1';
            submissionsTable.addEventListener('click', (ev) => {
              const btn = ev.target && ev.target.closest ? ev.target.closest('.js-delete-submission') : null;
              if (!btn) return;
              ev.preventDefault();
              deleteSubmissionRow(
                btn.getAttribute('data-submission-id') || '',
                btn.closest('tr'),
                btn.getAttribute('data-filename') || '',
                btn.getAttribute('data-project-id') || ''
              );
            });
          }
          const materialsTable = document.getElementById('materialsTable');
          if (materialsTable && !materialsTable.dataset.deleteBound) {
            materialsTable.dataset.deleteBound = '1';
            materialsTable.addEventListener('click', (ev) => {
              const btn = ev.target && ev.target.closest ? ev.target.closest('.js-delete-material') : null;
              if (!btn) return;
              ev.preventDefault();
              deleteMaterialRow(
                btn.getAttribute('data-material-id') || '',
                btn.closest('tr'),
                btn.getAttribute('data-filename') || '',
                btn.getAttribute('data-project-id') || ''
              );
            });
          }
        }
        async function refreshSubmissions(expectedProjectId=null, switchSeq=null) {
          const id = expectedProjectId || pid();
          const tbl = document.getElementById('submissionsTable');
          const tbody = tbl ? tbl.querySelector('tbody') : null;
          const emptyEl = document.getElementById('submissionsEmpty');
          if (tbody) tbody.innerHTML = '';
          if (!id) {
            if (emptyEl) {
              emptyEl.textContent = '暂无施组，请先选择项目。';
              emptyEl.style.display = 'block';
            }
            if (typeof clearMaterialUtilizationPanel === 'function') clearMaterialUtilizationPanel();
            if (typeof clearScoringReadinessPanel === 'function') clearScoringReadinessPanel();
            return;
          }
          let res;
          try {
            res = await fetch('/api/v1/projects/' + id + '/submissions?t=' + Date.now(), { cache: 'no-store' });
          } catch (err) {
            if (isStaleProjectResponse(id, switchSeq)) return;
            if (emptyEl) {
              emptyEl.textContent = '施组列表加载失败，请稍后重试。';
              emptyEl.style.display = 'block';
            }
            return;
          }
          if (isStaleProjectResponse(id, switchSeq)) return;
          const subs = await res.json().catch(() => []);
          if (!res.ok) {
            if (emptyEl) {
              emptyEl.textContent = '施组列表加载失败（HTTP ' + String(res.status || 0) + '）';
              emptyEl.style.display = 'block';
            }
            if (typeof clearMaterialUtilizationPanel === 'function') clearMaterialUtilizationPanel();
            if (typeof refreshScoringReadiness === 'function') await refreshScoringReadiness(id, switchSeq);
            return;
          }
          if (!Array.isArray(subs) || subs.length === 0) {
            if (emptyEl) {
              emptyEl.textContent = '暂无施组，请下方添加。';
              emptyEl.style.display = 'block';
            }
            if (typeof clearMaterialUtilizationPanel === 'function') clearMaterialUtilizationPanel();
            if (typeof refreshScoringReadiness === 'function') await refreshScoringReadiness(id, switchSeq);
            return;
          }
          if (emptyEl) emptyEl.style.display = 'none';
          subs.forEach(s => {
            const tr = document.createElement('tr');
            const rep = (s && typeof s === 'object') ? (s.report || {}) : {};
            const pred = (rep && rep.pred_total_score != null) ? rep.pred_total_score : null;
            const rule = (rep && rep.rule_total_score != null) ? rep.rule_total_score : null;
            const llm = (rep && rep.llm_total_score != null) ? rep.llm_total_score : null;
            const scoringStatus = String((rep && rep.scoring_status) || '').toLowerCase();
            const isPending = scoringStatus === 'pending';
            const isBlocked = scoringStatus === 'blocked';
            const repMeta = (rep && typeof rep.meta === 'object') ? rep.meta : {};
            const utilGate = (repMeta && typeof repMeta.material_utilization_gate === 'object')
              ? repMeta.material_utilization_gate
              : {};
            const utilBlocked = !!(utilGate && utilGate.blocked);
            const evidenceTrace = (repMeta && typeof repMeta.evidence_trace === 'object')
              ? repMeta.evidence_trace
              : {};
            let scoreHtml = '-';
            if (isPending) {
              scoreHtml = '<span class="note">待评分</span>';
            } else if (isBlocked) {
              scoreHtml = '<span class="error">待补资料后重评分</span>';
            } else if (pred != null) {
              scoreHtml = escapeHtmlText(String(pred));
              const notes = [];
              if (rule != null) notes.push('规则: ' + escapeHtmlText(String(rule)));
              if (llm != null) notes.push('LLM: ' + escapeHtmlText(String(llm)));
              if (notes.length) scoreHtml += '<div class="note">' + notes.join(' / ') + '</div>';
            } else if (s && s.total_score != null) {
              scoreHtml = escapeHtmlText(String(s.total_score));
            }
            const evidenceCount = Number(evidenceTrace.total_hits || 0);
            const evidenceFileCount = Number(evidenceTrace.source_files_hit_count || 0);
            if (!isPending && evidenceCount > 0) {
              scoreHtml += '<div class="note">证据命中: ' + escapeHtmlText(evidenceCount) + ' 条 / 文件覆盖: ' + escapeHtmlText(evidenceFileCount) + ' 份</div>';
            }
            const utilSummary = (repMeta && typeof repMeta.material_utilization === 'object')
              ? repMeta.material_utilization
              : {};
            const utilByType = (utilSummary && typeof utilSummary.by_type === 'object')
              ? utilSummary.by_type
              : {};
            const utilAvailableTypes = Array.isArray(utilSummary.available_types) ? utilSummary.available_types : [];
            const typeLabelShort = (t) => {
              const key = String(t || '').trim();
              if (key === 'tender_qa') return '招答';
              if (key === 'boq') return '清单';
              if (key === 'drawing') return '图纸';
              if (key === 'site_photo') return '照片';
              return key || '-';
            };
            const hasTypeEvidence = (t) => {
              const row = (utilByType && typeof utilByType[t] === 'object') ? utilByType[t] : {};
              const retrievalHit = Number((row && row.retrieval_hit) || 0);
              const consistencyHit = Number((row && row.consistency_hit) || 0);
              return (retrievalHit + consistencyHit) > 0;
            };
            const orderedTypes = ['tender_qa', 'boq', 'drawing', 'site_photo'];
            const coverageTokens = orderedTypes.map((t) => {
              const inScope = utilAvailableTypes.includes(t);
              if (!inScope) return typeLabelShort(t) + '·';
              return typeLabelShort(t) + (hasTypeEvidence(t) ? '✓' : '×');
            });
            if (!isPending && coverageTokens.length) {
              scoreHtml += '<div class="note">类型覆盖: ' + escapeHtmlText(coverageTokens.join(' / ')) + '</div>';
            }
            const evidenceFiles = Array.isArray(evidenceTrace.source_files_hit) ? evidenceTrace.source_files_hit : [];
            if (!isPending && evidenceFiles.length) {
              scoreHtml += '<div class="note">命中文件: ' + escapeHtmlText(evidenceFiles.slice(0, 2).join('；')) + (evidenceFiles.length > 2 ? ' 等' : '') + '</div>';
            }
            if (utilBlocked) {
              scoreHtml += '<div class="error">资料利用门禁未达标（建议补齐资料后重评分）</div>';
            }
            tr.innerHTML =
              '<td>' + escapeHtmlText(s.filename || '') + '</td>' +
              '<td>' + scoreHtml + '</td>' +
              '<td>' + escapeHtmlText((s.created_at || '').slice(0,19)) + '</td>' +
              '<td><button type="button" class="btn-danger js-delete-submission" data-submission-id="' + escapeHtmlText(String(s.id || '')) + '" data-project-id="' + escapeHtmlText(String(id || '')) + '" data-filename="' + escapeHtmlText(String(s.filename || '')) + '">删除</button></td>';
            if (tbody) tbody.appendChild(tr);
          });
          let utilPayload = null;
          for (const s of subs) {
            const rep = (s && typeof s === 'object') ? (s.report || {}) : {};
            const meta = (rep && typeof rep.meta === 'object') ? rep.meta : {};
            const util = (meta && typeof meta.material_utilization === 'object') ? meta.material_utilization : null;
            if (!util) continue;
            utilPayload = {
              material_utilization: util,
              material_utilization_alerts: Array.isArray(meta.material_utilization_alerts)
                ? meta.material_utilization_alerts
                : [],
              material_utilization_gate: (meta && typeof meta.material_utilization_gate === 'object')
                ? meta.material_utilization_gate
                : {},
            };
            break;
          }
          if (typeof renderMaterialUtilizationPanel === 'function') {
            if (utilPayload) renderMaterialUtilizationPanel(utilPayload);
            else clearMaterialUtilizationPanel();
          }
          updateTableEmptyState('submissionsTable', 'submissionsEmpty');
          if (typeof refreshScoringReadiness === 'function') await refreshScoringReadiness(id, switchSeq);
          await refreshGroundTruthSubmissionOptions(id, switchSeq);
        }
        async function refreshGroundTruthSubmissionOptions(expectedProjectId=null, switchSeq=null) {
          const id = expectedProjectId || pid();
          const sel = document.getElementById('groundTruthSubmissionSelect');
          if (!sel) return;
          const prev = String(sel.value || '');
          sel.innerHTML = '';
          const pendingOpt = document.createElement('option');
          pendingOpt.value = '';
          if (!id) {
            pendingOpt.textContent = '-- 请先选择项目 --';
            sel.appendChild(pendingOpt);
            sel.value = '';
            return;
          }
          pendingOpt.textContent = '-- 加载步骤4施组中... --';
          sel.appendChild(pendingOpt);
          sel.disabled = true;
          let res;
          try {
            res = await fetch('/api/v1/projects/' + id + '/submissions?t=' + Date.now(), { cache: 'no-store' });
          } catch (_) {
            sel.innerHTML = '';
            const errOpt = document.createElement('option');
            errOpt.value = '';
            errOpt.textContent = '-- 施组列表加载失败，请稍后重试 --';
            sel.appendChild(errOpt);
            sel.value = '';
            sel.disabled = false;
            return;
          }
          if (isStaleProjectResponse(id, switchSeq)) {
            sel.disabled = false;
            return;
          }
          const subs = await res.json().catch(() => []);
          sel.innerHTML = '';
          const leadOpt = document.createElement('option');
          leadOpt.value = '';
          if (!res.ok || !Array.isArray(subs)) {
            leadOpt.textContent = '-- 施组列表加载失败，请稍后重试 --';
            sel.appendChild(leadOpt);
            sel.value = '';
            sel.disabled = false;
            return;
          }
          if (!subs.length) {
            leadOpt.textContent = '-- 暂无施组，请先在步骤4上传 --';
            sel.appendChild(leadOpt);
            sel.value = '';
            sel.disabled = false;
            return;
          }
          leadOpt.textContent = '-- 请选择步骤4已上传施组文件 --';
          sel.appendChild(leadOpt);
          subs.forEach((s) => {
            const opt = document.createElement('option');
            opt.value = String((s && s.id) || '');
            const report = (s && s.report) || {};
            const status = String((report && report.scoring_status) || '').toLowerCase();
            const statusLabel = status === 'pending'
              ? '待评分'
              : (status === 'blocked' ? '已阻断' : (status === 'scored' ? '已评分' : ''));
            const createdAt = String((s && s.created_at) || '').slice(0, 19);
            const suffix = [createdAt, statusLabel].filter(Boolean).join(' / ');
            opt.textContent = String((s && s.filename) || '未命名施组') + (suffix ? ('（' + suffix + '）') : '');
            sel.appendChild(opt);
          });
          if (prev && subs.some((s) => String((s && s.id) || '') === prev)) {
            sel.value = prev;
          } else {
            sel.value = '';
          }
          sel.disabled = false;
        }
        async function refreshMaterials(expectedProjectId=null, switchSeq=null) {
          const id = expectedProjectId || pid();
          const tbl = document.getElementById('materialsTable');
          const tbody = tbl ? tbl.querySelector('tbody') : null;
          const emptyEl = document.getElementById('materialsEmpty');
          if (tbody) tbody.innerHTML = '';
          if (!id) {
            if (emptyEl) {
              emptyEl.textContent = '暂无资料，请先选择项目。';
              emptyEl.style.display = 'block';
            }
            if (typeof clearScoringReadinessPanel === 'function') clearScoringReadinessPanel();
            return;
          }
          let res;
          try {
            res = await fetch('/api/v1/projects/' + id + '/materials?t=' + Date.now(), { cache: 'no-store' });
          } catch (err) {
            if (isStaleProjectResponse(id, switchSeq)) return;
            if (emptyEl) {
              emptyEl.textContent = '资料列表加载失败，请稍后重试。';
              emptyEl.style.display = 'block';
            }
            return;
          }
          if (isStaleProjectResponse(id, switchSeq)) return;
          const mats = await res.json().catch(() => []);
          if (!res.ok) {
            if (emptyEl) {
              emptyEl.textContent = '资料列表加载失败（HTTP ' + String(res.status || 0) + '）';
              emptyEl.style.display = 'block';
            }
            if (typeof refreshScoringReadiness === 'function') await refreshScoringReadiness(id, switchSeq);
            return;
          }
          if (!Array.isArray(mats) || mats.length === 0) {
            if (emptyEl) {
              emptyEl.textContent = '暂无资料，请下方添加。';
              emptyEl.style.display = 'block';
            }
            if (typeof refreshScoringReadiness === 'function') await refreshScoringReadiness(id, switchSeq);
            return;
          }
          if (emptyEl) emptyEl.style.display = 'none';
          mats.forEach(m => {
            const tr = document.createElement('tr');
            const typeLabel = materialTypeLabel((m && m.material_type) || 'tender_qa');
            tr.innerHTML =
              '<td>' + escapeHtmlText(typeLabel) + '</td>' +
              '<td>' + escapeHtmlText(m.filename || '') + '</td>' +
              '<td>' + escapeHtmlText((m.created_at || '').slice(0,19)) + '</td>' +
              '<td><button type="button" class="btn-danger js-delete-material" data-material-id="' + escapeHtmlText(String(m.id || '')) + '" data-project-id="' + escapeHtmlText(String(id || '')) + '" data-filename="' + escapeHtmlText(String(m.filename || '')) + '">删除</button></td>';
            if (tbody) tbody.appendChild(tr);
          });
          updateTableEmptyState('materialsTable', 'materialsEmpty');
          if (typeof refreshScoringReadiness === 'function') await refreshScoringReadiness(id, switchSeq);
        }
        bindDeleteRowHandlers();
        const btnRefSub = document.getElementById('btnRefreshSubmissions');
        if (btnRefSub) btnRefSub.onclick = refreshSubmissions;
        const btnRefMat = document.getElementById('btnRefreshMaterials');
        if (btnRefMat) btnRefMat.onclick = refreshMaterials;
        safeClick('btnMaterialDepthReport', async () => {
          if (!ensureProjectForAction('materialDepthReportResult')) return;
          const id = pid();
          setResultLoading('materialDepthReportResult', '资料深读体检生成中...');
          let res;
          let data = {};
          try {
            res = await fetch('/api/v1/projects/' + encodeURIComponent(id) + '/materials/depth_report', {
              method: 'GET',
              headers: apiHeaders(false),
            });
            data = await res.json().catch(() => ({}));
          } catch (err) {
            setResultError('materialDepthReportResult', '体检失败：' + String((err && err.message) || err || '网络异常'));
            return;
          }
          if (!res.ok) {
            const detail = (data && data.detail) ? String(data.detail) : ('HTTP ' + String(res.status || 0));
            setResultError('materialDepthReportResult', '体检失败：' + detail);
            return;
          }
          renderMaterialDepthReportPanel(data);
          const out = document.getElementById('output');
          if (out) out.textContent = JSON.stringify(data, null, 2);
        });
        safeClick('btnMaterialDepthReportDownload', async () => {
          if (!ensureProjectForAction('materialDepthReportResult')) return;
          const id = pid();
          const url = '/api/v1/projects/' + encodeURIComponent(id) + '/materials/depth_report.md';
          const a = document.createElement('a');
          a.href = url;
          a.download = 'material_depth_report_' + id + '.md';
          document.body.appendChild(a);
          a.click();
          a.remove();
          setResultSuccess('materialDepthReportResult', '资料深读体检报告下载已触发。');
        });
        safeClick('btnMaterialKnowledgeProfile', async () => {
          if (!ensureProjectForAction('materialKnowledgeProfileResult')) return;
          const id = pid();
          setResultLoading('materialKnowledgeProfileResult', '资料知识画像生成中...');
          let res;
          let data = {};
          try {
            res = await fetch('/api/v1/projects/' + encodeURIComponent(id) + '/materials/knowledge_profile', {
              method: 'GET',
              headers: apiHeaders(false),
            });
            data = await res.json().catch(() => ({}));
          } catch (err) {
            setResultError('materialKnowledgeProfileResult', '画像生成失败：' + String((err && err.message) || err || '网络异常'));
            return;
          }
          if (!res.ok) {
            const detail = (data && data.detail) ? String(data.detail) : ('HTTP ' + String(res.status || 0));
            setResultError('materialKnowledgeProfileResult', '画像生成失败：' + detail);
            return;
          }
          renderMaterialKnowledgeProfilePanel(data);
          const out = document.getElementById('output');
          if (out) out.textContent = JSON.stringify(data, null, 2);
        });
        safeClick('btnMaterialKnowledgeProfileDownload', async () => {
          if (!ensureProjectForAction('materialKnowledgeProfileResult')) return;
          const id = pid();
          const url = '/api/v1/projects/' + encodeURIComponent(id) + '/materials/knowledge_profile.md';
          const a = document.createElement('a');
          a.href = url;
          a.download = 'material_knowledge_profile_' + id + '.md';
          document.body.appendChild(a);
          a.click();
          a.remove();
          setResultSuccess('materialKnowledgeProfileResult', '资料知识画像报告下载已触发。');
        });

        async function refreshFeedMaterials(expectedProjectId=null, switchSeq=null) {
          const id = expectedProjectId || pid();
          const tbl = document.getElementById('feedMaterialsTable');
          const tbody = tbl ? tbl.querySelector('tbody') : null;
          const emptyEl = document.getElementById('feedMaterialsEmpty');
          if (tbody) tbody.innerHTML = '';
          if (!id) {
            if (emptyEl) {
              emptyEl.textContent = '暂无投喂包，请先选择项目。';
              emptyEl.style.display = 'block';
            }
            return;
          }
          let res;
          try {
            res = await fetch('/api/v1/projects/' + id + '/materials');
          } catch (err) {
            if (isStaleProjectResponse(id, switchSeq)) return;
            if (emptyEl) {
              emptyEl.textContent = '投喂包列表加载失败，请稍后重试。';
              emptyEl.style.display = 'block';
            }
            return;
          }
          if (isStaleProjectResponse(id, switchSeq)) return;
          const mats = await res.json().catch(() => []);
          if (!res.ok) {
            if (emptyEl) {
              emptyEl.textContent = '投喂包列表加载失败（HTTP ' + String(res.status || 0) + '）';
              emptyEl.style.display = 'block';
            }
            return;
          }
          if (!Array.isArray(mats) || mats.length === 0) {
            if (emptyEl) {
              emptyEl.textContent = '暂无投喂包，请在上方或「3) 项目资料」上传。';
              emptyEl.style.display = 'block';
            }
            return;
          }
          if (emptyEl) emptyEl.style.display = 'none';
          mats.forEach(m => {
            const tr = document.createElement('tr');
            tr.innerHTML = '<td>' + m.filename + '</td><td>' + (m.created_at || '').slice(0,19) + '</td><td><button type="button" class="btn-danger js-delete-material" data-material-id="' + String(m.id || '') + '" data-project-id="' + String(id || '') + '" data-filename="' + String(m.filename || '').replace(/"/g, '&quot;') + '">删除</button></td>';
            const btn = tr.querySelector('button');
            if (btn) btn.onclick = async () => {
              const r = await fetch('/api/v1/projects/' + id + '/materials/' + m.id, { method: 'DELETE', headers: apiHeaders() });
              if (r.ok) { refreshMaterials(); refreshFeedMaterials(); }
              else { const o = document.getElementById('output'); if (o) o.textContent = await r.text(); }
            };
            if (tbody) tbody.appendChild(tr);
          });
        }
        const btnRefFeed = document.getElementById('btnRefreshFeedMaterials');
        if (btnRefFeed) btnRefFeed.onclick = refreshFeedMaterials;

        function groundTruthListProjectId() {
          const scope = document.getElementById('groundTruthScope').value;
          if (scope === 'current') return pid();
          if (scope === 'other') return document.getElementById('groundTruthOtherProject').value || '';
          return '';
        }
        async function refreshGroundTruth(expectedProjectId=null, switchSeq=null) {
          const listId = expectedProjectId || groundTruthListProjectId();
          const tbl = document.getElementById('groundTruthTable');
          const tbody = tbl ? tbl.querySelector('tbody') : null;
          const emptyEl = document.getElementById('groundTruthEmpty');
          if (tbody) tbody.innerHTML = '';
          if (!listId) {
            if (emptyEl) {
              emptyEl.textContent = '暂无真实评标，请先选择项目。';
              emptyEl.style.display = 'block';
            }
            setResultError('evolveResult', '请先选择项目后再刷新真实评标列表');
            return;
          }
          setResultLoading('evolveResult', '正在刷新真实评标列表...');
          let res;
          try {
            res = await fetch('/api/v1/projects/' + listId + '/ground_truth');
          } catch (err) {
            if (isStaleProjectResponse(expectedProjectId || listId, switchSeq)) return;
            if (emptyEl) {
              emptyEl.textContent = '真实评标列表加载失败，请稍后重试。';
              emptyEl.style.display = 'block';
            }
            setResultError('evolveResult', '真实评标列表加载失败：' + String((err && err.message) || err || '网络异常'));
            return;
          }
          if (isStaleProjectResponse(expectedProjectId || listId, switchSeq)) return;
          const list = await res.json().catch(() => []);
          if (!res.ok) {
            if (emptyEl) {
              emptyEl.textContent = '真实评标列表加载失败（HTTP ' + String(res.status || 0) + '）';
              emptyEl.style.display = 'block';
            }
            setResultError('evolveResult', '真实评标列表加载失败（HTTP ' + String(res.status || 0) + '）');
            return;
          }
          const scope = document.getElementById('groundTruthScope').value;
          const isCurrent = scope === 'current' && listId === selectedProjectIdStrict();
          if (list.length === 0) {
            if (emptyEl) {
              emptyEl.textContent = '暂无真实评标，请下方录入。';
              emptyEl.style.display = 'block';
            }
            setResultSuccess('evolveResult', '刷新完成：当前范围暂无真实评标记录');
            return;
          }
          if (emptyEl) emptyEl.style.display = 'none';
          list.forEach((r, idx) => {
            const summary = (r.shigong_text || '').slice(0, 50);
            const scores = Array.isArray(r.judge_scores) ? r.judge_scores : [];
            const scoresStr = scores.length
              ? (scores.map(s => Number(s).toFixed(1)).join(', ') + '（' + scores.length + '人）')
              : '-';
            const tr = document.createElement('tr');
            const st = r.shigong_text || '';
            const actionCell = isCurrent
              ? '<td><button type="button" class="btn-danger js-delete-ground-truth" data-gt-id="' + escapeHtmlText(String(r.id || '')) + '">删除</button></td>'
              : '<td></td>';
            tr.innerHTML = '<td>' + (idx + 1) + '</td><td title="' + st.slice(0, 200).replace(/"/g, '&quot;') + '">' + (summary ? summary + (st.length > 50 ? '…' : '') : '-') + '</td><td>' + scoresStr + '</td><td>' + (r.final_score != null ? r.final_score : '-') + '</td><td>' + (r.source || '-') + '</td>' + actionCell;
            if (isCurrent) {
              const delBtn = tr.querySelector('button');
              if (delBtn) delBtn.onclick = async () => {
                const delRes = await fetch('/api/v1/projects/' + listId + '/ground_truth/' + r.id, { method: 'DELETE', headers: apiHeaders() });
                if (delRes.status === 204) refreshGroundTruth();
                else { const o = document.getElementById('output'); if (o) o.textContent = await delRes.text(); }
              };
            }
            if (tbody) tbody.appendChild(tr);
          });
          setResultSuccess('evolveResult', '刷新完成：共 ' + list.length + ' 条真实评标记录');
        }
        safeChange('groundTruthScope', function() {
          const otherSel = document.getElementById('groundTruthOtherProject');
          if (otherSel && this.value === 'other') {
            otherSel.style.display = 'inline';
            fetch('/api/v1/projects').then(r => r.json()).then(list => {
              const cur = pid();
              otherSel.innerHTML = '<option value="">-- 选择要查看的项目 --</option>';
              (list || []).forEach(p => {
                if (p.id === cur) return;
                const o = document.createElement('option');
                o.value = p.id;
                o.textContent = (p.name || p.id) + ' (' + (p.id || '').slice(0, 8) + '…)';
                otherSel.appendChild(o);
              });
            });
          } else if (otherSel) {
            otherSel.style.display = 'none';
            otherSel.value = '';
          }
          refreshGroundTruth();
        });
        safeChange('gtJudgeCount', function() {
          syncGroundTruthJudgeInputs();
        });
        safeChange('groundTruthOtherProject', refreshGroundTruth);
        safeClick('btnRefreshGroundTruth', refreshGroundTruth);
        safeClick('btnRefreshGroundTruthSubmissionOptions', async () => {
          if (!ensureProjectForAction('evolveResult')) return;
          setResultLoading('evolveResult', '施组选项刷新中...');
          await refreshGroundTruthSubmissionOptions();
          setResultSuccess('evolveResult', '施组选项已刷新：请在下拉框中选择步骤4已上传施组。');
        });

        function showJson(id, data) {
          const out = document.getElementById('output');
          if (out) out.textContent = typeof data === 'string' ? data : JSON.stringify(data, null, 2);
        }
        function escapeHtmlText(v) {
          return String(v == null ? '' : v)
            .replace(/&/g, '&amp;')
            .replace(/</g, '&lt;')
            .replace(/>/g, '&gt;')
            .replace(/\"/g, '&quot;')
            .replace(/'/g, '&#39;');
        }
        function ollamaPreviewUpdatedAt(data) {
          const preview = (data && typeof data.preview === 'object') ? data.preview : {};
          return String((preview && preview.updated_at) || (data && data.updated_at) || '-');
        }
        function ollamaPreviewStatusText(data) {
          return data && data.fallback ? '已回退' : '增强成功';
        }
        function setOllamaPreviewActionStatus(text, isError=false) {
          const el = document.getElementById('ollamaPreviewActionStatus');
          if (!el) return;
          el.textContent = text || '';
          el.style.color = isError ? '#b91c1c' : '#64748b';
        }
        function setOllamaPreviewActionButtons(enabled) {
          ['btnOllamaPreviewCopy', 'btnOllamaPreviewExport'].forEach((id) => {
            const btn = document.getElementById(id);
            if (!btn) return;
            btn.disabled = !enabled;
            btn.title = enabled ? '' : '请先生成 Ollama 增强预览';
          });
        }
        function resetOllamaPreviewActions() {
          latestOllamaPreviewPayload = null;
          latestOllamaPreviewProjectId = '';
          setOllamaPreviewActionButtons(false);
          setOllamaPreviewActionStatus('生成预览后可复制或导出。', false);
        }
        function storeOllamaPreviewPayload(projectId, data) {
          latestOllamaPreviewProjectId = String(projectId || '');
          latestOllamaPreviewPayload = (data && typeof data === 'object') ? data : null;
          setOllamaPreviewActionButtons(!!latestOllamaPreviewPayload);
          setOllamaPreviewActionStatus(
            latestOllamaPreviewPayload ? '预览已生成，可复制或导出 JSON。' : '生成预览后可复制或导出。',
            false
          );
        }
        function formatOllamaPreviewPlainText(data) {
          const preview = (data && typeof data.preview === 'object') ? data.preview : {};
          const logic = Array.isArray(preview.high_score_logic) ? preview.high_score_logic : [];
          const guidance = Array.isArray(preview.writing_guidance) ? preview.writing_guidance : [];
          const lines = [
            'Ollama 增强预览',
            '状态：' + ollamaPreviewStatusText(data),
            'enhanced_by：' + String((data && data.enhanced_by) || 'rules'),
            'fallback：' + String(!!(data && data.fallback)),
            'error_summary：' + String((data && data.error_summary) || '-'),
            '更新时间：' + ollamaPreviewUpdatedAt(data),
            '',
            '高分逻辑预览',
          ];
          if (logic.length) logic.forEach((x, idx) => lines.push(String(idx + 1) + '. ' + String(x)));
          else lines.push('暂无');
          lines.push('', '编制指导预览');
          if (guidance.length) guidance.forEach((x, idx) => lines.push(String(idx + 1) + '. ' + String(x)));
          else lines.push('暂无');
          lines.push('', '仅预览，不写入正式学习进化结果；不影响评分，不进入核心评分主链。');
          return lines.join(NL);
        }
        async function copyTextToClipboard(text) {
          if (navigator.clipboard && typeof navigator.clipboard.writeText === 'function') {
            await navigator.clipboard.writeText(text);
            return;
          }
          const area = document.createElement('textarea');
          area.value = text;
          area.setAttribute('readonly', 'readonly');
          area.style.position = 'fixed';
          area.style.left = '-9999px';
          document.body.appendChild(area);
          area.select();
          document.execCommand('copy');
          area.remove();
        }
        function downloadOllamaPreviewJson(projectId, data) {
          const safeProjectId = String(projectId || 'project').replace(/[^a-zA-Z0-9_-]+/g, '_');
          const stamp = new Date().toISOString().replace(/[:.]/g, '-');
          const blob = new Blob([JSON.stringify(data || {}, null, 2)], {
            type: 'application/json;charset=utf-8',
          });
          const url = URL.createObjectURL(blob);
          const a = document.createElement('a');
          a.href = url;
          a.download = 'ollama_preview_' + safeProjectId + '_' + stamp + '.json';
          document.body.appendChild(a);
          a.click();
          a.remove();
          URL.revokeObjectURL(url);
        }
        function renderOllamaPreviewHtml(data) {
          const preview = (data && typeof data.preview === 'object') ? data.preview : {};
          const logic = Array.isArray(preview.high_score_logic) ? preview.high_score_logic : [];
          const guidance = Array.isArray(preview.writing_guidance) ? preview.writing_guidance : [];
          const enhancedBy = (data && data.enhanced_by) || 'rules';
          const fallback = !!(data && data.fallback);
          const errorSummary = (data && data.error_summary) || '-';
          const updatedAt = ollamaPreviewUpdatedAt(data);
          let html = '<strong>Ollama 增强预览</strong>';
          html += '<p style="margin:6px 0">状态：'
            + (fallback ? '<span class="error">已回退</span>' : '<span class="success">增强成功</span>')
            + '</p>';
          html += '<table><tr><th>enhanced_by</th><th>fallback</th><th>error_summary</th><th>更新时间</th></tr>'
            + '<tr><td>' + escapeHtmlText(enhancedBy) + '</td>'
            + '<td>' + escapeHtmlText(String(fallback)) + '</td>'
            + '<td>' + escapeHtmlText(errorSummary) + '</td>'
            + '<td>' + escapeHtmlText(updatedAt) + '</td></tr></table>';
          if (data && data.error_summary) {
            html += '<p class="error">' + escapeHtmlText(data.error_summary) + '</p>';
          }
          if (logic.length) {
            html += '<strong>高分逻辑预览</strong><ul>'
              + logic.map((x) => '<li>' + escapeHtmlText(x) + '</li>').join('')
              + '</ul>';
          }
          if (guidance.length) {
            html += '<strong>编制指导预览</strong><ul>'
              + guidance.map((x) => '<li>' + escapeHtmlText(x) + '</li>').join('')
              + '</ul>';
          }
          if (!logic.length && !guidance.length) {
            html += '<p style="color:#64748b">暂无可展示的预览内容。</p>';
          }
          html += '<p style="font-size:12px;color:#64748b;margin-top:8px">仅预览，不写入正式学习进化结果；不影响评分，不进入核心评分主链。</p>';
          return html;
        }
        function materialTypeDisplayName(materialType) {
          const t = String(materialType || '').trim();
          if (t === 'tender_qa') return '招标文件和答疑';
          if (t === 'boq') return '清单';
          if (t === 'drawing') return '图纸';
          if (t === 'site_photo') return '现场照片';
          return t || '项目资料';
        }
        function renderMaterialDepthReportPanel(payload) {
          const el = document.getElementById('materialDepthReportResult');
          if (!el) return;
          const data = (payload && typeof payload === 'object') ? payload : {};
          const byType = Array.isArray(data.by_type) ? data.by_type : [];
          const quality = (data.quality_summary && typeof data.quality_summary === 'object')
            ? data.quality_summary
            : {};
          const capabilities = (data.capabilities && typeof data.capabilities === 'object')
            ? data.capabilities
            : {};
          const recs = Array.isArray(data.recommendations) ? data.recommendations : [];
          const ready = !!data.ready_to_score;
          let html = '<strong>资料深读体检（评分前）</strong>';
          html += '<p style="margin:6px 0">评分就绪：'
            + (ready ? '<span class="success">是</span>' : '<span class="error">否</span>')
            + '；解析分块 ' + escapeHtmlText(quality.total_parsed_chunks || 0)
            + '；数字约束项 ' + escapeHtmlText(quality.total_numeric_terms || 0)
            + '；解析失败率 ' + escapeHtmlText(((Number(quality.parse_fail_ratio || 0) * 100).toFixed(1)) + '%')
            + '</p>';
          html += '<p style="margin:4px 0 8px 0;font-size:12px;color:#334155">解析能力：OCR '
            + (capabilities.ocr_available ? '<span class="success">可用</span>' : '<span class="error">不可用</span>')
            + '（现场照片 ' + escapeHtmlText(capabilities.site_photo_file_count || 0) + '）'
            + '；DWG转换器 '
            + (capabilities.dwg_converter_available ? '<span class="success">可用</span>' : '<span class="error">不可用</span>')
            + '（DWG ' + escapeHtmlText(capabilities.dwg_file_count || 0) + '）'
            + '</p>';
          html += '<table><tr><th>资料类型</th><th>文件数</th><th>成功/失败</th><th>字数</th><th>分块</th><th>数字约束</th><th>目标(字数/分块/数字)</th></tr>';
          html += byType.length
            ? byType.map((row) => {
              const targets = (row && typeof row.targets === 'object') ? row.targets : {};
              return '<tr>'
                + '<td>' + escapeHtmlText((row && row.material_type_label) || (row && row.material_type) || '-') + '</td>'
                + '<td>' + escapeHtmlText((row && row.files) || 0) + '</td>'
                + '<td>' + escapeHtmlText((row && row.parsed_ok_files) || 0) + '/' + escapeHtmlText((row && row.parsed_failed_files) || 0) + '</td>'
                + '<td>' + escapeHtmlText((row && row.parsed_chars) || 0) + '</td>'
                + '<td>' + escapeHtmlText((row && row.parsed_chunks) || 0) + '</td>'
                + '<td>' + escapeHtmlText((row && row.numeric_terms) || 0) + '</td>'
                + '<td>' + escapeHtmlText((targets.min_chars || 0) + '/' + (targets.min_chunks || 0) + '/' + (targets.min_numeric_terms || 0)) + '</td>'
                + '</tr>';
            }).join('')
            : '<tr><td colspan="7">暂无资料体检数据</td></tr>';
          html += '</table>';
          if (recs.length) {
            html += '<ul style="margin:6px 0 0 18px;color:#92400e">'
              + recs.slice(0, 10).map((x) => '<li>' + escapeHtmlText(x) + '</li>').join('')
              + '</ul>';
          }
          el.style.display = 'block';
          el.innerHTML = html;
        }
        function renderMaterialKnowledgeProfilePanel(payload) {
          const el = document.getElementById('materialKnowledgeProfileResult');
          if (!el) return;
          const data = (payload && typeof payload === 'object') ? payload : {};
          const summary = (data.summary && typeof data.summary === 'object') ? data.summary : {};
          const byDimension = Array.isArray(data.by_dimension) ? data.by_dimension : [];
          const byType = Array.isArray(data.by_type) ? data.by_type : [];
          const recs = Array.isArray(data.recommendations) ? data.recommendations : [];
          const topDims = byDimension
            .slice()
            .sort((a, b) => Number((b && b.coverage_score) || 0) - Number((a && a.coverage_score) || 0))
            .slice(0, 8);
          let html = '<strong>资料知识画像（按维度覆盖）</strong>';
          html += '<p style="margin:6px 0">维度覆盖率 '
            + escapeHtmlText(((Number(summary.dimension_coverage_rate || 0) * 100).toFixed(1)) + '%')
            + '；低覆盖维度 ' + escapeHtmlText(summary.low_coverage_dimensions || 0)
            + '；解析字数 ' + escapeHtmlText(summary.total_parsed_chars || 0)
            + '</p>';
          html += '<table><tr><th>维度</th><th>关键词命中</th><th>来源类型数</th><th>覆盖评分</th><th>等级</th></tr>';
          html += topDims.length
            ? topDims.map((row) => '<tr>'
              + '<td>' + escapeHtmlText((row && row.dimension_id) || '-') + ' ' + escapeHtmlText((row && row.dimension_name) || '') + '</td>'
              + '<td>' + escapeHtmlText((row && row.keyword_hits) || 0) + '</td>'
              + '<td>' + escapeHtmlText(((row && row.source_types) || []).length) + '</td>'
              + '<td>' + escapeHtmlText((row && row.coverage_score) || 0) + '</td>'
              + '<td>' + escapeHtmlText((row && row.coverage_level) || '-') + '</td>'
              + '</tr>').join('')
            : '<tr><td colspan="5">暂无维度覆盖数据</td></tr>';
          html += '</table>';
          html += '<table style="margin-top:8px"><tr><th>资料类型</th><th>文件数</th><th>字数</th><th>分块</th><th>词项数</th></tr>';
          html += byType.length
            ? byType.slice(0, 6).map((row) => '<tr>'
              + '<td>' + escapeHtmlText((row && row.material_type_label) || (row && row.material_type) || '-') + '</td>'
              + '<td>' + escapeHtmlText((row && row.files) || 0) + '</td>'
              + '<td>' + escapeHtmlText((row && row.parsed_chars) || 0) + '</td>'
              + '<td>' + escapeHtmlText((row && row.parsed_chunks) || 0) + '</td>'
              + '<td>' + escapeHtmlText((row && row.unique_terms) || 0) + '</td>'
              + '</tr>').join('')
            : '<tr><td colspan="5">暂无资料类型画像</td></tr>';
          html += '</table>';
          if (recs.length) {
            html += '<ul style="margin:6px 0 0 18px;color:#92400e">'
              + recs.slice(0, 8).map((x) => '<li>' + escapeHtmlText(x) + '</li>').join('')
              + '</ul>';
          }
          el.style.display = 'block';
          el.innerHTML = html;
        }
        function renderEvolutionHealthPanel(payload) {
          const el = document.getElementById('evolutionHealthResult');
          if (!el) return;
          const data = (payload && typeof payload === 'object') ? payload : {};
          const summary = (data.summary && typeof data.summary === 'object') ? data.summary : {};
          const windows = (data.windows && typeof data.windows === 'object') ? data.windows : {};
          const wAll = (windows.all && typeof windows.all === 'object') ? windows.all : {};
          const w30 = (windows.recent_30d && typeof windows.recent_30d === 'object') ? windows.recent_30d : {};
          const w90 = (windows.recent_90d && typeof windows.recent_90d === 'object') ? windows.recent_90d : {};
          const drift = (data.drift && typeof data.drift === 'object') ? data.drift : {};
          const recs = Array.isArray(data.recommendations) ? data.recommendations : [];
          let html = '<strong>进化健康度（误差趋势 / 漂移）</strong>';
          html += '<p style="margin:6px 0">漂移等级：'
            + escapeHtmlText(drift.level || 'insufficient_data')
            + '；近30天 MAE=' + escapeHtmlText((w30.mae != null) ? w30.mae : '-')
            + '；近90天 MAE=' + escapeHtmlText((w90.mae != null) ? w90.mae : '-')
            + '；全量 MAE=' + escapeHtmlText((wAll.mae != null) ? wAll.mae : '-')
            + '</p>';
          html += '<table><tr><th>指标</th><th>值</th></tr>';
          html += '<tr><td>真实评分样本</td><td>' + escapeHtmlText(summary.ground_truth_count || 0) + '</td></tr>';
          html += '<tr><td>已匹配预测</td><td>' + escapeHtmlText(summary.matched_prediction_count || 0) + '</td></tr>';
          html += '<tr><td>未匹配样本</td><td>' + escapeHtmlText(summary.unmatched_ground_truth_count || 0) + '</td></tr>';
          html += '<tr><td>当前权重来源</td><td>' + escapeHtmlText(summary.current_weights_source || '-') + '</td></tr>';
          html += '<tr><td>进化权重状态</td><td>' + (summary.has_evolved_multipliers ? '<span class="success">已生效</span>' : '<span class="error">未生效</span>') + '</td></tr>';
          html += '</table>';
          if (recs.length) {
            html += '<ul style="margin:6px 0 0 18px;color:#92400e">'
              + recs.slice(0, 8).map((x) => '<li>' + escapeHtmlText(x) + '</li>').join('')
              + '</ul>';
          }
          el.style.display = 'block';
          el.innerHTML = html;
        }
        function applyScoringReadiness(payload) {
          const data = (payload && typeof payload === 'object') ? payload : {};
          const currentProjectId = pid();
          const payloadProjectId = String(data.project_id || currentProjectId || '');
          const ready = !!data.ready;
          const gatePassed = !!data.gate_passed;
          const issues = Array.isArray(data.issues) ? data.issues : [];
          scoringReadinessState = {
            project_id: payloadProjectId,
            ready: ready,
            gate_passed: gatePassed,
            issues: issues,
          };
          const btn = document.getElementById('btnScoreShigong');
          if (!btn) return;
          const sameProject = payloadProjectId && currentProjectId && payloadProjectId === currentProjectId;
          if (!sameProject) {
            btn.disabled = !currentProjectId;
            btn.title = currentProjectId ? '' : '请先选择项目';
            return;
          }
          if (ready) {
            btn.disabled = false;
            btn.title = '';
            return;
          }
          btn.disabled = true;
          const reason = issues.length ? String(issues[0]) : '评分前置条件未满足';
          btn.title = reason;
        }
        function clearScoringReadinessPanel() {
          const el = document.getElementById('scoringReadinessResult');
          if (!el) return;
          el.style.display = 'none';
          el.innerHTML = '';
          applyScoringReadiness({ project_id: pid(), ready: false, gate_passed: false, issues: ['评分前置检查未完成'] });
        }
        function renderScoringReadinessPanel(payload) {
          const el = document.getElementById('scoringReadinessResult');
          if (!el) return;
          const source = (payload && typeof payload === 'object') ? payload : {};
          const issues = Array.isArray(source.issues) ? source.issues : [];
          const warnings = Array.isArray(source.warnings) ? source.warnings : [];
          const materialQuality = (source.material_quality && typeof source.material_quality === 'object')
            ? source.material_quality
            : {};
          const gate = (source.material_gate && typeof source.material_gate === 'object')
            ? source.material_gate
            : {};
          const depthGate = (source.material_depth_gate && typeof source.material_depth_gate === 'object')
            ? source.material_depth_gate
            : {};
          const submissions = (source.submissions && typeof source.submissions === 'object')
            ? source.submissions
            : {};
          const depthChunks = Number(materialQuality.total_parsed_chunks || 0);
          const depthNumeric = Number(materialQuality.total_numeric_terms || 0);
          const depthPassed = depthGate && depthGate.passed !== false;
          const depthEnforced = !!(depthGate && depthGate.enforce);
          const toPct = (v) => {
            const n = Number(v);
            return Number.isFinite(n) ? (n * 100).toFixed(1) + '%' : '-';
          };
          let html = '<strong>评分前置检查</strong>';
          if (source.ready) {
            html += '<p class="success" style="margin:6px 0">已满足评分条件，可点击“评分施组”。</p>';
          } else {
            html += '<p class="error" style="margin:6px 0">暂不满足评分条件，请先补齐以下问题。</p>';
          }
          html += '<table><tr><th>检查项</th><th>状态</th><th>说明</th></tr>';
          html += '<tr><td>资料门禁</td><td>' + (source.gate_passed ? '<span class="success">通过</span>' : '<span class="error">未通过</span>') + '</td><td>必需资料类型、解析字数、失败率</td></tr>';
          html += '<tr><td>施组上传</td><td>' + (Number(submissions.non_empty || 0) > 0 ? '<span class="success">已上传</span>' : '<span class="error">缺失</span>') + '</td><td>已上传 ' + escapeHtmlText(submissions.non_empty || 0) + ' 份施组</td></tr>';
          html += '<tr><td>资料解析质量</td><td>' + escapeHtmlText(materialQuality.parsed_ok_files || 0) + ' / ' + escapeHtmlText(materialQuality.total_files || 0) + '</td><td>失败率 ' + escapeHtmlText(toPct(materialQuality.parse_fail_ratio)) + '</td></tr>';
          html += '<tr><td>资料深读质量' + (depthEnforced ? '（门禁）' : '（预警）') + '</td><td>'
            + (depthPassed ? '<span class="success">达标</span>' : (depthEnforced ? '<span class="error">未达标</span>' : '<span style="color:#9a3412">待增强</span>'))
            + '</td><td>分块 ' + escapeHtmlText(depthChunks) + ' 段 / 数字约束 ' + escapeHtmlText(depthNumeric) + ' 项</td></tr>';
          html += '</table>';
          if (issues.length) {
            html += '<ul style="margin:6px 0 0 18px;color:#9f1239">'
              + issues.map((x) => '<li>' + escapeHtmlText(x) + '</li>').join('')
              + '</ul>';
          }
          if (warnings.length) {
            html += '<ul style="margin:6px 0 0 18px;color:#92400e">'
              + warnings.map((x) => '<li>' + escapeHtmlText(x) + '</li>').join('')
              + '</ul>';
          }
          if (gate && gate.required_types && Array.isArray(gate.required_types)) {
            html += '<p style="margin-top:8px;font-size:12px;color:#334155">必需资料类型：'
              + escapeHtmlText(gate.required_types.map(materialTypeDisplayName).join('、'))
              + '</p>';
          }
          el.style.display = 'block';
          el.innerHTML = html;
          applyScoringReadiness(source);
        }
        async function refreshScoringReadiness(expectedProjectId=null, switchSeq=null) {
          const id = expectedProjectId || pid();
          if (!id) {
            clearScoringReadinessPanel();
            return;
          }
          let res;
          try {
            res = await fetch('/api/v1/projects/' + id + '/scoring_readiness?t=' + Date.now(), { cache: 'no-store' });
          } catch (_) {
            if (isStaleProjectResponse(id, switchSeq)) return;
            renderScoringReadinessPanel({
              project_id: id,
              ready: false,
              gate_passed: false,
              issues: ['评分前置检查失败：无法连接服务'],
              warnings: [],
              material_quality: {},
              material_gate: {},
              submissions: {},
            });
            return;
          }
          if (isStaleProjectResponse(id, switchSeq)) return;
          const data = await res.json().catch(() => ({}));
          if (!res.ok || !data || typeof data !== 'object') {
            renderScoringReadinessPanel({
              project_id: id,
              ready: false,
              gate_passed: false,
              issues: ['评分前置检查失败（HTTP ' + String((res && res.status) || 0) + '）'],
              warnings: [],
              material_quality: {},
              material_gate: {},
              submissions: {},
            });
            return;
          }
          renderScoringReadinessPanel(data);
        }
        function clearMaterialUtilizationPanel() {
          const el = document.getElementById('materialUtilizationResult');
          if (!el) return;
          el.style.display = 'none';
          el.innerHTML = '';
        }
        function renderMaterialUtilizationPanel(payload) {
          const el = document.getElementById('materialUtilizationResult');
          if (!el) return;
          const source = (payload && typeof payload === 'object') ? payload : {};
          const summary = (source.material_utilization && typeof source.material_utilization === 'object')
            ? source.material_utilization
            : (source.by_type ? source : null);
          if (!summary) {
            clearMaterialUtilizationPanel();
            return;
          }
          const retrievalTotal = Number(summary.retrieval_total || 0);
          const retrievalHit = Number(summary.retrieval_hit || 0);
          const consistencyTotal = Number(summary.consistency_total || 0);
          const consistencyHit = Number(summary.consistency_hit || 0);
          const fallbackTotal = Number(summary.fallback_total || 0);
          const fallbackHit = Number(summary.fallback_hit || 0);
          const retrievalFileTotal = Number(summary.retrieval_file_total || 0);
          const retrievalFileHit = Number(summary.retrieval_file_hit || 0);
          const retrievalUnhitFileCount = Number(summary.retrieval_unhit_file_count || 0);
          const retrievalUnhitFilenames = Array.isArray(summary.retrieval_unhit_filenames)
            ? summary.retrieval_unhit_filenames
            : [];
          const retrievalTopK = Number(summary.retrieval_top_k || 0);
          const retrievalPerTypeQuota = Number(summary.retrieval_per_type_quota || 0);
          const retrievalPerFileQuota = Number(summary.retrieval_per_file_quota || 0);
          const retrievalBaseTopK = Number(summary.retrieval_base_top_k || 0);
          const retrievalBasePerTypeQuota = Number(summary.retrieval_base_per_type_quota || 0);
          const retrievalBasePerFileQuota = Number(summary.retrieval_base_per_file_quota || 0);
          const retrievalBudgetReasons = Array.isArray(summary.retrieval_budget_reasons)
            ? summary.retrieval_budget_reasons
            : [];
          const materialTotalSizeMb = Number(summary.material_total_size_mb || 0);
          const materialTypeCount = Number(summary.material_type_count || 0);
          const materialFileCount = Number(summary.material_file_count || 0);
          const selectedViaCounts = (summary.retrieval_selected_via_counts && typeof summary.retrieval_selected_via_counts === 'object')
            ? summary.retrieval_selected_via_counts
            : {};
          const totalViaCounts = (summary.retrieval_total_via_counts && typeof summary.retrieval_total_via_counts === 'object')
            ? summary.retrieval_total_via_counts
            : {};
          const hitViaCounts = (summary.retrieval_hit_via_counts && typeof summary.retrieval_hit_via_counts === 'object')
            ? summary.retrieval_hit_via_counts
            : {};
          const queryTermsCount = Number(summary.query_terms_count || 0);
          const queryNumericTermsCount = Number(summary.query_numeric_terms_count || 0);
          const byType = (summary.by_type && typeof summary.by_type === 'object') ? summary.by_type : {};
          const availableTypes = Array.isArray(summary.available_types) ? summary.available_types : [];
          const uncoveredTypes = Array.isArray(summary.uncovered_types) ? summary.uncovered_types : [];
          const alerts = Array.isArray(source.material_utilization_alerts)
            ? source.material_utilization_alerts
            : (Array.isArray(source.alerts) ? source.alerts : []);
          const gate = (source.material_utilization_gate && typeof source.material_utilization_gate === 'object')
            ? source.material_utilization_gate
            : (source.gate && typeof source.gate === 'object' ? source.gate : null);
          const bySubmission = Array.isArray(source.material_utilization_by_submission)
            ? source.material_utilization_by_submission
            : [];
          const toPct = (v) => {
            const n = Number(v);
            return Number.isFinite(n) ? (n * 100).toFixed(1) + '%' : '-';
          };

          const orderedTypes = [];
          availableTypes.forEach((t) => {
            const key = String(t || '').trim();
            if (key && !orderedTypes.includes(key)) orderedTypes.push(key);
          });
          Object.keys(byType).forEach((t) => {
            const key = String(t || '').trim();
            if (key && !orderedTypes.includes(key)) orderedTypes.push(key);
          });

          let html =
            '<strong>资料利用审计（本次评分）</strong>'
            + (gate && gate.enabled
              ? (
                  (gate.blocked_submissions > 0 || gate.blocked)
                    ? '<p class="error" style="margin:6px 0 8px 0"><strong>门禁状态：阻断</strong>（存在资料利用阈值未达标的施组）</p>'
                    : (gate.warn_submissions > 0 || gate.warned)
                      ? '<p style="margin:6px 0 8px 0;color:#9a3412"><strong>门禁状态：预警</strong>（建议补齐资料关联后重评分）</p>'
                      : '<p class="success" style="margin:6px 0 8px 0"><strong>门禁状态：通过</strong></p>'
                )
              : '')
            + '<table><tr><th>维度</th><th>命中/总数</th><th>命中率</th></tr>'
            + '<tr><td>资料检索锚点</td><td>' + escapeHtmlText(retrievalHit) + ' / ' + escapeHtmlText(retrievalTotal) + '</td><td>' + escapeHtmlText(toPct(summary.retrieval_hit_rate)) + '</td></tr>'
            + '<tr><td>检索文件覆盖</td><td>' + escapeHtmlText(retrievalFileHit) + ' / ' + escapeHtmlText(retrievalFileTotal) + '</td><td>' + escapeHtmlText(toPct(summary.retrieval_file_coverage_rate)) + '</td></tr>'
            + '<tr><td>未命中资料文件</td><td>' + escapeHtmlText(retrievalUnhitFileCount) + ' / ' + escapeHtmlText(retrievalFileTotal) + '</td><td>建议优先补齐这些文件的章节引用</td></tr>'
            + '<tr><td>跨资料一致性</td><td>' + escapeHtmlText(consistencyHit) + ' / ' + escapeHtmlText(consistencyTotal) + '</td><td>' + escapeHtmlText(toPct(summary.consistency_hit_rate)) + '</td></tr>'
            + '<tr><td>关键词兜底</td><td>' + escapeHtmlText(fallbackHit) + ' / ' + escapeHtmlText(fallbackTotal) + '</td><td>' + escapeHtmlText(toPct(summary.fallback_hit_rate)) + '</td></tr>'
            + '<tr><td>查询特征规模</td><td>' + escapeHtmlText(queryTermsCount) + ' + ' + escapeHtmlText(queryNumericTermsCount) + '</td><td>文本词 + 数字约束</td></tr>'
            + '</table>';

          if (gate && gate.enabled) {
            const blockedCount = Number(gate.blocked_submissions || 0);
            const warnCount = Number(gate.warn_submissions || 0);
            const passCount = Number(gate.pass_submissions || 0);
            html += '<p style="margin:8px 0 0 0;font-size:12px;color:#334155">'
              + '门禁统计：通过 ' + escapeHtmlText(passCount)
              + '，预警 ' + escapeHtmlText(warnCount)
              + '，阻断 ' + escapeHtmlText(blockedCount)
              + '</p>';
          }
          if (retrievalTopK > 0 || materialFileCount > 0) {
            html += '<details style="margin-top:8px"><summary>检索预算与资料体量</summary>';
            html += '<table><tr><th>项</th><th>当前</th><th>基线</th></tr>';
            html += '<tr><td>top_k</td><td>' + escapeHtmlText(retrievalTopK || 0) + '</td><td>' + escapeHtmlText(retrievalBaseTopK || 0) + '</td></tr>';
            html += '<tr><td>每类型配额</td><td>' + escapeHtmlText(retrievalPerTypeQuota || 0) + '</td><td>' + escapeHtmlText(retrievalBasePerTypeQuota || 0) + '</td></tr>';
            html += '<tr><td>每文件配额</td><td>' + escapeHtmlText(retrievalPerFileQuota || 0) + '</td><td>' + escapeHtmlText(retrievalBasePerFileQuota || 0) + '</td></tr>';
            html += '<tr><td>资料体量</td><td colspan="2">文件 ' + escapeHtmlText(materialFileCount || 0)
              + '，类型 ' + escapeHtmlText(materialTypeCount || 0)
              + '，总大小约 ' + escapeHtmlText((materialTotalSizeMb || 0).toFixed(2)) + ' MB</td></tr>';
            html += '</table>';
            if (retrievalBudgetReasons.length) {
              html += '<p style="margin:6px 0 0 0;font-size:12px;color:#475569">预算调整：'
                + escapeHtmlText(retrievalBudgetReasons.join('；'))
                + '</p>';
            }
            html += '</details>';
          }
          const viaKeys = Object.keys(selectedViaCounts)
            .concat(Object.keys(totalViaCounts))
            .concat(Object.keys(hitViaCounts))
            .filter((v, idx, arr) => v && arr.indexOf(v) === idx);
          if (viaKeys.length) {
            const viaLabel = (k) => {
              const key = String(k || '').trim();
              if (key === 'type_quota') return '类型配额命中';
              if (key === 'type_backfill') return '类型回填命中';
              if (key === 'global_rank') return '全局排序命中';
              if (key === 'global_backfill') return '全局回填命中';
              if (key === 'fallback_keywords') return '关键词兜底命中';
              return key || 'unknown';
            };
            html += '<details style="margin-top:8px"><summary>检索策略命中分布</summary>';
            html += '<table><tr><th>策略</th><th>选中块</th><th>有效锚点命中</th><th>命中率</th></tr>';
            html += viaKeys.map((k) => {
              const selectedCnt = Number(selectedViaCounts[k] || 0);
              const totalCnt = Number(totalViaCounts[k] || 0);
              const hitCnt = Number(hitViaCounts[k] || 0);
              const rate = totalCnt > 0 ? ((hitCnt / totalCnt) * 100).toFixed(1) + '%' : '-';
              return '<tr>'
                + '<td>' + escapeHtmlText(viaLabel(k)) + '</td>'
                + '<td>' + escapeHtmlText(selectedCnt) + '</td>'
                + '<td>' + escapeHtmlText(hitCnt) + ' / ' + escapeHtmlText(totalCnt) + '</td>'
                + '<td>' + escapeHtmlText(rate) + '</td>'
                + '</tr>';
            }).join('');
            html += '</table></details>';
          }

          if (orderedTypes.length) {
            html += '<details open style="margin-top:8px"><summary>按资料类型查看命中情况</summary>';
            html += '<table><tr><th>资料类型</th><th>检索命中/总数</th><th>一致性命中/总数</th><th>兜底命中/总数</th></tr>';
            html += orderedTypes.map((t) => {
              const row = (byType[t] && typeof byType[t] === 'object') ? byType[t] : {};
              const rt = Number(row.retrieval_total || 0);
              const rh = Number(row.retrieval_hit || 0);
              const ct = Number(row.consistency_total || 0);
              const ch = Number(row.consistency_hit || 0);
              const ft = Number(row.fallback_total || 0);
              const fh = Number(row.fallback_hit || 0);
              return '<tr>'
                + '<td>' + escapeHtmlText(materialTypeDisplayName(t)) + '</td>'
                + '<td>' + escapeHtmlText(rh) + ' / ' + escapeHtmlText(rt) + '</td>'
                + '<td>' + escapeHtmlText(ch) + ' / ' + escapeHtmlText(ct) + '</td>'
                + '<td>' + escapeHtmlText(fh) + ' / ' + escapeHtmlText(ft) + '</td>'
                + '</tr>';
            }).join('');
            html += '</table></details>';
          }

          if (uncoveredTypes.length) {
            html += '<p class="error" style="margin-top:8px">未形成有效评分证据的资料类型：'
              + escapeHtmlText(uncoveredTypes.map(materialTypeDisplayName).join('、'))
              + '</p>';
          }
          if (alerts.length) {
            html += '<ul style="margin:6px 0 0 18px;color:#9f1239">'
              + alerts.map((a) => '<li>' + escapeHtmlText(a) + '</li>').join('')
              + '</ul>';
          }
          if (retrievalUnhitFilenames.length) {
            html += '<details style="margin-top:8px"><summary>未形成命中证据的资料文件（Top 10）</summary>';
            html += '<ul style="margin:6px 0 0 18px;color:#92400e">'
              + retrievalUnhitFilenames.slice(0, 10).map((name) => '<li>' + escapeHtmlText(name) + '</li>').join('')
              + '</ul></details>';
          }
          if (bySubmission.length) {
            html += '<details style="margin-top:8px"><summary>按施组查看门禁结果</summary>';
            html += '<table><tr><th>施组文件</th><th>门禁状态</th><th>关键原因</th></tr>';
            html += bySubmission.slice(0, 20).map((row) => {
              const gateRow = (row && typeof row.gate === 'object') ? row.gate : {};
              const level = String((gateRow && gateRow.level) || 'unknown');
              const reasons = Array.isArray(gateRow.reasons)
                ? gateRow.reasons
                : (Array.isArray(row.alerts) ? row.alerts : []);
              const levelLabel = level === 'blocked'
                ? '<span class="error">阻断</span>'
                : (level === 'warn' ? '<span style="color:#9a3412">预警</span>' : '<span class="success">通过</span>');
              return '<tr>'
                + '<td>' + escapeHtmlText(String((row && row.filename) || '-')) + '</td>'
                + '<td>' + levelLabel + '</td>'
                + '<td>' + escapeHtmlText((reasons || []).slice(0, 2).join('；') || '-') + '</td>'
                + '</tr>';
            }).join('');
            html += '</table></details>';
          }
          el.style.display = 'block';
          el.innerHTML = html;
        }
        window.renderMaterialUtilizationPanel = renderMaterialUtilizationPanel;
        function setResultLoading(resultId, label) {
          const el = document.getElementById(resultId);
          if (!el) return;
          el.style.display = 'block';
          el.innerHTML = '<span style="color:#334155">' + escapeHtmlText(label || '处理中，请稍候...') + '</span>';
        }
        function setResultError(resultId, msg) {
          const text = msg || '请求失败';
          const el = document.getElementById(resultId);
          if (el) {
            el.style.display = 'block';
            el.innerHTML = '<span class="error">' + escapeHtmlText(text) + '</span>';
          }
          const out = document.getElementById('output');
          if (out) out.textContent = text;
        }
        function ensureProjectForAction(resultId) {
          if (pid()) return true;
          setResultError(resultId, '请先在「2) 选择项目」中选择项目');
          return false;
        }
        function actionProjectId() {
          return pid() || '__NO_PROJECT__';
        }
        function setResultSuccess(resultId, msg) {
          const text = msg || '操作成功';
          const el = document.getElementById(resultId);
          if (el) {
            el.style.display = 'block';
            el.innerHTML = '<span class="success">' + escapeHtmlText(text) + '</span>';
          }
        }
        function formatApiOutput(res, data, fallback='请求失败') {
          if (res && res.ok) return data;
          if (data && typeof data === 'object') {
            if (Object.prototype.hasOwnProperty.call(data, 'detail')) return data;
            if (Object.keys(data).length > 0) return data;
          } else if (typeof data === 'string' && data.trim()) {
            return data;
          }
          const status = res && typeof res.status === 'number' ? res.status : 0;
          return { detail: (status ? ('HTTP ' + status + ' ') : '') + fallback };
        }

        safeClick('btnCompare', async () => {
          if (!ensureProjectForAction('compareResult')) return;
          setResultLoading('compareResult', '对比排名加载中...');
          const projectId = actionProjectId();
          const res = await fetch('/api/v1/projects/' + projectId + '/compare');
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          const el = document.getElementById('compareResult');
          el.style.display = 'block';
          if (res.ok && data.rankings) {
            const scoreSource = (s) => (s === 'pred' ? '预测' : '规则');
            el.innerHTML = '<strong>排名</strong><table><tr><th>文件名</th><th>总分(优先预测)</th><th>规则分(追溯)</th><th>来源</th><th>时间</th></tr>' +
              data.rankings.map(r => '<tr><td>' + r.filename + '</td><td>' + r.total_score + '</td><td>' + (r.rule_total_score ?? '-') + '</td><td>' + scoreSource(r.score_source) + '</td><td>' + r.created_at + '</td></tr>').join('') + '</table>';
          } else {
            el.innerHTML = '<span class="error">' + (data.detail || '请求失败') + '</span>';
          }
        });

        safeClick('btnCompareReport', async () => {
          if (!ensureProjectForAction('compareReportResult')) return;
          setResultLoading('compareReportResult', '对比报告生成中...');
          const projectId = actionProjectId();
          const res = await fetch('/api/v1/projects/' + projectId + '/compare_report');
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          const el = document.getElementById('compareReportResult');
          el.style.display = 'block';
          document.getElementById('insightsResult').style.display = 'none';
          document.getElementById('learningResult').style.display = 'none';
          if (res.ok) {
            const esc = (v) => String(v == null ? '' : v)
              .replace(/&/g, '&amp;')
              .replace(/</g, '&lt;')
              .replace(/>/g, '&gt;')
              .replace(/\"/g, '&quot;')
              .replace(/'/g, '&#39;');
            const escMultiline = (v) => esc(v).replace(/\\n/g, '<br/>');
            const scoreText = (row) => {
              if (!row) return '-';
              const source = row.score_source === 'pred' ? '预测' : '规则';
              const rule = row.rule_total_score == null ? '-' : row.rule_total_score;
              return esc(row.total_score) + ' 分（规则 ' + esc(rule) + '，来源 ' + source + '）';
            };
            let html = '<p><strong>摘要</strong>: ' + (data.summary || '') + '</p>';
            if (data.top_submission && data.top_submission.filename)
              html += '<p>最高: ' + data.top_submission.filename + ' — ' + scoreText(data.top_submission) + '</p>';
            if (data.bottom_submission && data.bottom_submission.filename)
              html += '<p>最低: ' + data.bottom_submission.filename + ' — ' + scoreText(data.bottom_submission) + '</p>';
            if (data.submission_scorecards && data.submission_scorecards.length) {
              html += '<strong>逐份施组得分项/失分项（按文件）</strong>' +
                data.submission_scorecards.map((card, idx) => {
                  const dimRows = Array.isArray(card.dimension_score_items) ? card.dimension_score_items : [];
                  const lossRows = Array.isArray(card.loss_items) ? card.loss_items : [];
                  const gainRows = Array.isArray(card.gain_items) ? card.gain_items : [];
                  const dedRows = Array.isArray(card.deduction_items) ? card.deduction_items : [];
                  const title = esc(card.filename || '') +
                    '（排名 ' + esc(card.rank_desc || '-') +
                    '，总分 ' + esc(card.total_score) +
                    '，距满分 ' + esc(card.gap_to_full_total) +
                    '，累计扣分 ' + esc(card.total_deduction_points) + '）';

                  const lossTable = '<strong>主要失分项（Top5）</strong><table><tr><th>维度</th><th>得分/满分</th><th>失分</th><th>定位页码</th><th>证据片段</th></tr>' +
                    (lossRows.length ? lossRows.map(r => '<tr>' +
                      '<td>' + esc((r.dimension || '') + ' ' + (r.dimension_name || '')) + '</td>' +
                      '<td>' + esc(r.score) + '/' + esc(r.max_score) + '</td>' +
                      '<td>' + esc(r.gap_to_full) + '</td>' +
                      '<td>' + esc(r.page_hint || '页码未知') + '</td>' +
                      '<td>' + esc(r.evidence || '') + '<br/><span style="color:#64748b;font-size:12px">' + escMultiline(r.evidence_context || '') + '</span></td>' +
                    '</tr>').join('') : '<tr><td colspan="5">暂无失分项。</td></tr>') +
                    '</table>';

                  const gainTable = '<strong>主要得分项（Top5）</strong><table><tr><th>维度</th><th>得分/满分</th><th>定位页码</th><th>证据片段</th></tr>' +
                    (gainRows.length ? gainRows.map(r => '<tr>' +
                      '<td>' + esc((r.dimension || '') + ' ' + (r.dimension_name || '')) + '</td>' +
                      '<td>' + esc(r.score) + '/' + esc(r.max_score) + '</td>' +
                      '<td>' + esc(r.page_hint || '页码未知') + '</td>' +
                      '<td>' + esc(r.evidence || '') + '<br/><span style="color:#64748b;font-size:12px">' + escMultiline(r.evidence_context || '') + '</span></td>' +
                    '</tr>').join('') : '<tr><td colspan="4">暂无得分项。</td></tr>') +
                    '</table>';

                  const deductionTable = '<strong>扣分项（按该施组）</strong><table><tr><th>扣分码</th><th>扣分</th><th>定位页码</th><th>原因</th><th>证据片段</th></tr>' +
                    (dedRows.length ? dedRows.map(d => '<tr>' +
                      '<td>' + esc(d.code || '') + '</td>' +
                      '<td>' + esc(d.points) + '</td>' +
                      '<td>' + esc(d.page_hint || '页码未知') + '</td>' +
                      '<td>' + esc(d.reason || '') + '</td>' +
                      '<td>' + esc(d.evidence || '') + '<br/><span style="color:#64748b;font-size:12px">' + escMultiline(d.evidence_context || '') + '</span></td>' +
                    '</tr>').join('') : '<tr><td colspan="5">该施组暂无扣分项。</td></tr>') +
                    '</table>';

                  const fullDimTable = '<details style="margin-top:6px"><summary>查看该施组16维完整得分表</summary><table><tr><th>维度</th><th>名称</th><th>模块</th><th>得分</th><th>满分</th><th>失分</th><th>定位页码</th></tr>' +
                    (dimRows.length ? dimRows.map(r => '<tr>' +
                      '<td>' + esc(r.dimension || '') + '</td>' +
                      '<td>' + esc(r.dimension_name || '') + '</td>' +
                      '<td>' + esc(r.module || '') + '</td>' +
                      '<td>' + esc(r.score) + '</td>' +
                      '<td>' + esc(r.max_score) + '</td>' +
                      '<td>' + esc(r.gap_to_full) + '</td>' +
                      '<td>' + esc(r.page_hint || '页码未知') + '</td>' +
                    '</tr>').join('') : '<tr><td colspan="7">暂无维度数据。</td></tr>') +
                    '</table></details>';

                  return '<details style="margin-top:8px"' + (idx === 0 ? ' open' : '') + '><summary>' + title + '</summary><div style="margin-top:8px">' + lossTable + gainTable + deductionTable + fullDimTable + '</div></details>';
                }).join('');
            }
            if (data.submission_optimization_cards && data.submission_optimization_cards.length) {
              html += '<strong>逐文件优化清单（你要的直接执行版）</strong>' +
                data.submission_optimization_cards.map(card => {
                  const rows = Array.isArray(card.recommendations) ? card.recommendations : [];
                  const table = '<table><tr><th>优先级</th><th>类别</th><th>建议章节</th><th>定位页码</th><th>预计提分</th><th>优先理由</th><th>问题</th><th>原文内容</th><th>直接替换文本 / 原位补充内容</th><th>证据片段</th><th>证据窗口（前后文）</th><th>改写前后示例</th><th>改写指导</th><th>验收标准</th><th>执行检查表</th></tr>' +
                    rows.map(r => '<tr>' +
                      '<td>' + esc(r.priority || '') + '</td>' +
                      '<td>' + esc(r.category || '') + '</td>' +
                      '<td>' + esc(r.chapter_hint || '') + '</td>' +
                      '<td>' + esc(r.page_hint || '页码未知') + '</td>' +
                      '<td>' + esc(r.target_delta_reduction == null ? '' : r.target_delta_reduction) + '</td>' +
                      '<td>' + escMultiline(r.priority_reason || '') + '</td>' +
                      '<td>' + escMultiline(r.issue || '') + '</td>' +
                      '<td>' + escMultiline(r.original_text || r.evidence || '') + '</td>' +
                      '<td>' + escMultiline(r.direct_apply_text || r.replacement_text || r.insertion_content || '') + '</td>' +
                      '<td>' + escMultiline(r.evidence || '') + '</td>' +
                      '<td><span style="font-size:12px;color:#334155">' + escMultiline(r.evidence_context || '') + '</span></td>' +
                      '<td><details><summary>展开</summary><span style="font-size:12px;color:#0f172a">' + escMultiline(r.before_after_example || '') + '</span></details></td>' +
                      '<td><details open><summary>怎么改</summary>' + escMultiline(r.insertion_guidance || r.rewrite_instruction || '') + '</details></td>' +
                      '<td><details><summary>验收标准</summary>' + escMultiline(r.acceptance_check || '') + '</details></td>' +
                      '<td><details><summary>检查表</summary>' + escMultiline(r.execution_checklist || '') + '</details></td>' +
                    '</tr>').join('') + '</table>';
                  const refTop = card.reference_top_score == null ? '' : ('，项目最高 ' + esc(card.reference_top_score) + ' 分');
                  const title = esc(card.filename || '') + '（当前 ' + esc(card.total_score) + ' 分，目标 ' + esc(card.target_score) + ' 分，差距 ' + esc(card.target_gap) + refTop + '）';
                  return '<details style="margin-top:8px" open><summary>' + title + '</summary><div style="margin-top:8px">' + table + '</div></details>';
                }).join('');
            }
            if (data.score_overview && Object.keys(data.score_overview).length) {
              html += '<strong>总体分布</strong><table><tr><th>施组数</th><th>最高分</th><th>最低分</th><th>分差</th><th>项目均分</th><th>波动(标准差)</th></tr><tr>' +
                '<td>' + esc(data.score_overview.submission_count) + '</td>' +
                '<td>' + esc(data.score_overview.top_score) + '</td>' +
                '<td>' + esc(data.score_overview.bottom_score) + '</td>' +
                '<td>' + esc(data.score_overview.score_gap) + '</td>' +
                '<td>' + esc(data.score_overview.project_avg_score) + '</td>' +
                '<td>' + esc(data.score_overview.project_std_score) + '</td>' +
                '</tr></table>';
            }
            if (data.key_diffs && data.key_diffs.length) {
              html += '<strong>主要差距维度</strong><table><tr><th>维度</th><th>名称</th><th>模块</th><th>最高分施组</th><th>最低分施组</th><th>分差</th></tr>' +
                data.key_diffs.map(d =>
                  '<tr><td>' + esc(d.dimension || d.dim_id) + '</td><td>' + esc(d.dimension_name || '') + '</td><td>' + esc(d.module || '') + '</td><td>' + esc((d.top_filename || '-') + '（' + (d.top_dimension_score == null ? '-' : d.top_dimension_score) + '）') + '</td><td>' + esc((d.bottom_filename || '-') + '（' + (d.bottom_dimension_score == null ? '-' : d.bottom_dimension_score) + '）') + '</td><td>' + esc(d.delta) + '</td></tr>'
                ).join('') + '</table>';
            }
            if (data.dimension_diagnostics && data.dimension_diagnostics.length) {
              html += '<strong>维度诊断（编制依据）</strong><table><tr><th>维度</th><th>均分</th><th>最高/最低</th><th>分差</th><th>定位页码</th><th>弱势文件（含该维得分）</th><th>建议动作</th></tr>' +
                data.dimension_diagnostics.map(d => {
                  const actions = Array.isArray(d.actions) ? d.actions.slice(0, 2).map(x => esc(x)).join('<br/>') : '';
                  const weakFiles = Array.isArray(d.weak_files_with_scores) && d.weak_files_with_scores.length
                    ? d.weak_files_with_scores.slice(0, 4).map(x => esc(x)).join('、')
                    : (Array.isArray(d.weak_filenames) && d.weak_filenames.length
                      ? d.weak_filenames.slice(0, 4).map(x => esc(x)).join('、')
                      : ('共 ' + esc(d.weak_file_count || 0) + ' 份'));
                  return '<tr><td>' + esc((d.dimension || '') + ' ' + (d.dimension_name || '')) + '</td><td>' + esc(d.project_avg) + '</td><td>' + esc(d.top_score) + ' / ' + esc(d.bottom_score) + '</td><td>' + esc(d.delta) + '</td><td>' + esc(d.bottom_page_hint || d.top_page_hint || '页码未知') + '</td><td>' + weakFiles + '</td><td>' + actions + '</td></tr>';
                }).join('') + '</table>';
              html += '<details style="margin-top:6px"><summary>查看维度证据与改写模板</summary>' +
                data.dimension_diagnostics.map(d => {
                  const topRow = Array.isArray(d.top_evidence_rows) && d.top_evidence_rows.length ? d.top_evidence_rows[0] : null;
                  const botRow = Array.isArray(d.bottom_evidence_rows) && d.bottom_evidence_rows.length ? d.bottom_evidence_rows[0] : null;
                  const te = topRow ? esc(topRow.snippet || '') : (Array.isArray(d.top_evidence) && d.top_evidence.length ? esc(d.top_evidence[0]) : '无');
                  const be = botRow ? esc(botRow.snippet || '') : (Array.isArray(d.bottom_evidence) && d.bottom_evidence.length ? esc(d.bottom_evidence[0]) : '无');
                  const tp = topRow ? esc(topRow.page_hint || '页码未知') : esc(d.top_page_hint || '页码未知');
                  const bp = botRow ? esc(botRow.page_hint || '页码未知') : esc(d.bottom_page_hint || '页码未知');
                  const tc = topRow ? escMultiline(topRow.context_window || '') : '无';
                  const bc = botRow ? escMultiline(botRow.context_window || '') : '无';
                  const topFile = esc(d.top_filename || '高分施组');
                  const bottomFile = esc(d.bottom_filename || '低分施组');
                  return '<div style="margin-top:8px;padding:8px;border:1px solid #e2e8f0;border-radius:6px"><strong>' + esc((d.dimension || '') + ' ' + (d.dimension_name || '')) + '</strong>' +
                    '<p style="margin:4px 0"><strong>' + topFile + ' 证据片段（' + tp + '）：</strong>' + te + '</p>' +
                    '<p style="margin:4px 0;color:#64748b;font-size:12px"><strong>高分证据窗口：</strong><br/>' + tc + '</p>' +
                    '<p style="margin:4px 0"><strong>' + bottomFile + ' 证据片段（' + bp + '）：</strong>' + be + '</p>' +
                    '<p style="margin:4px 0;color:#64748b;font-size:12px"><strong>低分证据窗口：</strong><br/>' + bc + '</p>' +
                    '<p style="margin:4px 0"><strong>建议改写模板：</strong>' + esc(d.rewrite_template || '') + '</p></div>';
                }).join('') + '</details>';
            }
            if (data.penalty_diagnostics && data.penalty_diagnostics.length) {
              html += '<strong>扣分项诊断（编制风险）</strong><table><tr><th>扣分码</th><th>出现次数</th><th>影响文件数</th><th>累计扣分</th><th>高风险页码</th><th>建议动作</th></tr>' +
                data.penalty_diagnostics.map(p => {
                  const actions = Array.isArray(p.actions) ? p.actions.slice(0, 2).map(x => esc(x)).join('<br/>') : '';
                  const pageHint = Array.isArray(p.page_hints) && p.page_hints.length ? esc(p.page_hints[0]) : '页码未知';
                  return '<tr><td>' + esc(p.code) + '</td><td>' + esc(p.count) + '</td><td>' + esc(p.affected_submission_count) + '</td><td>' + esc(p.total_points) + '</td><td>' + pageHint + '</td><td>' + actions + '</td></tr>';
                }).join('') + '</table>';
              html += '<details style="margin-top:6px"><summary>查看扣分原因样本与证据片段</summary>' +
                data.penalty_diagnostics.map(p => {
                  const reason = Array.isArray(p.reason_samples) && p.reason_samples.length ? esc(p.reason_samples[0]) : '无';
                  const ev = Array.isArray(p.evidence_samples) && p.evidence_samples.length ? esc(p.evidence_samples[0]) : '无';
                  const evCtx = Array.isArray(p.evidence_context_samples) && p.evidence_context_samples.length ? escMultiline(p.evidence_context_samples[0]) : '无';
                  const pageHint = Array.isArray(p.page_hints) && p.page_hints.length ? esc(p.page_hints[0]) : '页码未知';
                  return '<div style="margin-top:8px;padding:8px;border:1px solid #e2e8f0;border-radius:6px"><strong>' + esc(p.code) + '</strong>' +
                    '<p style="margin:4px 0"><strong>高频原因：</strong>' + reason + '</p>' +
                    '<p style="margin:4px 0"><strong>证据片段（' + pageHint + '）：</strong>' + ev + '</p>' +
                    '<p style="margin:4px 0;color:#64748b;font-size:12px"><strong>证据窗口：</strong><br/>' + evCtx + '</p></div>';
                }).join('') + '</details>';
            }
            if (data.priority_actions && data.priority_actions.length) {
              html += '<strong>优先优化动作清单（可直接下发编制）</strong><ol>' +
                data.priority_actions.map(a =>
                  '<li><strong>' + esc(a.priority) + ' - ' + esc(a.theme) + '</strong><br/>' +
                  '依据：' + esc(a.reason || '') + '<br/>' +
                  '证据：' + esc(a.evidence || '') + '（' + esc(a.page_hint || '页码未知') + '）<br/>' +
                  '动作：' + esc(a.action || '') + '<br/>' +
                  '预期效果：' + esc(a.expected_impact || '') + '</li>'
                ).join('') + '</ol>';
            }
            if (data.submission_diagnostics && data.submission_diagnostics.length) {
              html += '<details style="margin-top:6px"><summary>分文件诊断（逐份施组）</summary><table><tr><th>文件</th><th>总分</th><th>弱项维度</th><th>主要扣分</th><th>建议</th></tr>' +
                data.submission_diagnostics.map(r => {
                  const weak = Array.isArray(r.weakest_dimensions) ? r.weakest_dimensions.map(w => esc((w.dimension || '') + (w.dimension_name ? (' ' + w.dimension_name) : ''))).join('、') : '-';
                  const penalties = Array.isArray(r.major_penalties) ? r.major_penalties.map(p => esc(p.code + '×' + p.count)).join('，') : '-';
                  return '<tr><td>' + esc(r.filename) + '</td><td>' + esc(r.total_score) + '</td><td>' + weak + '</td><td>' + penalties + '</td><td>' + esc(r.actionable_summary || '') + '</td></tr>';
                }).join('') + '</table></details>';
            }
            el.innerHTML = html;
          } else {
            el.innerHTML = '<span class="error">' + (data.detail || '请求失败') + '</span>';
          }
        });

        safeClick('btnInsights', async () => {
          if (!ensureProjectForAction('insightsResult')) return;
          setResultLoading('insightsResult', '洞察分析中...');
          const projectId = actionProjectId();
          const res = await fetch('/api/v1/projects/' + projectId + '/insights');
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          const el = document.getElementById('insightsResult');
          el.style.display = 'block';
          document.getElementById('compareReportResult').style.display = 'none';
          document.getElementById('learningResult').style.display = 'none';
          if (res.ok) {
            let html = '';
            if (data.weakest_dims && data.weakest_dims.length)
              html += '<strong>弱项维度</strong><ul>' + data.weakest_dims.map(d => '<li>' + (d.dimension || d.dimension_id) + ': ' + valueOrDefault(d.avg_score, d.avg) + '</li>').join('') + '</ul>';
            if (data.frequent_penalties && data.frequent_penalties.length)
              html += '<strong>常见扣分</strong><ul>' + data.frequent_penalties.map(p => '<li>' + (p.code || '') + ': ' + (p.count || 0) + ' 次</li>').join('') + '</ul>';
            if (data.recommendations && data.recommendations.length)
              html += '<strong>建议</strong><ul>' + data.recommendations.map(r => '<li>' + (r.reason || '') + ' — ' + (r.action || '') + '</li>').join('') + '</ul>';
            el.innerHTML = html || '<pre>' + JSON.stringify(data, null, 2) + '</pre>';
          } else {
            el.innerHTML = '<span class="error">' + (data.detail || '请求失败') + '</span>';
          }
        });

        safeClick('btnLearning', async () => {
          if (!ensureProjectForAction('learningResult')) return;
          setResultLoading('learningResult', '学习画像生成中...');
          const projectId = actionProjectId();
          const res = await fetch('/api/v1/projects/' + projectId + '/learning', { method: 'POST' });
          const text = await res.text();
          document.getElementById('output').textContent = text;
          const el = document.getElementById('learningResult');
          el.style.display = 'block';
          document.getElementById('compareReportResult').style.display = 'none';
          document.getElementById('insightsResult').style.display = 'none';
          try {
            const data = JSON.parse(text);
            el.innerHTML = '<p class="success">学习画像已生成/更新</p><pre>' + JSON.stringify(data.dimension_multipliers || {}, null, 2) + '</pre>';
          } catch (_) {
            el.innerHTML = '<pre>' + text + '</pre>';
          }
        });

        safeClick('btnEvidenceTrace', async () => {
          if (!ensureProjectForAction('evidenceTraceResult')) return;
          setResultLoading('evidenceTraceResult', '证据追溯生成中...');
          const projectId = actionProjectId();
          const res = await fetch('/api/v1/projects/' + projectId + '/evidence_trace/latest');
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          const el = document.getElementById('evidenceTraceResult');
          el.style.display = 'block';
          if (!res.ok) {
            el.innerHTML = '<span class="error">' + (data.detail || '请求失败') + '</span>';
            return;
          }
          const summary = (data.summary && typeof data.summary === 'object') ? data.summary : {};
          const byDim = Array.isArray(data.by_dimension) ? data.by_dimension : [];
          const conflicts = (data.material_conflicts && typeof data.material_conflicts === 'object')
            ? data.material_conflicts
            : {};
          const recommendations = Array.isArray(data.recommendations) ? data.recommendations : [];
          const rows = byDim.slice(0, 12);
          let html = '<strong>证据追溯（最新施组）</strong>';
          html += '<p style="margin:6px 0">文件：' + escapeHtmlText(data.filename || '-') + '</p>';
          html += '<table><tr><th>要求总数</th><th>命中总数</th><th>整体命中率</th><th>强制项命中率</th><th>命中文件数</th></tr>'
            + '<tr><td>' + escapeHtmlText(summary.total_requirements || 0) + '</td><td>' + escapeHtmlText(summary.total_hits || 0) + '</td><td>' + escapeHtmlText(summary.overall_hit_rate ?? '-') + '</td><td>' + escapeHtmlText(summary.mandatory_hit_rate ?? '-') + '</td><td>' + escapeHtmlText(summary.source_files_hit_count || 0) + '</td></tr></table>';
          html += '<p style="margin:6px 0">一致性冲突：' + escapeHtmlText(conflicts.conflict_count || 0)
            + '（高风险 ' + escapeHtmlText(conflicts.high_severity_count || 0) + '）</p>';
          html += '<details style="margin-top:6px"><summary>按维度命中（Top12）</summary><table><tr><th>维度</th><th>total</th><th>hit</th><th>mandatory</th><th>hit_rate</th></tr>'
            + (rows.length
              ? rows.map((r) => '<tr><td>' + escapeHtmlText((r.dimension_id || '') + ' ' + (r.dimension_name || '')) + '</td><td>' + escapeHtmlText(r.total || 0) + '</td><td>' + escapeHtmlText(r.hit || 0) + '</td><td>' + escapeHtmlText((r.mandatory_hit || 0) + '/' + (r.mandatory_total || 0)) + '</td><td>' + escapeHtmlText(r.hit_rate ?? '-') + '</td></tr>').join('')
              : '<tr><td colspan="5">暂无维度证据数据</td></tr>')
            + '</table></details>';
          if (recommendations.length) {
            html += '<strong>建议动作</strong><ul>' + recommendations.slice(0, 8).map((x) => '<li>' + escapeHtmlText(x) + '</li>').join('') + '</ul>';
          }
          html += '<div style="margin-top:8px"><button type="button" class="secondary" id="btnEvidenceTraceDownload">下载 Markdown</button></div>';
          el.innerHTML = html;
          const dlBtn = document.getElementById('btnEvidenceTraceDownload');
          if (dlBtn) {
            dlBtn.onclick = () => {
              const sid = String(data.submission_id || '').trim();
              if (!sid) return;
              const a = document.createElement('a');
              a.href = '/api/v1/projects/' + encodeURIComponent(projectId) + '/submissions/' + encodeURIComponent(sid) + '/evidence_trace.md';
              a.download = 'evidence_trace_' + projectId + '_' + sid + '.md';
              document.body.appendChild(a);
              a.click();
              a.remove();
            };
          }
        });

        safeClick('btnScoringBasis', async () => {
          if (!ensureProjectForAction('scoringBasisResult')) return;
          setResultLoading('scoringBasisResult', '评分依据审计生成中...');
          const projectId = actionProjectId();
          const res = await fetch('/api/v1/projects/' + projectId + '/scoring_basis/latest');
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          const el = document.getElementById('scoringBasisResult');
          el.style.display = 'block';
          if (!res.ok) {
            el.innerHTML = '<span class="error">' + (data.detail || '请求失败') + '</span>';
            return;
          }
          const mece = (data.mece_inputs && typeof data.mece_inputs === 'object') ? data.mece_inputs : {};
          const util = (data.material_utilization && typeof data.material_utilization === 'object') ? data.material_utilization : {};
          const gate = (data.material_utilization_gate && typeof data.material_utilization_gate === 'object') ? data.material_utilization_gate : {};
          const trace = (data.evidence_trace && typeof data.evidence_trace === 'object') ? data.evidence_trace : {};
          const recommendations = Array.isArray(data.recommendations) ? data.recommendations : [];
          let html = '<strong>评分依据审计（最新施组）</strong>';
          html += '<p style="margin:6px 0">文件：' + escapeHtmlText(data.filename || '-') + '；评分状态：' + escapeHtmlText(data.scoring_status || '-') + '</p>';
          html += '<table><tr><th>资料门禁</th><th>资料检索命中率</th><th>文件覆盖率</th><th>强制项命中率</th><th>命中文件数</th></tr>'
            + '<tr><td>' + (mece.materials_quality_gate_passed ? '<span class="success">通过</span>' : '<span class="error">未通过</span>') + '</td><td>' + escapeHtmlText(util.retrieval_hit_rate ?? '-') + '</td><td>' + escapeHtmlText(util.retrieval_file_coverage_rate ?? '-') + '</td><td>' + escapeHtmlText(trace.mandatory_hit_rate ?? '-') + '</td><td>' + escapeHtmlText(trace.source_files_hit_count || 0) + '</td></tr></table>';
          const hitFiles = Array.isArray(trace.source_files_hit) ? trace.source_files_hit : [];
          if (hitFiles.length) {
            html += '<p style="margin:6px 0">命中文件：' + escapeHtmlText(hitFiles.slice(0, 8).join('；')) + (hitFiles.length > 8 ? ' 等' : '') + '</p>';
          } else {
            html += '<p style="margin:6px 0;color:#92400e">尚未命中项目资料文件，请检查资料深读体检与一致性。</p>';
          }
          const gateReasons = Array.isArray(gate.reasons) ? gate.reasons : [];
          if (gateReasons.length) {
            html += '<details style="margin-top:6px"><summary>门禁原因</summary><ul>' + gateReasons.slice(0, 8).map((x) => '<li>' + escapeHtmlText(x) + '</li>').join('') + '</ul></details>';
          }
          if (recommendations.length) {
            html += '<strong>建议动作</strong><ul>' + recommendations.slice(0, 8).map((x) => '<li>' + escapeHtmlText(x) + '</li>').join('') + '</ul>';
          }
          el.innerHTML = html;
        });

        safeClick('btnAdaptive', async () => {
          if (!ensureProjectForAction('adaptiveResult')) return;
          setResultLoading('adaptiveResult', '自适应建议生成中...');
          const projectId = actionProjectId();
          const res = await fetch('/api/v1/projects/' + projectId + '/adaptive');
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          const el = document.getElementById('adaptiveResult');
          el.style.display = 'block';
          document.getElementById('adaptivePatchResult').style.display = 'none';
          document.getElementById('adaptiveValidateResult').style.display = 'none';
          document.getElementById('adaptiveApplyResult').style.display = 'none';
          if (res.ok) {
            let html = '';
            if (data.penalty_stats && Object.keys(data.penalty_stats).length)
              html += '<strong>扣分统计</strong><ul>' + Object.entries(data.penalty_stats).map(([k,v]) => '<li>' + k + ': ' + v + '</li>').join('') + '</ul>';
            if (data.suggestions && data.suggestions.length)
              html += '<strong>建议</strong><ul>' + data.suggestions.map(s => '<li>' + (s.message || JSON.stringify(s)) + '</li>').join('') + '</ul>';
            el.innerHTML = html || '<pre>' + JSON.stringify(data, null, 2) + '</pre>';
          } else {
            el.innerHTML = '<span class="error">' + (data.detail || '请求失败') + '</span>';
          }
        });

        safeClick('btnAdaptivePatch', async () => {
          if (!ensureProjectForAction('adaptivePatchResult')) return;
          setResultLoading('adaptivePatchResult', '补丁生成中...');
          const projectId = actionProjectId();
          const res = await fetch('/api/v1/projects/' + projectId + '/adaptive_patch');
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          const el = document.getElementById('adaptivePatchResult');
          el.style.display = 'block';
          document.getElementById('adaptiveValidateResult').style.display = 'none';
          document.getElementById('adaptiveApplyResult').style.display = 'none';
          if (res.ok) {
            let html = '';
            if (data.lexicon_additions && Object.keys(data.lexicon_additions).length)
              html += '<strong>词库补丁</strong><details><summary>展开</summary><pre>' + JSON.stringify(data.lexicon_additions, null, 2) + '</pre></details>';
            if (data.rubric_adjustments && Object.keys(data.rubric_adjustments).length)
              html += '<strong>规则补丁</strong><details><summary>展开</summary><pre>' + JSON.stringify(data.rubric_adjustments, null, 2) + '</pre></details>';
            el.innerHTML = html || '<pre>' + JSON.stringify(data, null, 2) + '</pre>';
          } else {
            el.innerHTML = '<span class="error">' + (data.detail || '请求失败') + '</span>';
          }
        });

        safeClick('btnAdaptiveValidate', async () => {
          if (!ensureProjectForAction('adaptiveValidateResult')) return;
          setResultLoading('adaptiveValidateResult', '验证效果计算中...');
          const projectId = actionProjectId();
          const res = await fetch('/api/v1/projects/' + projectId + '/adaptive_validate');
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          const el = document.getElementById('adaptiveValidateResult');
          el.style.display = 'block';
          document.getElementById('adaptiveApplyResult').style.display = 'none';
          if (res.ok && data.comparisons) {
            let html = '<p><strong>平均分差（新-旧）</strong>: ' + data.avg_delta + '</p>';
            html += '<table><tr><th>文件名</th><th>旧分</th><th>新分</th><th>变化</th></tr>' +
              data.comparisons.map(c => '<tr><td>' + c.filename + '</td><td>' + c.old_score + '</td><td>' + c.new_score + '</td><td>' + c.delta + '</td></tr>').join('') + '</table>';
            el.innerHTML = html;
          } else {
            el.innerHTML = res.ok ? '<pre>' + JSON.stringify(data, null, 2) + '</pre>' : '<span class="error">' + (data.detail || '请求失败') + '</span>';
          }
        });

        safeClick('btnAdaptiveApply', async () => {
          if (!ensureProjectForAction('adaptiveApplyResult')) return;
          setResultLoading('adaptiveApplyResult', '应用补丁中...');
          const projectId = actionProjectId();
          const storedApiKey = storageGet('api_key');
          let apiKey = storedApiKey || '';
          if (!apiKey) {
            const prompted = prompt('应用补丁将修改 lexicon 配置，需要 API Key。请输入 X-API-Key（无则留空）：');
            apiKey = prompted == null ? '' : prompted;
          }
          const headers = {};
          if (apiKey) headers['X-API-Key'] = apiKey;
          const res = await fetch('/api/v1/projects/' + projectId + '/adaptive_apply', { method: 'POST', headers });
          const text = await res.text();
          document.getElementById('output').textContent = text;
          const el = document.getElementById('adaptiveApplyResult');
          el.style.display = 'block';
          try {
            const data = JSON.parse(text);
            el.innerHTML = data.applied ? '<p class="success">已应用。变更: ' + (data.changes || []).join(', ') + '</p><p>备份: ' + (data.backup_path || '') + '</p>' : '<p class="error">' + (data.detail || text) + '</p>';
          } catch (_) {
            el.innerHTML = '<pre>' + text + '</pre>';
          }
        });

        safeClick('btnUploadFeed', async () => {
          if (!ensureProjectForAction('evolveResult')) return;
          const projectId = actionProjectId();
          const feedInput = document.getElementById('feedFile');
          const files = Array.from((feedInput && feedInput.files) || []);
          const headers = {};
          const apiKey = storageGet('api_key');
          if (apiKey) headers['X-API-Key'] = apiKey;
          const requestCount = files.length > 0 ? files.length : 1;
          setResultLoading('evolveResult', '投喂包上传中（请求 ' + requestCount + ' 次）...');
          document.getElementById('output').textContent = '投喂包上传中（请求 ' + requestCount + ' 次）...';
          setActionStatus('feedActionStatus', '投喂包上传中（请求 ' + requestCount + ' 次）...', false);
          let okCount = 0;
          let failCount = 0;
          const details = [];
          const uploadTargets = files.length > 0 ? files : [null];
          for (const f of uploadTargets) {
            const fd = new FormData();
            if (f) fd.append('file', f);
            const label = f ? f.name : '（空文件请求）';
            try {
              const res = await fetch('/api/v1/projects/' + projectId + '/materials', { method: 'POST', headers, body: fd });
              const text = await res.text();
              if (res.ok) {
                okCount += 1;
                details.push('[成功] ' + label);
              } else {
                failCount += 1;
                let detail = text || '';
                try { const j = JSON.parse(text || '{}'); detail = (j && j.detail) || detail; } catch (_) {}
                details.push('[失败] ' + label + ' -> HTTP ' + res.status + ' ' + String(detail).slice(0, 120));
              }
            } catch (err) {
              failCount += 1;
              details.push('[失败] ' + label + ' -> ' + String((err && err.message) || err || '网络异常'));
            }
          }
          document.getElementById('output').textContent = '投喂包上传完成：成功 ' + okCount + '，失败 ' + failCount + NL + details.join(NL);
          setActionStatus(
            'feedActionStatus',
            '上传完成：成功 ' + okCount + '，失败 ' + failCount + '。',
            failCount > 0
          );
          const evolveEl = document.getElementById('evolveResult');
          if (okCount > 0) {
            evolveEl.innerHTML = '<p class="success">投喂包已保存：成功 ' + okCount + ' 个，失败 ' + failCount + ' 个。</p>';
            evolveEl.style.display = 'block';
            refreshMaterials();
            refreshFeedMaterials();
          } else {
            evolveEl.innerHTML = '<p class="error">投喂包上传失败，请检查文件格式或网络。</p>';
            evolveEl.style.display = 'block';
          }
          if (feedInput && failCount === 0) feedInput.value = '';
        });
        safeClick('btnAddGroundTruth', async () => {
          if (!ensureProjectForAction('evolveResult')) return;
          const projectId = actionProjectId();
          const submissionSelect = document.getElementById('groundTruthSubmissionSelect');
          const submissionId = String((submissionSelect && submissionSelect.value) || '').trim();
          if (!submissionId) {
            setResultError('evolveResult', '请先在“施组文件”下拉框选择步骤4已上传施组。');
            return;
          }
          const judgeScores = collectGroundTruthJudgeScores();
          const finalScore = parseFloat(document.getElementById('gtFinal').value) || 0;
          setResultLoading('evolveResult', '真实评标录入中（基于步骤4已上传施组）...');
          document.getElementById('output').textContent = '真实评标录入中（基于步骤4已上传施组）...';
          const payload = {
            submission_id: submissionId,
            judge_scores: judgeScores,
            final_score: finalScore,
            source: '青天大模型',
          };
          let res, data;
          try {
            res = await fetch('/api/v1/projects/' + projectId + '/ground_truth/from_submission', {
              method: 'POST',
              headers: apiHeaders(true),
              body: JSON.stringify(payload),
            });
            data = await res.json().catch(() => ({}));
          } catch (err) {
            document.getElementById('output').textContent = '真实评标录入失败：' + String((err && err.message) || err || '网络异常');
            return;
          }
          showJson('output', res.ok ? data : (data.detail || data));
          if (!res.ok) {
            const evolveErr = document.getElementById('evolveResult');
            evolveErr.innerHTML = '<p class="error">真实评标录入失败：' + (data.detail || ('HTTP ' + res.status)) + '</p>';
            evolveErr.style.display = 'block';
            return;
          }
          const sourceName = String((submissionSelect && submissionSelect.options && submissionSelect.selectedIndex >= 0 && submissionSelect.options[submissionSelect.selectedIndex] && submissionSelect.options[submissionSelect.selectedIndex].textContent) || submissionId);
          const evolveEl = document.getElementById('evolveResult');
          evolveEl.innerHTML =
            '<p class="success">真实评标录入完成：已记录 1 条。</p>' +
            '<p style="margin:6px 0 0 0"><strong>施组：</strong>' + escapeHtmlText(sourceName) + '</p>' +
            '<p style="margin:4px 0 0 0"><strong>评委人数：</strong>' + escapeHtmlText(String(judgeScores.length)) + ' 位</p>' +
            '<p style="margin:4px 0 0 0"><strong>最终分：</strong>' + escapeHtmlText(String(data.final_score != null ? data.final_score : finalScore)) + '</p>';
          evolveEl.style.display = 'block';
          if (submissionSelect) submissionSelect.value = '';
          refreshGroundTruth();
        });
        safeClick('btnEvolve', async () => {
          if (!ensureProjectForAction('evolveResult')) return;
          const projectId = actionProjectId();
          setResultLoading('evolveResult', '学习进化执行中...');
          const res = await fetch('/api/v1/projects/' + projectId + '/evolve', { method: 'POST', headers: apiHeaders(false) });
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data, '请求失败，若需认证请设置 API Key'));
          const el = document.getElementById('evolveResult');
          el.style.display = 'block';
          if (res.ok) {
            let html = '<p class="success">学习完成（基于 ' + (data.sample_count || 0) + ' 条真实评标）</p>';
            if (data.high_score_logic && data.high_score_logic.length) html += '<strong>高分逻辑</strong><ul>' + data.high_score_logic.map(l => '<li>' + l + '</li>').join('') + '</ul>';
            if (data.writing_guidance && data.writing_guidance.length) html += '<strong>编制指导</strong><ul>' + data.writing_guidance.map(l => '<li>' + l + '</li>').join('') + '</ul>';
            if (data.scoring_evolution && data.scoring_evolution.dimension_multipliers && Object.keys(data.scoring_evolution.dimension_multipliers).length) {
              html += '<strong>评分进化（贴近青天）</strong><p style="font-size:12px;color:#64748b">' + (data.scoring_evolution.goal || '') + '</p>';
              html += '<details><summary>维度权重建议</summary><pre style="margin:4px 0">' + JSON.stringify(data.scoring_evolution.dimension_multipliers, null, 2) + '</pre></details>';
            }
            html += '<p style="font-size:12px;margin-top:8px">点击「编制系统指令」可查看/导出编制约束（必备章节、图表、禁止表述等）。</p>';
            el.innerHTML = html;
          } else { el.innerHTML = '<span class="error">' + (data.detail || '请求失败，若需认证请设置 API Key') + '</span>'; }
        });
        safeClick('btnOllamaPreview', async () => {
          if (!ensureProjectForAction('ollamaPreviewResult')) return;
          const projectId = actionProjectId();
          resetOllamaPreviewActions();
          setResultLoading('ollamaPreviewResult', 'Ollama 增强预览生成中...');
          let res;
          let data = {};
          try {
            res = await fetch('/api/v1/projects/' + encodeURIComponent(projectId) + '/evolve/ollama_preview', {
              method: 'POST',
              headers: apiHeaders(false),
            });
            data = await res.json().catch(() => ({}));
          } catch (err) {
            setResultError('ollamaPreviewResult', 'Ollama 增强预览失败：' + String((err && err.message) || err || '网络异常'));
            return;
          }
          showJson('output', formatApiOutput(res, data, 'Ollama 增强预览失败'));
          const el = document.getElementById('ollamaPreviewResult');
          if (!el) return;
          el.style.display = 'block';
          if (res.ok) {
            storeOllamaPreviewPayload(projectId, data);
            el.innerHTML = renderOllamaPreviewHtml(data);
          } else {
            resetOllamaPreviewActions();
            el.innerHTML = '<span class="error">' + escapeHtmlText(data.detail || 'Ollama 增强预览失败') + '</span>';
          }
        });
        safeClick('btnOllamaPreviewCopy', async () => {
          if (!latestOllamaPreviewPayload) {
            setOllamaPreviewActionStatus('请先生成 Ollama 增强预览。', true);
            return;
          }
          await copyTextToClipboard(formatOllamaPreviewPlainText(latestOllamaPreviewPayload));
          setOllamaPreviewActionStatus('预览结果已复制。', false);
        });
        safeClick('btnOllamaPreviewExport', async () => {
          if (!latestOllamaPreviewPayload) {
            setOllamaPreviewActionStatus('请先生成 Ollama 增强预览。', true);
            return;
          }
          downloadOllamaPreviewJson(latestOllamaPreviewProjectId, latestOllamaPreviewPayload);
          setOllamaPreviewActionStatus('JSON 导出已触发。', false);
        });
        safeClick('btnEvolutionHealth', async () => {
          if (!ensureProjectForAction('evolutionHealthResult')) return;
          const projectId = actionProjectId();
          setResultLoading('evolutionHealthResult', '进化健康度分析中...');
          let res;
          let data = {};
          try {
            res = await fetch('/api/v1/projects/' + encodeURIComponent(projectId) + '/evolution/health', {
              method: 'GET',
              headers: apiHeaders(false),
            });
            data = await res.json().catch(() => ({}));
          } catch (err) {
            setResultError('evolutionHealthResult', '分析失败：' + String((err && err.message) || err || '网络异常'));
            return;
          }
          showJson('output', formatApiOutput(res, data));
          if (!res.ok) {
            const detail = (data && data.detail) ? String(data.detail) : ('HTTP ' + String(res.status || 0));
            setResultError('evolutionHealthResult', '分析失败：' + detail);
            return;
          }
          renderEvolutionHealthPanel(data);
        });
        safeClick('btnWritingGuidance', async () => {
          if (!ensureProjectForAction('guidanceResult')) return;
          const projectId = actionProjectId();
          setResultLoading('guidanceResult', '正在生成编制指导...');
          const res = await fetch('/api/v1/projects/' + projectId + '/writing_guidance');
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          const el = document.getElementById('guidanceResult');
          el.style.display = 'block';
          if (res.ok) {
            let html = '';
            if (data.high_score_logic && data.high_score_logic.length) html += '<strong>高分逻辑</strong><ul>' + data.high_score_logic.map(l => '<li>' + l + '</li>').join('') + '</ul>';
            if (data.guidance && data.guidance.length) html += '<strong>编制指导</strong><ul>' + data.guidance.map(l => '<li>' + l + '</li>').join('') + '</ul>';
            el.innerHTML = html || '<pre>' + JSON.stringify(data, null, 2) + '</pre>';
          } else { el.innerHTML = '<span class="error">' + (data.detail || '') + '</span>'; }
        });
        safeClick('btnCompilationInstructions', async () => {
          if (!ensureProjectForAction('compilationInstructionsResult')) return;
          const projectId = actionProjectId();
          setResultLoading('compilationInstructionsResult', '正在生成编制系统指令...');
          const res = await fetch('/api/v1/projects/' + projectId + '/compilation_instructions');
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          const el = document.getElementById('compilationInstructionsResult');
          el.style.display = 'block';
          if (res.ok) {
            let html = '<strong>编制系统指令</strong>（可导出为编制施组时的系统约束）<br>';
            if ((data.required_sections || []).length) html += '<p><b>必备章节：</b>' + data.required_sections.join('；') + '</p>';
            if ((data.required_charts_images || []).length) html += '<p><b>必备图表/图片：</b>' + data.required_charts_images.join('；') + '</p>';
            if ((data.mandatory_elements || []).length) html += '<p><b>必备要素：</b>' + data.mandatory_elements.join('；') + '</p>';
            if ((data.forbidden_patterns || []).length) html += '<p><b>禁止表述：</b>' + data.forbidden_patterns.join('；') + '</p>';
            if ((data.guidance_items || []).length) html += '<p><b>编制指导：</b><ul>' + data.guidance_items.map(l => '<li>' + l + '</li>').join('') + '</ul></p>';
            html += '<button type="button" class="secondary" id="btnExportInstructions" style="margin-top:8px">导出为文本（复制到剪贴板）</button>';
            el.innerHTML = html;
            (document.getElementById('btnExportInstructions')||{}).onclick = () => {
              const L = (arr, prefix) => (arr || []).map(s => prefix + s).join(String.fromCharCode(10));
              const lines = ['# 施组编制系统指令', '', '## 必备章节', L(data.required_sections || [], '- '), '', '## 必备图表/图片', L(data.required_charts_images || [], '- '), '', '## 必备要素', L(data.mandatory_elements || [], '- '), '', '## 禁止表述', L(data.forbidden_patterns || [], '- '), '', '## 编制指导', L(data.guidance_items || [], '- ')];
              const plain = lines.join(String.fromCharCode(10));
              navigator.clipboard.writeText(plain).then(() => alert('已复制到剪贴板')).catch(() => prompt('请手动复制：', plain));
            };
          } else { el.innerHTML = '<span class="error">' + (data.detail || '') + '</span>'; }
        });
        async function latestPatchId(projectId) {
          const res = await fetch('/api/v1/projects/' + projectId + '/patches');
          const data = await res.json().catch(() => ([]));
          if (!res.ok || !Array.isArray(data) || !data.length) return '';
          return data[0].id || '';
        }
        function showBlock(id, ok, title, payload) {
          const el = document.getElementById(id);
          if (!el) return;
          el.style.display = 'block';
          const klass = ok ? 'success' : 'error';
          el.innerHTML = '<p class="' + klass + '">' + title + '</p><pre style="margin:0">' + JSON.stringify(payload, null, 2) + '</pre>';
        }
        function fmtMetric(v) {
          if (typeof v !== 'number' || Number.isNaN(v)) return '-';
          return Number(v).toFixed(4);
        }
        function showCalibratorSummaryBlock(id, ok, title, payload) {
          const el = document.getElementById(id);
          if (!el) return;
          el.style.display = 'block';
          const klass = ok ? 'success' : 'error';
          let html = '<p class="' + klass + '">' + title + '</p>';
          const data = payload && typeof payload === 'object' ? payload : {};
          if (ok) {
            const summary = (data.calibrator_summary && typeof data.calibrator_summary === 'object') ? data.calibrator_summary : {};
            const metrics = data.metrics || {};
            const cv = summary.cv_metrics || data.calibrator_cv_metrics || {};
            const baseline = summary.baseline_metrics || data.calibrator_baseline_metrics || {};
            const gate = summary.gate || data.calibrator_gate || {};
            const modelType = summary.model_type || data.calibrator_model_type || data.model_type || '-';
            const version = summary.calibrator_version || data.calibrator_version || '-';
            const gatePassed = summary.gate_passed ?? data.calibrator_gate_passed ?? metrics.gate_passed;
            const sampleCount = summary.sample_count;
            const skippedReason = summary.skipped_reason || '';
            const cvMae = cv.mae ?? metrics.cv_mae;
            const cvRmse = cv.rmse ?? metrics.cv_rmse;
            const cvSpearman = cv.spearman ?? metrics.cv_spearman;
            const baselineMae = baseline.mae ?? metrics.baseline_mae;
            const baselineRmse = baseline.rmse ?? metrics.baseline_rmse;
            const baselineSpearman = baseline.spearman ?? metrics.baseline_spearman;
            const improveThreshold = gate.improve_threshold ?? metrics.gate_improve_threshold;
            const spearmanTolerance = gate.spearman_tolerance ?? metrics.gate_spearman_tolerance;
            const cvMode = cv.mode ?? metrics.cv_mode;
            const cvPredCount = cv.pred_count ?? metrics.cv_pred_count;
            const gateText = typeof gatePassed === 'boolean' ? (gatePassed ? '通过' : '未通过') : '-';
            html += '<p style="margin:4px 0"><b>校准摘要</b>：模型=' + modelType + '；版本=' + version + '；闸门=' + gateText + (sampleCount != null ? ('；样本=' + sampleCount) : '') + '</p>';
            if (skippedReason) {
              html += '<p style="margin:4px 0;color:#9a3412"><b>提示</b>：本次未训练校准器（' + skippedReason + '）。</p>';
            }
            if (cvMae !== undefined || baselineMae !== undefined) {
              html += '<p style="margin:4px 0"><b>CV</b> MAE=' + fmtMetric(cvMae) + ' RMSE=' + fmtMetric(cvRmse) + ' Spearman=' + fmtMetric(cvSpearman) + '（' + (cvMode || '-') + ', n=' + (cvPredCount || 0) + '）</p>';
              html += '<p style="margin:4px 0"><b>Baseline</b> MAE=' + fmtMetric(baselineMae) + ' RMSE=' + fmtMetric(baselineRmse) + ' Spearman=' + fmtMetric(baselineSpearman) + '</p>';
              html += '<p style="margin:4px 0"><b>闸门阈值</b> MAE改进≥' + fmtMetric(improveThreshold) + '，Spearman不下降超过' + fmtMetric(spearmanTolerance) + '</p>';
            }
            const candidates = Array.isArray(summary.auto_candidates) ? summary.auto_candidates : (Array.isArray(data.calibrator_auto_candidates) ? data.calibrator_auto_candidates : []);
            if (candidates.length) {
              const rows = candidates.map((c) => {
                const gt = c.gate_passed === true ? '通过' : '未通过';
                return '<tr><td style="padding:2px 8px;border:1px solid #dbe3ef">' + (c.model_type || '-') + '</td><td style="padding:2px 8px;border:1px solid #dbe3ef">' + gt + '</td><td style="padding:2px 8px;border:1px solid #dbe3ef">' + fmtMetric(c.cv_mae) + '</td><td style="padding:2px 8px;border:1px solid #dbe3ef">' + fmtMetric(c.cv_spearman) + '</td></tr>';
              }).join('');
              html += '<details style="margin:6px 0"><summary>候选模型对比（auto）</summary><table style="border-collapse:collapse;font-size:12px;margin-top:4px"><thead><tr><th style="padding:2px 8px;border:1px solid #dbe3ef">模型</th><th style="padding:2px 8px;border:1px solid #dbe3ef">闸门</th><th style="padding:2px 8px;border:1px solid #dbe3ef">CV MAE</th><th style="padding:2px 8px;border:1px solid #dbe3ef">CV Spearman</th></tr></thead><tbody>' + rows + '</tbody></table></details>';
            }
            const govern = (data.patch_auto_govern && typeof data.patch_auto_govern === 'object')
              ? data.patch_auto_govern
              : {};
            if (Object.keys(govern).length) {
              html += '<details style="margin:6px 0"><summary>补丁自动治理</summary>';
              html += '<p style="margin:4px 0"><b>状态</b>：'
                + (govern.checked ? '已检查' : '未检查')
                + '；动作=' + String(govern.action || '-')
                + '；原因=' + String(govern.reason || '-')
                + '；闸门=' + String(govern.gate_passed === true ? '通过' : (govern.gate_passed === false ? '未通过' : '-'))
                + '</p>';
              html += '<p style="margin:4px 0"><b>样本</b>：' + String(govern.sample_count || 0)
                + '；回滚目标=' + String(govern.rollback_to_patch_id || '-') + '</p>';
              html += '</details>';
            }
          }
          html += '<pre style="margin:0">' + JSON.stringify(payload, null, 2) + '</pre>';
          el.innerHTML = html;
        }
        safeClick('btnRebuildDelta', async () => {
          const projectId = actionProjectId();
          setResultLoading('deltaResult', '正在重建 DELTA_CASE...');
          const res = await fetch('/api/v1/projects/' + projectId + '/delta_cases/rebuild', { method: 'POST', headers: apiHeaders(false) });
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          showBlock('deltaResult', res.ok, res.ok ? 'DELTA_CASE 重建完成' : 'DELTA_CASE 重建失败', data);
        });
        safeClick('btnRebuildSamples', async () => {
          const projectId = actionProjectId();
          setResultLoading('sampleResult', '正在重建 FEATURE_ROW...');
          const res = await fetch('/api/v1/projects/' + projectId + '/calibration_samples/rebuild', { method: 'POST', headers: apiHeaders(false) });
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          showBlock('sampleResult', res.ok, res.ok ? 'FEATURE_ROW 重建完成' : 'FEATURE_ROW 重建失败', data);
        });
        safeClick('btnTrainCalibratorV2', async () => {
          const projectId = actionProjectId();
          setResultLoading('calibTrainResult', '正在训练校准器...');
          const body = { project_id: projectId, model_type: 'auto', alpha: 1.0, auto_deploy: true };
          const res = await fetch('/api/v1/calibration/train', { method: 'POST', headers: apiHeaders(true), body: JSON.stringify(body) });
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          showCalibratorSummaryBlock('calibTrainResult', res.ok, res.ok ? '校准器训练完成' : '校准器训练失败', data);
        });
        safeClick('btnApplyCalibPredict', async () => {
          const projectId = actionProjectId();
          setResultLoading('calibTrainResult', '正在回填预测分...');
          const res = await fetch('/api/v1/projects/' + projectId + '/calibration/predict', { method: 'POST', headers: apiHeaders(false) });
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          showBlock('calibTrainResult', res.ok, res.ok ? '预测分回填完成' : '预测分回填失败', data);
        });
        safeClick('btnAutoRunReflection', async () => {
          const projectId = actionProjectId();
          setResultLoading('calibTrainResult', '正在执行闭环流程...');
          const res = await fetch('/api/v1/projects/' + projectId + '/reflection/auto_run', { method: 'POST', headers: apiHeaders(false) });
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          showCalibratorSummaryBlock('calibTrainResult', res.ok, res.ok ? '一键闭环执行完成' : '一键闭环执行失败', data);
        });
        safeClick('btnEvalMetricsV2', async () => {
          const projectId = actionProjectId();
          setResultLoading('evalResult', '正在评估项目指标...');
          const res = await fetch('/api/v1/projects/' + projectId + '/evaluation');
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          if (!res.ok) { showBlock('evalResult', false, '指标评估失败', data); return; }
          const a = data.acceptance || {};
          const summary = {
            sample_count_qt: data.sample_count_qt,
            acceptance: a,
            variants: data.variants || {},
          };
          showBlock('evalResult', true, '项目指标评估完成', summary);
        });
        safeClick('btnEvalSummaryV2', async () => {
          const res = await fetch('/api/v1/evaluation/summary');
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          if (!res.ok) { showBlock('evalResult', false, '跨项目汇总评估失败', data); return; }
          showBlock('evalResult', true, '跨项目汇总评估完成', data);
        });
        safeClick('btnMinePatchV2', async () => {
          const projectId = actionProjectId();
          setResultLoading('patchResult', '正在挖掘补丁...');
          const patchType = (document.getElementById('patchType') || {}).value || 'threshold';
          const body = { patch_type: patchType, top_k: 5 };
          const res = await fetch('/api/v1/projects/' + projectId + '/patches/mine', { method: 'POST', headers: apiHeaders(true), body: JSON.stringify(body) });
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          if (res.ok && data && data.id) { const input = document.getElementById('patchIdInput'); if (input) input.value = data.id; }
          showBlock('patchResult', res.ok, res.ok ? 'PATCH_PACKAGE 挖掘完成' : 'PATCH_PACKAGE 挖掘失败', data);
        });
        safeClick('btnShadowPatchV2', async () => {
          const projectId = actionProjectId();
          setResultLoading('patchShadowResult', '正在执行影子评估...');
          let patchId = ((document.getElementById('patchIdInput') || {}).value || '').trim();
          if (!patchId) patchId = await latestPatchId(projectId);
          if (!patchId) patchId = '__NO_PATCH__';
          const res = await fetch('/api/v1/patches/' + patchId + '/shadow_eval', { method: 'POST', headers: apiHeaders(false) });
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          showBlock('patchShadowResult', res.ok, res.ok ? '补丁影子评估完成' : '补丁影子评估失败', data);
        });
        safeClick('btnDeployPatchV2', async () => {
          const projectId = actionProjectId();
          setResultLoading('patchDeployResult', '正在发布补丁...');
          let patchId = ((document.getElementById('patchIdInput') || {}).value || '').trim();
          if (!patchId) patchId = await latestPatchId(projectId);
          if (!patchId) patchId = '__NO_PATCH__';
          const res = await fetch('/api/v1/patches/' + patchId + '/deploy', { method: 'POST', headers: apiHeaders(true), body: JSON.stringify({ action: 'deploy' }) });
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          showBlock('patchDeployResult', res.ok, res.ok ? '补丁已发布' : '补丁发布失败', data);
        });
        safeClick('btnRollbackPatchV2', async () => {
          const projectId = actionProjectId();
          setResultLoading('patchDeployResult', '正在回滚补丁...');
          let patchId = ((document.getElementById('patchIdInput') || {}).value || '').trim();
          if (!patchId) patchId = await latestPatchId(projectId);
          if (!patchId) patchId = '__NO_PATCH__';
          const res = await fetch('/api/v1/patches/' + patchId + '/deploy', { method: 'POST', headers: apiHeaders(true), body: JSON.stringify({ action: 'rollback' }) });
          const data = await res.json().catch(() => ({}));
          showJson('output', formatApiOutput(res, data));
          showBlock('patchDeployResult', res.ok, res.ok ? '补丁已回滚' : '补丁回滚失败', data);
        });

        // 关闭“硬接管”兜底，避免覆盖 safeClick 的详细渲染结果。
        initWeightsSection();
        syncGroundTruthJudgeInputs();
        updateProjectBoundControlsState();
        refreshProjects();

      </script>
    </body>
    </html>
    """
    html = html.replace("__PROJECT_OPTIONS__", project_options_html)
    html = html.replace("__CREATE_NOTICE_HTML__", create_notice_html)
    html = html.replace("__GLOBAL_NOTICE_HTML__", global_notice_html)
    html = html.replace("__SELECTED_PROJECT_ID__", html_lib.escape(selected_project_id))
    html = html.replace("__EXPERT_PROFILE_STATUS__", initial_profile_status)
    html = html.replace("__EXPERT_WEIGHTS_ROWS__", initial_weights_rows_html)
    html = html.replace("__EXPERT_WEIGHTS_SUMMARY__", html_lib.escape(initial_weights_summary))
    html = html.replace("__MATERIAL_ROWS__", initial_material_rows_html)
    html = html.replace("__MATERIALS_EMPTY_DISPLAY__", initial_materials_empty_display)
    html = html.replace("__SUBMISSION_ROWS__", initial_submission_rows_html)
    html = html.replace("__SUBMISSIONS_EMPTY_DISPLAY__", initial_submissions_empty_display)
    html = html.replace("__PROJECT_SCORE_SCALE_MAX__", str(score_scale_initial))
    return Response(
        content=html.encode("utf-8"),
        media_type="text/html; charset=utf-8",
        headers={
            # Prevent stale cached UI when frontend script is updated.
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


def create_app() -> FastAPI:
    return app


if __name__ == "__main__":
    import os
    import sys
    import threading
    import time
    import webbrowser

    port = int(os.environ.get("PORT", "8000"))

    def _open_browser() -> None:
        time.sleep(2.5)
        try:
            webbrowser.open(f"http://127.0.0.1:{port}/")
        except Exception:
            pass

    if "--no-browser" not in sys.argv:
        threading.Thread(target=_open_browser, daemon=True).start()
    print(f"浏览器将自动打开: http://127.0.0.1:{port}/")
    print("按 Ctrl+C 停止")

    import uvicorn

    uvicorn.run("app.main:app", host="0.0.0.0", port=port, reload=False)
